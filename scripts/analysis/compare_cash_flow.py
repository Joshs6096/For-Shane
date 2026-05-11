#!/usr/bin/env python3
"""
Cash Flow Sheet Comparison: XLSB (team file) vs XLSX (corrected file)
Compares every non-null column for May 8, 2026 and a ±5 business day window.
"""

from pyxlsb import open_workbook
from openpyxl import load_workbook
from datetime import date, timedelta, datetime
import sys

XLSB_PATH = "/Users/josh/Desktop/Daily Cash Fcst - 05.08.26.xlsb"
XLSX_PATH = "/Users/josh/Downloads/Daily Cash Fcst - 5.8.26_CORRECTED.xlsx"

SHEET_NAME_XLSB = "Cash Flow"
SHEET_NAME_XLSX = "Cash Flow"

MAY8_SERIAL = 46150  # Excel serial for May 8, 2026
MAY8_DATE = date(2026, 5, 8)

DIFF_THRESHOLD = 0.0001  # $100 in millions scale

# Business day window: May 1 – May 15, 2026
WINDOW_START = date(2026, 5, 1)
WINDOW_END = date(2026, 5, 15)

# Key columns to compare in window (0-based)
KEY_COLS = {
    7: "Net CF",
    11: "Available Cash",
    15: "Total Cash",
    85: "Inflows",
    123: "Disbursements Total"
}


def serial_to_date(serial):
    """Convert Excel serial date to Python date."""
    return date(1899, 12, 30) + timedelta(days=int(serial))


def normalize_xlsx_date(val):
    """Normalize xlsx date value to Python date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Try string parsing
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                pass
    # Could be a float/int serial
    if isinstance(val, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(val))
    return None


def is_error_value(val):
    """Check if a value is an error (e.g., #N/A, #REF!, etc.)."""
    if isinstance(val, str):
        errors = ["#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!"]
        return any(val.startswith(e) for e in errors)
    # openpyxl represents errors as special types
    return False


# ─────────────────────────────────────────────
# 1. READ XLSB
# ─────────────────────────────────────────────
print("=" * 80)
print("READING XLSB FILE...")
print("=" * 80)

xlsb_rows = []
xlsb_headers = None
xlsb_sheet_found = False

with open_workbook(XLSB_PATH) as wb:
    print(f"XLSB sheets available: {wb.sheets}")
    for sheet_name in wb.sheets:
        if sheet_name.strip() == SHEET_NAME_XLSB:
            xlsb_sheet_found = True
            with wb.get_sheet(sheet_name) as sheet:
                for i, row in enumerate(sheet.rows()):
                    row_vals = [cell.v for cell in row]
                    if i == 0:
                        xlsb_headers = row_vals
                    xlsb_rows.append(row_vals)
            break

if not xlsb_sheet_found:
    print(f"ERROR: Sheet '{SHEET_NAME_XLSB}' not found in XLSB")
    print("Available sheets:", wb.sheets)
    sys.exit(1)

print(f"XLSB rows read: {len(xlsb_rows)}")
print(f"XLSB col count: {len(xlsb_headers) if xlsb_headers else 'N/A'}")


# ─────────────────────────────────────────────
# 2. READ XLSX
# ─────────────────────────────────────────────
print("\n" + "=" * 80)
print("READING XLSX FILE...")
print("=" * 80)

xlsx_rows = []
xlsx_headers = None
xlsx_sheet_found = False
xlsx_error_cells = []

wb_xlsx = load_workbook(XLSX_PATH, data_only=True, read_only=True)
print(f"XLSX sheets available: {wb_xlsx.sheetnames}")

for sname in wb_xlsx.sheetnames:
    if sname.strip() == SHEET_NAME_XLSX:
        xlsx_sheet_found = True
        ws = wb_xlsx[sname]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_list = list(row)
            if i == 0:
                xlsx_headers = row_list
            # Check for error values
            for j, val in enumerate(row_list):
                if is_error_value(val):
                    xlsx_error_cells.append((i + 1, j, val))
            xlsx_rows.append(row_list)
        break

wb_xlsx.close()

if not xlsx_sheet_found:
    print(f"ERROR: Sheet '{SHEET_NAME_XLSX}' not found in XLSX")
    sys.exit(1)

print(f"XLSX rows read: {len(xlsx_rows)}")
print(f"XLSX col count: {len(xlsx_headers) if xlsx_headers else 'N/A'}")


# ─────────────────────────────────────────────
# 3. STRUCTURAL COMPARISON
# ─────────────────────────────────────────────
print("\n" + "=" * 80)
print("STRUCTURAL COMPARISON")
print("=" * 80)
print(f"{'Metric':<35} {'XLSB':>15} {'XLSX':>15}")
print("-" * 65)
print(f"{'Row count':<35} {len(xlsb_rows):>15} {len(xlsx_rows):>15}")
print(f"{'Col count':<35} {len(xlsb_headers) if xlsb_headers else 0:>15} {len(xlsx_headers) if xlsx_headers else 0:>15}")


# Find date rows in xlsb
xlsb_date_rows = []
for i, row in enumerate(xlsb_rows):
    if row and row[0] is not None:
        val = row[0]
        if isinstance(val, (int, float)) and 40000 < val < 60000:
            xlsb_date_rows.append((i, val, serial_to_date(val)))

# Find date rows in xlsx
xlsx_date_rows = []
for i, row in enumerate(xlsx_rows):
    if row and row[0] is not None:
        d = normalize_xlsx_date(row[0])
        if d is not None and date(2020, 1, 1) <= d <= date(2035, 1, 1):
            xlsx_date_rows.append((i, row[0], d))

if xlsb_date_rows:
    print(f"{'XLSB first date row':<35} {str(xlsb_date_rows[0][2]):>15}")
    print(f"{'XLSB last date row':<35} {str(xlsb_date_rows[-1][2]):>15}")
    print(f"{'XLSB total date rows':<35} {len(xlsb_date_rows):>15}")
else:
    print("XLSB: No date rows found")

if xlsx_date_rows:
    print(f"{'XLSX first date row':<35} {str(xlsx_date_rows[0][2]):>15}")
    print(f"{'XLSX last date row':<35} {str(xlsx_date_rows[-1][2]):>15}")
    print(f"{'XLSX total date rows':<35} {len(xlsx_date_rows):>15}")
else:
    print("XLSX: No date rows found")


# Error values in xlsx
print(f"\n{'XLSX error cells (#N/A etc)':<35} {len(xlsx_error_cells):>15}")
if xlsx_error_cells:
    print("  First 20 error cells (row, col, value):")
    for r, c, v in xlsx_error_cells[:20]:
        hdr = xlsx_headers[c] if xlsx_headers and c < len(xlsx_headers) else f"col{c}"
        print(f"    Row {r}, Col {c} ({hdr}): {v}")


# ─────────────────────────────────────────────
# 4. FIND MAY 8 ROWS
# ─────────────────────────────────────────────
print("\n" + "=" * 80)
print("LOCATING MAY 8, 2026 ROWS")
print("=" * 80)

xlsb_may8_idx = None
for i, d, dt in xlsb_date_rows:
    if dt == MAY8_DATE:
        xlsb_may8_idx = i
        print(f"XLSB May 8 row index: {i} (serial={xlsb_rows[i][0]})")
        break

if xlsb_may8_idx is None:
    print("XLSB: May 8, 2026 row NOT FOUND")

xlsx_may8_idx = None
for i, raw, dt in xlsx_date_rows:
    if dt == MAY8_DATE:
        xlsx_may8_idx = i
        print(f"XLSX May 8 row index: {i} (raw value={raw})")
        break

if xlsx_may8_idx is None:
    print("XLSX: May 8, 2026 row NOT FOUND")


# ─────────────────────────────────────────────
# 5. COLUMN-BY-COLUMN COMPARISON FOR MAY 8
# ─────────────────────────────────────────────
if xlsb_may8_idx is not None and xlsx_may8_idx is not None:
    print("\n" + "=" * 80)
    print("MAY 8, 2026 — COLUMN-BY-COLUMN COMPARISON")
    print("=" * 80)

    xlsb_row = xlsb_rows[xlsb_may8_idx]
    xlsx_row = xlsx_rows[xlsx_may8_idx]

    max_cols = max(len(xlsb_row), len(xlsx_row))

    # Extend shorter row with Nones
    xlsb_row_ext = xlsb_row + [None] * (max_cols - len(xlsb_row))
    xlsx_row_ext = xlsx_row + [None] * (max_cols - len(xlsx_row))

    flags = []
    all_diffs = []

    print(f"\n{'Col':>5} {'Header':<40} {'XLSB Value':>18} {'XLSX Value':>18} {'Diff':>15} {'FLAG'}")
    print("-" * 110)

    for col_idx in range(max_cols):
        xb_val = xlsb_row_ext[col_idx]
        xl_val = xlsx_row_ext[col_idx]

        # Skip col 0 (date column)
        if col_idx == 0:
            continue

        # Skip if both None
        if xb_val is None and xl_val is None:
            continue

        # Get header
        hdr = ""
        if xlsb_headers and col_idx < len(xlsb_headers) and xlsb_headers[col_idx] is not None:
            hdr = str(xlsb_headers[col_idx])[:40]
        elif xlsx_headers and col_idx < len(xlsx_headers) and xlsx_headers[col_idx] is not None:
            hdr = str(xlsx_headers[col_idx])[:40]
        else:
            hdr = f"(col {col_idx})"

        # Skip text/non-numeric columns where both are strings or None
        xb_is_num = isinstance(xb_val, (int, float))
        xl_is_num = isinstance(xl_val, (int, float))

        if not xb_is_num and not xl_is_num:
            # Both non-numeric — check if they differ
            if str(xb_val) != str(xl_val):
                print(f"{col_idx:>5} {hdr:<40} {str(xb_val):>18} {str(xl_val):>18} {'N/A':>15} TEXT_DIFF")
            continue

        # Compute difference
        xb_num = float(xb_val) if xb_is_num else None
        xl_num = float(xl_val) if xl_is_num else None

        if xb_num is not None and xl_num is not None:
            diff = xb_num - xl_num
            flag = "*** FLAG ***" if abs(diff) > DIFF_THRESHOLD else ""
            if abs(diff) > DIFF_THRESHOLD:
                flags.append((col_idx, hdr, xb_num, xl_num, diff))
            all_diffs.append((col_idx, hdr, xb_num, xl_num, diff))
            print(f"{col_idx:>5} {hdr:<40} {xb_num:>18.6f} {xl_num:>18.6f} {diff:>15.6f} {flag}")
        elif xb_num is not None and xl_num is None:
            flag = "*** XLSB_ONLY ***"
            flags.append((col_idx, hdr, xb_num, None, None))
            print(f"{col_idx:>5} {hdr:<40} {xb_num:>18.6f} {'None':>18} {'N/A':>15} {flag}")
        elif xb_num is None and xl_num is not None:
            flag = "*** XLSX_ONLY ***"
            flags.append((col_idx, hdr, None, xl_num, None))
            print(f"{col_idx:>5} {hdr:<40} {'None':>18} {xl_num:>18.6f} {'N/A':>15} {flag}")

    print("\n" + "=" * 80)
    print(f"SUMMARY — MAY 8 DISCREPANCIES (|diff| > {DIFF_THRESHOLD})")
    print("=" * 80)
    if flags:
        print(f"{'Col':>5} {'Header':<40} {'XLSB':>18} {'XLSX':>18} {'Diff':>15}")
        print("-" * 100)
        for col_idx, hdr, xb, xl, diff in flags:
            xb_s = f"{xb:.6f}" if xb is not None else "None"
            xl_s = f"{xl:.6f}" if xl is not None else "None"
            diff_s = f"{diff:.6f}" if diff is not None else "N/A"
            print(f"{col_idx:>5} {hdr:<40} {xb_s:>18} {xl_s:>18} {diff_s:>15}")
        print(f"\nTotal flagged columns: {len(flags)}")
    else:
        print("NO discrepancies found on May 8, 2026 — files match perfectly.")


# ─────────────────────────────────────────────
# 6. WINDOW COMPARISON: May 1–15, 2026
# ─────────────────────────────────────────────
print("\n" + "=" * 80)
print("WINDOW COMPARISON: May 1–15, 2026")
print("Key columns: 7=Net CF, 11=Available Cash, 15=Total Cash, 85=Inflows, 123=Disbursements Total")
print("=" * 80)

# Build lookup dicts keyed by date
xlsb_by_date = {}
for i, serial, dt in xlsb_date_rows:
    xlsb_by_date[dt] = xlsb_rows[i]

xlsx_by_date = {}
for i, raw, dt in xlsx_date_rows:
    xlsx_by_date[dt] = xlsx_rows[i]

# Generate business days in window
def is_business_day(d):
    return d.weekday() < 5  # Mon–Fri

window_dates = []
cur = WINDOW_START
while cur <= WINDOW_END:
    if is_business_day(cur):
        window_dates.append(cur)
    cur += timedelta(days=1)

print(f"\nBusiness days in window: {[str(d) for d in window_dates]}")

for col_idx, col_name in sorted(KEY_COLS.items()):
    print(f"\n{'─'*80}")
    print(f"COLUMN {col_idx}: {col_name}")
    print(f"{'─'*80}")
    print(f"{'Date':<15} {'XLSB':>18} {'XLSX':>18} {'Diff':>15} {'FLAG'}")
    print("-" * 75)

    any_flag = False
    for dt in window_dates:
        xb_row = xlsb_by_date.get(dt)
        xl_row = xlsx_by_date.get(dt)

        xb_val = None
        xl_val = None

        if xb_row and col_idx < len(xb_row):
            xb_val = xb_row[col_idx]
        if xl_row and col_idx < len(xl_row):
            xl_val = xl_row[col_idx]

        xb_is_num = isinstance(xb_val, (int, float))
        xl_is_num = isinstance(xl_val, (int, float))

        if xb_is_num and xl_is_num:
            diff = float(xb_val) - float(xl_val)
            flag = "*** FLAG ***" if abs(diff) > DIFF_THRESHOLD else ""
            if flag:
                any_flag = True
            print(f"{str(dt):<15} {float(xb_val):>18.6f} {float(xl_val):>18.6f} {diff:>15.6f} {flag}")
        elif xb_is_num and not xl_is_num:
            print(f"{str(dt):<15} {float(xb_val):>18.6f} {'None/Missing':>18} {'N/A':>15} *** XLSX_MISSING ***")
            any_flag = True
        elif not xb_is_num and xl_is_num:
            print(f"{str(dt):<15} {'None/Missing':>18} {float(xl_val):>18.6f} {'N/A':>15} *** XLSB_MISSING ***")
            any_flag = True
        else:
            xb_s = str(xb_val) if xb_val is not None else "None"
            xl_s = str(xl_val) if xl_val is not None else "None"
            print(f"{str(dt):<15} {xb_s:>18} {xl_s:>18} {'N/A':>15}")

    if not any_flag:
        print("  [All values match within threshold for this column]")


print("\n" + "=" * 80)
print("COMPARISON COMPLETE")
print("=" * 80)
