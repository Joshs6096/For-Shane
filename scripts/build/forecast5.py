"""6-Month forecast — formula-driven, P&L cash flow, buys=listings=sales locked at 40% sell-through."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

HDR_FILL = PatternFill('solid', fgColor='2C3E50')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='34495E')
SECTION_FONT = Font(bold=True, color='FFFFFF', size=12)
INPUT_FILL = PatternFill('solid', fgColor='FFF3CD')
INPUT_FONT = Font(bold=True, color='8A6D00')
SEED_FILL = PatternFill('solid', fgColor='F39C12')
SEED_FONT = Font(bold=True, color='FFFFFF')
TOTAL_FILL = PatternFill('solid', fgColor='ECF0F1')
TOTAL_FONT = Font(bold=True)
INCOME_FILL = PatternFill('solid', fgColor='D5F5E3')
EXPENSE_FILL = PatternFill('solid', fgColor='FADBD8')
NET_FILL = PatternFill('solid', fgColor='AED6F1')
THIN = Side(border_style='thin', color='BDC3C7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()

# ============================================================
# TAB 1: SUMMARY (with INPUTS)
# ============================================================
ws = wb.active
ws.title = 'Summary'
ws.column_dimensions['A'].width = 38
for c in 'BCDEFGH': ws.column_dimensions[c].width = 16

ws['A1'] = 'Curate Coins — 6-Month Outlook'
ws['A1'].font = Font(bold=True, size=16)
ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d")}  |  Edit yellow cells in INPUTS — everything recalculates.'
ws['A2'].font = Font(italic=True, size=10, color='7F8C8D')

ws['A4'] = 'INPUTS'
ws['A4'].fill = SECTION_FILL; ws['A4'].font = SECTION_FONT
ws.merge_cells('A4:C4')

inputs = [
    ('Drafts / week (your lever)',    5,     '0.0'),                # B5 — primary input
    ('Avg sale price ($)',            125,   '$#,##0.00'),         # B6
    ('eBay fee %',                    0.13,  '0.0%'),               # B7
    ('Ship + supplies per item ($)',  2.50,  '$#,##0.00'),         # B8
    ('Reinvestment % of net',         0.40,  '0.0%'),               # B9
    ('Silver-buy weekly ($)',         120,   '$#,##0.00'),         # B10
    ('Silver-buy monthly ($)',        '=B10*4.33', '$#,##0.00'),   # B11
    ('Board seed (Month 1) ($)',      1000,  '$#,##0'),             # B12
    ('M1 ramp factor',                0.80,  '0%'),                 # B13
    ('M2 ramp factor',                0.95,  '0%'),                 # B14
    ('M3-M6 ramp factor',             1.00,  '0%'),                 # B15
    ('Silver spot ($/oz)',            81.88, '$#,##0.00'),         # B16
    ('Sell-through % (per month)',    0.40,  '0.0%'),               # B17
    ('Avg flippable cost / unit ($)', 38.09, '$#,##0.00'),         # B18
]
for i, (label, val, fmt) in enumerate(inputs, 5):
    ws.cell(row=i, column=1, value=label)
    c = ws.cell(row=i, column=2, value=val)
    c.number_format = fmt
    c.fill = INPUT_FILL; c.font = INPUT_FONT; c.border = BORDER

# Cell refs
DRAFTS_WK = 'Summary!$B$5'
SALES = '(Summary!$B$5*4.33)'  # Sales/mo derived from drafts/wk (1 listing = 1 sale at steady state)
PRICE = 'Summary!$B$6'
FEE_PCT = 'Summary!$B$7'
SHIP = 'Summary!$B$8'
REINV_PCT = 'Summary!$B$9'
SILVER_MO = 'Summary!$B$11'
SEED = 'Summary!$B$12'
M1_R = 'Summary!$B$13'
M2_R = 'Summary!$B$14'
M3_R = 'Summary!$B$15'
SELL_THRU = 'Summary!$B$17'
UNIT_COST = 'Summary!$B$18'

# 6-Month Headline
ws['A20'] = '6-MONTH HEADLINE'
ws['A20'].fill = SECTION_FILL; ws['A20'].font = SECTION_FONT
ws.merge_cells('A20:C20')

headline = [
    ("Total sales (6mo)",            "='Cash Flow (P&L)'!H6",  '0'),
    ("Gross revenue",                "='Cash Flow (P&L)'!H7",  '$#,##0'),
    ("+ Board seed inflow",          "='Cash Flow (P&L)'!H8",  '$#,##0'),
    ("TOTAL INCOME (6mo)",           "='Cash Flow (P&L)'!H9",  '$#,##0'),
    ("- eBay fees",                  "='Cash Flow (P&L)'!H12", '$#,##0'),
    ("- Shipping + supplies",        "='Cash Flow (P&L)'!H13", '$#,##0'),
    ("- Inventory reinvestment",     "='Cash Flow (P&L)'!H14", '$#,##0'),
    ("- Silver-buy spend",           "='Cash Flow (P&L)'!H15", '$#,##0'),
    ("TOTAL EXPENSES (6mo)",         "='Cash Flow (P&L)'!H16", '$#,##0'),
    ("FREE CASH TO JOSH",            "='Cash Flow (P&L)'!H18", '$#,##0'),
]
for i, (label, formula, fmt) in enumerate(headline, 21):
    ws.cell(row=i, column=1, value=label)
    c = ws.cell(row=i, column=2, value=formula)
    c.number_format = fmt; c.border = BORDER
    if 'INCOME' in label or 'EXPENSES' in label:
        ws.cell(row=i, column=1).font = Font(bold=True)
        c.font = Font(bold=True); c.fill = TOTAL_FILL
    if 'CASH TO JOSH' in label:
        ws.cell(row=i, column=1).font = Font(bold=True)
        c.font = Font(bold=True); c.fill = NET_FILL

# Notes
ws['A33'] = 'KEY TAKEAWAYS'
ws['A33'].fill = SECTION_FILL; ws['A33'].font = SECTION_FONT
ws.merge_cells('A33:C33')
takeaways = [
    "• Lever = drafts/week (B5). Sales/mo = drafts/wk × 4.33 — every listing eventually sells (100% sell-through observed).",
    "• Buys, listings, and sales scale 1:1 — all move together with B5.",
    "• 40% sell-through (B17) sets the active inventory pile size (= sales / 40%), not the sales rate itself.",
    "• Free cash = Total Income − Total Expenses (Cash Flow P&L tab).",
    "• Core (188oz inherited + numismatic) never touched in any scenario.",
    "• To flex the plan, edit B5 — every downstream number recalculates.",
]
for i, t in enumerate(takeaways, 34):
    ws.cell(row=i, column=1, value=t)
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)

# ============================================================
# TAB 2: CASH FLOW (P&L)
# ============================================================
ws2 = wb.create_sheet('Cash Flow (P&L)')
for col in range(1, 11): ws2.column_dimensions[get_column_letter(col)].width = 15
ws2.column_dimensions['A'].width = 32

ws2['A1'] = '6-Month P&L — Income vs Expenses'
ws2['A1'].font = Font(bold=True, size=14)

months_labels = []
start = datetime.now().replace(day=1)
for i in range(6):
    m = start.replace(month=((start.month + i - 1) % 12) + 1, year=start.year + ((start.month + i - 1) // 12))
    months_labels.append(m.strftime('%b %Y'))

hdr = ['Line item'] + months_labels + ['6-Mo Total']
for i, v in enumerate(hdr, 1):
    c = ws2.cell(row=4, column=i, value=v)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER

def ramp_for(col_letter):
    if col_letter == 'B': return M1_R
    if col_letter == 'C': return M2_R
    return M3_R

# INCOME
ws2['A5'] = 'INCOME'
ws2['A5'].fill = INCOME_FILL; ws2['A5'].font = Font(bold=True, color='1E5F2C')
ws2.merge_cells('A5:H5')

ws2.cell(row=6, column=1, value='Sales (units)')
for col_i, col_letter in enumerate(['B','C','D','E','F','G'], 2):
    ws2.cell(row=6, column=col_i, value=f'=ROUND({SALES}*{ramp_for(col_letter)}, 0)')
    ws2.cell(row=6, column=col_i).number_format = '0'
ws2.cell(row=6, column=8, value='=SUM(B6:G6)').number_format = '0'
ws2.cell(row=6, column=8).font = Font(bold=True)

ws2.cell(row=7, column=1, value='Gross revenue')
for col_i, col_letter in enumerate(['B','C','D','E','F','G'], 2):
    ws2.cell(row=7, column=col_i, value=f'={col_letter}6*{PRICE}')
    ws2.cell(row=7, column=col_i).number_format = '$#,##0'
ws2.cell(row=7, column=8, value='=SUM(B7:G7)').number_format = '$#,##0'
ws2.cell(row=7, column=8).font = Font(bold=True)

ws2.cell(row=8, column=1, value='Board seed inflow')
ws2.cell(row=8, column=1).fill = SEED_FILL; ws2.cell(row=8, column=1).font = SEED_FONT
ws2.cell(row=8, column=2, value=f'={SEED}')
for col_i in range(3, 8):
    ws2.cell(row=8, column=col_i, value=0)
for col_i in range(2, 8):
    ws2.cell(row=8, column=col_i).number_format = '$#,##0'
    ws2.cell(row=8, column=col_i).fill = SEED_FILL
    ws2.cell(row=8, column=col_i).font = SEED_FONT
ws2.cell(row=8, column=8, value='=SUM(B8:G8)').number_format = '$#,##0'
ws2.cell(row=8, column=8).font = Font(bold=True)

ws2.cell(row=9, column=1, value='TOTAL INCOME').font = Font(bold=True)
ws2.cell(row=9, column=1).fill = INCOME_FILL
for col_i in range(2, 9):
    ws2.cell(row=9, column=col_i, value=f'={get_column_letter(col_i)}7+{get_column_letter(col_i)}8')
    ws2.cell(row=9, column=col_i).number_format = '$#,##0'
    ws2.cell(row=9, column=col_i).font = Font(bold=True)
    ws2.cell(row=9, column=col_i).fill = INCOME_FILL

# EXPENSES
ws2['A11'] = 'EXPENSES'
ws2['A11'].fill = EXPENSE_FILL; ws2['A11'].font = Font(bold=True, color='8B1A1A')
ws2.merge_cells('A11:H11')

ws2.cell(row=12, column=1, value='eBay fees (13%)')
for col_i, col_letter in enumerate(['B','C','D','E','F','G'], 2):
    ws2.cell(row=12, column=col_i, value=f'={col_letter}7*{FEE_PCT}')
    ws2.cell(row=12, column=col_i).number_format = '$#,##0'
ws2.cell(row=12, column=8, value='=SUM(B12:G12)').number_format = '$#,##0'
ws2.cell(row=12, column=8).font = Font(bold=True)

ws2.cell(row=13, column=1, value='Shipping + supplies')
for col_i, col_letter in enumerate(['B','C','D','E','F','G'], 2):
    ws2.cell(row=13, column=col_i, value=f'={col_letter}6*{SHIP}')
    ws2.cell(row=13, column=col_i).number_format = '$#,##0'
ws2.cell(row=13, column=8, value='=SUM(B13:G13)').number_format = '$#,##0'
ws2.cell(row=13, column=8).font = Font(bold=True)

ws2.cell(row=14, column=1, value='Inventory reinvestment')
for col_i, col_letter in enumerate(['B','C','D','E','F','G'], 2):
    ws2.cell(row=14, column=col_i, value=f'=({col_letter}7-{col_letter}12-{col_letter}13)*{REINV_PCT}')
    ws2.cell(row=14, column=col_i).number_format = '$#,##0'
ws2.cell(row=14, column=8, value='=SUM(B14:G14)').number_format = '$#,##0'
ws2.cell(row=14, column=8).font = Font(bold=True)

ws2.cell(row=15, column=1, value='Silver-buy budget')
for col_i in range(2, 8):
    ws2.cell(row=15, column=col_i, value=f'={SILVER_MO}')
    ws2.cell(row=15, column=col_i).number_format = '$#,##0'
ws2.cell(row=15, column=8, value='=SUM(B15:G15)').number_format = '$#,##0'
ws2.cell(row=15, column=8).font = Font(bold=True)

ws2.cell(row=16, column=1, value='TOTAL EXPENSES').font = Font(bold=True)
ws2.cell(row=16, column=1).fill = EXPENSE_FILL
for col_i in range(2, 9):
    L = get_column_letter(col_i)
    ws2.cell(row=16, column=col_i, value=f'={L}12+{L}13+{L}14+{L}15')
    ws2.cell(row=16, column=col_i).number_format = '$#,##0'
    ws2.cell(row=16, column=col_i).font = Font(bold=True)
    ws2.cell(row=16, column=col_i).fill = EXPENSE_FILL

# NET
ws2['A18'] = 'FREE CASH TO JOSH (Income − Expenses)'
ws2.cell(row=18, column=1).font = Font(bold=True)
ws2.cell(row=18, column=1).fill = NET_FILL
for col_i in range(2, 9):
    L = get_column_letter(col_i)
    ws2.cell(row=18, column=col_i, value=f'={L}9-{L}16')
    ws2.cell(row=18, column=col_i).number_format = '$#,##0'
    ws2.cell(row=18, column=col_i).font = Font(bold=True)
    ws2.cell(row=18, column=col_i).fill = NET_FILL

for r in range(4, 19):
    for col_i in range(1, 9):
        ws2.cell(row=r, column=col_i).border = BORDER

# ============================================================
# TAB 3: SEED DEPLOYMENT
# ============================================================
ws3 = wb.create_sheet('Seed Deployment')
for col in 'ABCDEF': ws3.column_dimensions[col].width = 24

ws3['A1'] = '$1,000 Board Seed — Deployment Plan'
ws3['A1'].font = Font(bold=True, size=14)

ws3['A4'] = 'PROPOSED ALLOCATION'
ws3['A4'].fill = SECTION_FILL; ws3['A4'].font = SECTION_FONT
ws3.merge_cells('A4:F4')

hdr = ['Bucket', 'Allocation %', '$', 'Avg unit cost', 'Units', 'Expected hold']
for i, h in enumerate(hdr, 1):
    c = ws3.cell(row=5, column=i, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER

deploy = [
    ('High-velocity bullion premium (NORFEDs, Libertads, ASE rolls)', 0.40, 100, '30-45 days'),
    ('Common silver upgrades (90% halves, dimes, Walker fills)',       0.30, 25,  '45-60 days'),
    ('Numismatic plays (semi-key dates, slabbed semi-keys)',           0.20, 75,  '60-90 days'),
    ('Reserve / tactical buys (Whatnot, show floor)',                   0.10, 50,  'opportunistic'),
]
for i, (b, pct, ucost, hold) in enumerate(deploy, 6):
    ws3.cell(row=i, column=1, value=b)
    c = ws3.cell(row=i, column=2, value=pct); c.number_format = '0%'; c.fill = INPUT_FILL; c.font = INPUT_FONT
    ws3.cell(row=i, column=3, value=f'={SEED}*B{i}').number_format = '$#,##0'
    c = ws3.cell(row=i, column=4, value=ucost); c.number_format = '$#,##0'; c.fill = INPUT_FILL; c.font = INPUT_FONT
    ws3.cell(row=i, column=5, value=f'=C{i}/D{i}').number_format = '0.0'
    ws3.cell(row=i, column=6, value=hold)
    for col_i in range(1, 7):
        ws3.cell(row=i, column=col_i).border = BORDER

ws3.cell(row=10, column=1, value='TOTAL').font = Font(bold=True)
ws3.cell(row=10, column=2, value='=SUM(B6:B9)').number_format = '0%'
ws3.cell(row=10, column=3, value='=SUM(C6:C9)').number_format = '$#,##0'
ws3.cell(row=10, column=5, value='=SUM(E6:E9)').number_format = '0.0'
for col_i in range(1, 7):
    ws3.cell(row=10, column=col_i).border = BORDER
    ws3.cell(row=10, column=col_i).fill = TOTAL_FILL
    ws3.cell(row=10, column=col_i).font = Font(bold=True)

ws3['A12'] = 'EXPECTED RETURN ON SEED'
ws3['A12'].fill = SECTION_FILL; ws3['A12'].font = SECTION_FONT
ws3.merge_cells('A12:F12')

ret_hdr = ['Bucket', 'Investment $', 'Margin %', 'Expected gross', 'Expected net (after 13% fees)', 'ROI %']
for i, h in enumerate(ret_hdr, 1):
    c = ws3.cell(row=13, column=i, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER

returns = [('Bullion premium',6,0.15),('Common silver upgrades',7,0.40),('Numismatic plays',8,0.60),('Tactical reserve',9,0.50)]
for i, (lbl, src_row, mgn) in enumerate(returns, 14):
    ws3.cell(row=i, column=1, value=lbl)
    ws3.cell(row=i, column=2, value=f'=C{src_row}').number_format = '$#,##0'
    c = ws3.cell(row=i, column=3, value=mgn); c.number_format = '0%'; c.fill = INPUT_FILL; c.font = INPUT_FONT
    ws3.cell(row=i, column=4, value=f'=B{i}*(1+C{i})').number_format = '$#,##0'
    ws3.cell(row=i, column=5, value=f'=D{i}*(1-{FEE_PCT})').number_format = '$#,##0'
    ws3.cell(row=i, column=6, value=f'=(E{i}-B{i})/B{i}').number_format = '0%'
    for col_i in range(1, 7):
        ws3.cell(row=i, column=col_i).border = BORDER

ws3.cell(row=18, column=1, value='TOTAL').font = Font(bold=True)
ws3.cell(row=18, column=2, value='=SUM(B14:B17)').number_format = '$#,##0'
ws3.cell(row=18, column=4, value='=SUM(D14:D17)').number_format = '$#,##0'
ws3.cell(row=18, column=5, value='=SUM(E14:E17)').number_format = '$#,##0'
ws3.cell(row=18, column=6, value='=(E18-B18)/B18').number_format = '0%'
for col_i in range(1, 7):
    ws3.cell(row=18, column=col_i).border = BORDER
    ws3.cell(row=18, column=col_i).fill = TOTAL_FILL
    ws3.cell(row=18, column=col_i).font = Font(bold=True)
ws3.cell(row=18, column=5).fill = INCOME_FILL

# ============================================================
# TAB 4: INVENTORY PLAN — buys=listings=sales locked
# ============================================================
ws4 = wb.create_sheet('Inventory Plan')
for col in 'ABCDEFG': ws4.column_dimensions[col].width = 22

ws4['A1'] = 'Inventory Plan — Drafts/week is the lever; buys, listings, sales scale together'
ws4['A1'].font = Font(bold=True, size=14)
ws4['A2'] = 'Edit Drafts/week (Summary!B5). Revenue ramps proportionally — every listing eventually sells.'
ws4['A2'].font = Font(italic=True, color='7F8C8D')

# A. Current state
ws4['A4'] = 'A. CURRENT INVENTORY STATE'
ws4['A4'].fill = SECTION_FILL; ws4['A4'].font = SECTION_FONT
ws4.merge_cells('A4:G4')

state_hdr = ['Segment', 'Description', 'Units', 'Cost basis', 'Low value', 'High value', 'Notes']
for i, h in enumerate(state_hdr, 1):
    c = ws4.cell(row=5, column=i, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER

state = [
    ('Core (inherited)',  '$0 cost basis — heritage holdings', 155, 0,    8000,   31877, 'NEVER liquidate. Appreciation engine.'),
    ('Core (numismatic)', 'Proof sets, key dates, premium graded', 550, 4187, 14000, 13221, 'Hold for grading or premium retail.'),
    ('Flippable bullion', 'Common ASE, Maple Leafs, rounds, bars', 130,  6500, 12000, 18000, 'Primary listing rotation.'),
    ('Flippable silver',  '90% halves/quarters/dimes, Morgans', 212,  6528,  10000, 12293, 'Bullion premium + numismatic blend.'),
]
for i, (seg, desc, units, cost, lo, hi, note) in enumerate(state, 6):
    ws4.cell(row=i, column=1, value=seg)
    ws4.cell(row=i, column=2, value=desc)
    ws4.cell(row=i, column=3, value=units).number_format = '0'
    ws4.cell(row=i, column=4, value=cost).number_format = '$#,##0'
    ws4.cell(row=i, column=5, value=lo).number_format = '$#,##0'
    ws4.cell(row=i, column=6, value=hi).number_format = '$#,##0'
    ws4.cell(row=i, column=7, value=note)
    for col_i in range(1, 8):
        ws4.cell(row=i, column=col_i).border = BORDER

ws4.cell(row=10, column=1, value='TOTAL').font = Font(bold=True)
ws4.cell(row=10, column=3, value='=SUM(C6:C9)').number_format = '0'
ws4.cell(row=10, column=4, value='=SUM(D6:D9)').number_format = '$#,##0'
ws4.cell(row=10, column=5, value='=SUM(E6:E9)').number_format = '$#,##0'
ws4.cell(row=10, column=6, value='=SUM(F6:F9)').number_format = '$#,##0'
for col_i in range(1, 8):
    ws4.cell(row=10, column=col_i).border = BORDER
    ws4.cell(row=10, column=col_i).fill = TOTAL_FILL
    ws4.cell(row=10, column=col_i).font = Font(bold=True)

ws4.cell(row=11, column=1, value='Flippable subtotal').font = Font(bold=True)
ws4.cell(row=11, column=3, value='=C8+C9').number_format = '0'
ws4.cell(row=11, column=4, value='=D8+D9').number_format = '$#,##0'
ws4.cell(row=11, column=6, value='=F8+F9').number_format = '$#,##0'
for col_i in range(1, 8):
    ws4.cell(row=11, column=col_i).border = BORDER

# B. Locked ratios — Buys = Listings = Sales
ws4['A13'] = 'B. THE LOCKED RATIO — BUYS = LISTINGS = SALES (steady-state)'
ws4['A13'].fill = SECTION_FILL; ws4['A13'].font = SECTION_FONT
ws4.merge_cells('A13:G13')
ws4['A14'] = 'Sales/mo = Drafts/wk × 4.33 (every listing sells). Buys = Listings = Sales. Active pile = Sales ÷ 40% sell-through.'
ws4['A14'].font = Font(italic=True, size=10, color='7F8C8D')
ws4.merge_cells('A14:G14')

ws4.cell(row=15, column=1, value='Metric').font = HDR_FONT; ws4.cell(row=15, column=1).fill = HDR_FILL
ws4.cell(row=15, column=2, value='Per Month').font = HDR_FONT; ws4.cell(row=15, column=2).fill = HDR_FILL
ws4.cell(row=15, column=3, value='Per Week').font = HDR_FONT; ws4.cell(row=15, column=3).fill = HDR_FILL
ws4.cell(row=15, column=4, value='Formula').font = HDR_FONT; ws4.cell(row=15, column=4).fill = HDR_FILL
for col_i in [1,2,3,4]: ws4.cell(row=15, column=col_i).border = BORDER

locked = [
    ('Drafts (input)',  f'={DRAFTS_WK}*4.33',         f'={DRAFTS_WK}',  'Input: Summary!B5 = drafts/week'),
    ('Sales',           '=B16',                        '=B16/4.33',     '= Drafts (every listing eventually sells)'),
    ('Buys (units)',    '=B16',                        '=B16/4.33',     '= Sales (1:1 inventory replacement)'),
    ('Buys ($)',        f'=B18*{UNIT_COST}',           '=B19/4.33',     '= Buys × avg unit cost'),
]
for i, (lbl, mo, wk, formula_text) in enumerate(locked, 16):
    ws4.cell(row=i, column=1, value=lbl)
    c = ws4.cell(row=i, column=2, value=mo); c.number_format = '0.0' if 'unit' in lbl.lower() or 'Sales' in lbl or 'Listings' in lbl else '$#,##0'
    c = ws4.cell(row=i, column=3, value=wk); c.number_format = '0.0' if 'unit' in lbl.lower() or 'Sales' in lbl or 'Listings' in lbl else '$#,##0'
    if 'Buys ($)' in lbl:
        ws4.cell(row=i, column=2).number_format = '$#,##0'
        ws4.cell(row=i, column=3).number_format = '$#,##0'
    else:
        ws4.cell(row=i, column=2).number_format = '0'
        ws4.cell(row=i, column=3).number_format = '0.0'
    ws4.cell(row=i, column=4, value=formula_text).font = Font(italic=True, size=9, color='7F8C8D')
    for col_i in [1,2,3,4]: ws4.cell(row=i, column=col_i).border = BORDER

# Active listings pile
ws4.cell(row=20, column=1, value='Active listings (pile)').font = Font(bold=True)
ws4.cell(row=20, column=2, value=f'={SALES}/{SELL_THRU}').number_format = '0'
ws4.cell(row=20, column=4, value=f'= Sales / Sell-through % (40%)').font = Font(italic=True, size=9, color='7F8C8D')
for col_i in [1,2,3,4]: ws4.cell(row=20, column=col_i).border = BORDER
ws4.cell(row=20, column=2).fill = TOTAL_FILL

# C. Replenishment cash flow
ws4['A22'] = 'C. REPLENISHMENT CASH FLOW (does it pay for itself?)'
ws4['A22'].fill = SECTION_FILL; ws4['A22'].font = SECTION_FONT
ws4.merge_cells('A22:G22')

ws4.cell(row=23, column=1, value='Metric').font = HDR_FONT; ws4.cell(row=23, column=1).fill = HDR_FILL
ws4.cell(row=23, column=2, value='Value').font = HDR_FONT; ws4.cell(row=23, column=2).fill = HDR_FILL
for col_i in [1,2]: ws4.cell(row=23, column=col_i).border = BORDER

cash_rows = [
    ('Replenishment $/mo NEEDED (= Buys $)',          '=B19',                                        '$#,##0'),
    ('Replenishment $/mo from 40% reinvest',          f"='Cash Flow (P&L)'!H14/6",                  '$#,##0'),
    ('Silver-buy budget ($/mo)',                       f'={SILVER_MO}',                                '$#,##0'),
    ('Total replenishment $/mo AVAILABLE',             '=B25+B26',                                    '$#,##0'),
    ('SURPLUS / SHORTFALL',                            '=B27-B24',                                    '$#,##0'),
]
for i, (lbl, formula, fmt) in enumerate(cash_rows, 24):
    ws4.cell(row=i, column=1, value=lbl)
    c = ws4.cell(row=i, column=2, value=formula)
    c.number_format = fmt; c.border = BORDER
    ws4.cell(row=i, column=1).border = BORDER
    if 'SURPLUS' in lbl:
        ws4.cell(row=i, column=1).font = Font(bold=True)
        c.font = Font(bold=True); c.fill = INCOME_FILL

# D. Monthly buying split
ws4['A30'] = 'D. MONTHLY BUYING SPLIT (where the $ goes)'
ws4['A30'].fill = SECTION_FILL; ws4['A30'].font = SECTION_FONT
ws4.merge_cells('A30:G30')

buy_hdr = ['Bucket', '% of buy', '$/mo', 'Units/mo', 'Channels', 'Vetting']
for i, h in enumerate(buy_hdr, 1):
    c = ws4.cell(row=31, column=i, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER

buy_buckets = [
    ('Bullion premium plays', 0.45, 100, 'eBay, Whatnot, dealers', 'Sheldon-vet (CUR-42)'),
    ('Common silver upgrades', 0.30, 25, 'eBay BIN, Whatnot bulk', 'Self-vet vs melt+10%'),
    ('Numismatic semi-keys', 0.20, 75, 'eBay auctions, GreatCollections', 'Sheldon-vet mandatory'),
    ('Tactical / show floor', 0.05, 50, 'Coin shows, dealers', 'On-the-spot judgment'),
]
for i, (b, pct, ucost, ch, vet) in enumerate(buy_buckets, 32):
    ws4.cell(row=i, column=1, value=b)
    c = ws4.cell(row=i, column=2, value=pct); c.number_format = '0%'; c.fill = INPUT_FILL; c.font = INPUT_FONT
    ws4.cell(row=i, column=3, value=f'=B19*B{i}').number_format = '$#,##0'  # B19 = Buys $/mo
    ws4.cell(row=i, column=4, value=f'=C{i}/{ucost}').number_format = '0.0'
    ws4.cell(row=i, column=5, value=ch)
    ws4.cell(row=i, column=6, value=vet)
    for col_i in range(1, 7):
        ws4.cell(row=i, column=col_i).border = BORDER

ws4.cell(row=36, column=1, value='TOTAL').font = Font(bold=True)
ws4.cell(row=36, column=2, value='=SUM(B32:B35)').number_format = '0%'
ws4.cell(row=36, column=3, value='=SUM(C32:C35)').number_format = '$#,##0'
ws4.cell(row=36, column=4, value='=SUM(D32:D35)').number_format = '0.0'
for col_i in range(1, 7):
    ws4.cell(row=36, column=col_i).border = BORDER
    ws4.cell(row=36, column=col_i).fill = TOTAL_FILL
    ws4.cell(row=36, column=col_i).font = Font(bold=True)

# E. Guardrails
ws4['A38'] = 'E. GUARDRAILS'
ws4['A38'].fill = SECTION_FILL; ws4['A38'].font = SECTION_FONT
ws4.merge_cells('A38:G38')
guards = [
    "• Core (rows 6-7) is NEVER liquidated. Appreciation engine.",
    "• Buys/listings/sales lock together — if buying budget grows X%, listings grow X%, sales grow X%.",
    "• If SURPLUS/SHORTFALL goes negative, raise reinvest %, lower sales target, or accept core drawdown.",
    "• All buys Sheldon-vetted (CUR-42).",
    "• Bullion buys >$200/unit need board check-in.",
]
for i, g in enumerate(guards, 39):
    ws4.cell(row=i, column=1, value=g)
    ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)

# Save
out = '/Users/josh/.openclaw/workspace/curate-coins-6mo-forecast.xlsx'
wb.save(out)
print(f"Saved: {out}")
