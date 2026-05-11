import warnings; warnings.filterwarnings('ignore')
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, copy

SRC = '/Users/josh/Desktop/Daily Cash Fcst - 3.5.25_BU view_SC payroll true up_$2M SC payroll hedge.xlsx'
DST = '/Users/josh/Downloads/Daily Cash Fcst - 5.8.26_10tab.xlsx'

# ── Styles ────────────────────────────────────────────────────────────────
NAVY  = "1F3864"; WHITE = "FFFFFF"; LBLUE = "D6E4F0"; LGREY = "F2F2F2"
DGREY = "595959"; AMBER = "FFF2CC"; GREEN = "E2EFDA"; RED   = "FFD5D5"

def hdr(sz=9,bold=True,color=WHITE): return Font(name='Arial',size=sz,bold=bold,color=color)
def body(sz=9,bold=False,color="000000"): return Font(name='Arial',size=sz,bold=bold,color=color)
def fill(h): return PatternFill("solid",fgColor=h)
def ctr(wrap=False): return Alignment(horizontal='center',vertical='center',wrap_text=wrap)
def lft(i=0): return Alignment(horizontal='left',vertical='center',indent=i)
def rgt(): return Alignment(horizontal='right',vertical='center')
def thin():
    s=Side(style='thin',color='BBBBBB')
    return Border(left=s,right=s,top=s,bottom=s)
def med_bottom():
    m=Side(style='medium',color=NAVY); t=Side(style='thin',color='BBBBBB')
    return Border(left=t,right=t,top=t,bottom=m)

def ph(ws,row,col,text,bg=NAVY,fg=WHITE,sz=9,wrap=True):
    c=ws.cell(row,col,text)
    c.font=hdr(sz,True,fg); c.fill=fill(bg)
    c.alignment=ctr(wrap); c.border=med_bottom()
    return c

def cv(ws,row,col,val=None,bg=WHITE,bold=False,fmt=None,align=None):
    c=ws.cell(row,col)
    if val is not None: c.value=val
    c.font=body(bold=bold); c.fill=fill(bg); c.border=thin()
    if fmt: c.number_format=fmt
    if align: c.alignment=align
    return c

def copy_sheet(ws_src, ws_dst, max_r=None, max_c=None):
    mr = max_r or ws_src.max_row
    mc = max_c or ws_src.max_column
    for r in range(1, mr+1):
        for c in range(1, min(mc+1, mc+1)):
            v = ws_src.cell(r,c).value
            if v is not None:
                ws_dst.cell(r,c).value = v
    for col,dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col].width = dim.width or 10
    for rn,dim in ws_src.row_dimensions.items():
        if rn<=mr and dim.height:
            ws_dst.row_dimensions[rn].height=dim.height

# ── Load source ───────────────────────────────────────────────────────────
print("Loading source …")
wb_s = load_workbook(SRC, data_only=True)
print(f"  Source sheets: {len(wb_s.sheetnames)}")

# ── Build output workbook ─────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Sheet order matches actual file (visible first, then hidden)
# VISIBLE (10) then HIDDEN (8)

# ════════════════════════════════════════════════════════════════════════
# 1. CASH FLOW  (visible)
# ════════════════════════════════════════════════════════════════════════
print("Building Cash Flow …")
ws = wb.create_sheet('Cash Flow')
ws_s = wb_s['Cash Flow']

# Row 1: internal col-index row (actual file pattern)
# Row 2: header labels matching actual file
# Row 3: grand-total summary row
# Row 4+: daily data

# Full header map (1-based col → label) matching actual file exactly
CF = {
    1:'Date', 4:'Week',
    5:'Operating Cash Flows', 6:'Non-Operating Cash Flows', 7:'Net Cash Flow',
    10:'Net Change in Cash', 11:'Available Cash', 12:'AC (No HTS)',
    13:'Segregated Account', 14:'Unavailable Cash', 15:'Total Cash',
    16:'TopCo Undrawn Equity',
    17:'April 2022 $110M PPM - Uncommitted Equity',
    18:'June 2022 $140M offering - Uncommitted Equity',
    19:'August 2022 $120M offering - Uncommitted Equity',
    20:'Revolver - DIP',
    21:'Cash Surplus / (Deficit) including Asset Sales',
    22:'Cash Surplus / (Deficit)', 23:'LC Opportunities',
    24:'Cash Surplus / (Deficit) + LC Opportunities',
    25:'Cash Requirement', 26:'Previous Surplus / (Deficit)',
    27:'Variance vs. Previous Forecast',
    28:'Tranche 1 Principal',29:'Tranche 2 Principal',
    30:'Tranche 3 Principal',31:'Tranche 4 Principal',
    32:'Cyrus Term Loan',33:'Tranche 6 Principal',
    34:'Tranche 7 Principal',35:'Tranche 8 Principal',
    36:'Tranche 9 Principal',37:'Tranche 10 Principal',
    38:'Tranche 11 Principal',39:'Cyrus $55M Term Loan',
    40:'ESL $15M Term Loan',41:'ESL $16.5M Term Loan',
    42:'Cyrus $16.5M Term Loan',43:'ESL $27M Term Loan',
    44:'ESL $50M Term Loan',45:'Cyrus $25M Term Loan',
    46:'ESL $70M Term Loan',47:'Cyrus $8M Term Loan',
    48:'ESL $175M Term Loan (T22)',49:'Cyrus 7th Term Loan',
    50:'ESL 15th Term Loan',
    51:'Sept. 2025 Term Loans',          # present in actual, missing from source
    52:'Cyrus Eighth Incremental Term Loans',  # present in actual, missing from source
    58:'Bohemia Loan (Mortgage)',59:'Lease Opco Loan',
    60:'Lacey, WA Loan (Mortgage) [$11.2M Facility]',
    61:'New Brunswick Loan',62:'Hackensack, NJ Loan (2022)',
    63:'Hackensack, NJ Loan (2026)',64:'Wintrust/Manteno Loan',
    65:'Durham, NC Loan',
    66:'UBS Loan Note A',67:'UBS Loan Note B',
    68:'Proceeds from Asset Sales ($250M Balance)',69:'Guam Loan',
    70:'Cash Collateralization of BAML LC (5%)',
    71:'Citi LC',72:'Citi to PNC (Back-to-Back LC 4% cash collateralization)',
    73:'PNC LC',74:'Cash Collateralization of PNC LC (2%)',
    75:'Total Cash Collateralization of PNC/Citi LC',
    76:'UBS LC',77:'Cash Collateralization of UBS LC (2%)',
    78:'Balance Sheet Inventory',
    79:'2019 Term Loan',80:'2020 Term Loan',81:'2020 Term Loan - LO',
    82:'ESL Note',83:'HPS RE Loan',
    85:'Inflows',86:'CITI Reimbursement',87:'Debt/Financing',
    88:'Supply Chain',89:'Asset Sales',90:'Rx',91:'Misc Inflows',
    92:'Payroll/Bens',93:'Merch (CIA)',94:'Merch (On Terms)',
    95:'Rent',96:'Logistics',97:'Advertising',
    98:'Other Non-Merch Disbursements',99:'RiskMgt/Ins',100:'Taxes',
    101:'RealEstate Transactions',102:'HW Reserves & Deposits',
    103:'PropCo 2025',104:'WU',105:'P-Card',106:'SHI',
    107:'Collateral',108:'Interest',109:'Fees',110:'PropCo',
    111:'Cstc',112:'Debt Repayment',113:'Misc',
    115:'SG&A',116:'Home Services',117:'Store Closures',
    119:'Total Operating Receipts',
    120:'Merch/Non-Merch Disbursements',
    121:'Segregated Account Available Cash',
    122:'Inflows Total',123:'Disbursements Total',124:'Financing',
    125:'Cash Flow + Financing',
    126:'Unavail Cash Lookup from Model',
    127:'Change in BAML/PNC Cash Collateralization',
    128:'Excess Availability+ Need',129:'Junior DIP Borrowing Need',
    130:'Total EA',131:'Change in PNC/ Cash Collateralization',
    137:'Cash Flows no Asset Sales',138:'Adj Available Cash',
    139:'Adj Cash Requirement',
    143:'Property taxes',144:'Utilities',145:'FACILITIES/CAPEX',
    146:'RE MISC',147:'RE TOAL',
}

# Row 1: col index reference
for col,lbl in CF.items():
    c=ws.cell(1,col,col)
    c.font=body(sz=7,color=DGREY); c.fill=fill(LGREY)

# Row 2: headers
for col,lbl in CF.items():
    ph(ws,2,col,lbl,bg=NAVY,sz=8)

# Row 3: grand-total spacer (single space label)
ws.cell(3,1,' ').fill=fill(AMBER)
ws.cell(3,1).font=body()

# Copy data rows from source (source rows 4+ → our rows 4+)
# Source col mapping (openpyxl 1-based) → our col
SRC_TO_DST = {
    # Date+context
    1:1, 2:2, 3:3, 4:4,
    # Cash flows
    5:5, 6:6, 7:7, 10:10, 11:11, 12:12, 13:13, 14:14, 15:15,
    # Equity/surplus
    16:16, 17:17, 18:18, 19:19, 20:20, 21:21, 22:22,
    24:24, 25:25, 26:26, 27:27,
    # Term loans (tranches)
    28:28,29:29,30:30,31:31,32:32,33:33,34:34,35:35,36:36,37:37,
    38:38,39:39,40:40,41:41,42:42,43:43,44:44,45:45,46:46,47:47,
    48:48,49:49,50:50,
    # Real estate / debt cols (source may be offset — use 1:1 for available)
    # Bohemia onward: source BF=58, BG=59 etc
    58:58,59:59,60:60,61:61,62:62,63:63,
    66:66,67:67,68:68,
    73:73,74:74,75:75,76:76,77:77,78:78,
    79:79,80:80,81:81,82:82,83:83,
    # Inflows / disbursements
    85:85,86:86,87:87,88:88,89:89,90:90,91:91,
    92:92,93:93,94:94,95:95,96:96,97:97,98:98,99:99,100:100,
    101:101,102:102,103:103,104:104,105:105,106:106,
    107:107,108:108,109:109,110:110,111:111,112:112,113:113,
    # Summaries
    122:122,123:123,124:124,125:125,126:126,130:130,
    137:137,138:138,139:139,143:143,144:144,145:145,147:147,
}
FMT_NUM = '#,##0.000;(#,##0.000);"-"'
FMT_DATE = 'mmm d, yyyy'

MAX_R = min(ws_s.max_row, 1720)
for sr in range(4, MAX_R+1):
    date_v = ws_s.cell(sr,1).value
    if date_v is None: continue
    # determine row bg
    is_monthly = isinstance(date_v, str)
    bg = AMBER if is_monthly else WHITE
    ws.cell(sr,1).fill=fill(bg)
    for sc,dc in SRC_TO_DST.items():
        v = ws_s.cell(sr,sc).value
        if v is None: continue
        c=ws.cell(sr,dc)
        c.value=v; c.fill=fill(bg); c.border=thin()
        if sc==1 and not is_monthly:
            c.number_format=FMT_DATE
            c.alignment=lft()
        elif isinstance(v,(int,float)) and sc>=5:
            c.number_format=FMT_NUM
            c.alignment=rgt()
        c.font=body(bold=is_monthly)

ws.column_dimensions['A'].width=12
for ci in range(2,148):
    ws.column_dimensions[get_column_letter(ci)].width=9
ws.freeze_panes='E3'
ws.row_dimensions[1].height=12
ws.row_dimensions[2].height=32
print(f"  Cash Flow: {ws.max_row} rows, {ws.max_column} cols")

# ════════════════════════════════════════════════════════════════════════
# 2. LC FORECAST CHANGES  (visible)
# ════════════════════════════════════════════════════════════════════════
print("Building LC Forecast Changes …")
ws2=wb.create_sheet('LC Forecast Changes')
copy_sheet(wb_s['LC Forecast Changes'],ws2)
# Section labels
for row,lbl in {5:'Completed Prior:',71:'Completed 2022:',87:'Completed 2023:',
                123:'Completed 2024:',155:'Completed 2025:',171:'Completed 2026:'}.items():
    c=ws2.cell(row,11,lbl)
    c.font=body(bold=True,color='1F3864'); c.fill=fill(LBLUE)
# Zone A headers
ph(ws2,2,13,'Current',bg=LBLUE,fg='1F3864',sz=9)
ph(ws2,2,17,'New',bg=LBLUE,fg='1F3864',sz=9)
for col,lbl in {11:'Beneficiary',12:'LC #',13:'Facility',14:'Face Amount',
                15:'Collateral Return/Draw Date',17:'Facility',18:'Face Amount',
                19:'Collateral Post Date',21:'NET CHANGE IN COLLATERAL BETTER/(WORSE)',
                22:'PENDING CHANGE IN COLLATERAL BETTER/(WORSE)',
                28:'Comments'}.items():
    ph(ws2,3,col,lbl,bg=NAVY,sz=8)
ws2.column_dimensions['A'].width=8
for ci in range(11,29):
    ws2.column_dimensions[get_column_letter(ci)].width=14
ws2.freeze_panes='K4'
print(f"  LC Forecast Changes: {ws2.max_row} rows")

# ════════════════════════════════════════════════════════════════════════
# 3. TRAPPED CASH  (visible)
# ════════════════════════════════════════════════════════════════════════
print("Building Trapped Cash …")
ws3=wb.create_sheet('Trapped Cash')
copy_sheet(wb_s['Trapped Cash'],ws3)
for col,lbl in {3:'Cash — Other than Available Cash',6:'Restricted Cash',
                8:'Accounts Receivable / Other Assets',10:'Cash and Cash Equivalents',
                12:'Owner',13:'Collection Date',14:'Comments'}.items():
    ph(ws3,2,col,lbl,bg=NAVY,sz=8)
ws3.column_dimensions['A'].width=6
ws3.column_dimensions['C'].width=44
for ci in [4,6,8,10]: ws3.column_dimensions[get_column_letter(ci)].width=18
ws3.freeze_panes='C3'
print(f"  Trapped Cash: {ws3.max_row} rows")

# ════════════════════════════════════════════════════════════════════════
# 4. UBS RESERVE  (visible)
# ════════════════════════════════════════════════════════════════════════
print("Building UBS reserve …")
ws4=wb.create_sheet('UBS reserve')
copy_sheet(wb_s['UBS reserve'],ws4)
for col,lbl in {2:'Maturity',3:'Types',4:'Lender',5:'Property Count',
                6:'Loan Amount',7:'Lit Value',8:'Loan/Lit',9:'Pricing',
                11:'Interest Expense',12:'Working Capital',13:'ADA & Life Safety',
                14:'Deferred Maintenance',15:'General Reserve',
                16:'Insurance Holdback',17:'Total Reserves'}.items():
    ph(ws4,2,col,lbl,bg=NAVY,sz=8)
ws4.cell(1,1,'UBS Loan Reserve Detail').font=body(bold=True,color='1F3864',sz=11)
ws4.column_dimensions['A'].width=6
for ci in range(2,18): ws4.column_dimensions[get_column_letter(ci)].width=15
print(f"  UBS reserve: {ws4.max_row} rows")

# ════════════════════════════════════════════════════════════════════════
# 5. 2020 TERM LOANS  (visible)
# ════════════════════════════════════════════════════════════════════════
print("Building 2020 Term Loans …")
ws5=wb.create_sheet('2020 Term Loans')
copy_sheet(wb_s['2020 Term Loans'],ws5)
ws5.cell(1,3,'2020 Term Loan & 2021 Term Loan').font=body(bold=True,color='1F3864',sz=11)
TL_HDRS=['Tranche 1 Principal','Tranche 2 Principal','SKIPPED',
         'Tranche 3 Principal','Tranche 4 Principal','Cyrus Term Loan',
         'Tranche 6 Principal','Tranche 7 Principal','Tranche 8 Principal',
         'Tranche 9 Principal','Tranche 10 Principal','Tranche 11 Principal',
         'Cyrus $55M Term Loan','ESL $15M Term Loan','ESL $16.5M Term Loan',
         'Cyrus $16.5M Term Loan','ESL $27M Term Loan','ESL $50M Term Loan',
         'Cyrus $25M Term Loan','ESL $70M Term Loan','Cyrus $8M Term Loan',
         'ESL $175M Term Loan (T22)','Cyrus 7th Term Loan','ESL 15th Term Loan',
         'Sept. 2025 Term Loans','Cyrus Eighth Incremental Term Loans']
for i,h in enumerate(TL_HDRS,4): ph(ws5,3,i,h,bg=NAVY,sz=8)
ws5.column_dimensions['C'].width=36
for ci in range(4,31): ws5.column_dimensions[get_column_letter(ci)].width=12
ws5.freeze_panes='D4'
print(f"  2020 Term Loans: {ws5.max_row} rows")

# ════════════════════════════════════════════════════════════════════════
# 6. INFLOWS FORECASTING  (visible)
# ════════════════════════════════════════════════════════════════════════
print("Building Inflows Forecasting …")
ws6=wb.create_sheet('Inflows Forecasting')
ws6_s=wb_s['Inflows Forecasting']
copy_sheet(ws6_s,ws6,max_c=60)  # cap at 60 cols for performance
ws6.cell(1,2,'Current Actual/Forecast').font=body(bold=True,color='1F3864')
ws6.cell(1,2).fill=fill(LBLUE)
for i,h in enumerate(['Wk','Sears Stores','Kmart Stores','Home Services',
                       'KCD Wholesale','KCD (Royalty)','Online',
                       'Service Live Summary','Tenant Income','Total'],1):
    ph(ws6,2,i,h,bg=NAVY,sz=8)
ws6.column_dimensions['A'].width=6
for ci in range(2,12): ws6.column_dimensions[get_column_letter(ci)].width=14
ws6.freeze_panes='B3'
print(f"  Inflows Forecasting: {ws6.max_row} rows")

# ════════════════════════════════════════════════════════════════════════
# 7. INFLOWS ACTUALS  (visible)
# ════════════════════════════════════════════════════════════════════════
print("Building Inflows Actuals …")
ws7=wb.create_sheet('Inflows Actuals')
copy_sheet(wb_s['Inflows Actuals'],ws7,max_c=50)
for col,lbl in {1:'DATE',2:'Sears Stores',3:'Kmart Stores',4:'Home Services',
                5:'KCD',6:'Tenant Income',7:'Auto Centers',
                11:'Service Live Summary',12:'HTS',14:'Costco',
                15:'Supply Chain',16:'CITI Reimbursement',17:'CCHS',
                18:'Asset Sales',19:'Rx',20:'PropCo/Lease OpCo',21:'KES',
                22:'Sears Mexico (KCD)',23:'UBS Reserve',24:'SHI Dividend',
                25:'Debt/Financing',26:'Misc Inflows',27:'Daily Total'}.items():
    ph(ws7,1,col,lbl,bg=NAVY,sz=8)
ws7.column_dimensions['A'].width=12
for ci in range(2,48): ws7.column_dimensions[get_column_letter(ci)].width=11
ws7.freeze_panes='B2'
print(f"  Inflows Actuals: {ws7.max_row} rows")

# ════════════════════════════════════════════════════════════════════════
# 8. INFLOWS DETAIL  (visible) — formula-driven from FY File
# ════════════════════════════════════════════════════════════════════════
print("Building Inflows Detail …")
ws8=wb.create_sheet('Inflows Detail')
ws8_s=wb_s['Inflows Detail']
copy_sheet(ws8_s,ws8)

# Headers: 4-row structure matching actual file
for col,lbl in {1:'DATE',2:'Sears Stores',3:'Kmart Stores',4:'Home Services',
                5:'KCD (Royalty)',6:'Tenant Income',7:'Online',
                13:'Fidem',14:'NEXT2',15:'Supply Chain',16:'CITI Reimburse',
                17:'CCHS',18:'Asset Sales',19:'PH 2.0',20:'Lease OpCo',
                21:'KES',22:'Sears Mexico',23:'UBS Reserve',
                24:'SHI Dividend',25:'New Debt',26:'Misc Inflows'}.items():
    ph(ws8,1,col,lbl,bg=NAVY,sz=8)
ph(ws8,1,5,'KCD Wholesale',bg=LBLUE,fg='1F3864',sz=8)  # row 2 sub-label
for col,lbl in {2:'Sears Stores',3:'Kmart Stores',4:'Home Services',5:'KCD',
                6:'Tenant Income',7:'Online',8:'SHS Assurant Profit Sharing',
                13:'Fidem/Aress',14:'Risk/Insurance',15:'Supply Chain',
                16:'CITI Reimbursement',17:'Cross Country Home Services',
                18:'Asset Sales',19:'PropCo 2025',20:'PropCo/Lease OpCo',
                21:'KES',22:'Sears Mexico (KCD)',23:'UBS Reserve',
                24:'SHI Dividend',25:'Debt/Financing',26:'Misc Inflows',
                27:'Daily Total'}.items():
    ph(ws8,4,col,lbl,bg=LBLUE,fg='1F3864',sz=8)
ws8.column_dimensions['A'].width=12
for ci in range(2,30): ws8.column_dimensions[get_column_letter(ci)].width=12
ws8.freeze_panes='B5'
print(f"  Inflows Detail: {ws8.max_row} rows")

# ════════════════════════════════════════════════════════════════════════
# 9. DISBURSEMENT ACTUALS  (visible)
# ════════════════════════════════════════════════════════════════════════
print("Building Disbursement Actuals …")
ws9=wb.create_sheet('Disbursement Actuals')
copy_sheet(wb_s['Disbursement Actuals'],ws9,max_c=95)

# 5-row header structure
ph(ws9,1,1,'DATE',bg=NAVY,sz=8)
for col,lbl in {2:'Finance',3:'FinancialSvcs',4:'HoldingsCo',5:'Home Services',
                6:'HR',7:'KCD',8:'Legal',9:'Member Tech',10:'Monark',
                11:'MSO',12:'Real Estate',13:'Retail',14:'Retail Online',
                15:'Service Live',16:'SYWR',17:'Supply Chain',
                18:'PYRL/BENS TOTAL'}.items():
    ph(ws9,1,col,lbl,bg=NAVY,sz=7)
for col,lbl in {2:'PAYROLL',19:'HOME SERVICES',22:'Total Retail - Stores',
                25:'Online Business Unit',28:'KCD',31:'HTS',
                34:'Merch TOTALS'}.items():
    ph(ws9,3,col,lbl,bg=LBLUE,fg='1F3864',sz=8)
for col,lbl in {34:'Merch CIA',35:'Merch On Terms',36:'MERCH - TOTAL',
                37:'Google',38:'Other Advertising',39:'Total Advertising',
                40:'Logistics - UPS',41:'SHS Logistics',43:'Logistics - Other',
                44:'Total Logistics',51:'RiskMgt/Ins',52:'Taxes',53:'Western Union',
                54:'Rent',57:'Outside Serv/Assoc Exp',63:'Other Services',
                69:'Facilities',73:'SHI',79:'Total P-Card',
                82:'PropCo 2025',83:'PropCo',84:'RealEstate Transactions',
                85:'Collateral',86:'Misc',87:'Debt Repayment',
                88:'Cash Interest',89:'Cash Fees',90:'Daily Total'}.items():
    ph(ws9,5,col,lbl,bg=LBLUE,fg='1F3864',sz=7)
ws9.column_dimensions['A'].width=12
for ci in range(2,96): ws9.column_dimensions[get_column_letter(ci)].width=9
ws9.freeze_panes='B6'
print(f"  Disbursement Actuals: {ws9.max_row} rows")

# ════════════════════════════════════════════════════════════════════════
# 10. DISBURSEMENT DETAIL  (visible)
# ════════════════════════════════════════════════════════════════════════
print("Building Disbursement Detail …")
ws10=wb.create_sheet('Disbursement Detail')
copy_sheet(wb_s['Disbursement Detail'],ws10)

# Same 5-row header structure as Disbursement Actuals
ph(ws10,1,1,'DATE',bg=NAVY,sz=8)
for col,lbl in {2:'Finance',3:'FinancialSvcs',4:'HoldingsCo',5:'Home Services',
                6:'HR',7:'KCD',8:'Legal',9:'Member Tech',10:'Monark',
                11:'MSO',12:'Real Estate',13:'Retail',14:'Retail Online',
                15:'Service Live',16:'SYWR',17:'Supply Chain',
                18:'PYRL/BENS TOTAL'}.items():
    ph(ws10,1,col,lbl,bg=NAVY,sz=7)
for col,lbl in {2:'PAYROLL',19:'HOME SERVICES',22:'Total Retail - Stores',
                25:'Online Business Unit',28:'KCD',31:'HTS',
                34:'Merch TOTALS'}.items():
    ph(ws10,3,col,lbl,bg=LBLUE,fg='1F3864',sz=8)
for col,lbl in {34:'Merch CIA',35:'Merch On Terms',36:'MERCH - TOTAL',
                37:'Google',38:'Other Advertising',39:'Total Advertising',
                40:'Logistics - UPS',41:'SHS Logistics',43:'Logistics - Other',
                44:'Total Logistics',51:'RiskMgt/Ins',52:'Taxes',53:'Western Union',
                54:'Rent',57:'Outside Serv/Assoc Exp',63:'Other Services',
                69:'Facilities',73:'SHI',79:'Total P-Card',
                82:'PropCo 2025',83:'PropCo',84:'RealEstate Transactions',
                85:'Collateral',86:'Misc',87:'Debt Repayment',
                88:'Cash Interest',89:'Cash Fees',90:'Daily Total'}.items():
    ph(ws10,5,col,lbl,bg=LBLUE,fg='1F3864',sz=7)
ws10.column_dimensions['A'].width=12
for ci in range(2,102): ws10.column_dimensions[get_column_letter(ci)].width=9
ws10.freeze_panes='B6'
print(f"  Disbursement Detail: {ws10.max_row} rows")

# ════════════════════════════════════════════════════════════════════════
# HIDDEN SHEETS (8) — data backends for formula resolution
# ════════════════════════════════════════════════════════════════════════
hidden_map = [
    ('RE & Debt init with Manteno', 'RE & Debt init with Manteno'),
    ('KES & KCD Cash Flow',         'KES & KCD Cash Flow'),
    ('RC',                           'RC'),
    ('FY File',                      'FY File'),
    ('Merch',                        'MERCH'),
    ('Non-Merch',                    'Non-Merch'),
    ('Payroll',                      'Payroll'),
]

for src_name, dst_name in hidden_map:
    if src_name in wb_s.sheetnames:
        print(f"Building hidden: {dst_name} …")
        ws_h = wb.create_sheet(dst_name)
        copy_sheet(wb_s[src_name], ws_h,
                   max_r=min(wb_s[src_name].max_row, 500),
                   max_c=min(wb_s[src_name].max_column, 120))
        ws_h.sheet_state = 'veryHidden'
    else:
        print(f"  SKIPPED (not in source): {src_name}")

# SOP — recreate as hidden shell
print("Building hidden: SOP …")
ws_sop = wb.create_sheet('SOP')
ws_sop.cell(1,1,'Tab').font=body(bold=True,color='1F3864')
ws_sop.cell(1,2,'Update Instructions').font=body(bold=True,color='1F3864')
ws_sop.cell(1,3,'Source / Contact').font=body(bold=True,color='1F3864')
sop_rows = [
    ('Cash Flow','Manual changes to Cols AB–BZ only (Cap Table/term loan balances). All other cols formula-driven.','Cols AB–BY per Jon Goodin / CAP table. Col BZ per James Coutre.'),
    ('KES & KCD Cash Flow','Copy/paste from Brands finance team.',None),
    ('LC Forecast Changes','Per Jon Goodin.',None),
    ('Trapped Cash — PNC LCs','PNC Bank Account 4803431873',None),
    ('Trapped Cash — UBS LCs','BNY Mellon Account 507032',None),
    ('Trapped Cash — PNC Cash Mgmt','PNC Account 4645228406',None),
    ('Trapped Cash — Citi Pcard','Citibank Account 40784284',None),
    ('Trapped Cash — Discover Holdback','Per Brett Bassett',None),
    ('Trapped Cash — FDC Holdback','Per Brett Bassett',None),
    ('Trapped Cash — Utility Deposits','GL Account 14175',None),
    ('Trapped Cash — Landlord Advances','GL Account 14180',None),
    ('Trapped Cash — RE Segregated Acct','Per Real Estate finance team',None),
    ('Trapped Cash — Innovel Enviro Holdback','Per Luke Valentino',None),
    ('Trapped Cash — RE Escrows','Per Real Estate finance team',None),
    ('Trapped Cash — Safelite (Hackensack)','Per Real Estate finance team',None),
    ('Trapped Cash — Drawn LC restricted','Per Jon Goodin',None),
    ('Trapped Cash — Hackensack','Per Real Estate finance team',None),
    ('Trapped Cash — Bohemia','Per Real Estate finance team',None),
    ('Trapped Cash — HW Reserves','Sum of PNC warranty reserve bank accounts',None),
    ('UBS reserve','Per Zach Straebel.',None),
    ('2020 Term Loans','Per Jon Goodin.',None),
    ('Inflows Forecasting','Each Monday: update ABZ1 to prior Friday; copy/paste B:J → ABZ:ACG.',None),
    ('Inflows Actuals','From cash position Daily Detail file, named range INFLOWS.',None),
    ('Inflows Detail','All formula driven.',None),
    ('Disbursement Actuals','All formula driven.',None),
    ('Disbursement Detail','All formula driven.',None),
    ('FY File','HIDDEN — actual/forecast file for year.',None),
    ('MERCH','HIDDEN — copy/paste from James Coutre.',None),
    ('Non-Merch','HIDDEN — copy/paste from Adharvana.',None),
    ('Payroll','HIDDEN — copy/paste from Payroll finance team.',None),
]
for i,(tab,instr,src_note) in enumerate(sop_rows,2):
    ws_sop.cell(i,1,tab).font=body(bold=True,color='1F3864')
    ws_sop.cell(i,2,instr).font=body()
    if src_note: ws_sop.cell(i,3,src_note).font=body()
ws_sop.column_dimensions['A'].width=30
ws_sop.column_dimensions['B'].width=60
ws_sop.column_dimensions['C'].width=36
ws_sop.sheet_state='hidden'  # SOP is hidden (not veryHidden — can be unhidden by user)

# ════════════════════════════════════════════════════════════════════════
# Reorder: VISIBLE first, HIDDEN last (matches actual file tab order)
# ════════════════════════════════════════════════════════════════════════
TAB_ORDER = [
    # VISIBLE (10)
    'Cash Flow','LC Forecast Changes','Trapped Cash','UBS reserve',
    '2020 Term Loans','Inflows Forecasting','Inflows Actuals',
    'Inflows Detail','Disbursement Actuals','Disbursement Detail',
    # HIDDEN (8)
    'RE & Debt init with Manteno','KES & KCD Cash Flow','SOP','RC',
    'FY File','MERCH','Non-Merch','Payroll',
]
present = {s:wb[s] for s in wb.sheetnames}
wb._sheets = [present[t] for t in TAB_ORDER if t in present]

# ── Verify ────────────────────────────────────────────────────────────────
print("\nFinal tab inventory:")
for i,name in enumerate(wb.sheetnames,1):
    state=wb[name].sheet_state
    icon='👁 ' if state=='visible' else '🔒'
    print(f"  {icon} {i:2}. {name:<42} [{state}]")

wb.save(DST)
sz=os.path.getsize(DST)
print(f"\nSaved → {DST}")
print(f"Size: {sz:,} bytes ({sz/1024/1024:.1f} MB)")
