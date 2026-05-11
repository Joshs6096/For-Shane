"""6-Month Curate Coins Forecast — single realistic scenario with $1,000 board seed."""

import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SPREADSHEET = '/Users/josh/.openclaw/media/inbound/a49b353f-c14f-44a8-a5e2-f8a583d8e4d9.xlsx'
SOLD = json.load(open('/tmp/sold.json'))

# ---- Single realistic scenario ----
SPOT = 81.88
EBAY_FEE_PCT = 0.13
SHIPPING_PER_ITEM = 2.0
SUPPLIES_PER_ITEM = 0.50
WEEKLY_BUDGET = 120.0
MONTHS = 6
BOARD_SEED = 1000.0
SALES_PER_MONTH = 18           # realistic with seed
AVG_PRICE = 125
REINVEST_PCT = 0.40
RAMP = [0.80, 0.95, 1.00, 1.00, 1.00, 1.00]  # M1 ramp because seed deploys M1

def f(x):
    if x is None: return 0.0
    s = str(x).replace('$','').replace(',','').strip()
    try: return float(s)
    except: return 0.0

# Read collection
wb_in = openpyxl.load_workbook(SPREADSHEET, data_only=True)
ws_in = wb_in.active
header = [c.value for c in ws_in[1]]
inv = [dict(zip(header, row)) for row in ws_in.iter_rows(min_row=2, values_only=True) if any(c is not None for c in row)]

# Realized
sold_total_gross = sum(f(s['price']) * f(s['qty_sold']) for s in SOLD)
sold_count = sum(int(f(s['qty_sold']) or 1) for s in SOLD)
fees_total = sold_total_gross * EBAY_FEE_PCT
ship_supply_total = sold_count * (SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM)
sold_net = sold_total_gross - fees_total - ship_supply_total
sold_avg = sold_total_gross / max(sold_count, 1)

# Inventory segmentation
NUMISMATIC_CATS = {'Proof Set', 'Silver Proof Set', 'Mint Set', 'Ancient Coin', 'Indian Head Cent',
                   'Indian Head Nickel', 'Liberty Head Nickel', 'Buffalo Nickel', 'Standing Liberty Quarter',
                   'Barber Coin', 'Confederate Money', 'Silver Certificate', 'Currency',
                   'Shipwreck', 'Medal', 'Diamond', 'Picture'}
core_value, core_cost, core_count = 0, 0, 0
flip_value, flip_cost, flip_count = 0, 0, 0
for r in inv:
    cat = r.get('Category') or 'Unknown'
    cost = f(r.get('Cost Paid'))
    high = f(r.get('High Value'))
    qty = f(r.get('Qty'))
    if cost == 0 or cat in NUMISMATIC_CATS:
        core_value += high; core_cost += cost; core_count += qty
    else:
        flip_value += high; flip_cost += cost; flip_count += qty
flip_avg_cost = flip_cost / max(flip_count, 1)
silver_buy_budget = WEEKLY_BUDGET * 4.33

# Compute monthly projection
def project():
    months = []
    for i, r in enumerate(RAMP):
        m_sales = SALES_PER_MONTH * r
        m_gross = m_sales * AVG_PRICE
        m_fees = m_gross * EBAY_FEE_PCT
        m_ship = m_sales * (SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM)
        m_net = m_gross - m_fees - m_ship
        m_reinvest = m_net * REINVEST_PCT
        m_seed = BOARD_SEED if i == 0 else 0
        m_cash = m_net - m_reinvest - silver_buy_budget
        months.append({
            'sales': m_sales, 'gross': m_gross, 'fees': m_fees, 'ship': m_ship,
            'net': m_net, 'reinvest': m_reinvest, 'silver_buy': silver_buy_budget,
            'seed_in': m_seed, 'cash': m_cash
        })
    return months

monthly = project()
total_sales = sum(m['sales'] for m in monthly)
total_gross = sum(m['gross'] for m in monthly)
total_net = sum(m['net'] for m in monthly)
total_reinvest = sum(m['reinvest'] for m in monthly)
total_silver_buy = sum(m['silver_buy'] for m in monthly)
total_cash = sum(m['cash'] for m in monthly)

# Build workbook
wb = openpyxl.Workbook()
HDR_FILL = PatternFill('solid', fgColor='2C3E50')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='34495E')
SECTION_FONT = Font(bold=True, color='FFFFFF', size=12)
SEED_FILL = PatternFill('solid', fgColor='F39C12')
SEED_FONT = Font(bold=True, color='FFFFFF')
TOTAL_FILL = PatternFill('solid', fgColor='ECF0F1')
TOTAL_FONT = Font(bold=True)
GOOD_FILL = PatternFill('solid', fgColor='D5F5E3')
BAD_FILL = PatternFill('solid', fgColor='FADBD8')
THIN = Side(border_style='thin', color='BDC3C7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_row(ws, row, vals, font=None, fill=None, border=True):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        if font: c.font = font
        if fill: c.fill = fill
        if border: c.border = BORDER

# ---- Tab 1: Summary ----
ws = wb.active
ws.title = 'Summary'
ws.column_dimensions['A'].width = 38
for c in 'BCDEFGH': ws.column_dimensions[c].width = 16

ws['A1'] = 'Curate Coins — 6-Month Outlook'
ws['A1'].font = Font(bold=True, size=16)
ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M MT")}  |  Spot ${SPOT}/oz  |  $1,000 board seed deployed Month 1'
ws['A2'].font = Font(italic=True, size=10, color='7F8C8D')

# Assumptions block
ws['A4'] = 'ASSUMPTIONS'
ws['A4'].fill = SECTION_FILL; ws['A4'].font = SECTION_FONT
ws.merge_cells('A4:C4')
assumptions = [
    ('Sales / month (steady-state)', f'{SALES_PER_MONTH}'),
    ('Avg sale price', f'${AVG_PRICE}'),
    ('eBay fees', f'{EBAY_FEE_PCT*100:.0f}%'),
    ('Ship + supplies per item', f'${SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM:.2f}'),
    ('Reinvestment from net', f'{REINVEST_PCT*100:.0f}%'),
    ('Silver-buy budget', f'${WEEKLY_BUDGET}/wk = ${silver_buy_budget:.0f}/mo'),
    ('M1 ramp (seed in hand)', '80% → 95% by M2 → 100% steady-state'),
    ('Board seed (Month 1)', f'${BOARD_SEED:,.0f}'),
]
for i, (k, v) in enumerate(assumptions, 5):
    ws.cell(row=i, column=1, value=k)
    ws.cell(row=i, column=2, value=v)

# Headline
ws['A14'] = '6-MONTH HEADLINE'
ws['A14'].fill = SECTION_FILL; ws['A14'].font = SECTION_FONT
ws.merge_cells('A14:C14')
headline = [
    ('Total sales (6mo)', total_sales, '0'),
    ('Gross revenue', total_gross, '$#,##0'),
    ('Net (after fees + ship)', total_net, '$#,##0'),
    ('Inventory reinvestment (40%)', -total_reinvest, '$#,##0'),
    ('Silver-buy spend ($120/wk)', -total_silver_buy, '$#,##0'),
    ('FREE CASH TO JOSH', total_cash, '$#,##0'),
]
for i, (k, v, fmt) in enumerate(headline, 15):
    ws.cell(row=i, column=1, value=k)
    c = ws.cell(row=i, column=2, value=v)
    c.number_format = fmt
    if 'CASH' in k:
        ws.cell(row=i, column=1).font = Font(bold=True)
        c.font = Font(bold=True)
        c.fill = GOOD_FILL if v > 0 else BAD_FILL

# Key takeaways
ws['A23'] = 'KEY TAKEAWAYS'
ws['A23'].fill = SECTION_FILL; ws['A23'].font = SECTION_FONT
ws.merge_cells('A23:C23')
takeaways = [
    "• Bottleneck = listing throughput, not capital or inventory.",
    f"• Current pace = 3 sales/mo at $152 avg. Target = {SALES_PER_MONTH} sales/mo at ${AVG_PRICE} avg.",
    f"• Free cash to Josh: ${total_cash:,.0f} over 6 months = ~${total_cash/6:,.0f}/mo run rate.",
    f"• Flippable inventory grows during the period — core (188oz, $32k profit potential) never touched.",
    f"• Seed buys ~26 flippable units at ${flip_avg_cost:.0f} avg → unlocks faster M1 ramp.",
    f"• Reality check: hitting {SALES_PER_MONTH}/mo needs ~45 rotating active listings (vs 5 today).",
]
for i, t in enumerate(takeaways, 24):
    ws.cell(row=i, column=1, value=t)
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)

# ---- Tab 2: Cash Flow Month by Month ----
ws3 = wb.create_sheet('Cash Flow 6mo')
for col in range(1, 11): ws3.column_dimensions[get_column_letter(col)].width = 14
ws3.column_dimensions['A'].width = 30

ws3['A1'] = '6-Month Cash Flow Timeline'
ws3['A1'].font = Font(bold=True, size=14)

months_labels = []
start = datetime.now().replace(day=1)
for i in range(MONTHS):
    m = start.replace(month=((start.month + i - 1) % 12) + 1, year=start.year + ((start.month + i - 1) // 12))
    months_labels.append(m.strftime('%b %Y'))

write_row(ws3, 3, ['Line item'] + months_labels + ['6-Mo Total'], font=HDR_FONT, fill=HDR_FILL)

rows = [
    ('Sales / month', [int(m['sales']) for m in monthly], '#,##0'),
    ('Gross revenue', [m['gross'] for m in monthly], '$#,##0'),
    ('eBay fees (13%)', [-m['fees'] for m in monthly], '$#,##0'),
    ('Shipping + supplies', [-m['ship'] for m in monthly], '$#,##0'),
    ('NET REVENUE', [m['net'] for m in monthly], '$#,##0'),
    ('Inventory reinvestment (40%)', [-m['reinvest'] for m in monthly], '$#,##0'),
    ('Silver-buy budget ($120/wk)', [-m['silver_buy'] for m in monthly], '$#,##0'),
    ('BOARD SEED INFLOW', [m['seed_in'] for m in monthly], '$#,##0'),
    ('FREE CASH TO JOSH', [m['cash'] for m in monthly], '$#,##0'),
]
for ridx, (label, vals, fmt) in enumerate(rows, 4):
    ws3.cell(row=ridx, column=1, value=label)
    if 'NET' in label or 'CASH' in label or 'SEED' in label:
        ws3.cell(row=ridx, column=1).font = Font(bold=True)
    if 'SEED' in label:
        ws3.cell(row=ridx, column=1).fill = SEED_FILL
        ws3.cell(row=ridx, column=1).font = SEED_FONT
    for col_i, v in enumerate(vals, 2):
        c = ws3.cell(row=ridx, column=col_i, value=v)
        c.number_format = fmt
        if 'CASH' in label and isinstance(v, (int, float)):
            c.font = Font(bold=True)
            c.fill = GOOD_FILL if v > 0 else BAD_FILL
        if 'SEED' in label and v != 0:
            c.fill = SEED_FILL; c.font = SEED_FONT
    total = sum(vals)
    c = ws3.cell(row=ridx, column=8, value=total)
    c.number_format = fmt
    c.font = Font(bold=True)
    if 'CASH' in label:
        c.fill = GOOD_FILL if total > 0 else BAD_FILL

# Note about payouts
ws3['A14'] = 'Notes:'
ws3['A14'].font = Font(bold=True)
notes = [
    "• eBay payouts release ~7-10 days post-sale (delivery confirmation triggers funds release).",
    "• M1 cash is light — listing engine ramping, seed deployed but inventory still being photographed.",
    "• M2 onward = steady state at full cadence; this is where the real run rate shows.",
    "• 'Free cash to Josh' = net revenue − reinvestment − silver-buy budget. Available to extract or roll into reserves.",
]
for i, n in enumerate(notes, 15):
    ws3.cell(row=i, column=1, value=n)

# ---- Tab 3: Seed Deployment ----
ws4 = wb.create_sheet('Seed Deployment')
for col in 'ABCDEF': ws4.column_dimensions[col].width = 22

ws4['A1'] = '$1,000 Board Seed — Deployment Plan'
ws4['A1'].font = Font(bold=True, size=14)

ws4['A4'] = 'PROPOSED ALLOCATION'
ws4['A4'].fill = SECTION_FILL; ws4['A4'].font = SECTION_FONT
ws4.merge_cells('A4:F4')
write_row(ws4, 5, ['Bucket', 'Allocation', '$', 'Avg unit cost', 'Units', 'Expected hold'], font=HDR_FONT, fill=HDR_FILL)
deploy = [
    ('High-velocity bullion premium (NORFEDs, Libertads, ASE rolls)', '40%', 400, 100, 4, '30-45 days'),
    ('Common silver upgrades (90% halves, dimes, Walker fills)',      '30%', 300, 25, 12, '45-60 days'),
    ('Numismatic plays (semi-key dates, slabbed semi-keys)',          '20%', 200, 75, 2.5, '60-90 days'),
    ('Reserve / tactical buys (Whatnot, show floor)',                  '10%', 100, 50, 2, 'opportunistic'),
]
total_units = 0
for i, (b, pct, dollars, ucost, units, hold) in enumerate(deploy, 6):
    write_row(ws4, i, [b, pct, dollars, ucost, units, hold])
    ws4.cell(row=i, column=3).number_format = '$#,##0'
    ws4.cell(row=i, column=4).number_format = '$#,##0'
    total_units += units
write_row(ws4, 10, ['TOTAL', '100%', BOARD_SEED, '', round(total_units, 1), ''], font=TOTAL_FONT, fill=TOTAL_FILL)
ws4.cell(row=10, column=3).number_format = '$#,##0'

ws4['A12'] = 'EXPECTED RETURN ON SEED'
ws4['A12'].fill = SECTION_FILL; ws4['A12'].font = SECTION_FONT
ws4.merge_cells('A12:F12')
write_row(ws4, 13, ['Bucket', 'Investment', 'Margin target', 'Expected gross', 'Expected net', 'ROI'], font=HDR_FONT, fill=HDR_FILL)
returns = [
    ('Bullion premium', 400, '15%', 460, 400, '0%'),
    ('Common silver upgrades', 300, '40%', 420, 365, '22%'),
    ('Numismatic plays', 200, '60%', 320, 278, '39%'),
    ('Tactical reserve', 100, '50%', 150, 130, '30%'),
]
total_inv = total_gross_seed = total_net_seed = 0
for i, (b, inv_amt, mgn, gross, net, roi) in enumerate(returns, 14):
    write_row(ws4, i, [b, inv_amt, mgn, gross, net, roi])
    ws4.cell(row=i, column=2).number_format = '$#,##0'
    ws4.cell(row=i, column=4).number_format = '$#,##0'
    ws4.cell(row=i, column=5).number_format = '$#,##0'
    total_inv += inv_amt; total_gross_seed += gross; total_net_seed += net
total_roi = (total_net_seed - total_inv) / total_inv * 100
write_row(ws4, 18, ['TOTAL', total_inv, '', total_gross_seed, total_net_seed, f'{total_roi:.0f}%'],
          font=TOTAL_FONT, fill=TOTAL_FILL)
ws4.cell(row=18, column=2).number_format = '$#,##0'
ws4.cell(row=18, column=4).number_format = '$#,##0'
ws4.cell(row=18, column=5).number_format = '$#,##0'
ws4.cell(row=18, column=5).fill = GOOD_FILL

ws4['A20'] = 'NOTES'
ws4['A20'].fill = SECTION_FILL; ws4['A20'].font = SECTION_FONT
ws4.merge_cells('A20:F20')
notes = [
    "• Seed treated as company equity — not a Josh-payback obligation.",
    "• Catalytic value > direct ROI: $1k accelerates listing cadence and unlocks the steady-state run rate.",
    "• All buys must be Sheldon-vetted per CUR-42 before execution.",
    "• Bullion buys >$200/unit need board check-in.",
    "• Track seed deployment as a separate cost center for clean Month-6 attribution.",
]
for i, n in enumerate(notes, 21):
    ws4.cell(row=i, column=1, value=n)
    ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

# ---- Tab 4: Inventory Plan ----
ws5 = wb.create_sheet('Inventory Plan')
for col in 'ABCDEF': ws5.column_dimensions[col].width = 28

ws5['A1'] = 'Inventory Replenishment'
ws5['A1'].font = Font(bold=True, size=14)

ws5['A3'] = 'CURRENT STATE'
ws5['A3'].fill = SECTION_FILL; ws5['A3'].font = SECTION_FONT
ws5.merge_cells('A3:C3')
write_row(ws5, 4, ['Segment', 'Units', 'Cost basis', 'High value'], font=HDR_FONT, fill=HDR_FILL)
write_row(ws5, 5, ['Core (inherited + numismatic — DO NOT TOUCH)', core_count, core_cost, core_value])
write_row(ws5, 6, ['Flippable bullion + common silver', flip_count, flip_cost, flip_value])
write_row(ws5, 7, ['TOTAL', core_count + flip_count, core_cost + flip_cost, core_value + flip_value], font=TOTAL_FONT, fill=TOTAL_FILL)
for r in [5,6,7]:
    for col in [3,4]:
        ws5.cell(row=r, column=col).number_format = '$#,##0'

ws5['A9'] = 'BURN-DOWN MATH'
ws5['A9'].fill = SECTION_FILL; ws5['A9'].font = SECTION_FONT
ws5.merge_cells('A9:C9')
net_per_mo = total_net / 6
replenish_need = SALES_PER_MONTH * flip_avg_cost
replenish_avail = net_per_mo * REINVEST_PCT + silver_buy_budget
shortfall_surplus = replenish_avail - replenish_need

rows = [
    ('Flippable units on hand', flip_count, '0'),
    (f'Sales / month (steady-state)', SALES_PER_MONTH, '0'),
    ('Months runway with zero replenishment', round(flip_count / SALES_PER_MONTH, 1), '0.0'),
    ('Avg cost per flippable unit', flip_avg_cost, '$#,##0.00'),
    ('Replenishment $/mo needed', replenish_need, '$#,##0'),
    ('Replenishment $/mo from 40% reinvest', net_per_mo * REINVEST_PCT, '$#,##0'),
    ('Silver-buy budget ($120/wk)', silver_buy_budget, '$#,##0'),
    ('Total replenishment available', replenish_avail, '$#,##0'),
    ('SURPLUS / SHORTFALL', shortfall_surplus, '$#,##0'),
]
for i, (k, v, fmt) in enumerate(rows, 10):
    ws5.cell(row=i, column=1, value=k)
    c = ws5.cell(row=i, column=2, value=v)
    c.number_format = fmt
    if 'SURPLUS' in k:
        ws5.cell(row=i, column=1).font = Font(bold=True)
        c.font = Font(bold=True)
        c.fill = GOOD_FILL if v > 0 else BAD_FILL

ws5['A20'] = 'GUARDRAILS'
ws5['A20'].fill = SECTION_FILL; ws5['A20'].font = SECTION_FONT
ws5.merge_cells('A20:C20')
guards = [
    "• Core (inherited + numismatic premium) is NEVER liquidated. That's the appreciation engine.",
    "• If flippable units drop below 30% of total inventory, pause sales and reset listing strategy.",
    "• Lean replenishment toward bullion premium plays (NORFEDs, Libertads, ASE rolls) — fastest turn.",
    "• Numismatic plays (semi-key dates) for higher margin per unit but slower turnover.",
    "• All buys Sheldon-vetted (CUR-42) before purchase.",
]
for i, g in enumerate(guards, 21):
    ws5.cell(row=i, column=1, value=g)
    ws5.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

# ---- Tab 5: Inventory by Category ----
ws6 = wb.create_sheet('Inventory by Category')
for col in 'ABCDEFGHIJ': ws6.column_dimensions[col].width = 16
ws6.column_dimensions['A'].width = 28

ws6['A1'] = 'Inventory Detail by Category'
ws6['A1'].font = Font(bold=True, size=14)
write_row(ws6, 3, ['Category', 'Segment', 'Entries', 'Units', 'Silver oz', 'Cost', 'Low value', 'High value', 'Profit ($)', 'ROI %'], font=HDR_FONT, fill=HDR_FILL)

from collections import defaultdict
agg = defaultdict(lambda: {'entries':0,'qty':0,'oz':0,'cost':0,'low':0,'high':0,'profit':0})
for r in inv:
    cat = r.get('Category') or 'Unknown'
    a = agg[cat]
    a['entries'] += 1
    a['qty'] += f(r.get('Qty'))
    a['oz'] += f(r.get('Total Silver Oz'))
    a['cost'] += f(r.get('Cost Paid'))
    a['low'] += f(r.get('Low Value'))
    a['high'] += f(r.get('High Value'))
    a['profit'] += f(r.get('Profit ($)'))

row_i = 4
for cat in sorted(agg.keys()):
    a = agg[cat]
    if a['cost'] == 0: seg = 'Core (inherited)'
    elif cat in NUMISMATIC_CATS: seg = 'Core (numismatic)'
    else: seg = 'Flippable'
    roi = (a['profit'] / a['cost'] * 100) if a['cost'] > 0 else 0
    write_row(ws6, row_i, [cat, seg, a['entries'], int(a['qty']), round(a['oz'],1), a['cost'], a['low'], a['high'], a['profit'], roi])
    for col in [6,7,8,9]: ws6.cell(row=row_i, column=col).number_format = '$#,##0'
    ws6.cell(row=row_i, column=10).number_format = '0.0%'
    if seg == 'Flippable': ws6.cell(row=row_i, column=2).fill = GOOD_FILL
    row_i += 1

# ---- Save (overwriting the old files) ----
out = '/Users/josh/.openclaw/workspace/curate-coins-6mo-forecast.xlsx'
wb.save(out)
print(f"Saved: {out}\n")

# Print headlines
print(f"=== REALISTIC SCENARIO (with $1,000 seed) ===")
print(f"  {SALES_PER_MONTH} sales/mo at ${AVG_PRICE} avg, 40% reinvestment")
print(f"  Total sales (6mo): {total_sales:.0f} units")
print(f"  Gross: ${total_gross:,.0f}")
print(f"  Net: ${total_net:,.0f}")
print(f"  Reinvest: ${total_reinvest:,.0f}")
print(f"  Silver-buy: ${total_silver_buy:,.0f}")
print(f"  FREE CASH TO JOSH: ${total_cash:,.0f}")
print(f"\n  Inventory: {flip_count} flippable units, {flip_count/SALES_PER_MONTH:.1f} months runway")
print(f"  Replenishment surplus: ${shortfall_surplus:,.0f}/mo")
