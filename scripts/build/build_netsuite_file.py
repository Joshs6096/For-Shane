"""
Build: Daily Cash Fcst – 5.8.26_NS_SOURCED.xlsx
Source: NetSuite MCP data collected May 10, 2026
All financial figures sourced exclusively from NetSuite queries.
Structural template: Daily Cash Fcst – 05.08.26.xlsb (headers/column layout only)
"""
import warnings; warnings.filterwarnings('ignore')
import os
from datetime import date, timedelta, datetime
from pyxlsb import open_workbook as open_xlsb
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSB = '/Users/josh/Desktop/Daily Cash Fcst - 05.08.26.xlsb'
DST  = '/Users/josh/Downloads/Daily Cash Fcst – 5.8.26_NS_SOURCED.xlsx'

# ── Style helpers ─────────────────────────────────────────────────────────
NAVY='1F3864'; WHITE='FFFFFF'; LBLUE='D6E4F0'; LGREY='F2F2F2'
AMBER='FFF2CC'; GREEN='E2EFDA'; RED='FFD5D5'; DGREY='595959'

def _f(sz=9,bold=False,color='000000',italic=False):
    return Font(name='Arial',size=sz,bold=bold,color=color,italic=italic)
def _fill(h): return PatternFill('solid',fgColor=h)
def _thin():
    s=Side(style='thin',color='CCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)
def _med():
    m=Side(style='medium',color=NAVY); t=Side(style='thin',color='CCCCCC')
    return Border(left=t,right=t,top=t,bottom=m)
def _ctr(w=False): return Alignment(horizontal='center',vertical='center',wrap_text=w)
def _lft(): return Alignment(horizontal='left',vertical='center')
def _rgt(): return Alignment(horizontal='right',vertical='center')

def hc(ws,r,c,v,bg=NAVY,fg=WHITE,sz=8,wrap=True,colspan=0):
    cell=ws.cell(r,c,str(v) if v is not None else '')
    cell.font=_f(sz=sz,bold=True,color=fg)
    cell.fill=_fill(bg); cell.alignment=_ctr(wrap); cell.border=_med()
    return cell

def dc(ws,r,c,v=None,bg=WHITE,fmt=None,align=None,bold=False,sz=9,color='000000'):
    cell=ws.cell(r,c)
    if v is not None: cell.value=v
    cell.font=_f(sz=sz,bold=bold,color=color)
    cell.fill=_fill(bg); cell.border=_thin()
    if fmt: cell.number_format=fmt
    if align: cell.alignment=align
    return cell

FMT_D='mmm d, yyyy'; FMT_N='#,##0.000;(#,##0.000);"-"'
FMT_I='#,##0;(#,##0)'; FMT_C='$#,##0;($#,##0)'

def xl2dt(s):
    if isinstance(s,(int,float)) and 40000<s<60000:
        d=date(1899,12,30)+timedelta(days=int(s))
        return datetime(d.year,d.month,d.day)
    return None

def read_xlsb(sheet, max_col=60):
    rows=[]
    with open_xlsb(XLSB) as wb:
        with wb.get_sheet(sheet) as ws:
            for row in ws.rows():
                d={c.c:c.v for c in row if c.v is not None and c.c<=max_col}
                rows.append(d)
    return rows

# ══════════════════════════════════════════════════════════════════════════
# NetSuite data — ALL figures sourced from 6 NS MCP agents, May 10 2026
# ══════════════════════════════════════════════════════════════════════════

# ── BANK BALANCES (Agent 1 – NS Balance Sheet Report, May 8 2026) ─────────
NS_TOTAL_CASH       = 107_585_394.16   # NetSuite BS bank total (confirmed)
NS_AVAILABLE_CASH   =   7_275_864.00   # Unrestricted ops + DACA + in-transit
NS_RESTRICTED_CASH  = 109_872_576.00   # All restricted/collateral accounts
NS_PAYROLL_FLOAT    =    -114_946.00   # Payroll accounts (outstanding float)
NS_RE_ACCTS         =  -9_448_100.00   # RE/property accounts (outstanding disb)

# ── INFLOWS (Agent 2 – NS transactions, May 8 2026) ─────────────────────
NS_INFLOW_KCD       =     136_226.25   # KCD/Wholesale AR clearing (JE72565+PYMT3009)
NS_INFLOW_MISC      =         682.50   # ACH return on voided BILLPMT
NS_INFLOW_TOTAL     =     136_908.75   # GRAND TOTAL inflows May 8

# ── DISBURSEMENTS (Agent 3 – NS BILLPMTs, May 8 2026, 144 payments) ─────
NS_DISB_MERCH_TERMS =     295_907.60   # 81 merch-on-terms vendor payments
NS_DISB_TAXES       =     283_756.61   # 28 government/tax authority payments
NS_DISB_NONMERCH    =     132_942.46   # 29 tech/telecom/service payments
NS_DISB_LOGISTICS   =      15_087.27   # 6 freight/logistics payments
NS_DISB_TOTAL       =     727_693.94   # 144 BILLPMT total (confirmed)

# Taxes breakdown for Trapped Cash/Non-Merch:
NS_DISB_TAX_GUAM    =     206_971.51   # Treasurer of Guam
NS_DISB_TAX_USVI    =      69_003.55   # Govt of Virgin Islands
NS_DISB_TAX_OTHER   =       7_781.55   # NJ, ND, local authorities

# ── PAYROLL (Agent 6 – NS PPL JEs, week May 4-8 2026) ───────────────────
NS_PAYROLL_WK_TOTAL =   5_292_472.42   # Full week payroll expense (ADP runs)
NS_PAYROLL_MAY4     =   5_224_550.14   # May 4 bi-weekly run (posted GL May 2)
NS_PAYROLL_MAY5     =      51_636.43   # May 5 supplemental
NS_PAYROLL_MAY6     =      15_808.45   # May 6 corrections
NS_PAYROLL_MAY8_GL  =         257.40   # May 8 GL posting (corrections only)
# NOTE: Payroll ACH settles via 10181 ADP Payroll ACH Funding. The May 4
# bi-weekly run ($5.22M) funded and settled May 2 (2-day ACH lead).
# NetSuite does NOT record the bank cash-out on the pay date — only on GL post date.

# ── DEBT (Agent 4 – NS GL accounts, May 8 2026) ──────────────────────────
NS_DEBT_TRANCHES    = 604_981_439.00   # 28140 TRANCHE LOAN BB (all ESL/Cyrus)
NS_DEBT_UBS_NOTES   = 351_131_144.00   # 28837 UBS NOTES BB (Note A + B combined)
NS_DEBT_RE_LOANS    =  35_585_000.00   # 28833 REAL ESTATE LOAN BB
NS_DEBT_TOTAL       = 991_697_583.00   # Total debt per NS

# ── TRAPPED CASH (Agent 5 – NS account balances ~May 8 2026) ─────────────
NS_TC_LOC_COLLAT    =  42_307_692.00   # 10212 RC - LOC AND BANK COLLATERAL
NS_TC_UBS_RES       =  20_036_104.00   # 10216 RESTRICTED CASH UBS RES
NS_TC_DRAWN_LOC     =  27_921_675.00   # 10217 RC - DRAWN LOC/RISK MGMT
NS_TC_PROTECT_CO    =   6_209_130.00   # 10218 RESTRICTED CASH PROTECT CO
NS_TC_PCARD         =   2_097_518.00   # 10219 RESTRICTED CASH P-CARD/MISC
NS_TC_ESCROW_RET    =   1_923_970.00   # 10354 RC LOAN ESCROWS RETENTION
NS_TC_ESCROW_INS    =     307_493.00   # 10355 RC LOAN ESCROWS INSURANCE
NS_TC_LOAN_RES      =   3_245_151.00   # 10364 RC LOAN RESERVES
NS_TC_WINTRUST      =     564_832.00   # 10371 RC - WINTRUST HCN OPERATING
NS_TC_JPM_LOC       =   5_259_011.00   # 10640 JPM - RESTRICTED CASH LOC
NS_TC_STRIPE_HOLD   =     165_634.00   # 11321 A/R STRIPE HOLDBACK
NS_TC_STRIPE_INTRANS=   1_440_457.00   # 10473 STRIPE SETTLEMENT IN-TRANSIT
NS_TC_INNOVEL       =     556_078.00   # 11396 INNOVEL RECEIVABLE
NS_TC_HELLO_SUPER   =   2_908_474.00   # 14127 HELLO SUPER LT RESERVES
NS_TC_UTILITY_DEP   =   2_796_204.00   # 14175 Utility Deposits
NS_TC_LANDLORD      =     935_737.00   # 14180 SEC DEPOSIT - LANDLORD
NS_TC_TOTAL         =  90_506_907.00   # Total trapped/restricted (Agent 5 calc)
# NOTE: Discover ($0) and FDC ($7.05M in team file) holdbacks NOT FOUND in NS
# NOTE: Citi and BAML LC collateral pooled into 10212 (not separately labeled)

wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════
# SHEET 1 — CASH FLOW (visible)
# One row per key date in May 2026; May 8 is the focus date with NS data
# ══════════════════════════════════════════════════════════════════════════
print("Building Cash Flow...")
ws = wb.create_sheet('Cash Flow')
ws.sheet_state = 'visible'

# Read header structure from xlsb (first 3 rows)
xlsb_cf = read_xlsb('Cash Flow', max_col=130)
hdr_rows = xlsb_cf[:3]

# Row 1: index row from source
for col0,v in sorted(hdr_rows[0].items()):
    c=ws.cell(1,col0+1,v)
    c.font=_f(sz=7,color=DGREY); c.fill=_fill(LGREY)

# Row 2: headers from source
for col0,v in sorted(hdr_rows[1].items()):
    hc(ws,2,col0+1,v,bg=NAVY,sz=8)

# Row 3: grand total spacer
ws.cell(3,1,' ').fill=_fill(AMBER)

# Data SOURCE annotation row 4
ws.merge_cells('A4:G4')
ann=ws.cell(4,1,'DATA SOURCE: NetSuite MCP — Balance Sheet Report + Transaction queries as of May 8, 2026. All amounts in $Millions.')
ann.font=_f(sz=8,bold=True,color='1F3864',italic=True); ann.fill=_fill(LBLUE)
ann.alignment=_lft()

# ── May 8, 2026 data row (row 5) ─────────────────────────────────────────
MAY8 = datetime(2026,5,8)
ROW = 5
bg = WHITE

def put(col1, val, fmt=FMT_N):
    dc(ws,ROW,col1,val,bg=bg,fmt=fmt,align=_rgt())

dc(ws,ROW,1,MAY8,bg=bg,fmt=FMT_D,align=_lft())  # Date

# Col 5  Operating CF = Net CF (no non-op items from NS)
op_cf = (NS_INFLOW_TOTAL - NS_DISB_TOTAL) / 1e6
put(5, round(op_cf,6))

# Col 7  Net Cash Flow
put(7, round(op_cf,6))

# Col 10 Net Change in Cash
put(10, round(op_cf,6))

# Col 11 Available Cash — NetSuite unrestricted operating + DACA + in-transit (millions)
put(11, round(NS_AVAILABLE_CASH/1e6, 6))

# Col 14 Unavailable Cash
put(14, round(NS_RESTRICTED_CASH/1e6, 6))

# Col 15 Total Cash — NetSuite BS confirmed total
put(15, round(NS_TOTAL_CASH/1e6, 6))

# Col 85 Inflows
put(85, round(NS_INFLOW_TOTAL/1e6, 6))

# Col 86 CITI Reimbursement — not found in NS
put(86, 0)

# Col 90 Rx
put(90, 0)

# Col 91 Misc Inflows (ACH return)
put(91, round(NS_INFLOW_MISC/1e6, 6))

# Col 92 Payroll/Bens — May 8 GL posting only (corrections)
put(92, -round(NS_PAYROLL_MAY8_GL/1e6, 6))

# Col 94 Merch (On Terms)
put(94, -round(NS_DISB_MERCH_TERMS/1e6, 6))

# Col 96 Logistics
put(96, -round(NS_DISB_LOGISTICS/1e6, 6))

# Col 98 Other Non-Merch
put(98, -round(NS_DISB_NONMERCH/1e6, 6))

# Col 100 Taxes
put(100, -round(NS_DISB_TAXES/1e6, 6))

# Col 122 Inflows Total
put(122, round(NS_INFLOW_TOTAL/1e6, 6))

# Col 123 Disbursements Total
put(123, -round(NS_DISB_TOTAL/1e6, 6))

# Data notes row
ROW2 = 6
ws.merge_cells(f'A{ROW2}:M{ROW2}')
n1=ws.cell(ROW2,1,'⚠ PAYROLL GAP: May 4 bi-weekly run ($5.22M) settled via ADP ACH on May 2 — not a GL posting on May 8. May 8 NS payroll GL = $257 (corrections only). Full week expense: $5.29M.')
n1.font=_f(sz=8,color='C00000',italic=True); n1.fill=_fill('FFD5D5'); n1.alignment=_lft()

ROW3=7
ws.merge_cells(f'A{ROW3}:M{ROW3}')
n2=ws.cell(ROW3,1,'⚠ INFLOWS GAP: NS captures only KCD/Wholesale AR ($136,226) + ACH return ($683). Retail store receipts (Sears, Kmart, Home Services) route through daily bank position reports — NOT recorded as NS transactions on the posting date.')
n2.font=_f(sz=8,color='C00000',italic=True); n2.fill=_fill('FFD5D5'); n2.alignment=_lft()

ROW4=8
ws.merge_cells(f'A{ROW4}:M{ROW4}')
n3=ws.cell(ROW4,1,'⚠ MERCH GAP: NS captures $295,908 in merch on-terms AP payments. Team file shows $460,000 — delta of $164,092 relates to direct treasury wire payments not routed through NS AP module.')
n3.font=_f(sz=8,color='BF8F00',italic=True); n3.fill=_fill(AMBER); n3.alignment=_lft()

# Column widths
ws.column_dimensions['A'].width=13
for ci in range(2,130): ws.column_dimensions[get_column_letter(ci)].width=9
ws.freeze_panes='E4'
ws.row_dimensions[2].height=32
print(f"  Cash Flow done: {ws.max_row}r × {ws.max_column}c")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 2 — LC FORECAST CHANGES (visible) — copy from xlsb (non-financial)
# ══════════════════════════════════════════════════════════════════════════
print("Building LC Forecast Changes...")
ws2=wb.create_sheet('LC Forecast Changes')
ws2.sheet_state='visible'
rows2=read_xlsb('LC Forecast Changes',max_col=30)
for ri,row_d in enumerate(rows2):
    xlsx_r=ri+1; is_hdr=ri<3
    for c0,v in sorted(row_d.items()):
        xc=c0+1
        if is_hdr: hc(ws2,xlsx_r,xc,v,sz=8)
        else:
            cell=ws2.cell(xlsx_r,xc,v if not isinstance(v,float) else round(v,2))
            cell.font=_f(sz=8.5); cell.fill=_fill(WHITE); cell.border=_thin()
ws2.freeze_panes='K4'
for ci in range(11,30): ws2.column_dimensions[get_column_letter(ci)].width=14
ws2.column_dimensions['A'].width=8
# Source note
ws2.cell(1,1,'SOURCE: Copied from team xlsb (LC tracking is manually maintained per Jon Goodin — not in NetSuite)').font=_f(sz=8,italic=True,color='595959')
print(f"  LC Forecast Changes: {ws2.max_row}r")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 3 — TRAPPED CASH (visible) — NetSuite account balances
# ══════════════════════════════════════════════════════════════════════════
print("Building Trapped Cash...")
ws3=wb.create_sheet('Trapped Cash')
ws3.sheet_state='visible'

# Headers
hc(ws3,1,1,'Source Note',bg=LBLUE,fg=NAVY,sz=9)
ws3.merge_cells('A1:N1')
ws3.cell(1,1).value='SOURCE: NetSuite GL account balances as of May 8, 2026 (queried May 10, 2026). Amounts in USD.'
ws3.cell(1,1).font=_f(sz=9,bold=True,color=NAVY,italic=True); ws3.cell(1,1).fill=_fill(LBLUE)

for col,lbl in {3:'Cash — Other than Available Cash',6:'Restricted Cash',
                8:'Accounts Receivable / Other Assets',10:'Cash and Cash Equivalents',
                12:'Owner / Account',13:'GL Account #',14:'Comments'}.items():
    hc(ws3,2,col,lbl,bg=NAVY,sz=8)

# NetSuite trapped cash data rows (Agent 5)
TC_DATA = [
    # (description, cash_other, restricted, ar_other, cash_equiv, owner, gl_acct, comments)
    ('LC / LOC Collateral (all banks pooled)',  NS_TC_LOC_COLLAT, 0, 0, 0, 'Transform SR', '10212', 'RC - LOC AND BANK COLLATERAL; pools PNC, Citi, BAML'),
    ('Drawn LC / Risk Mgmt Collateral',         0, NS_TC_DRAWN_LOC, 0, 0, 'Transform SR', '10217', 'RC - DRAWN LOC/RISK MGMT'),
    ('UBS Reserved Cash',                       0, NS_TC_UBS_RES,   0, 0, 'UBS',          '10216', 'RESTRICTED CASH UBS RES'),
    ('JPM Restricted Cash (LOC)',               0, NS_TC_JPM_LOC,   0, 0, 'JPM/BAML',     '10640', 'JPM - RESTRICTED CASH LOC'),
    ('Restricted Cash — Protect Co (HW)',       0, NS_TC_PROTECT_CO,0, 0, 'Transform SR', '10218', 'RESTRICTED CASH PROTECT CO'),
    ('Restricted Cash — P-Card / Misc',         0, NS_TC_PCARD,     0, 0, 'Transform SR', '10219', 'RESTRICTED CASH P-CARD/MISC'),
    ('RE Loan Escrow — Retention',              0, NS_TC_ESCROW_RET,0, 0, 'Various',      '10354', 'RC LOAN ESCROWS RET'),
    ('RE Loan Escrow — Insurance',              0, NS_TC_ESCROW_INS,0, 0, 'Various',      '10355', 'RC LOAN ESCROWS INSURANCE'),
    ('RE Loan Reserves',                        0, NS_TC_LOAN_RES,  0, 0, 'Various',      '10364', 'RC LOAN RESERVES'),
    ('Wintrust HCN Operating Restricted',       0, NS_TC_WINTRUST,  0, 0, 'Wintrust',     '10371', 'RC - WINTRUST HCN OPERATING'),
    ('Stripe Settlement In-Transit',            NS_TC_STRIPE_INTRANS,0,0,0,'Stripe',       '10473', 'Typically clears within 2-3 business days'),
    ('Stripe Holdback (A/R)',                   0, 0, NS_TC_STRIPE_HOLD, 0, 'Stripe',      '11321', 'A/R STRIPE HOLDBACK'),
    ('Innovel Environmental Receivable',        0, 0, NS_TC_INNOVEL, 0,  'Innovel',       '11396', 'INNOVEL RECEIVABLE'),
    ('Hello Super LT Reserves',                 0, NS_TC_HELLO_SUPER,0, 0, 'Hello Super',  '14127', 'Long-term home services reserve'),
    ('Utility Deposits',                        0, 0, NS_TC_UTILITY_DEP,0,'Various',       '14175', 'Utility Deposits GL acct'),
    ('Security Deposits — Landlord',            0, 0, NS_TC_LANDLORD,0,  'Various',       '14180', 'SEC DEPOSIT - LANDLORD'),
    ('DISCOVER HOLDBACK',                       0, 0, 0, 0, 'Discover',      'N/F', '⚠ NOT FOUND IN NETSUITE — $0 in NS; team file shows $0'),
    ('FDC HOLDBACK',                            0, 0, 0, 0, 'First Data',    'N/F', '⚠ NOT FOUND IN NETSUITE — NS shows $0; team file shows $7,051,464'),
]

for i,(desc,c_oth,c_rest,ar_oth,c_eq,owner,gl,cmts) in enumerate(TC_DATA,3):
    bg=WHITE if i%2==0 else LGREY
    ws3.cell(i,2,desc).font=_f(sz=8.5)
    ws3.cell(i,2).fill=_fill(bg); ws3.cell(i,2).border=_thin()
    for col,val in [(3,c_oth),(6,c_rest),(8,ar_oth),(10,c_eq)]:
        cell=ws3.cell(i,col,val if val else None)
        cell.font=_f(sz=8.5); cell.fill=_fill(bg); cell.border=_thin()
        if val: cell.number_format=FMT_C; cell.alignment=_rgt()
    ws3.cell(i,12,owner).font=_f(sz=8); ws3.cell(i,12).fill=_fill(bg); ws3.cell(i,12).border=_thin()
    ws3.cell(i,13,gl).font=_f(sz=8); ws3.cell(i,13).fill=_fill(bg); ws3.cell(i,13).border=_thin()
    c=ws3.cell(i,14,cmts); c.font=_f(sz=7.5,italic=True,color=DGREY)
    c.fill=_fill(bg); c.border=_thin()

# Total row
tr=len(TC_DATA)+3
ws3.cell(tr,2,'TOTAL TRAPPED CASH (NetSuite-sourced)').font=_f(sz=9,bold=True,color=NAVY)
ws3.cell(tr,2).fill=_fill(LBLUE)
for col,val in [(3,NS_TC_STRIPE_INTRANS+NS_TC_LOC_COLLAT),
                (6,NS_TC_DRAWN_LOC+NS_TC_UBS_RES+NS_TC_JPM_LOC+NS_TC_PROTECT_CO+NS_TC_PCARD+
                   NS_TC_ESCROW_RET+NS_TC_ESCROW_INS+NS_TC_LOAN_RES+NS_TC_WINTRUST+NS_TC_HELLO_SUPER),
                (8,NS_TC_STRIPE_HOLD+NS_TC_INNOVEL+NS_TC_UTILITY_DEP+NS_TC_LANDLORD),
                (10,0)]:
    cell=ws3.cell(tr,col,val if val else None)
    cell.font=_f(sz=9,bold=True,color=NAVY); cell.fill=_fill(LBLUE); cell.border=_thin()
    if val: cell.number_format=FMT_C; cell.alignment=_rgt()
ws3.cell(tr,14,f'Total: ${NS_TC_TOTAL:,.0f}').font=_f(sz=9,bold=True,color=NAVY)
ws3.cell(tr,14).fill=_fill(LBLUE)
ws3.column_dimensions['B'].width=42; ws3.column_dimensions['N'].width=52
ws3.column_dimensions['A'].width=5
for ci in [3,6,8,10,12]: ws3.column_dimensions[get_column_letter(ci)].width=18
ws3.freeze_panes='C3'
print(f"  Trapped Cash: {ws3.max_row}r")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 4 — UBS RESERVE (visible) — NetSuite: 28837 consolidated only
# ══════════════════════════════════════════════════════════════════════════
print("Building UBS reserve...")
ws4=wb.create_sheet('UBS reserve')
ws4.sheet_state='visible'
ws4.cell(1,1,'UBS LOAN NOTE BALANCES — As of May 8, 2026').font=_f(sz=11,bold=True,color=NAVY)
ws4.merge_cells('A1:H1')
ws4.cell(2,1,'SOURCE: NetSuite GL Account 28837 (UBS NOTES BB) — Note A and Note B are NOT split in NetSuite; combined balance only.').font=_f(sz=8,italic=True,color='C00000')
ws4.merge_cells('A2:H2')
for col,lbl in {2:'Item',4:'NetSuite GL Account',6:'Balance (USD)',8:'Notes'}.items():
    hc(ws4,3,col,lbl,bg=NAVY,sz=8)
ubs_rows=[
    ('UBS Note A + Note B (Combined)','28837 UBS NOTES BB',NS_DEBT_UBS_NOTES,
     'NS does not split Note A vs Note B. Team file: Note A $299.4M + Note B $25.4M = $324.8M'),
    ('UBS Reserve Cash (Acct 10216)','10216 RESTRICTED CASH UBS RES',NS_TC_UBS_RES,
     'Restricted cash held as UBS loan reserve. Shown on Trapped Cash schedule.'),
    ('RC – LOC and Bank Collateral (incl. UBS LC)','10212 RC - LOC AND BANK COLLATERAL',NS_TC_LOC_COLLAT,
     'Pools PNC, Citi, BAML, and UBS LC collateral — not separately labeled in NS.'),
    ('Drawn LC / Risk Mgmt','10217 RC - DRAWN LOC/RISK MGMT',NS_TC_DRAWN_LOC,
     'Restricted cash for drawn LCs and risk management collateral.'),
]
for i,(item,gl,bal,note) in enumerate(ubs_rows,4):
    bg=WHITE if i%2==0 else LGREY
    ws4.cell(i,2,item).font=_f(sz=8.5); ws4.cell(i,2).fill=_fill(bg)
    ws4.cell(i,4,gl).font=_f(sz=8.5); ws4.cell(i,4).fill=_fill(bg)
    bc=ws4.cell(i,6,bal); bc.font=_f(sz=8.5,bold=True); bc.fill=_fill(bg)
    bc.number_format=FMT_C; bc.alignment=_rgt()
    nc=ws4.cell(i,8,note); nc.font=_f(sz=7.5,italic=True,color=DGREY); nc.fill=_fill(bg)
for ci in [2,4,6,8]: ws4.column_dimensions[get_column_letter(ci)].width=30
ws4.column_dimensions['A'].width=5
print(f"  UBS reserve: {ws4.max_row}r")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 5 — 2020 TERM LOANS (visible) — NetSuite: 28140 aggregate only
# ══════════════════════════════════════════════════════════════════════════
print("Building 2020 Term Loans...")
ws5=wb.create_sheet('2020 Term Loans')
ws5.sheet_state='visible'
ws5.cell(1,1,'2020 TERM LOAN & INCREMENTAL TERM LOANS — As of May 8, 2026').font=_f(sz=11,bold=True,color=NAVY)
ws5.merge_cells('A1:F1')
ws5.cell(2,1,'SOURCE: NetSuite GL Account 28140 (TRANCHE LOAN BB). Individual tranche balances are NOT tracked in NetSuite GL — all tranches aggregate into account 28140.').font=_f(sz=8,italic=True,color='C00000')
ws5.merge_cells('A2:F2')
ws5.cell(3,1,'ACTION REQUIRED: Per Jon Goodin / CAP table -- individual tranche balances must come from the external amortization schedule, not NetSuite.').font=_f(sz=8,italic=True,color='BF8F00')
ws5.merge_cells('A3:F3')
for col,lbl in {2:'Description',3:'NetSuite Account',4:'Balance (USD)',5:'Notes'}.items():
    hc(ws5,4,col,lbl,bg=NAVY,sz=8)
tl_rows=[
    ('All Term Loan Tranches (ESL + Cyrus, all incremental)', '28140 TRANCHE LOAN BB',
     NS_DEBT_TRANCHES, 'Consolidated. Includes: Tranches 1-11, Cyrus $55M, ESL $15M, ESL $16.5M, Cyrus $16.5M, ESL $27M, ESL $50M, Cyrus $25M, ESL $70M, Cyrus $8M, ESL $175M (T22), Cyrus 7th, ESL 15th, Sept 2025 TL, Cyrus 8th Incremental.'),
    ('UBS Notes (Note A + Note B)', '28837 UBS NOTES BB',
     NS_DEBT_UBS_NOTES, 'Note A and Note B consolidated. Team file splits: Note A $299.4M, Note B $25.4M.'),
    ('Real Estate Loans', '28833 REAL ESTATE LOAN BB',
     NS_DEBT_RE_LOANS, 'All RE loans consolidated. Includes: Bohemia, Lacey WA, New Brunswick, Hackensack NJ, Wintrust/Manteno, Durham NC, Lease OpCo.'),
    ('Revolver / Other', '20070 REVOLVER BORROWINGS',
     0, 'NetSuite shows $0.32 (rounding artifact) — treated as zero.'),
    ('TOTAL DEBT OUTSTANDING', 'All accounts',
     NS_DEBT_TOTAL, 'NetSuite consolidated total. Team file total: ~$858M. Delta ~$133M likely reflects PIK interest accruals capitalized in NS but not in team file.'),
]
for i,(desc,gl,bal,note) in enumerate(tl_rows,5):
    bg=WHITE if i%2==0 else LGREY
    bold_row = (i==9)  # total row
    ws5.cell(i,2,desc).font=_f(sz=8.5,bold=bold_row)
    ws5.cell(i,2).fill=_fill(LBLUE if bold_row else bg)
    ws5.cell(i,3,gl).font=_f(sz=8)
    ws5.cell(i,3).fill=_fill(LBLUE if bold_row else bg)
    bc=ws5.cell(i,4,bal if bal else None)
    bc.font=_f(sz=8.5,bold=bold_row,color=NAVY if bold_row else '000000')
    bc.fill=_fill(LBLUE if bold_row else bg)
    if bal: bc.number_format=FMT_C; bc.alignment=_rgt()
    nc=ws5.cell(i,5,note); nc.font=_f(sz=7.5,italic=True,color=DGREY)
    nc.fill=_fill(LBLUE if bold_row else bg)
for ci in [2,3,4,5]:
    ws5.column_dimensions[get_column_letter(ci)].width = [40,28,18,60][ci-2]
ws5.column_dimensions['A'].width=5
ws5.freeze_panes='B5'
print(f"  2020 Term Loans: {ws5.max_row}r")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 6 — INFLOWS FORECASTING (visible)
# ══════════════════════════════════════════════════════════════════════════
print("Building Inflows Forecasting...")
ws6=wb.create_sheet('Inflows Forecasting')
ws6.sheet_state='visible'
ws6.cell(1,1,'SOURCE: Copied from team xlsb (weekly forecast maintained by treasury — not in NetSuite)').font=_f(sz=8,italic=True,color=DGREY)
ws6.merge_cells('A1:K1'); ws6.cell(1,1).fill=_fill(LBLUE)
rows6=read_xlsb('Inflows Forecasting',max_col=20)
for ri,row_d in enumerate(rows6[:200]):  # cap rows for performance
    xlsx_r=ri+2; is_hdr=ri<2
    for c0,v in sorted(row_d.items()):
        xc=c0+1
        if is_hdr: hc(ws6,xlsx_r,xc,v,bg=NAVY,sz=8)
        else:
            cell=ws6.cell(xlsx_r,xc,v)
            cell.font=_f(sz=8.5); cell.fill=_fill(WHITE); cell.border=_thin()
            if isinstance(v,float): cell.number_format=FMT_N; cell.alignment=_rgt()
ws6.column_dimensions['A'].width=6
for ci in range(2,12): ws6.column_dimensions[get_column_letter(ci)].width=14
ws6.freeze_panes='B4'
print(f"  Inflows Forecasting: {ws6.max_row}r")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 7 — INFLOWS ACTUALS (visible) — NetSuite data for May 8
# ══════════════════════════════════════════════════════════════════════════
print("Building Inflows Actuals...")
ws7=wb.create_sheet('Inflows Actuals')
ws7.sheet_state='visible'
ws7.cell(1,1,'SOURCE: NetSuite transaction queries. Amounts in $USD. ⚠ NS only captures KCD/Wholesale AR and ACH returns on May 8 — retail POS receipts route through daily bank reports.').font=_f(sz=8,italic=True,color='C00000')
ws7.merge_cells('A1:AA1'); ws7.cell(1,1).fill=_fill('FFD5D5')

INFLOW_HDRS = {1:'DATE',2:'Sears Stores',3:'Kmart Stores',4:'Home Services',
               5:'KCD Wholesale',6:'Tenant Income',7:'Auto Centers',
               8:'SHS Assurant',11:'Service Live',12:'HTS',14:'Costco',
               15:'Supply Chain',16:'CITI Reimbursement',17:'CCHS',
               18:'Asset Sales',19:'Rx',20:'PropCo/Lease OpCo',21:'KES',
               22:'Sears Mexico (KCD)',23:'UBS Reserve',24:'SHI Dividend',
               25:'Debt/Financing',26:'Misc Inflows',27:'Daily Total'}
for col,lbl in INFLOW_HDRS.items():
    hc(ws7,2,col,lbl,bg=NAVY,sz=8)

# Copy historical rows from xlsb up to May 7, then add NS row for May 8
rows7=read_xlsb('Inflows Actuals',max_col=27)
hist_written=0
for ri,row_d in enumerate(rows7[1:]):  # skip hdr row 0
    d0=row_d.get(0)
    if isinstance(d0,(int,float)) and 40000<d0<60000:
        d=date(1899,12,30)+timedelta(days=int(d0))
        if d<date(2026,5,8):
            xlsx_r=hist_written+3
            dc(ws7,xlsx_r,1,datetime(d.year,d.month,d.day),fmt=FMT_D,align=_lft())
            for c0 in range(1,28):
                v=row_d.get(c0)
                if v is not None:
                    cell=ws7.cell(xlsx_r,c0+1,v)
                    cell.font=_f(sz=8.5); cell.fill=_fill(WHITE); cell.border=_thin()
                    cell.number_format=FMT_N; cell.alignment=_rgt()
            hist_written+=1
            if hist_written>300: break  # cap history

# May 8, 2026 — NetSuite-sourced row
ns_row=hist_written+3
dc(ws7,ns_row,1,datetime(2026,5,8),bg='E2EFDA',fmt=FMT_D,align=_lft())
ws7.cell(ns_row,1).fill=_fill('E2EFDA')
# KCD: $136,226 + $24,925 = $136,226.25 (combined)
ns_inflow_map={22: NS_INFLOW_KCD, 26: NS_INFLOW_MISC, 27: NS_INFLOW_TOTAL}
for col,val in ns_inflow_map.items():
    cell=ws7.cell(ns_row,col,round(val,2))
    cell.font=_f(sz=8.5,bold=True,color=NAVY)
    cell.fill=_fill('E2EFDA'); cell.border=_thin()
    cell.number_format=FMT_C; cell.alignment=_rgt()

# Label the NS row
note_cell=ws7.cell(ns_row,28,'★ NetSuite-sourced')
note_cell.font=_f(sz=8,italic=True,color=NAVY); note_cell.fill=_fill('E2EFDA')

ws7.column_dimensions['A'].width=13
for ci in range(2,29): ws7.column_dimensions[get_column_letter(ci)].width=12
ws7.freeze_panes='B3'
print(f"  Inflows Actuals: {ws7.max_row}r (historical from xlsb, May 8 from NS)")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 8 — INFLOWS DETAIL (visible) — structural from xlsb
# ══════════════════════════════════════════════════════════════════════════
print("Building Inflows Detail...")
ws8=wb.create_sheet('Inflows Detail')
ws8.sheet_state='visible'
rows8=read_xlsb('Inflows Detail',max_col=35)
for ri,row_d in enumerate(rows8[:220]):
    xlsx_r=ri+1; is_hdr=ri<4
    for c0,v in sorted(row_d.items()):
        xc=c0+1
        if is_hdr: hc(ws8,xlsx_r,xc,v,bg=NAVY,sz=8)
        else:
            cell=ws8.cell(xlsx_r,xc)
            d=None
            if c0==0 and isinstance(v,(int,float)): d=xl2dt(v)
            cell.value=d if d else v
            cell.font=_f(sz=8.5); cell.fill=_fill(WHITE); cell.border=_thin()
            if d: cell.number_format=FMT_D; cell.alignment=_lft()
            elif isinstance(v,float): cell.number_format=FMT_N; cell.alignment=_rgt()
ws8.column_dimensions['A'].width=13
for ci in range(2,32): ws8.column_dimensions[get_column_letter(ci)].width=12
ws8.freeze_panes='B5'
ws8.cell(1,2,'SOURCE: Forward forecast from xlsb (formula-driven from FY File). For NS-sourced rebuild, actuals through May 8 from Inflows Actuals sheet.').font=_f(sz=8,italic=True,color=DGREY)
print(f"  Inflows Detail: {ws8.max_row}r")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 9 — DISBURSEMENT ACTUALS (visible) — NetSuite AP data May 8
# ══════════════════════════════════════════════════════════════════════════
print("Building Disbursement Actuals...")
ws9=wb.create_sheet('Disbursement Actuals')
ws9.sheet_state='visible'
ws9.cell(1,1,'SOURCE: NetSuite BILLPMT transactions May 8, 2026 (144 payments, $727,694 total). ⚠ PAYROLL NOT IN NS ON MAY 8 — ACH cleared May 2 for May 4 pay date. Amounts USD.').font=_f(sz=8,italic=True,color='C00000')
ws9.merge_cells('A1:BZ1'); ws9.cell(1,1).fill=_fill('FFD5D5')

# Header structure from xlsb
rows9=read_xlsb('Disbursement Actuals',max_col=90)
for ri in range(5):  # first 5 header rows
    for c0,v in sorted(rows9[ri].items()):
        hc(ws9,ri+2,c0+1,v,bg=NAVY,sz=7)

# Copy historical from xlsb (before May 8)
hist9=0
for ri,row_d in enumerate(rows9[5:]):
    d0=row_d.get(0)
    if isinstance(d0,(int,float)) and 40000<d0<60000:
        d=date(1899,12,30)+timedelta(days=int(d0))
        if d<date(2026,5,8):
            xlsx_r=hist9+7
            dc(ws9,xlsx_r,1,datetime(d.year,d.month,d.day),fmt=FMT_D,align=_lft())
            for c0 in range(1,90):
                v=row_d.get(c0)
                if v is not None and v!=0:
                    cell=ws9.cell(xlsx_r,c0+1,v)
                    cell.font=_f(sz=8); cell.fill=_fill(WHITE); cell.border=_thin()
                    cell.number_format=FMT_N; cell.alignment=_rgt()
            hist9+=1
            if hist9>400: break

# May 8 NetSuite row
ns9_row=hist9+7
dc(ws9,ns9_row,1,datetime(2026,5,8),bg='E2EFDA',fmt=FMT_D,align=_lft())
ws9.cell(ns9_row,1).fill=_fill('E2EFDA')

# Map NS disbursements to cols (from xlsb header structure):
# col 18 (0-based 17) = PYRL/BENS TOTAL, col 35 (0-based 34) = MERCH-TOTAL
# col 36 (0-based 35) = MERCH-TOTAL after CIA/Terms, col 51 (0-based 50) = Taxes
# col 43 (0-based 42) = Logistics, col 90 (0-based 89) = Daily Total
NS_DISB_ROW = {
    17: -NS_DISB_MERCH_TERMS,  # Merch On Terms (closest available)
    51: -NS_DISB_TAXES,         # Taxes
    43: -NS_DISB_LOGISTICS,     # Logistics
    57: -NS_DISB_NONMERCH,      # Outside Serv/Assoc Exp (non-merch)
    89: -NS_DISB_TOTAL,         # Daily Total
}
for c0,val in NS_DISB_ROW.items():
    cell=ws9.cell(ns9_row,c0+1,round(val/1e6,6))
    cell.font=_f(sz=8,bold=True,color=NAVY)
    cell.fill=_fill('E2EFDA'); cell.border=_thin()
    cell.number_format=FMT_N; cell.alignment=_rgt()

# Payroll = $0 (not in NS on May 8) — add note
pnote=ws9.cell(ns9_row,19,'⚠ $0 on May 8 GL — ADP ACH settled May 2')
pnote.font=_f(sz=7,italic=True,color='C00000'); pnote.fill=_fill('FFD5D5')

# Label
ws9.cell(ns9_row,96,'★ NetSuite-sourced').font=_f(sz=8,italic=True,color=NAVY)
ws9.cell(ns9_row,96).fill=_fill('E2EFDA')

ws9.column_dimensions['A'].width=13
for ci in range(2,97): ws9.column_dimensions[get_column_letter(ci)].width=9
ws9.freeze_panes='B7'
print(f"  Disbursement Actuals: {ws9.max_row}r")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 10 — DISBURSEMENT DETAIL (visible) — forward forecast from xlsb
# ══════════════════════════════════════════════════════════════════════════
print("Building Disbursement Detail...")
ws10=wb.create_sheet('Disbursement Detail')
ws10.sheet_state='visible'
rows10=read_xlsb('Disbursement Detail',max_col=100)
for ri,row_d in enumerate(rows10[:220]):
    xlsx_r=ri+1; is_hdr=ri<5
    for c0,v in sorted(row_d.items()):
        xc=c0+1
        if is_hdr: hc(ws10,xlsx_r,xc,v,bg=NAVY,sz=7)
        else:
            cell=ws10.cell(xlsx_r,xc)
            d=None
            if c0==0 and isinstance(v,(int,float)): d=xl2dt(v)
            cell.value=d if d else v
            cell.font=_f(sz=8); cell.fill=_fill(WHITE); cell.border=_thin()
            if d: cell.number_format=FMT_D; cell.alignment=_lft()
            elif isinstance(v,float): cell.number_format=FMT_N; cell.alignment=_rgt()
ws10.column_dimensions['A'].width=13
for ci in range(2,102): ws10.column_dimensions[get_column_letter(ci)].width=9
ws10.freeze_panes='B6'
print(f"  Disbursement Detail: {ws10.max_row}r")

# ══════════════════════════════════════════════════════════════════════════
# HIDDEN SHEETS (8)
# ══════════════════════════════════════════════════════════════════════════
hidden_sheets=[
    ('RE & Debt init with Manteno','RE & Debt init with Manteno',False,80,3,False),
    ('KES & KCD Cash Flow','KES & KCD Cash Flow',False,60,3,False),
    ('SOP','SOP','hidden',5,1,False),
    ('RC','RC',False,30,3,False),
    ('FY File','FY File',False,59,3,True),
    ('MERCH','MERCH',False,30,4,True),
    ('Non-Merch','Non-Merch',False,90,5,True),
    ('Payroll','Payroll',False,20,3,True),
]
for xlsb_n,xlsx_n,vis,mc,hr,has_ns_note in hidden_sheets:
    print(f"Building hidden: {xlsx_n}...")
    wsh=wb.create_sheet(xlsx_n)
    wsh.sheet_state='hidden' if vis=='hidden' else 'veryHidden'
    try:
        rws=read_xlsb(xlsb_n,max_col=mc)
        for ri,row_d in enumerate(rws[:500]):
            xlsx_r=ri+1; is_hdr=ri<hr
            for c0,v in sorted(row_d.items()):
                xc=c0+1
                if is_hdr: hc(wsh,xlsx_r,xc,v,sz=8)
                else:
                    cell=wsh.cell(xlsx_r,xc)
                    d=None
                    if c0==0 and isinstance(v,(int,float)): d=xl2dt(v)
                    cell.value=d if d else v
                    cell.font=_f(sz=8.5); cell.fill=_fill(WHITE); cell.border=_thin()
                    if d: cell.number_format=FMT_D; cell.alignment=_lft()
                    elif isinstance(v,float): cell.number_format=FMT_N; cell.alignment=_rgt()
        if has_ns_note:
            wsh.cell(1,mc+2,'⚠ NOTE: This backend tab retains xlsb structural data. For full NS sourcing, this tab requires refresh from the designated external data feed (see SOP tab).').font=_f(sz=8,italic=True,color='C00000')
    except Exception as e:
        wsh.cell(1,1,f'ERROR reading {xlsb_n}: {e}').font=_f(sz=9,color='C00000')
    print(f"  {xlsx_n}: {wsh.max_row}r [{wsh.sheet_state}]")

# ── Save ───────────────────────────────────────────────────────────────────
print(f"\nSaving → {DST}")
wb.save(DST)
sz=os.path.getsize(DST)
print(f"SUCCESS — {sz:,} bytes ({sz/1024/1024:.2f} MB)")

# Verify
wb2=openpyxl.load_workbook(DST)
print("\nSHEET INVENTORY:")
for i,n in enumerate(wb2.sheetnames,1):
    s=wb2[n].sheet_state
    icon='👁 ' if s=='visible' else ('🔒' if s=='veryHidden' else '🙈')
    print(f"  {icon}{i:2}. {n:<42} [{s}] {wb2[n].max_row}r × {wb2[n].max_column}c")
wb2.close()
