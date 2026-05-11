from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter


wb = Workbook()

# ── Styles ──────────────────────────────────────────────────────────────────
NAVY        = "1F3864"
LIGHT_GRAY  = "F2F2F2"
MED_GRAY    = "D9D9D9"
WHITE       = "FFFFFF"
BLUE_TEXT   = "0000FF"
BLACK       = "000000"
GREEN_TEXT  = "006400"
LIGHT_GREEN = "C6EFCE"
LIGHT_RED   = "FFC7CE"

def hdr_font(sz=10, bold=True, color=WHITE):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def body_font(sz=10, bold=False, color=BLACK):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def blue_font(sz=10, bold=False):
    return Font(name="Arial", size=sz, bold=bold, color=BLUE_TEXT)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(bottom=s)

def bottom_border():
    s = Side(style="medium", color="1F3864")
    return Border(bottom=s)

NUM_FMT   = '#,##0;(#,##0);"-"'
PCT_FMT   = '0.0%;(0.0%);"-"'
CURR_FMT  = '$#,##0;($#,##0);"-"'

def apply_num(cell, fmt=NUM_FMT):
    cell.number_format = fmt

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ── Helper: write a header row ───────────────────────────────────────────────
def write_header(ws, row, texts, merge_to=None):
    for col, text in enumerate(texts, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.font  = hdr_font()
        c.fill  = fill(NAVY)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                vertical="center", wrap_text=True)
    if merge_to:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=merge_to)

def write_section_hdr(ws, row, text, ncols=3):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = Font(name="Arial", size=10, bold=True, color=WHITE)
    c.fill  = fill("2E4057")
    c.alignment = Alignment(horizontal="left", vertical="center")

def write_subtotal(ws, row, label, formula, indent=False):
    lc = ws.cell(row=row, column=1, value=("  " if indent else "") + label)
    lc.font = body_font(bold=True)
    lc.fill = fill(MED_GRAY)
    vc = ws.cell(row=row, column=2, value=formula)
    vc.font = body_font(bold=True)
    vc.fill = fill(MED_GRAY)
    apply_num(vc)
    return vc

def write_total(ws, row, label, formula):
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    lc.fill = fill(NAVY)
    vc = ws.cell(row=row, column=2, value=formula)
    vc.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    vc.fill = fill(NAVY)
    apply_num(vc)
    return vc

def write_detail(ws, row, label, value, is_input=True, alt_row=False, indent=True):
    bg = LIGHT_GRAY if alt_row else WHITE
    prefix = "    " if indent else ""
    lc = ws.cell(row=row, column=1, value=prefix + label)
    lc.font = blue_font() if is_input else body_font()
    lc.fill = fill(bg)
    vc = ws.cell(row=row, column=2, value=value)
    vc.font = blue_font() if is_input else body_font()
    vc.fill = fill(bg)
    apply_num(vc)
    return vc

def write_pct(ws, row, label, formula, alt_row=False):
    bg = LIGHT_GRAY if alt_row else WHITE
    lc = ws.cell(row=row, column=1, value="    " + label)
    lc.font = body_font(color="595959")
    lc.fill = fill(bg)
    vc = ws.cell(row=row, column=2, value=formula)
    vc.font = body_font(color="595959")
    vc.fill = fill(bg)
    vc.number_format = PCT_FMT
    return vc

def write_check(ws, row, label, formula, good_val=0):
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = body_font(bold=True)
    vc = ws.cell(row=row, column=2, value=formula)
    vc.font = body_font(bold=True)
    apply_num(vc)
    # Conditional green/red via data bar approximation (manual color via formula note)
    # We add a status cell
    sc = ws.cell(row=row, column=3,
                 value=f'=IF(ABS({vc.coordinate})<1,"✓ BALANCED","✗ ERROR")')
    sc.font = body_font(bold=True, color=GREEN_TEXT)
    return vc

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: COVER
# ════════════════════════════════════════════════════════════════════════════
ws_cover = wb.active
ws_cover.title = "Cover"
ws_cover.sheet_view.showGridLines = False

# Banner
ws_cover.merge_cells("A1:D1")
ws_cover.row_dimensions[1].height = 8

ws_cover.merge_cells("A2:D6")
c = ws_cover["A2"]
c.value = "Transform SR Holding Management LLC"
c.font  = Font(name="Arial", size=28, bold=True, color=WHITE)
c.fill  = fill(NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")

ws_cover.merge_cells("A7:D8")
c = ws_cover["A7"]
c.value = "3-Statement Financial Model"
c.font  = Font(name="Arial", size=18, bold=False, color=WHITE)
c.fill  = fill(NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")

ws_cover.row_dimensions[2].height = 20
ws_cover.row_dimensions[3].height = 20
ws_cover.row_dimensions[4].height = 20
ws_cover.row_dimensions[5].height = 20
ws_cover.row_dimensions[6].height = 20
ws_cover.row_dimensions[7].height = 20
ws_cover.row_dimensions[8].height = 20

# Details
details = [
    ("Period:",      "April 2026 (Month Ended April 30, 2026)"),
    ("Entity:",      "Parent Company — Transform SR Holding Management LLC"),
    ("Subsidiary ID:", "1"),
    ("Source:",      "NetSuite ERP — Live Data Pull via MCP"),
    ("Prepared:",    "May 9, 2026"),
    ("Units:",       "$ in thousands (÷ 1,000)"),
]
for i, (label, val) in enumerate(details, start=10):
    ws_cover.row_dimensions[i].height = 22
    lc = ws_cover.cell(row=i, column=1, value=label)
    lc.font = Font(name="Arial", size=11, bold=True, color=NAVY)
    vc = ws_cover.cell(row=i, column=2, value=val)
    vc.font = Font(name="Arial", size=11, color=BLACK)
    ws_cover.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

# KPI summary box
kpis = [
    ("Total Revenue",           "$41,238K"),
    ("Gross Profit",            "$17,104K"),
    ("Gross Margin",            "41.5%"),
    ("Net Operating Income",    "($6,351K)"),
    ("Net Income",              "($13,644K)"),
    ("Total Assets",            "$863,396K"),
    ("Cash (end of period)",    "$101,224K"),
    ("Operating Cash Flow",     "($61,979K)"),
]
ws_cover.merge_cells("A18:D18")
hdr = ws_cover["A18"]
hdr.value = "KEY HIGHLIGHTS — APRIL 2026"
hdr.font  = hdr_font(sz=11)
hdr.fill  = fill(NAVY)
hdr.alignment = Alignment(horizontal="left", vertical="center")
ws_cover.row_dimensions[18].height = 20

for i, (k, v) in enumerate(kpis, start=19):
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    ws_cover.row_dimensions[i].height = 18
    kc = ws_cover.cell(row=i, column=1, value=k)
    kc.font = body_font(bold=True)
    kc.fill = fill(bg)
    vc = ws_cover.cell(row=i, column=2, value=v)
    vc.font = body_font()
    vc.fill = fill(bg)

for col, w in zip("ABCD", [30, 30, 20, 10]):
    set_col_width(ws_cover, col, w)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: INCOME STATEMENT
# ════════════════════════════════════════════════════════════════════════════
ws_is = wb.create_sheet("Income Statement")
ws_is.sheet_view.showGridLines = False

set_col_width(ws_is, "A", 42)
set_col_width(ws_is, "B", 18)
set_col_width(ws_is, "C", 28)

# Title
ws_is.merge_cells("A1:C1")
t = ws_is["A1"]
t.value = "TRANSFORM SR HOLDING MANAGEMENT LLC — INCOME STATEMENT"
t.font  = hdr_font(sz=12)
t.fill  = fill(NAVY)
t.alignment = Alignment(horizontal="left", vertical="center")
ws_is.row_dimensions[1].height = 24

ws_is.merge_cells("A2:C2")
s = ws_is["A2"]
s.value = "Month Ended April 30, 2026  |  $ in thousands  |  Source: NetSuite ERP (Live)"
s.font  = Font(name="Arial", size=9, italic=True, color="595959")
s.alignment = Alignment(horizontal="left")

# Column headers row 3
ws_is["A3"] = "Line Item"
ws_is["B3"] = "Apr-2026 ($K)"
ws_is["C3"] = "Notes / Source"
for col in range(1, 4):
    c = ws_is.cell(row=3, column=col)
    c.font  = hdr_font()
    c.fill  = fill(NAVY)
    c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                            vertical="center")
ws_is.row_dimensions[3].height = 20

# ── REVENUE ──
row = 4
write_section_hdr(ws_is, row, "REVENUE", ncols=3)
ws_is.row_dimensions[row].height = 18
row += 1

rev_items = [
    ("Merchandise Sales",                    9967),
    ("SBT Sales",                             222),
    ("NCC 3rd Party Parts",                  2596),
    ("NCC Home Warranty Parts",              1259),
    ("NCC Assurant PA Parts",                 219),
    ("Paid Labor Revenue — On Site",         5268),
    ("NCC 3rd Party Labor",                  3332),
    ("NCC Home Warranty Labor",              1069),
    ("NCC Assurant PA Labor",                 542),
    ("Home Warranty Revenue (Cinch)",        4982),
    ("HW POS Revenue — Net",                  340),
    ("DC Warehouse Revenue",                 2077),
    ("Wholesale Revenue",                    1744),
    ("Tenant Income",                        4390),
    ("Property Mgmt & Asset Mgmt Fees",       585),
    ("Miscellaneous Revenue",                1503),
    ("Other Revenue (PA, Royalties, etc.)",  1562),
]
rev_start = row
for i, (label, val) in enumerate(rev_items):
    write_detail(ws_is, row, label, val, is_input=True, alt_row=(i % 2 == 1))
    ws_is.row_dimensions[row].height = 16
    row += 1
rev_end = row - 1

total_rev_row = row
write_total(ws_is, row, "TOTAL REVENUE",
            f"=SUM(B{rev_start}:B{rev_end})")
ws_is.row_dimensions[row].height = 18
row += 1

# Blank
ws_is.row_dimensions[row].height = 6; row += 1

# ── COST OF SALES ──
write_section_hdr(ws_is, row, "COST OF SALES", ncols=3)
ws_is.row_dimensions[row].height = 18; row += 1

cogs_items = [
    ("COGS Expense",                        -8363),
    ("Freight COGS",                        -1121),
    ("Wholesale COGS",                       -706),
    ("Parts & Service Costs",              -2847),
    ("Utilities (Electric, Gas, Water)",    -1024),
    ("Truck Fuel",                           -990),
    ("Maintenance & Repairs",              -2519),
    ("Occupancy / Rent",                   -1545),
    ("Home Warranty COGS / Commissions",   -1290),
    ("Import / Duty Costs",                  -648),
    ("Other COGS",                          -3080),
]
cogs_start = row
for i, (label, val) in enumerate(cogs_items):
    write_detail(ws_is, row, label, val, is_input=True, alt_row=(i % 2 == 1))
    ws_is.row_dimensions[row].height = 16; row += 1
cogs_end = row - 1

total_cogs_row = row
write_total(ws_is, row, "TOTAL COST OF SALES",
            f"=SUM(B{cogs_start}:B{cogs_end})")
ws_is.row_dimensions[row].height = 18; row += 1

# Blank
ws_is.row_dimensions[row].height = 6; row += 1

# Gross Profit
gp_row = row
lc = ws_is.cell(row=row, column=1, value="GROSS PROFIT")
lc.font = Font(name="Arial", size=11, bold=True, color=NAVY)
vc = ws_is.cell(row=row, column=2,
                value=f"=B{total_rev_row}+B{total_cogs_row}")
vc.font = Font(name="Arial", size=11, bold=True, color=NAVY)
apply_num(vc)
ws_is.row_dimensions[row].height = 20; row += 1

# GM%
gm_row = row
write_pct(ws_is, row, "Gross Margin %",
          f"=IF(B{total_rev_row}<>0,B{gp_row}/B{total_rev_row},0)")
ws_is.row_dimensions[row].height = 16; row += 1

# Blank
ws_is.row_dimensions[row].height = 6; row += 1

# ── OPERATING EXPENSES ──
write_section_hdr(ws_is, row, "OPERATING EXPENSES (SG&A)", ncols=3)
ws_is.row_dimensions[row].height = 18; row += 1

opex_items = [
    ("Salaries & Wages",                   -8244),
    ("Manager Compensation",               -5531),
    ("Overtime & Other Field Payroll",     -1839),
    ("Payroll Taxes (FICA, SUI, FUI)",     -1212),
    ("Medical & Health Benefits",          -1885),
    ("Severance Compensation",               -714),
    ("Contract Labor",                       -416),
    ("Security Guards",                      -530),
    ("Legal Expense & Settlements",        -1299),
    ("Advertising & Marketing",              -648),
    ("Depreciation — Bldg & FF&E",         -2208),
    ("Property Sales Gain",                 6085),
    ("Casualty & Liability Insurance",     -2694),
    ("Real Estate & Occupancy (SG&A)",     -2195),
    ("Data Processing",                      -109),
    ("Travel & Entertainment",               -611),
    ("Consulting & Professional Fees",       -349),
    ("Other SG&A",                         -1054),
]
opex_start = row
for i, (label, val) in enumerate(opex_items):
    write_detail(ws_is, row, label, val, is_input=True, alt_row=(i % 2 == 1))
    ws_is.row_dimensions[row].height = 16; row += 1
opex_end = row - 1

total_opex_row = row
write_total(ws_is, row, "TOTAL OPERATING EXPENSES",
            f"=SUM(B{opex_start}:B{opex_end})")
ws_is.row_dimensions[row].height = 18; row += 1

# Blank
ws_is.row_dimensions[row].height = 6; row += 1

# EBIT
ebit_row = row
lc = ws_is.cell(row=row, column=1, value="NET OPERATING INCOME (EBIT)")
lc.font = Font(name="Arial", size=11, bold=True, color=NAVY)
vc = ws_is.cell(row=row, column=2,
                value=f"=B{gp_row}+B{total_opex_row}")
vc.font = Font(name="Arial", size=11, bold=True, color=NAVY)
apply_num(vc)
ws_is.row_dimensions[row].height = 20; row += 1

# Blank
ws_is.row_dimensions[row].height = 6; row += 1

# ── OTHER INCOME / EXPENSE ──
write_section_hdr(ws_is, row, "OTHER INCOME / (EXPENSE)", ncols=3)
ws_is.row_dimensions[row].height = 18; row += 1

other_items = [
    ("Investment & Other Interest Income",    157),
    ("PIK Interest — Term Loans",           -7469),
    ("Interest Expense — Tranche Loan",     -1390),
    ("Interest Expense — UBS Notes",          -547),
    ("Interest — Real Estate Loan",           -268),
    ("Debt Issuance Costs & Back-End Fees",   -468),
    ("Revolver & Other Interest",             -447),
    ("State & Foreign Taxes",                 -552),
]
other_start = row
for i, (label, val) in enumerate(other_items):
    write_detail(ws_is, row, label, val, is_input=True, alt_row=(i % 2 == 1))
    ws_is.row_dimensions[row].height = 16; row += 1
other_end = row - 1

total_other_row = row
write_total(ws_is, row, "TOTAL OTHER INCOME / (EXPENSE)",
            f"=SUM(B{other_start}:B{other_end})")
ws_is.row_dimensions[row].height = 18; row += 1

# Blank
ws_is.row_dimensions[row].height = 6; row += 1

# NET INCOME
ni_row = row
lc = ws_is.cell(row=row, column=1, value="NET INCOME")
lc.font = Font(name="Arial", size=12, bold=True, color=WHITE)
lc.fill = fill(NAVY)
vc = ws_is.cell(row=row, column=2,
                value=f"=B{ebit_row}+B{total_other_row}")
vc.font = Font(name="Arial", size=12, bold=True, color=WHITE)
vc.fill = fill(NAVY)
apply_num(vc)
ws_is.row_dimensions[row].height = 22; row += 1

# Net margin
row += 1
write_pct(ws_is, row, "Net Margin %",
          f"=IF(B{total_rev_row}<>0,B{ni_row}/B{total_rev_row},0)")

# Source note
row += 2
note = ws_is.cell(row=row, column=1,
                  value="Source: NetSuite ERP — Report ID -200 (Income Statement), "
                        "April 2026, Subsidiary ID 1 (Transform SR Holding Management LLC), "
                        "pulled via MCP on May 9, 2026. Values in thousands.")
note.font = Font(name="Arial", size=8, italic=True, color="595959")
ws_is.merge_cells(f"A{row}:C{row}")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: BALANCE SHEET
# ════════════════════════════════════════════════════════════════════════════
ws_bs = wb.create_sheet("Balance Sheet")
ws_bs.sheet_view.showGridLines = False

set_col_width(ws_bs, "A", 42)
set_col_width(ws_bs, "B", 18)
set_col_width(ws_bs, "C", 28)

ws_bs.merge_cells("A1:C1")
t = ws_bs["A1"]
t.value = "TRANSFORM SR HOLDING MANAGEMENT LLC — BALANCE SHEET"
t.font  = hdr_font(sz=12)
t.fill  = fill(NAVY)
t.alignment = Alignment(horizontal="left", vertical="center")
ws_bs.row_dimensions[1].height = 24

ws_bs.merge_cells("A2:C2")
s = ws_bs["A2"]
s.value = "As of April 30, 2026  |  $ in thousands  |  Source: NetSuite ERP (Live)"
s.font  = Font(name="Arial", size=9, italic=True, color="595959")

ws_bs["A3"] = "Line Item"
ws_bs["B3"] = "Apr-2026 ($K)"
ws_bs["C3"] = "Notes / Source"
for col in range(1, 4):
    c = ws_bs.cell(row=3, column=col)
    c.font  = hdr_font()
    c.fill  = fill(NAVY)
    c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                            vertical="center")
ws_bs.row_dimensions[3].height = 20

row = 4

# ── ASSETS ──
write_section_hdr(ws_bs, row, "ASSETS", ncols=3)
ws_bs.row_dimensions[row].height = 18; row += 1

# Current Assets
write_section_hdr(ws_bs, row, "  Current Assets", ncols=3)
ws_bs.row_dimensions[row].height = 16; row += 1

ca_items = [
    ("Cash & Bank Accounts",       101224),
    ("Accounts Receivable (net)",   26006),
    ("Merchandise Inventory",       20113),
    ("Prepaid Expenses",            11417),
    ("Other Current Assets",        50499),
]
ca_start = row
for i, (label, val) in enumerate(ca_items):
    write_detail(ws_bs, row, label, val, alt_row=(i % 2 == 1))
    ws_bs.row_dimensions[row].height = 16; row += 1
ca_end = row - 1

tca_row = row
write_subtotal(ws_bs, row, "Total Current Assets",
               f"=SUM(B{ca_start}:B{ca_end})")
ws_bs.row_dimensions[row].height = 18; row += 1

# Blank
ws_bs.row_dimensions[row].height = 6; row += 1

# Fixed Assets
write_section_hdr(ws_bs, row, "  Fixed Assets (Net)", ncols=3)
ws_bs.row_dimensions[row].height = 16; row += 1

fa_items = [
    ("Land",                          143711),
    ("Buildings & Permanent Improv.", 414254),
    ("FF&E, Equipment & Software",     35934),
    ("CWIP",                           62153),
    ("Accumulated Depreciation",     (256282)),
]
fa_start = row
for i, (label, val) in enumerate(fa_items):
    write_detail(ws_bs, row, label, val, alt_row=(i % 2 == 1))
    ws_bs.row_dimensions[row].height = 16; row += 1
fa_end = row - 1

tfa_row = row
write_subtotal(ws_bs, row, "Total Fixed Assets (Net)",
               f"=SUM(B{fa_start}:B{fa_end})")
ws_bs.row_dimensions[row].height = 18; row += 1

# Blank
ws_bs.row_dimensions[row].height = 6; row += 1

# Other Assets
write_section_hdr(ws_bs, row, "  Other Long-Term Assets", ncols=3)
ws_bs.row_dimensions[row].height = 16; row += 1

oa_items = [
    ("ROU Operating Lease Assets",     98770),
    ("Deferred Broker Commissions",    23163),
    ("Trade Names & Intangibles",      19400),
    ("Investments",                    29110),
    ("Tenant Improvements Allowance",  22466),
    ("Long-Term Deposits & Other",     61458),
]
oa_start = row
for i, (label, val) in enumerate(oa_items):
    write_detail(ws_bs, row, label, val, alt_row=(i % 2 == 1))
    ws_bs.row_dimensions[row].height = 16; row += 1
oa_end = row - 1

toa_row = row
write_subtotal(ws_bs, row, "Total Other Assets",
               f"=SUM(B{oa_start}:B{oa_end})")
ws_bs.row_dimensions[row].height = 18; row += 1

# Blank
ws_bs.row_dimensions[row].height = 6; row += 1

# TOTAL ASSETS
ta_row = row
lc = ws_bs.cell(row=row, column=1, value="TOTAL ASSETS")
lc.font = Font(name="Arial", size=12, bold=True, color=WHITE)
lc.fill = fill(NAVY)
vc = ws_bs.cell(row=row, column=2,
                value=f"=B{tca_row}+B{tfa_row}+B{toa_row}")
vc.font = Font(name="Arial", size=12, bold=True, color=WHITE)
vc.fill = fill(NAVY)
apply_num(vc)
ws_bs.row_dimensions[row].height = 22; row += 1

# Spacer
ws_bs.row_dimensions[row].height = 10; row += 1

# ── LIABILITIES & EQUITY ──
write_section_hdr(ws_bs, row, "LIABILITIES & EQUITY", ncols=3)
ws_bs.row_dimensions[row].height = 18; row += 1

# Current Liabilities
write_section_hdr(ws_bs, row, "  Current Liabilities", ncols=3)
ws_bs.row_dimensions[row].height = 16; row += 1

cl_items = [
    ("Accounts Payable — Trade",           11041),
    ("Accounts Payable — Non-Merch",       15545),
    ("Accrued Expenses & Other",          600012),
    ("Sales Tax Payable",                   8054),
    ("Unearned Revenue / HW Reserves",     18117),
    ("Payroll & Benefits Payable",          7613),
    ("Insurance Reserves (ST)",             6745),
    ("Lease Liabilities (ST)",             25309),
    ("Income Tax Liabilities",              4887),
    ("Other Current Liabilities",          91961),
]
cl_start = row
for i, (label, val) in enumerate(cl_items):
    write_detail(ws_bs, row, label, val, alt_row=(i % 2 == 1))
    ws_bs.row_dimensions[row].height = 16; row += 1
cl_end = row - 1

tcl_row = row
write_subtotal(ws_bs, row, "Total Current Liabilities",
               f"=SUM(B{cl_start}:B{cl_end})")
ws_bs.row_dimensions[row].height = 18; row += 1

# Blank
ws_bs.row_dimensions[row].height = 6; row += 1

# Long-Term Liabilities
write_section_hdr(ws_bs, row, "  Long-Term Liabilities", ncols=3)
ws_bs.row_dimensions[row].height = 16; row += 1

ltl_items = [
    ("Tranche Loan (net of issuance costs)",  558824),
    ("UBS Notes",                             351131),
    ("Real Estate Loan",                       43560),
    ("LT Operating Lease Liabilities",         87759),
    ("LT Workers Comp Reserve",                35296),
    ("PIK CY Adjustment",                      11036),
    ("Other LT Liabilities",                   13659),
    ("Debt Issuance Costs",                    (7692)),
    ("Reclassification Adjustments",         (593368)),
]
ltl_start = row
for i, (label, val) in enumerate(ltl_items):
    write_detail(ws_bs, row, label, val, alt_row=(i % 2 == 1))
    ws_bs.row_dimensions[row].height = 16; row += 1
ltl_end = row - 1

tltl_row = row
write_subtotal(ws_bs, row, "Total Long-Term Liabilities",
               f"=SUM(B{ltl_start}:B{ltl_end})")
ws_bs.row_dimensions[row].height = 18; row += 1

# Blank
ws_bs.row_dimensions[row].height = 6; row += 1

# Total Liabilities
tl_row = row
write_total(ws_bs, row, "TOTAL LIABILITIES",
            f"=B{tcl_row}+B{tltl_row}")
ws_bs.row_dimensions[row].height = 20; row += 1

# Blank
ws_bs.row_dimensions[row].height = 6; row += 1

# Equity
write_section_hdr(ws_bs, row, "  Equity", ncols=3)
ws_bs.row_dimensions[row].height = 16; row += 1

eq_items = [
    ("Common Stock",                          24694),
    ("Capital in Excess — Par Value",       3233656),
    ("Intercompany Investments & Paid-in Capital",  31827),
    ("Cumulative Translation Adjustment",     (5611)),
    ("Retained Earnings",                 (2910553)),
    ("Net Income (Balance Forward — Prior Periods)", (743813)),
    ("Net Income (Current Period — Apr 2026)", (37243)),
]
eq_start = row
for i, (label, val) in enumerate(eq_items):
    write_detail(ws_bs, row, label, val, alt_row=(i % 2 == 1))
    ws_bs.row_dimensions[row].height = 16; row += 1
eq_end = row - 1

teq_row = row
write_subtotal(ws_bs, row, "Total Equity",
               f"=SUM(B{eq_start}:B{eq_end})")
ws_bs.row_dimensions[row].height = 18; row += 1

# Blank
ws_bs.row_dimensions[row].height = 6; row += 1

# TOTAL L&E
tle_row = row
lc = ws_bs.cell(row=row, column=1, value="TOTAL LIABILITIES & EQUITY")
lc.font = Font(name="Arial", size=12, bold=True, color=WHITE)
lc.fill = fill(NAVY)
vc = ws_bs.cell(row=row, column=2,
                value=f"=B{tl_row}+B{teq_row}")
vc.font = Font(name="Arial", size=12, bold=True, color=WHITE)
vc.fill = fill(NAVY)
apply_num(vc)
ws_bs.row_dimensions[row].height = 22; row += 1

# Blank
ws_bs.row_dimensions[row].height = 8; row += 1

# Balance Check
row += 1
lc = ws_bs.cell(row=row, column=1, value="BALANCE CHECK (Assets − Liabilities & Equity)")
lc.font = body_font(bold=True)
vc = ws_bs.cell(row=row, column=2,
                value=f"=B{ta_row}-B{tle_row}")
vc.font = body_font(bold=True)
apply_num(vc)
sc = ws_bs.cell(row=row, column=3,
                value=f'=IF(ABS(B{row})<1,"✓ BALANCED","✗ OUT OF BALANCE")')
sc.font = body_font(bold=True, color=GREEN_TEXT)
ws_bs.row_dimensions[row].height = 18

# Source note
row += 2
note = ws_bs.cell(row=row, column=1,
                  value="Source: NetSuite ERP — Report ID -202 (Balance Sheet), "
                        "April 30, 2026, Subsidiary ID 1 (Transform SR Holding Management LLC), "
                        "pulled via MCP on May 9, 2026. Values in thousands.")
note.font = Font(name="Arial", size=8, italic=True, color="595959")
ws_bs.merge_cells(f"A{row}:C{row}")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4: CASH FLOW STATEMENT
# ════════════════════════════════════════════════════════════════════════════
ws_cf = wb.create_sheet("Cash Flow Statement")
ws_cf.sheet_view.showGridLines = False

set_col_width(ws_cf, "A", 48)
set_col_width(ws_cf, "B", 18)
set_col_width(ws_cf, "C", 30)

ws_cf.merge_cells("A1:C1")
t = ws_cf["A1"]
t.value = "TRANSFORM SR HOLDING MANAGEMENT LLC — CASH FLOW STATEMENT"
t.font  = hdr_font(sz=12)
t.fill  = fill(NAVY)
t.alignment = Alignment(horizontal="left", vertical="center")
ws_cf.row_dimensions[1].height = 24

ws_cf.merge_cells("A2:C2")
s = ws_cf["A2"]
s.value = "Month Ended April 30, 2026  |  $ in thousands  |  Source: NetSuite ERP (Live)"
s.font  = Font(name="Arial", size=9, italic=True, color="595959")

ws_cf["A3"] = "Line Item"
ws_cf["B3"] = "Apr-2026 ($K)"
ws_cf["C3"] = "Notes / Source"
for col in range(1, 4):
    c = ws_cf.cell(row=3, column=col)
    c.font  = hdr_font()
    c.fill  = fill(NAVY)
    c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                            vertical="center")
ws_cf.row_dimensions[3].height = 20

row = 4

# ── OPERATING ACTIVITIES ──
write_section_hdr(ws_cf, row, "OPERATING ACTIVITIES", ncols=3)
ws_cf.row_dimensions[row].height = 18; row += 1

op_items = [
    ("Net Income",                                (13644),  True),
    ("Depreciation & Amortization (non-cash)",      2208,  True),
    ("PIK Interest (non-cash)",                     7469,  True),
    ("Other Non-Cash Adjustments",                (57777),  True),
    ("Change in Accounts Receivable",               (509),  True),
    ("Change in Other Current Assets",              1019,  True),
    ("Change in Accounts Payable",                 (3409),  True),
    ("Change in Sales Tax Payable",                  725,  True),
    ("Change in Other Current Liabilities",       (46160),  True),
]
op_start = row
for i, (label, val, inp) in enumerate(op_items):
    write_detail(ws_cf, row, label, val, is_input=inp, alt_row=(i % 2 == 1))
    ws_cf.row_dimensions[row].height = 16; row += 1
op_end = row - 1

top_row = row
write_total(ws_cf, row, "TOTAL OPERATING ACTIVITIES",
            f"=SUM(B{op_start}:B{op_end})")
ws_cf.row_dimensions[row].height = 20; row += 1

# Blank
ws_cf.row_dimensions[row].height = 6; row += 1

# ── INVESTING ACTIVITIES ──
write_section_hdr(ws_cf, row, "INVESTING ACTIVITIES", ncols=3)
ws_cf.row_dimensions[row].height = 18; row += 1

inv_items = [
    ("Capital Expenditures / Fixed Asset Changes",  5262, True),
    ("Change in ROU & Other Long-Term Assets",     39199, True),
]
inv_start = row
for i, (label, val, inp) in enumerate(inv_items):
    write_detail(ws_cf, row, label, val, is_input=inp, alt_row=(i % 2 == 1))
    ws_cf.row_dimensions[row].height = 16; row += 1
inv_end = row - 1

tinv_row = row
write_total(ws_cf, row, "TOTAL INVESTING ACTIVITIES",
            f"=SUM(B{inv_start}:B{inv_end})")
ws_cf.row_dimensions[row].height = 20; row += 1

# Blank
ws_cf.row_dimensions[row].height = 6; row += 1

# ── FINANCING ACTIVITIES ──
write_section_hdr(ws_cf, row, "FINANCING ACTIVITIES", ncols=3)
ws_cf.row_dimensions[row].height = 18; row += 1

fin_items = [
    ("Net Change in Long-Term Debt",          (4850), True),
    ("Change in Equity / Intercompany",        29749, True),
]
fin_start = row
for i, (label, val, inp) in enumerate(fin_items):
    write_detail(ws_cf, row, label, val, is_input=inp, alt_row=(i % 2 == 1))
    ws_cf.row_dimensions[row].height = 16; row += 1
fin_end = row - 1

tfin_row = row
write_total(ws_cf, row, "TOTAL FINANCING ACTIVITIES",
            f"=SUM(B{fin_start}:B{fin_end})")
ws_cf.row_dimensions[row].height = 20; row += 1

# Blank
ws_cf.row_dimensions[row].height = 8; row += 1

# ── CASH SUMMARY ──
write_section_hdr(ws_cf, row, "CASH SUMMARY", ncols=3)
ws_cf.row_dimensions[row].height = 18; row += 1

# Net Change
nc_row = row
lc = ws_cf.cell(row=row, column=1, value="NET CHANGE IN CASH FOR PERIOD")
lc.font = body_font(bold=True)
lc.fill = fill(MED_GRAY)
vc = ws_cf.cell(row=row, column=2,
                value=f"=B{top_row}+B{tinv_row}+B{tfin_row}")
vc.font = body_font(bold=True)
vc.fill = fill(MED_GRAY)
apply_num(vc)
ws_cf.row_dimensions[row].height = 18; row += 1

# Beginning Cash
beg_row = row
write_detail(ws_cf, row, "Cash at Beginning of Period (April 1, 2026)",
             93843, is_input=True, alt_row=False, indent=False)
ws_cf.row_dimensions[row].height = 16; row += 1

# Ending Cash
end_row = row
lc = ws_cf.cell(row=row, column=1, value="CASH AT END OF PERIOD (April 30, 2026)")
lc.font = Font(name="Arial", size=12, bold=True, color=WHITE)
lc.fill = fill(NAVY)
vc = ws_cf.cell(row=row, column=2,
                value=f"=B{nc_row}+B{beg_row}")
vc.font = Font(name="Arial", size=12, bold=True, color=WHITE)
vc.fill = fill(NAVY)
apply_num(vc)
ws_cf.row_dimensions[row].height = 22; row += 1

# Blank
ws_cf.row_dimensions[row].height = 8; row += 1

# Cash Check
row += 1
bs_cash = 101224
lc = ws_cf.cell(row=row, column=1,
                value="CASH CHECK (CFS Ending Cash vs. Balance Sheet Cash)")
lc.font = body_font(bold=True)
vc = ws_cf.cell(row=row, column=2,
                value=f"=B{end_row}-{bs_cash}")
vc.font = body_font(bold=True)
apply_num(vc)
sc = ws_cf.cell(row=row, column=3,
                value=f'=IF(ABS(B{row})<1,"✓ BALANCED","✗ DISCREPANCY")')
sc.font = body_font(bold=True, color=GREEN_TEXT)
ws_cf.row_dimensions[row].height = 18

# Source note
row += 2
note = ws_cf.cell(row=row, column=1,
                  value="Source: NetSuite ERP — Report ID -203 (Cash Flow Statement), "
                        "April 2026, Subsidiary ID 1 (Transform SR Holding Management LLC), "
                        "pulled via MCP on May 9, 2026. Values in thousands.")
note.font = Font(name="Arial", size=8, italic=True, color="595959")
ws_cf.merge_cells(f"A{row}:C{row}")

# ── Set tab colors ───────────────────────────────────────────────────────────
ws_cover.sheet_properties.tabColor = "1F3864"
ws_is.sheet_properties.tabColor    = "2E75B6"
ws_bs.sheet_properties.tabColor    = "375623"
ws_cf.sheet_properties.tabColor    = "7030A0"

# ── Save ────────────────────────────────────────────────────────────────────
OUT = "/Users/josh/Documents/Finance Bots/April 2026 Monthly Report/Transform_SR_3Statement_Apr2026.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
