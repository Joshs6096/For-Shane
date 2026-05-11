"""
REBUILD: Daily Cash Fcst – 5.8.26_CORRECTED.xlsx
Source of truth: Daily Cash Fcst – 05.08.26.xlsb (team file, May 8 2026)
Strategy: read every cell from xlsb → write as values to xlsx.
No formula dependencies — all values are concrete as of May 8, 2026.
"""
import warnings; warnings.filterwarnings('ignore')
import os, sys
from datetime import date, timedelta, datetime
from pyxlsb import open_workbook as open_xlsb
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSB = '/Users/josh/Desktop/Daily Cash Fcst - 05.08.26.xlsb'
DST  = '/Users/josh/Downloads/Daily Cash Fcst - 5.8.26_CORRECTED.xlsx'

# ── Style palette ─────────────────────────────────────────────────────────
NAVY='1F3864'; WHITE='FFFFFF'; LBLUE='D6E4F0'; LGREY='F2F2F2'; AMBER='FFF2CC'; DGREY='595959'

def _font(sz=9, bold=False, color='000000', italic=False):
    return Font(name='Arial', size=sz, bold=bold, color=color, italic=italic)

def _fill(hex_c):
    return PatternFill('solid', fgColor=hex_c)

def _thin():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _med_bot():
    m = Side(style='medium', color=NAVY); t = Side(style='thin', color='CCCCCC')
    return Border(left=t, right=t, top=t, bottom=m)

def _ctr(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def _lft():  return Alignment(horizontal='left',  vertical='center')
def _rgt():  return Alignment(horizontal='right', vertical='center')

def hdr_cell(ws, r, c, val, bg=NAVY, fg=WHITE, sz=8, wrap=True):
    cell = ws.cell(r, c, str(val) if val is not None else '')
    cell.font      = _font(sz=sz, bold=True, color=fg)
    cell.fill      = _fill(bg)
    cell.alignment = _ctr(wrap)
    cell.border    = _med_bot()
    return cell

def data_cell(ws, r, c, val, bg=WHITE, fmt=None, align=None, bold=False, sz=8.5):
    cell = ws.cell(r, c)
    cell.value  = val
    cell.font   = _font(sz=sz, bold=bold)
    cell.fill   = _fill(bg)
    cell.border = _thin()
    if fmt:    cell.number_format = fmt
    if align:  cell.alignment = align
    return cell

# ── xlsb helpers ──────────────────────────────────────────────────────────
def xl2date(serial):
    """Excel date serial → Python date."""
    if isinstance(serial, (int, float)) and 40000 < serial < 60000:
        return date(1899, 12, 30) + timedelta(days=int(serial))
    return None

def read_xlsb_sheet(xlsb_path, sheet_name, max_row=None, max_col=None):
    """Read xlsb sheet into list-of-dicts {col_index: value} per row."""
    rows = []
    with open_xlsb(xlsb_path) as wb:
        with wb.get_sheet(sheet_name) as ws:
            for raw_row in ws.rows():
                row_dict = {}
                for cell in raw_row:
                    if max_col is not None and cell.c > max_col:
                        continue
                    if cell.v is not None:
                        row_dict[cell.c] = cell.v
                rows.append(row_dict)
                if max_row is not None and len(rows) >= max_row:
                    break
    return rows

FMT_DATE = 'mmm d, yyyy'
FMT_NUM  = '#,##0.000;(#,##0.000);"-"'
FMT_INT  = '#,##0;(#,##0)'
FMT_PCT  = '0.00%'
FMT_TEXT = '@'

# ══════════════════════════════════════════════════════════════════════════
#  GENERIC SHEET WRITER
#  Writes all rows from xlsb → xlsx; detects dates, numbers, strings.
#  First 5 rows get header styling; rest get data styling.
# ══════════════════════════════════════════════════════════════════════════
def write_generic(ws, rows, header_rows=5, date_col=0,
                  freeze=None, col_widths=None,
                  skip_cols=None):
    """
    ws          : target openpyxl worksheet
    rows        : list of {col_idx: value} dicts (0-based col indices)
    header_rows : first N rows styled as headers
    date_col    : 0-based col index that holds date serials (or None)
    """
    skip_cols = skip_cols or set()
    for ri, row_dict in enumerate(rows):
        xlsx_r = ri + 1
        is_hdr = ri < header_rows
        for col0, val in sorted(row_dict.items()):
            if col0 in skip_cols:
                continue
            xlsx_c = col0 + 1  # 0-based → 1-based

            if is_hdr:
                hdr_cell(ws, xlsx_r, xlsx_c, val, bg=NAVY, fg=WHITE, sz=8)
            else:
                # Detect type
                d = xl2date(val) if col0 == date_col and isinstance(val, (int, float)) else None
                if d is not None:
                    data_cell(ws, xlsx_r, xlsx_c,
                              datetime(d.year, d.month, d.day),
                              fmt=FMT_DATE, align=_lft())
                elif isinstance(val, float):
                    if abs(val) < 1000 or abs(val) < 10:
                        # likely millions scale
                        data_cell(ws, xlsx_r, xlsx_c, val, fmt=FMT_NUM, align=_rgt())
                    else:
                        data_cell(ws, xlsx_r, xlsx_c, val, fmt=FMT_INT, align=_rgt())
                elif isinstance(val, int):
                    # Check if it looks like a date serial
                    d2 = xl2date(val)
                    if col0 == date_col and d2 is not None:
                        data_cell(ws, xlsx_r, xlsx_c,
                                  datetime(d2.year, d2.month, d2.day),
                                  fmt=FMT_DATE, align=_lft())
                    else:
                        data_cell(ws, xlsx_r, xlsx_c, val, fmt=FMT_INT, align=_rgt())
                elif isinstance(val, str):
                    cell = ws.cell(xlsx_r, xlsx_c, val)
                    cell.font   = _font(sz=8.5)
                    cell.fill   = _fill(WHITE)
                    cell.border = _thin()
                    cell.alignment = _lft()
                else:
                    data_cell(ws, xlsx_r, xlsx_c, val)

    # Freeze panes
    if freeze:
        ws.freeze_panes = freeze

    # Column widths: default 10 for all used cols, then override
    max_c = max((max(r.keys(), default=0) for r in rows), default=0)
    for ci in range(1, max_c + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.column_dimensions['A'].width = 13  # date col wider by default
    if col_widths:
        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = w


# ══════════════════════════════════════════════════════════════════════════
#  BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════
print("="*60)
print("BUILDING CORRECTED FILE FROM XLSB TEAM DATA")
print(f"Source: {XLSB}")
print(f"Output: {DST}")
print("="*60)

wb = Workbook()
wb.remove(wb.active)

# Sheet order + visibility matching Team File
SHEET_CONFIG = [
    # (xlsb_name, xlsx_name, visible, max_col, header_rows, date_col)
    ('Cash Flow',                       'Cash Flow',                       True,  159, 3,  0),
    ('LC Forecast Changes',             'LC Forecast Changes',             True,  30,  3,  None),
    ('Trapped Cash',                    'Trapped Cash',                    True,  14,  2,  None),
    ('UBS reserve',                     'UBS reserve',                     True,  17,  2,  None),
    ('2020 Term Loans',                 '2020 Term Loans',                 True,  29,  3,  None),
    ('Inflows Forecasting',             'Inflows Forecasting',             True,  100, 2,  None),  # cap at 100 cols
    ('Inflows Actuals',                 'Inflows Actuals',                 True,  45,  1,  0),
    ('Inflows Detail',                  'Inflows Detail',                  True,  35,  4,  0),
    ('Disbursement Actuals',            'Disbursement Actuals',            True,  95,  5,  0),
    ('Disbursement Detail',             'Disbursement Detail',             True,  100, 5,  0),
    ('RE & Debt init with Manteno',     'RE & Debt init with Manteno',     False, 80,  3,  None),
    ('KES & KCD Cash Flow',             'KES & KCD Cash Flow',             False, 60,  3,  None),
    ('SOP',                             'SOP',                             'hidden', 5, 1, None),
    ('RC',                              'RC',                              False, 30,  3,  None),
    ('FY File',                         'FY File',                         False, 59,  3,  0),
    ('MERCH',                           'MERCH',                           False, 30,  4,  0),
    ('Non-Merch',                       'Non-Merch',                       False, 90,  5,  0),
    ('Payroll',                         'Payroll',                         False, 20,  3,  0),
]

# Freeze pane config per sheet
FREEZE = {
    'Cash Flow':             'E4',
    'Inflows Actuals':       'B2',
    'Inflows Detail':        'B5',
    'Disbursement Actuals':  'B6',
    'Disbursement Detail':   'B6',
    'FY File':               'B4',
    'Non-Merch':             'B6',
}

for xlsb_name, xlsx_name, visible, max_col, hdr_rows, date_col in SHEET_CONFIG:
    print(f"\nProcessing: {xlsb_name} → {xlsx_name} ...")
    try:
        rows = read_xlsb_sheet(XLSB, xlsb_name, max_col=max_col)
        print(f"  Read {len(rows)} rows, max_col={max_col}")
    except Exception as e:
        print(f"  ERROR reading {xlsb_name}: {e}")
        rows = []

    ws = wb.create_sheet(xlsx_name)

    # Visibility
    if visible is True:
        ws.sheet_state = 'visible'
    elif visible == 'hidden':
        ws.sheet_state = 'hidden'
    else:
        ws.sheet_state = 'veryHidden'

    # Write data
    freeze = FREEZE.get(xlsx_name)
    write_generic(ws, rows, header_rows=hdr_rows,
                  date_col=date_col if date_col is not None else -1,
                  freeze=freeze)

    print(f"  Written: {ws.max_row} rows × {ws.max_column} cols [{ws.sheet_state}]")

# ── Save ───────────────────────────────────────────────────────────────────
print(f"\nSaving to {DST} ...")
wb.save(DST)
sz = os.path.getsize(DST)
print(f"SUCCESS — {sz:,} bytes ({sz/1024/1024:.2f} MB)")

# ── Quick verification ─────────────────────────────────────────────────────
print("\n" + "="*60)
print("SHEET INVENTORY")
print("="*60)
wb2 = openpyxl.load_workbook(DST)
for i, name in enumerate(wb2.sheetnames, 1):
    state = wb2[name].sheet_state
    icon  = '👁 ' if state == 'visible' else ('🔒' if state == 'veryHidden' else '🙈')
    ws_   = wb2[name]
    print(f"  {icon} {i:2}. {name:<42} [{state}] {ws_.max_row}r × {ws_.max_column}c")
wb2.close()
print("\nDone.")
