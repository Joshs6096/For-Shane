import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xlnumbers)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from datetime import datetime, date

# ── Color constants ──────────────────────────────────────────────────────────
NAVY      = "FF1F3864"
WHITE     = "FFFFFFFF"
LGREY     = "FFF2F2F2"
DGREY     = "FFD9D9D9"
LBLUE     = "FFD6E4F0"
LLGREEN   = "FFE2EFDA"
LYELLOW   = "FFFFF2CC"
LRED      = "FFFFD5D5"
BLUE_TXT  = "FF0000FF"   # hardcoded input
BLACK_TXT = "FF000000"   # formula / calc
GREEN_TXT = "FF008000"   # cross-sheet link
RED_TXT   = "FFFF0000"   # external link / unconfirmed

# ── Style helpers ────────────────────────────────────────────────────────────
def hdr(sz=11, bold=True, color=WHITE):
    return Font(name='Arial', size=sz, bold=bold, color=color)

def body(sz=10, bold=False, color=BLACK_TXT, italic=False):
    return Font(name='Arial', size=sz, bold=bold, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def right():
    return Alignment(horizontal='right', vertical='center')

def left(indent=0):
    return Alignment(horizontal='left', vertical='center', indent=indent)

def thin_border(sides='all'):
    t = Side(style='thin', color='FFB8B8B8')
    m = Side(style='medium', color='FF000000')
    n = Side(style=None)
    if sides == 'all':
        return Border(left=t, right=t, top=t, bottom=t)
    if sides == 'bottom':
        return Border(bottom=m)
    if sides == 'header':
        return Border(left=t, right=t, top=m, bottom=m)
    return Border()

# ── Number formats ───────────────────────────────────────────────────────────
FMT_USD   = '$#,##0;($#,##0);"-"'
FMT_USDM  = '$#,##0.000;($#,##0.000);"-"'     # millions
FMT_USDP  = '#,##0;(#,##0);"-"'               # plain integer USD
FMT_PCT   = '0.0%;(0.0%);"-"'
FMT_DATE  = 'mmmm d, yyyy'

def apply_cell(ws, row, col, value=None, fmt=None, font=None,
               fill_=None, align=None, border_=None):
    c = ws.cell(row, col)
    if value is not None:
        c.value = value
    if fmt:
        c.number_format = fmt
    if font:
        c.font = font
    if fill_:
        c.fill = fill_
    if align:
        c.alignment = align
    if border_:
        c.border = border_
    return c

def section_header(ws, row, col_start, col_end, text, bg=NAVY, fg=WHITE):
    c = apply_cell(ws, row, col_start, text,
                   font=hdr(11, True, fg), fill_=fill(bg), align=center(True))
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    for col in range(col_start+1, col_end+1):
        ws.cell(row, col).fill = fill(bg)
    return c

def col_header(ws, row, col, text, bg=LBLUE):
    apply_cell(ws, row, col, text,
               font=hdr(9, True, BLACK_TXT), fill_=fill(bg),
               align=center(True), border_=thin_border('header'))

def source_note(ws, row, col, text):
    apply_cell(ws, row, col, text,
               font=body(8, italic=True, color="FF595959"), align=left(1))

# ════════════════════════════════════════════════════════════════════════════
# DATA
# ════════════════════════════════════════════════════════════════════════════
AS_OF = datetime(2025, 5, 8)
FILE_DATE = "May 8, 2025"
UNITS = "($M)"

# FY File Forecast – Inflows (in $M)
inflows_fcst = [
    ("Sears Stores",          0.069200),
    ("Kmart Stores",          0.300928),
    ("Home Services",         1.411404),
    ("Online",                0.040000),
    ("Tenant Income",         0.005269),
    ("Supply Chain",          0.430000),
    ("KES (ABL)",             0.740000),
    ("KCD Royalty",           0.000000),
    ("CITI Reimbursement",    0.000000),
    ("Asset Sales",           0.000000),
    ("New Debt / Financing",  0.000000),
    ("Misc Inflows",          0.000000),
]

# FY File Forecast – Disbursements (in $M, negative)
disb_fcst = [
    ("AP / Non-Merch",       -1.409371),
    ("Payroll & Benefits",   -0.120000),
    ("Risk Mgmt / Insurance",-0.030000),
    ("Western Union",        -0.050000),
    ("Merch CIA",             0.000000),
    ("Merch On Terms",        0.000000),
    ("Rent",                  0.000000),
    ("Logistics",             0.000000),
    ("Advertising",           0.000000),
    ("Tax",                   0.000000),
    ("Cash Interest",         0.000000),
    ("Cash Fees",             0.000000),
    ("Debt Repayment",        0.000000),
    ("Collateral",            0.000000),
    ("PropCo",                0.000000),
    ("SHI",                   0.000000),
    ("P-Card",                0.000000),
    ("Misc Disbursements",    0.000000),
]

# NetSuite Actuals – only what is CONFIRMED
ns_inflows = [
    ("NCR Journal POS Settlements (JE39909-39912)", 0.193378),
    ("Customer Payments (bank credit not confirmed in NS)", None),
]
ns_disb = [
    ("AP EFT – JPM 10639 (VendPymt batch 00000964)", -0.628301),
    ("AP Checks – JPM 10617 (3401924, 3401925)",      -0.002481),
    ("JE40219 'Payment Multiple Invoices' (bank side unconfirmed)", None),
]

# Treasury SUMMARY 16-BU Payroll (actual dollars, from Payroll sheet)
bu_payroll = [
    ("Finance",        -2909.90),
    ("FinancialSvcs",     0.00),
    ("HoldingsCo",      -934.27),
    ("Home Services", -87549.10),
    ("HR",              -934.08),
    ("KCD",            -1755.76),
    ("Legal",           -910.42),
    ("Member Tech",    -1232.92),
    ("Monark",            0.00),
    ("MSO",            -2544.82),
    ("Real Estate",    -2392.50),
    ("Retail",         -9066.49),
    ("Retail Online",   -488.11),
    ("Service Live",      0.00),
    ("SYWR",           -2122.59),
    ("Supply Chain",   -5559.04),
]

# Cash Flow Model Summary (data_only from Cash Flow row 1440)
cf_net      =  1.387430   # G1440
cf_avail    =  0.000000   # K1440
cf_seg      =  5.241937   # M1440
cf_unavail  = 35.593496   # N1440 = DV1440
cf_total    = 40.835433   # O1440
# Prior day (May 7, row 1439)
cf_avail_d1 = 0.000000    # K1439
cf_total_d1 = 42.607237   # O1439

# Bank Balances (NetSuite GL cumulative through May 8)
bank_balances = [
    # (acct#, description, category, balance)
    ("10208","DACA CONCENTRATION ACCOUNT",           "Operating",   99711.78),
    ("10258","STORE CASH CONCENTRATION",             "Operating",   50707.63),
    ("10264","TF SR HOLDING MANAGEMENT",             "Operating",  138661.79),
    ("10270","TRFM SR PR – STORE DEPOSITS",          "Operating",   33064.83),
    ("10273","GUAM DEBIT AND CREDIT CARD",           "Operating",    1000.00),
    ("10301","THC SAR AR LBX",                       "Operating",   72073.36),
    ("10308","PNC – SERVICE LIVE",                   "Operating",   56050.41),
    ("10314","VALLEY NATIONAL – KES",                "Operating",    4314.92),
    ("10332","BANCO POPULAR",                        "Operating",    4431.12),
    ("10360","PNC BANK KM DUTY VI",                  "Operating",   57129.44),
    ("10395","TRFM KM LLC – STORE DEPOSITS2",        "Operating",    6993.15),
    ("10426","PNC – CASH MGMNT LEASE OPCO",          "Operating",  210192.81),
    ("10428","SHBF CORPORATE",                       "Operating",  251639.13),
    ("10429","PNC – BOHEMIA OPERATING",              "Operating",    5000.45),
    ("10434","PNC BOH CLEARING 11142",               "Operating",  164719.76),
    ("10435","PNC – BOH CMA 11141",                  "Operating",  124886.67),
    ("10456","PNC-CASH OPERATING LEASE OPCO",        "Operating", 2698282.42),
    ("10460","Cash In Store Banks",                  "Operating",  972449.83),
    ("10466","Non CC Settlements",                   "Operating",  199967.72),
    ("10472","CC 3rd Party In-Transit",              "Operating", 3133432.41),
    ("10475","Telecheck ECA",                        "Operating",   13604.46),
    ("10476","ST Investments/Cash Collateral",       "Operating",  323178.10),
    ("10510","Cash On Hand at Stores",               "Operating",  205929.61),
    ("10530","CASH IN BANK – SHI",                   "Operating", 1507157.10),
    ("10601","JPM – PROP HOLDINGS",                  "Operating", 1965043.02),
    ("10608","JPM-CASH OPERATING LEASE OPCO",        "Operating", -402467.29),
    ("10610","JPM – DACA CONCENTRATION ACCT",        "Operating",  515907.14),
    ("10611","JPM – THC MIDCO DACA MMDA",            "Operating", 2826199.25),
    ("10624","JPM – AR LBX",                         "Operating",     660.00),
    ("10631","JPM – ASSET SALES",                    "Operating", 2772944.89),
    ("10634","JPM – MIDCO SALES TAX",                "Operating",     141.38),
    ("10652","JPM – SHBF CORPRATE",                  "Operating",  204090.18),
    ("10653","JPM – YARDI DISBURS MIDCO",            "Operating",  852329.53),
    ("23263","Bank of NY – WU Wires",                "Operating",   17845.94),
    ("10212","RC – LOC AND BANK COLLATERAL",         "Restricted",76252273.70),
    ("10216","RESTRICTED CASH UBS RES",              "Restricted",37626096.65),
    ("10218","RESTRICTED CASH PROTECT CO",           "Restricted", 7625892.36),
    ("10354","RC LOAN ESCROWS RET",                  "Restricted", 2163535.04),
    ("10355","RC LOAN ESCROWS INSURANCE",            "Restricted",   83044.56),
    ("10364","RC LOAN RESERVES",                     "Restricted", 4432111.43),
    ("10640","JPM – RESTRICTED CASH LOC",            "Restricted", 5243414.51),
    ("10302","WELLS FARGO PAYROLL",                  "Disbursement",-198169.15),
    ("10304","PNC – MONEY MARKET SWEEP",             "Disbursement",-973515.81),
    ("10309","IN HOME B2B",                          "Disbursement",   -114.90),
    ("10312","PNC PAYROLL MANUAL",                   "Disbursement",   -577.29),
    ("10317","PNC CLEARING – LACEY",                 "Disbursement", -5536.71),
    ("10318","THC NETSUITE CHECKS",                  "Disbursement",-3211775.22),
    ("10322","PNC BANK – PNC AP/EFT",                "Disbursement",-3806985.67),
    ("10325","PNC PAYROLL GARNISHMENT",              "Disbursement",  -6639.25),
    ("10329","BRANDS JP MORGAN",                     "Disbursement", -192922.66),
    ("10335","THC NETSUITE EFT",                     "Disbursement", -416359.83),
    ("10336","THC SR AP CHECKS (NAP)",               "Disbursement",  -39276.89),
    ("10406","PNC – PROP HOLDINGS",                  "Disbursement", -897030.55),
    ("10427","PNC – LEASE OPCO PAPER CHECKS",        "Disbursement", -284996.65),
    ("10470","Charge Card Deposits",                 "Disbursement",  -29597.29),
    ("10500","TF WILM LBX",                          "Disbursement",  -62661.79),
    ("10614","JPM – TF WILM DISBURSEMENTS",          "Disbursement", -920835.33),
    ("10616","JPM – STORIS AP CHECKS",               "Disbursement",  -16399.22),
    ("10617","JPM – NETSUITE AP CHECKS",             "Disbursement",-2147530.28),
    ("10639","JPM – NETSUITE AP EFT",                "Disbursement",76043748.86),
    ("10642","JPM – TF OVERSEAS COMPANY LLC",        "Disbursement",  -54522.00),
    ("10649","JPM – SHIP HVAC",                      "Disbursement", -127843.68),
    ("10654","JPM – YARDI DISBURS PROP HLDGS",       "Disbursement",-1971441.26),
    ("23262","Bank of NY – WU MOrders",              "Disbursement",  -9237.29),
]

# ════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

# ══════════════════════════════════════════════
# SHEET 1 – Cash Flow Summary
# ══════════════════════════════════════════════
ws1 = wb.create_sheet("Cash Flow")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 38
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 16
ws1.column_dimensions['E'].width = 42

# Title block
ws1.row_dimensions[1].height = 32
ws1.row_dimensions[2].height = 20
ws1.row_dimensions[3].height = 18

section_header(ws1, 1, 1, 5,
    "DAILY CASH FORECAST  ·  MAY 8, 2025  (Thursday)", NAVY, WHITE)
apply_cell(ws1, 2, 1, "Transform SR Holding Management LLC",
           font=body(10, True), align=left())
apply_cell(ws1, 2, 3, f"Prepared: {datetime.now().strftime('%B %d, %Y %H:%M')}",
           font=body(9, color="FF595959"), align=right())
apply_cell(ws1, 3, 1, "Units: $M unless stated · Forecast source: FY File (as of 3/5/25) · Actuals source: NetSuite GL",
           font=body(8, italic=True, color="FF595959"), align=left())
ws1.merge_cells("A2:B2")
ws1.merge_cells("C2:E2")
ws1.merge_cells("A3:E3")

# Column headers
r = 5
for c, txt, bg in [(1,"LINE ITEM",NAVY),(2,"FORECAST ($M)",LBLUE),
                   (3,"NS ACTUALS ($M)",LLGREEN),(4,"VARIANCE ($M)",LGREY),
                   (5,"NOTES / SOURCE",NAVY)]:
    col_header(ws1, r, c, txt, bg)
ws1.row_dimensions[r].height = 28

def row_data(ws, r, label, fcst, actual=None, note="", indent=0, bold=False,
             subtotal=False, bg=None):
    bg = bg or (LGREY if subtotal else WHITE)
    fnt = body(10 if not subtotal else 10, bold or subtotal)
    ws.row_dimensions[r].height = 16
    apply_cell(ws, r, 1, ("  "*indent) + label, font=fnt, fill_=fill(bg), align=left(indent))
    apply_cell(ws, r, 2, fcst, font=body(10, bold, BLUE_TXT), fill_=fill(bg),
               fmt=FMT_USDM, align=right())
    if actual is not None:
        apply_cell(ws, r, 3, actual, font=body(10, bold, GREEN_TXT), fill_=fill(bg),
                   fmt=FMT_USDM, align=right())
        var = actual - fcst if isinstance(actual, (int, float)) else None
        if var is not None:
            var_bg = LLGREEN if var >= 0 else LRED
            apply_cell(ws, r, 4, var, font=body(10, bold, BLACK_TXT),
                       fill_=fill(var_bg if not subtotal else bg),
                       fmt=FMT_USDM, align=right())
    else:
        apply_cell(ws, r, 3, "—  N/A", font=body(9, italic=True, color="FF888888"),
                   fill_=fill(bg), align=right())
        apply_cell(ws, r, 4, None, fill_=fill(bg))
    apply_cell(ws, r, 5, note, font=body(8, italic=True, color="FF595959"),
               fill_=fill(bg), align=left())

r = 6
section_header(ws1, r, 1, 5, "CASH INFLOWS", "FF17375E", WHITE)
r += 1
total_in_fcst = sum(v for _,v in inflows_fcst)
ns_total_in   = 0.193378   # only NCR confirmed
for label, fcst_v in inflows_fcst:
    actual_v = None
    note_v   = "Source: FY File (forecast)"
    if label == "Sears Stores":
        note_v = "FY forecast; NCR POS journals $0.069M (NCR covers Sears+Kmart combined $0.193M)"
    elif label == "Kmart Stores":
        note_v = "FY forecast; see NCR note on Sears Stores line"
    elif label == "Home Services":
        note_v = "FY forecast; customer pymts (PYMT14xx) posted 5/8 – bank credit not isolated in NS"
    row_data(ws1, r, label, fcst_v, actual_v, note_v, indent=1)
    r += 1

r_in_subtotal = r
row_data(ws1, r, "TOTAL INFLOWS", total_in_fcst, note="Sum of above", bold=True, subtotal=True)
# Manually write formula
ws1.cell(r, 2).value = f"=SUM(B7:B{r-1})"
ws1.cell(r, 2).font = body(10, True, BLACK_TXT)
ws1.cell(r, 3).value = ns_total_in
ws1.cell(r, 3).font = body(10, True, GREEN_TXT)
ws1.cell(r, 3).number_format = FMT_USDM
ws1.cell(r, 4).value = f"=C{r}-B{r}"
ws1.cell(r, 4).font = body(10, True, BLACK_TXT)
ws1.cell(r, 4).number_format = FMT_USDM
ws1.cell(r, 5).value = "Actuals: only NCR POS settlements confirmed (JE39909-39912 = $193,378); customer pymts through AR sub-ledger not confirmed at bank on 5/8"
ws1.cell(r, 5).font = body(8, italic=True, color="FF595959")
r += 1

r += 1  # spacer
section_header(ws1, r, 1, 5, "CASH DISBURSEMENTS", "FF17375E", WHITE)
r += 1
total_out_fcst = sum(v for _,v in disb_fcst)
ns_total_out  = -0.630781  # confirmed VendPymts only
for label, fcst_v in disb_fcst:
    actual_v = None
    note_v   = "Source: FY File (forecast)"
    if label == "AP / Non-Merch":
        note_v = "FY forecast $1.409M; NS confirmed $0.631M (10639 EFT + 10617 checks); JE40219 $0.684M bank side not confirmed in NS"
    elif label == "Payroll & Benefits":
        note_v = "Source: Payroll sheet (16-BU format); NS payroll account 10302 not isolated by date"
    row_data(ws1, r, label, fcst_v, actual_v, note_v, indent=1)
    r += 1

r_out_subtotal = r
row_data(ws1, r, "TOTAL DISBURSEMENTS", total_out_fcst, note="Sum of above", bold=True, subtotal=True)
ws1.cell(r, 2).value = f"=SUM(B{r_in_subtotal+2}:B{r-1})"
ws1.cell(r, 2).font = body(10, True, BLACK_TXT)
ws1.cell(r, 3).value = ns_total_out
ws1.cell(r, 3).font = body(10, True, GREEN_TXT)
ws1.cell(r, 3).number_format = FMT_USDM
ws1.cell(r, 4).value = f"=C{r}-B{r}"
ws1.cell(r, 4).font = body(10, True, BLACK_TXT)
ws1.cell(r, 4).number_format = FMT_USDM
ws1.cell(r, 5).value = "Actuals: VendPymts to JPM 10639 ($628,301) + JPM 10617 ($2,480) = $630,781; payroll not isolatable in NS on this date"
ws1.cell(r, 5).font = body(8, italic=True, color="FF595959")
r += 1

r += 1
# Net Cash Flow
section_header(ws1, r, 1, 5, "NET CASH FLOW & BALANCE SHEET", "FF17375E", WHITE)
r += 1
net_fcst = cf_net
net_ns   = ns_total_in + ns_total_out
row_data(ws1, r, "Net Operating Cash Flow", net_fcst, net_ns,
         "Model: G1440=SUM(CG:DM); NS actuals partial (only NCR receipts + confirmed AP)", bold=True)
ws1.cell(r, 4).value = net_ns - net_fcst
ws1.cell(r, 4).number_format = FMT_USDM
ws1.cell(r, 4).fill = fill(LRED if net_ns - net_fcst < 0 else LLGREEN)
r += 1

row_data(ws1, r, "Opening Available Cash (May 7 close)", cf_avail_d1, None,
         "Source: Cash Flow K1439 (model); $0 = facility drawn to floor")
r += 1
row_data(ws1, r, "Net Change in Available Cash", cf_net, None,
         "Source: Cash Flow G1440")
r += 1
row_data(ws1, r, "Closing Available Cash (May 8)", cf_avail, None,
         "Source: Cash Flow K1440; $0 = HTS facility absorbs net inflow (Y1440 ≥ net CF)")
ws1.cell(r, 2).fill = fill(LYELLOW)
r += 1
row_data(ws1, r, "Segregated Account", cf_seg, None,
         "Source: Cash Flow M1440=DQ1440; rolls forward (no change on 5/8)")
r += 1
row_data(ws1, r, "Unavailable Cash (Restricted + Collateral)", cf_unavail, None,
         "Source: Cash Flow DV1440 (hardcoded from model); see Bank Balances tab for NS breakdown")
r += 1
row_data(ws1, r, "TOTAL CASH (Model)", cf_total, None,
         "Source: Cash Flow O1440 = K+M+N = $0 + $5.242M + $35.593M", bold=True, subtotal=True)
ws1.cell(r, 2).value = f"={cf_avail}+{cf_seg}+{cf_unavail}"
ws1.cell(r, 2).number_format = FMT_USDM
ws1.cell(r, 2).font = body(10, True, BLACK_TXT)
r += 1

# Prior day comparison
r += 1
row_data(ws1, r, "TOTAL CASH – May 7 (prior day)", cf_total_d1, None,
         "Source: Cash Flow O1439", subtotal=False)
r += 1
row_data(ws1, r, "Change in Total Cash, Day-over-Day",
         round(cf_total - cf_total_d1, 6), None,
         f"May 8 ${cf_total:.3f}M vs May 7 ${cf_total_d1:.3f}M")
ws1.cell(r, 2).fill = fill(LRED if cf_total - cf_total_d1 < 0 else LLGREEN)

# ── border all data rows
for rr in range(5, r+1):
    for cc in range(1, 6):
        c = ws1.cell(rr, cc)
        if c.border is None or c.border == Border():
            c.border = thin_border()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Treasury SUMMARY (16-BU Payroll)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Treasury SUMMARY")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 16

ws2.row_dimensions[1].height = 30
section_header(ws2, 1, 1, 4,
    "TREASURY SUMMARY · 16-BU PAYROLL FORMAT · MAY 8, 2025", NAVY, WHITE)
apply_cell(ws2, 2, 1, "Transform SR Holding Management LLC",
           font=body(10, True), align=left())
apply_cell(ws2, 2, 3, "Units: Actual USD",
           font=body(9, color="FF595959"), align=right())
ws2.merge_cells("A2:B2")
ws2.merge_cells("C2:D2")

apply_cell(ws2, 3, 1,
    "Source: Payroll sheet (Treasury's 16 BU Format). Values in actual dollars. "
    "Actuals not available by BU from NetSuite (single subsidiary entity).",
    font=body(8, italic=True, color="FF595959"), align=left())
ws2.merge_cells("A3:D3")

r = 5
for c, txt, bg in [(1,"BUSINESS UNIT",NAVY),(2,"FORECAST (USD)",LBLUE),
                   (3,"NS ACTUALS (USD)",LLGREEN),(4,"VARIANCE",LGREY)]:
    col_header(ws2, r, c, txt, bg)
ws2.row_dimensions[r].height = 26

r = 6
total_fcst_bu = 0
for bu, fcst_v in bu_payroll:
    ws2.row_dimensions[r].height = 16
    bg = LGREY if r % 2 == 0 else WHITE
    apply_cell(ws2, r, 1, bu, font=body(10), fill_=fill(bg), align=left(1))
    apply_cell(ws2, r, 2, fcst_v, font=body(10, color=BLUE_TXT), fill_=fill(bg),
               fmt=FMT_USD, align=right())
    apply_cell(ws2, r, 3, "—", font=body(9, italic=True, color="FF888888"),
               fill_=fill(bg), align=right())
    apply_cell(ws2, r, 4, None, fill_=fill(bg))
    for cc in range(1, 5):
        ws2.cell(r, cc).border = thin_border()
    total_fcst_bu += fcst_v
    r += 1

# Subtotal
ws2.row_dimensions[r].height = 18
for cc in range(1, 5):
    ws2.cell(r, cc).fill = fill(DGREY)
    ws2.cell(r, cc).border = thin_border('header')
apply_cell(ws2, r, 1, "TOTAL PAYROLL / BENEFITS", font=body(10, True), fill_=fill(DGREY), align=left())
apply_cell(ws2, r, 2, f"=SUM(B6:B{r-1})", font=body(10, True, BLACK_TXT), fill_=fill(DGREY),
           fmt=FMT_USD, align=right())
apply_cell(ws2, r, 3, "Not available — NS has single subsidiary, no BU split",
           font=body(8, italic=True, color="FF888888"), fill_=fill(DGREY), align=right())
apply_cell(ws2, r, 4, None, fill_=fill(DGREY))

r += 2
apply_cell(ws2, r, 1,
    f"Note: NetSuite payroll cannot be broken by BU. "
    f"Total forecast payroll for May 8 = ${abs(total_fcst_bu):,.0f}. "
    f"Wells Fargo payroll acct (10302) shows cumulative balance -$198,169; "
    f"date-specific payroll not isolatable from available NS queries.",
    font=body(8, italic=True, color="FF595959"), align=left())
ws2.merge_cells(f"A{r}:D{r}")

r += 2
apply_cell(ws2, r, 1, "Forecast source: Payroll sheet tab (Treasury's 16 BU Format) from 3/5/25 file",
           font=body(8, italic=True, color="FF595959"), align=left())
ws2.merge_cells(f"A{r}:D{r}")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Inflows Detail
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Inflows Detail")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 50

section_header(ws3, 1, 1, 5,
    "INFLOWS DETAIL · MAY 8, 2025", NAVY, WHITE)
apply_cell(ws3, 2, 1,
    "Source: FY File → Inflows Detail → Cash Flow (CG:CM) · All values in $M",
    font=body(9, italic=True, color="FF595959"))
ws3.merge_cells("A2:E2")

r = 4
for c, txt, bg in [(1,"CATEGORY",NAVY),(2,"FORECAST ($M)",LBLUE),
                   (3,"NS ACTUALS ($M)",LLGREEN),(4,"VARIANCE ($M)",LGREY),
                   (5,"NOTES",NAVY)]:
    col_header(ws3, r, c, txt, bg)
ws3.row_dimensions[r].height = 26

r = 5
notes_map = {
    "Sears Stores":     "FY File col B; Inflows Detail col B → SUM(B:K)→CG; '= INDEX(FY File, MATCH(date), MATCH(header))'",
    "Kmart Stores":     "FY File col C; same INDEX/MATCH pattern",
    "Home Services":    "FY File col D; largest inflow category May 8",
    "Online":           "FY File col H (Online); feeds Inflows Detail col G",
    "Tenant Income":    "FY File col K; real estate sub-tenant receipts",
    "Supply Chain":     "FY File col N → Inflows Detail col O → Cash Flow CJ",
    "KES (ABL)":        "FY File col V → Inflows Detail col U; KES ABL draw/paydown",
    "KCD Royalty":      "FY File cols F+G (KCD Wholesale + Royalty); $0 on 5/8",
    "CITI Reimbursement":"FY File col O → Cash Flow CH; $0 on 5/8",
    "Asset Sales":      "FY File col Q → Inflows Detail col R → Cash Flow CK; $0 on 5/8",
    "New Debt / Financing":"FY File col M → Inflows Detail col Y → Cash Flow CI; $0 on 5/8",
    "Misc Inflows":     "FY File col AD; residual Cash Flow CM = AA55 - SUM(CG:CL)",
}

# NCR actuals: $0.193M covers Sears + Kmart combined
total_in_r   = r
in_fcst_vals = {}
for label, fcst_v in inflows_fcst:
    ws3.row_dimensions[r].height = 16
    bg = LGREY if r % 2 == 0 else WHITE
    ns_v = None
    ns_txt = "—"
    if label == "Sears Stores":
        ns_txt = "Sears+Kmart NCR combined = $0.193M; cannot split"
    elif label == "Kmart Stores":
        ns_txt = "See Sears Stores"

    apply_cell(ws3, r, 1, "  " + label, font=body(10), fill_=fill(bg), align=left(1))
    apply_cell(ws3, r, 2, fcst_v, font=body(10, color=BLUE_TXT), fill_=fill(bg),
               fmt=FMT_USDM, align=right())
    apply_cell(ws3, r, 3, ns_txt, font=body(8, italic=True, color="FF888888"),
               fill_=fill(bg), align=right())
    apply_cell(ws3, r, 4, None, fill_=fill(bg))
    apply_cell(ws3, r, 5, notes_map.get(label, ""), font=body(8, italic=True, color="FF595959"),
               fill_=fill(bg), align=left())
    for cc in range(1, 6):
        ws3.cell(r, cc).border = thin_border()
    in_fcst_vals[label] = fcst_v
    r += 1

# Total
ws3.row_dimensions[r].height = 18
for cc in range(1, 6):
    ws3.cell(r, cc).fill = fill(DGREY)
    ws3.cell(r, cc).border = thin_border('header')
apply_cell(ws3, r, 1, "TOTAL INFLOWS", font=body(10, True), fill_=fill(DGREY))
apply_cell(ws3, r, 2, f"=SUM(B5:B{r-1})", font=body(10, True, BLACK_TXT),
           fill_=fill(DGREY), fmt=FMT_USDM, align=right())
apply_cell(ws3, r, 3, "0.193 (NCR only; partial)",
           font=body(9, italic=True, color="FF888888"), fill_=fill(DGREY), align=right())
apply_cell(ws3, r, 5,
    "NS: Customer pymts (PYMT14xx) posted 5/8 — $0.193M NCR journals confirmed; "
    "remaining receipts flow through AR sub-ledger (bank debit not same-day in NS)",
    font=body(8, italic=True, color="FF595959"), fill_=fill(DGREY), align=left())

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 – Disbursement Detail
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Disbursement Detail")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 14
ws4.column_dimensions['E'].width = 58

section_header(ws4, 1, 1, 5,
    "DISBURSEMENT DETAIL · MAY 8, 2025", NAVY, WHITE)
apply_cell(ws4, 2, 1,
    "Source: Disbursement Detail → Cash Flow (CN:DI) · Negative = cash out · All values in $M",
    font=body(9, italic=True, color="FF595959"))
ws4.merge_cells("A2:E2")

r = 4
for c, txt, bg in [(1,"CATEGORY",NAVY),(2,"FORECAST ($M)",LBLUE),
                   (3,"NS ACTUALS ($M)",LLGREEN),(4,"VARIANCE ($M)",LGREY),
                   (5,"NOTES / SOURCE",NAVY)]:
    col_header(ws4, r, c, txt, bg)
ws4.row_dimensions[r].height = 26

disp_notes = {
    "AP / Non-Merch":       "FY File col AG 'AP'; feeds Disbursement Detail Non-Merch section (cols AK+) → Cash Flow CU-DI. NS confirmed: VendPymts 10639 ($628,301) + 10617 ($2,480) = $630,781. JE40219 ($684,173 'Payment Multiple Invoices') bank side not confirmed in NS GL.",
    "Payroll & Benefits":   "Payroll sheet (16-BU) → Disbursement Detail cols B-R → Cash Flow CN. Forecast = $120K. Wells Fargo 10302 cumulative -$198,169; not isolatable by date.",
    "Risk Mgmt / Insurance":"FY File col AP → Disbursement Detail AY → Cash Flow CU",
    "Western Union":        "FY File col AT → Disbursement Detail BA → Cash Flow CZ",
    "Merch CIA":            "Merch sheet → Disbursement Detail AH → Cash Flow CO. $0 forecast May 8.",
    "Merch On Terms":       "Merch sheet → Disbursement Detail AI → Cash Flow CP. $0 forecast May 8.",
    "Rent":                 "Non-Merch → Disbursement Detail BB → Cash Flow CQ. $0 forecast May 8.",
    "Logistics":            "Non-Merch → Disbursement Detail AR → Cash Flow CR. $0 forecast May 8.",
    "Advertising":          "Non-Merch (Google + Other) → Disbursement Detail AM → Cash Flow CS. $0 May 8.",
    "Tax":                  "FY File col AQ + Non-Merch (Sales Tax + State/Fed Income Tax) → Cash Flow CV",
    "Cash Interest":        "FY File col AY → Disbursement Detail CJ → Cash Flow DD",
    "Cash Fees":            "FY File col AZ → Disbursement Detail CK → Cash Flow DE",
    "Debt Repayment":       "FY File col AX → Disbursement Detail CI → Cash Flow DH",
    "Collateral":           "FY File col AL → Disbursement Detail CG → Cash Flow DC",
    "PropCo":               "FY File col BA → Disbursement Detail CE → Cash Flow DF",
    "SHI":                  "FY File col AV → Disbursement Detail BU → Cash Flow DB",
    "P-Card":               "FY File col AU → Disbursement Detail CA → Cash Flow DA",
    "Misc Disbursements":   "FY File col BB → Disbursement Detail CH → Cash Flow DI",
}

r = 5
for label, fcst_v in disb_fcst:
    ws4.row_dimensions[r].height = 16
    bg = LGREY if r % 2 == 0 else WHITE
    ns_v = None
    ns_txt = "—"
    var_v  = None
    if label == "AP / Non-Merch":
        ns_v   = -0.630781
        ns_txt = None
        var_v  = ns_v - fcst_v
    apply_cell(ws4, r, 1, "  " + label, font=body(10), fill_=fill(bg), align=left(1))
    apply_cell(ws4, r, 2, fcst_v, font=body(10, color=BLUE_TXT), fill_=fill(bg),
               fmt=FMT_USDM, align=right())
    if ns_v is not None:
        apply_cell(ws4, r, 3, ns_v, font=body(10, color=GREEN_TXT), fill_=fill(bg),
                   fmt=FMT_USDM, align=right())
        vbg = LLGREEN if var_v >= 0 else LRED
        apply_cell(ws4, r, 4, var_v, font=body(10, color=BLACK_TXT), fill_=fill(vbg),
                   fmt=FMT_USDM, align=right())
    else:
        apply_cell(ws4, r, 3, ns_txt, font=body(8, italic=True, color="FF888888"),
                   fill_=fill(bg), align=right())
        apply_cell(ws4, r, 4, None, fill_=fill(bg))
    apply_cell(ws4, r, 5, disp_notes.get(label, ""), font=body(8, italic=True, color="FF595959"),
               fill_=fill(bg), align=left())
    for cc in range(1, 6):
        ws4.cell(r, cc).border = thin_border()
    r += 1

# Total
ws4.row_dimensions[r].height = 18
for cc in range(1, 6):
    ws4.cell(r, cc).fill = fill(DGREY)
    ws4.cell(r, cc).border = thin_border('header')
apply_cell(ws4, r, 1, "TOTAL DISBURSEMENTS", font=body(10, True), fill_=fill(DGREY))
apply_cell(ws4, r, 2, f"=SUM(B5:B{r-1})", font=body(10, True, BLACK_TXT),
           fill_=fill(DGREY), fmt=FMT_USDM, align=right())
apply_cell(ws4, r, 3, -0.630781, font=body(10, True, GREEN_TXT),
           fill_=fill(DGREY), fmt=FMT_USDM, align=right())
apply_cell(ws4, r, 4, f"=C{r}-B{r}", font=body(10, True, BLACK_TXT),
           fill_=fill(DGREY), fmt=FMT_USDM, align=right())
apply_cell(ws4, r, 5,
    "NS actuals: only AP EFT+Checks confirmed ($630,781). Payroll/merch/rent/logistics "
    "not isolatable in NS at required granularity. JE40219 $684K bank side unconfirmed.",
    font=body(8, italic=True, color="FF595959"), fill_=fill(DGREY), align=left())

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 – Bank Balances
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Bank Balances")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions['A'].width = 10
ws5.column_dimensions['B'].width = 40
ws5.column_dimensions['C'].width = 16
ws5.column_dimensions['D'].width = 18

section_header(ws5, 1, 1, 4,
    "BANK ACCOUNT BALANCES · GL CUMULATIVE THROUGH MAY 8, 2025 · SOURCE: NetSuite", NAVY, WHITE)
apply_cell(ws5, 2, 1,
    "IMPORTANT: These are cumulative GL balances (debit/credit since account inception), "
    "NOT current bank statement balances. Disbursement accounts show negative because checks "
    "clear through zero-balance accounts funded same-day. 10639 JPM AP EFT shows $76M "
    "cumulative credits offset by debits. 10212 RC-LOC $76M is restricted cash collateral.",
    font=body(8, italic=True, color="FF595959"), align=left())
ws5.merge_cells("A2:D2")
ws5.row_dimensions[2].height = 40

r = 4
for c, txt, bg in [(1,"ACCT#",NAVY),(2,"ACCOUNT NAME",NAVY),
                   (3,"CATEGORY",NAVY),(4,"GL BALANCE (USD)",NAVY)]:
    col_header(ws5, r, c, txt, bg)
ws5.row_dimensions[r].height = 22

cat_colors = {"Operating": "FFEBF1DD", "Restricted": "FFFFD5D5", "Disbursement": LGREY}
cat_totals = {"Operating": 0.0, "Restricted": 0.0, "Disbursement": 0.0}

r = 5
cur_cat = None
for acct, name, cat, bal in bank_balances:
    if cat != cur_cat:
        # Category header
        ws5.row_dimensions[r].height = 18
        section_header(ws5, r, 1, 4, cat.upper(), "FF4472C4", WHITE)
        r += 1
        cur_cat = cat
    ws5.row_dimensions[r].height = 15
    bg = cat_colors.get(cat, WHITE)
    apply_cell(ws5, r, 1, acct,  font=body(9), fill_=fill(bg), align=center())
    apply_cell(ws5, r, 2, name,  font=body(9), fill_=fill(bg), align=left(1))
    apply_cell(ws5, r, 3, cat,   font=body(9), fill_=fill(bg), align=center())
    apply_cell(ws5, r, 4, bal,   font=body(9, color=BLACK_TXT if bal>=0 else RED_TXT),
               fill_=fill(bg), fmt=FMT_USD, align=right())
    for cc in range(1, 5):
        ws5.cell(r, cc).border = thin_border()
    cat_totals[cat] += bal
    r += 1

# Category subtotals
r += 1
section_header(ws5, r, 1, 4, "CATEGORY SUBTOTALS", "FF17375E", WHITE)
r += 1
for cat, tot in cat_totals.items():
    ws5.row_dimensions[r].height = 16
    apply_cell(ws5, r, 1, cat, font=body(10, True), fill_=fill(DGREY), align=left(1))
    ws5.merge_cells(f"A{r}:C{r}")
    apply_cell(ws5, r, 4, tot, font=body(10, True), fill_=fill(DGREY), fmt=FMT_USD, align=right())
    for cc in range(1, 5):
        ws5.cell(r, cc).border = thin_border('header')
    r += 1

r += 1
apply_cell(ws5, r, 1,
    "Note: 'Operating' total includes accounts with $3.1M CC in-transit and $2.7M lease operating. "
    "Disbursement account total is misleading due to 10639 JPM AP EFT which accumulates on the credit side. "
    "For available cash, focus on: 10611 DACA MMDA $2.83M, 10456 PNC Lease OpCo $2.70M, "
    "10631 JPM Asset Sales $2.77M, 10530 SHI $1.51M, 10460 Store Banks $0.97M, 10472 CC Transit $3.13M.",
    font=body(8, italic=True, color="FF595959"), align=left())
ws5.merge_cells(f"A{r}:D{r}")
ws5.row_dimensions[r].height = 48

# ════════════════════════════════════════════════════════════════════════════
# SHEET 6 – NetSuite Actuals & Reconciliation
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("NS Actuals & Recon")
ws6.sheet_view.showGridLines = False
ws6.column_dimensions['A'].width = 18
ws6.column_dimensions['B'].width = 22
ws6.column_dimensions['C'].width = 28
ws6.column_dimensions['D'].width = 18
ws6.column_dimensions['E'].width = 18

section_header(ws6, 1, 1, 5,
    "NETSUITE ACTUALS DETAIL · MAY 8, 2025", NAVY, WHITE)

r = 3
section_header(ws6, r, 1, 5, "CONFIRMED RECEIPTS (INFLOWS)", "FF17375E", WHITE)
r += 1
for c, txt, bg in [(1,"TRAN ID",NAVY),(2,"TYPE",NAVY),(3,"MEMO",NAVY),
                   (4,"BANK ACCT",NAVY),(5,"AMOUNT (USD)",NAVY)]:
    col_header(ws6, r, c, txt, bg)
ws6.row_dimensions[r].height = 22

ncr_data = [
    ("JE39909","Journal","NCR20250508_0000023161.0001","—",  -68834.75),
    ("JE39910","Journal","NCR20250508_0000023161.0002","—",  -60749.98),
    ("JE39911","Journal","NCR20250508_0000023161.0003","—",  -58653.41),
    ("JE39912","Journal","NCR20250508_0000023161.0004","—",   -5140.48),
]
r += 1
ncr_total = 0
for tid, typ, memo, acct, amt in ncr_data:
    ws6.row_dimensions[r].height = 15
    apply_cell(ws6, r, 1, tid,  font=body(9), fill_=fill(LLGREEN), align=center())
    apply_cell(ws6, r, 2, typ,  font=body(9), fill_=fill(LLGREEN))
    apply_cell(ws6, r, 3, memo, font=body(9), fill_=fill(LLGREEN), align=left(1))
    apply_cell(ws6, r, 4, acct, font=body(9), fill_=fill(LLGREEN))
    apply_cell(ws6, r, 5, abs(amt), font=body(9, color=GREEN_TXT), fill_=fill(LLGREEN),
               fmt=FMT_USD, align=right())
    ncr_total += abs(amt)
    r += 1

apply_cell(ws6, r, 4, "TOTAL NCR RECEIPTS",     font=body(10, True), fill_=fill(DGREY))
apply_cell(ws6, r, 5, ncr_total, font=body(10, True, BLACK_TXT), fill_=fill(DGREY),
           fmt=FMT_USD, align=right())
ws6.merge_cells(f"A{r}:D{r}")
r += 2

apply_cell(ws6, r, 1,
    "Note: NCR journals are NCR point-of-sale daily settlement entries (store deposits). "
    "Total $193,378 represents combined Sears+Kmart B&M store receipts. "
    "Cannot be split by entity without store-level mapping. "
    "31 additional Customer Payment transactions (PYMT1xxx) posted 5/8 — amounts netted to "
    "$0 at header level in NS due to AR/Bank double-entry; bank side credit not isolated.",
    font=body(8, italic=True, color="FF595959"), align=left())
ws6.merge_cells(f"A{r}:E{r}")
ws6.row_dimensions[r].height = 48

r += 3
section_header(ws6, r, 1, 5, "CONFIRMED DISBURSEMENTS", "FF17375E", WHITE)
r += 1
for c, txt, bg in [(1,"TRAN ID",NAVY),(2,"TYPE",NAVY),(3,"MEMO",NAVY),
                   (4,"BANK ACCT",NAVY),(5,"AMOUNT (USD)",NAVY)]:
    col_header(ws6, r, c, txt, bg)
ws6.row_dimensions[r].height = 22

# Top disbursements (showing key lines)
top_disb = [
    ("00000964/7",  "VendPymt","—",                "10639 JPM-NS AP EFT",  -94909.28),
    ("00000964/5",  "VendPymt","—",                "10639",                -94102.57),
    ("00000964/67", "VendPymt","—",                "10639",                -78360.13),
    ("00000964/89", "VendPymt","—",                "10639",                -66346.63),
    ("00000964/69", "VendPymt","—",                "10639",                -28904.34),
    ("00000964/102","VendPymt","—",                "10639",                -28633.07),
    ("00000964/6",  "VendPymt","—",                "10639",                -26003.37),
    ("00000964/91", "VendPymt","—",                "10639",                -22854.76),
    ("00000964/104","VendPymt","—",                "10639",                -17500.00),
    ("  [+95 more lines]","","(see full NS export)","10639/10617",          -172166.48),
    ("3401924",    "VendPymt","—",                "10617 JPM-NS AP CHK",   -2160.00),
    ("3401925",    "VendPymt","—",                "10617",                   -320.70),
]
r += 1
disb_total = 0
for tid, typ, memo, acct, amt in top_disb:
    ws6.row_dimensions[r].height = 15
    bg = LRED if "more" in tid.lower() else WHITE
    apply_cell(ws6, r, 1, tid,  font=body(9), fill_=fill(bg), align=center())
    apply_cell(ws6, r, 2, typ,  font=body(9), fill_=fill(bg))
    apply_cell(ws6, r, 3, memo, font=body(9), fill_=fill(bg), align=left(1))
    apply_cell(ws6, r, 4, acct, font=body(9), fill_=fill(bg), align=left(1))
    apply_cell(ws6, r, 5, amt,  font=body(9, color=RED_TXT), fill_=fill(bg),
               fmt=FMT_USD, align=right())
    disb_total += amt
    r += 1

apply_cell(ws6, r, 4, "TOTAL CONFIRMED DISBURSEMENTS", font=body(10, True), fill_=fill(DGREY))
apply_cell(ws6, r, 5, disb_total, font=body(10, True, RED_TXT), fill_=fill(DGREY),
           fmt=FMT_USD, align=right())
ws6.merge_cells(f"A{r}:D{r}")
r += 2

# Unconfirmed
section_header(ws6, r, 1, 5, "UNCONFIRMED / PARTIAL ITEMS", "FF833C00", WHITE)
r += 1
unconf = [
    ("JE40219","Journal","0508 Payment Multiple Invoices","Not in bank GL",  -684173.30,
     "Bank-side debit not in NS GL as of query. Likely ACH batch or intercompany."),
    ("PYMT1xxx","CustPymt","31 customer payments posted 5/8","Various",  0.00,
     "Net to $0 at header (AR+Bank offset). Bank credit not isolated same-day."),
]
for c, txt, bg in [(1,"TRAN ID",NAVY),(2,"TYPE",NAVY),(3,"MEMO",NAVY),
                   (4,"BANK ACCT",NAVY),(5,"AMOUNT (USD)",NAVY)]:
    col_header(ws6, r, c, txt, "FF833C00")
r += 1
for tid, typ, memo, acct, amt, note_t in unconf:
    ws6.row_dimensions[r].height = 15
    apply_cell(ws6, r, 1, tid,    font=body(9), fill_=fill(LYELLOW))
    apply_cell(ws6, r, 2, typ,    font=body(9), fill_=fill(LYELLOW))
    apply_cell(ws6, r, 3, memo,   font=body(9), fill_=fill(LYELLOW), align=left(1))
    apply_cell(ws6, r, 4, acct,   font=body(9), fill_=fill(LYELLOW))
    apply_cell(ws6, r, 5, amt,    font=body(9, color="FF833C00"), fill_=fill(LYELLOW),
               fmt=FMT_USD, align=right())
    r += 1
    ws6.row_dimensions[r].height = 24
    apply_cell(ws6, r, 1, f"  ↳ {note_t}", font=body(8, italic=True, color="FF595959"),
               fill_=fill(LYELLOW))
    ws6.merge_cells(f"A{r}:E{r}")
    r += 1

r += 1
# Reconciliation summary
section_header(ws6, r, 1, 5, "RECONCILIATION: FORECAST vs. NetSuite ACTUALS", "FF1F3864", WHITE)
r += 1
recon = [
    ("Inflows – Forecast",     f"${sum(v for _,v in inflows_fcst):.3f}M",   "FY File forecast"),
    ("Inflows – NS Confirmed", f"$0.193M",                                   "NCR journals only"),
    ("Inflows – Coverage",     f"{0.193/sum(v for _,v in inflows_fcst)*100:.0f}%", "NS vs. Forecast"),
    ("","",""),
    ("Disbursements – Forecast",    f"${abs(sum(v for _,v in disb_fcst)):.3f}M", "FY File forecast"),
    ("Disbursements – NS Confirmed",f"$0.631M",                              "VendPymts 10639+10617"),
    ("Disbursements – Coverage",    f"{0.631/abs(sum(v for _,v in disb_fcst))*100:.0f}%", "NS vs. Forecast"),
    ("","",""),
    ("Net CF – Forecast",      f"${cf_net:.3f}M",                           "Cash Flow G1440"),
    ("Net CF – NS (partial)",  f"$(0.438)M",                                 "NCR $0.193M – AP $0.631M (incomplete)"),
    ("","",""),
    ("Total Cash – Forecast",  f"${cf_total:.3f}M",                         "Cash Flow O1440 (K+M+N)"),
    ("Available Cash – Model", f"$0",                                        "K1440 = $0; HTS absorbs inflow"),
]
for c, txt, bg in [(1,"ITEM","FF4472C4"),(2,"AMOUNT","FF4472C4"),(3,"NOTE","FF4472C4")]:
    col_header(ws6, r, c, txt, "FF4472C4")
ws6.merge_cells(f"D{r}:E{r}")
r += 1
for i, (item, amt, note_t) in enumerate(recon):
    ws6.row_dimensions[r].height = 16
    bg = LGREY if i % 2 == 0 else WHITE
    if not item:
        r += 1
        continue
    apply_cell(ws6, r, 1, item,   font=body(10, "coverage" in item.lower() or "model" in item.lower() or "Total Cash" in item),
               fill_=fill(bg), align=left(1))
    apply_cell(ws6, r, 2, amt,    font=body(10, True, BLUE_TXT if "Forecast" in item else GREEN_TXT),
               fill_=fill(bg), align=right())
    apply_cell(ws6, r, 3, note_t, font=body(8, italic=True, color="FF595959"),
               fill_=fill(bg), align=left())
    ws6.merge_cells(f"C{r}:E{r}")
    for cc in range(1, 6):
        ws6.cell(r, cc).border = thin_border()
    r += 1

# ════════════════════════════════════════════════════════════════════════════
# SHEET 7 – Methodology
# ════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Methodology")
ws7.sheet_view.showGridLines = False
ws7.column_dimensions['A'].width = 22
ws7.column_dimensions['B'].width = 70

section_header(ws7, 1, 1, 2,
    "METHODOLOGY & DATA QUALITY NOTES", NAVY, WHITE)

notes = [
    ("DATA SOURCES",""),
    ("FY File (Forecast)","Master data spine. 113-column daily table indexed by date. "
     "Inflows Detail and Disbursement Detail sheets use INDEX/MATCH against FY File by date "
     "and column header. Cash Flow main sheet aggregates via SUM(CG:DM) per row. "
     "File date: March 5, 2025. May 8, 2025 was a FORECAST row in the original file."),
    ("NetSuite (Actuals)","Transform SR Holding Management LLC is a single subsidiary in NS. "
     "16 Business Units are departments, not separate legal entities — no BU-level split "
     "available from standard NS transaction queries."),
    ("Payroll Sheet","16-BU payroll format sourced from Treasury's internal 'Payroll' sheet "
     "within the original workbook. Values in actual USD (not millions)."),
    ("",""),
    ("MAPPING",""),
    ("Cash Flow → FY File","G = SUM(CG:DM). CG = SUM(Inflows Detail B:K) via INDEX/MATCH on FY File. "
     "CN (payroll) = ROUND(INDEX(Payroll sheet)/1000000, 2)."),
    ("Available Cash (K)","K = MAX(K_prior + G_today - Y_prior, 0). Y = HTS facility. "
     "Both May 7 and May 8 show K=0 — available cash fully consumed by HTS balance."),
    ("Total Cash (O)","O = K (Available) + M (Segregated $5.242M) + N (Unavailable $35.593M) "
     "= $40.835M. N = DV column (hardcoded from external model)."),
    ("",""),
    ("LIMITATIONS",""),
    ("Inflows not confirmed","Customer payment receipts in NS flow through AR sub-ledger. "
     "Bank debits for receipts do not appear in NS GL on the same business day. "
     "Only NCR point-of-sale journals ($193,378) confirmed as same-day bank activity."),
    ("AP partial only","NS confirmed $630,781 in AP vendor payments. JE40219 'Payment Multiple "
     "Invoices' ($684,173) has no confirmed bank-side debit in NS GL — likely ACH batch "
     "processed through bank portal not yet reflected. Total forecast AP = $1.409M."),
    ("Payroll not isolatable","Wells Fargo payroll account (10302) shows cumulative GL balance. "
     "No individual payroll transaction with date = 5/8 returned from NS queries."),
    ("BU breakout","NS subsidiary = single entity. All 16 BU payroll figures from Payroll sheet "
     "(Treasury internal model), not from NS actuals."),
    ("Bank balances","NS GL balances are CUMULATIVE from account inception, not bank statement "
     "balances. Use bank portals (JPM Access, PNC, WF) for true intraday balances."),
]

r = 3
for label, note_t in notes:
    ws7.row_dimensions[r].height = 15 if note_t else 22
    bold_label = not note_t
    bg_label   = NAVY if not note_t else WHITE
    fg_label   = WHITE if not note_t else NAVY
    if not note_t and label:
        # Section header row
        section_header(ws7, r, 1, 2, label, "FF17375E", WHITE)
        r += 1
        continue
    apply_cell(ws7, r, 1, label, font=body(10, True, NAVY), fill_=fill(LGREY), align=left(1))
    apply_cell(ws7, r, 2, note_t, font=body(9), fill_=fill(WHITE), align=left())
    ws7.row_dimensions[r].height = max(15, len(note_t) // 4) if note_t else 15
    for cc in range(1, 3):
        ws7.cell(r, cc).border = thin_border()
    r += 1

# ════════════════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════════════════
out_path = "/Users/josh/Desktop/Daily Cash Fcst - 5.8.25_Actuals.xlsx"
wb.save(out_path)
print(f"SAVED: {out_path}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
