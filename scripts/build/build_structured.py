import warnings; warnings.filterwarnings('ignore')
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pyxlsb import open_workbook as open_xlsb
import os, copy

SRC_XLSX = '/Users/josh/Desktop/Daily Cash Fcst - 3.5.25_BU view_SC payroll true up_$2M SC payroll hedge.xlsx'
SRC_XLSB = '/Users/josh/Desktop/Daily Cash Fcst - 05.08.26.xlsb'
DST      = '/Users/josh/Downloads/Daily Cash Fcst - 5.8.26_Structured.xlsx'

# ── Colour palette (matches actual file tone) ─────────────────────────────
NAVY   = "1F3864";  WHITE = "FFFFFF";  LBLUE = "D6E4F0"
MGREY  = "D9D9D9";  LGREY = "F2F2F2";  DGREY = "595959"
GREEN  = "E2EFDA";  AMBER = "FFF2CC";  RED   = "FFD5D5"

def hfont(sz=9, bold=True, color="FFFFFF"):
    return Font(name='Arial', size=sz, bold=bold, color=color)
def dfont(sz=9, bold=False, color="000000"):
    return Font(name='Arial', size=sz, bold=bold, color=color)
def fill(hex_): return PatternFill("solid", fgColor=hex_)
def ctr(wrap=False):  return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
def lft(ind=0): return Alignment(horizontal='left',   vertical='center', indent=ind)
def rgt():      return Alignment(horizontal='right',  vertical='center')

def thin():
    s = Side(style='thin', color='BBBBBB')
    return Border(left=s, right=s, top=s, bottom=s)
def hdr_border():
    m = Side(style='medium', color='1F3864')
    t = Side(style='thin',   color='BBBBBB')
    return Border(left=t, right=t, top=m, bottom=m)

def paint_header(ws, row, col, text, bg=NAVY, fg=WHITE, sz=9, bold=True, align=ctr(True)):
    c = ws.cell(row, col, text)
    c.font      = Font(name='Arial', size=sz, bold=bold, color=fg)
    c.fill      = fill(bg)
    c.alignment = align
    c.border    = hdr_border()
    return c

def paint_data(ws, row, col, val=None, bg=WHITE, bold=False, fmt=None, align=None):
    c = ws.cell(row, col)
    if val is not None: c.value = val
    c.font      = dfont(bold=bold)
    c.fill      = fill(bg)
    c.border    = thin()
    if fmt:   c.number_format = fmt
    if align: c.alignment = align
    return c

# ── Read existing AI xlsx data ────────────────────────────────────────────
print("Loading source xlsx …")
wb_src = load_workbook(SRC_XLSX, data_only=True)
print(f"  Sheets: {wb_src.sheetnames}")

# ── Helper: copy a whole sheet's cell values ─────────────────────────────
def copy_sheet_values(ws_src, ws_dst, max_rows=None, max_cols=None):
    mr = max_rows or ws_src.max_row
    mc = max_cols or ws_src.max_column
    for r in range(1, mr+1):
        for c in range(1, mc+1):
            v = ws_src.cell(r, c).value
            if v is not None:
                ws_dst.cell(r, c).value = v
    # copy col widths
    for col_letter, dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col_letter].width = dim.width or 10
    # copy row heights
    for rn, dim in ws_src.row_dimensions.items():
        if rn <= mr and dim.height:
            ws_dst.row_dimensions[rn].height = dim.height

# ── Build output workbook ─────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

TAB_ORDER = [
    'RE & Debt init with Manteno', 'Cash Flow', 'SOP',
    'KES & KCD Cash Flow', 'LC Forecast Changes', 'Trapped Cash',
    'RC', 'UBS reserve', '2020 Term Loans', 'Inflows Forecasting',
    'Inflows Actuals', 'Inflows Detail', 'Disbursement Actuals',
    'Disbursement Detail', 'FY File', 'MERCH', 'Non-Merch', 'Payroll',
]

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Cash Flow
# ════════════════════════════════════════════════════════════════════════════
print("Building: Cash Flow …")
ws_cf = wb.create_sheet('Cash Flow')
ws_src_cf = wb_src['Cash Flow']

# Row 1 = column-index reference row (actual file pattern)
# Row 2 = header labels
# Row 3 = grand-total row (space)
# Row 4+ = data

# Full header map matching actual file (col index → label)
CF_HDRS = {
    1:'Date', 4:'Week',
    5:'Operating Cash Flows', 6:'Non-Operating Cash Flows', 7:'Net Cash Flow',
    10:'Net Change in Cash', 11:'Available Cash', 12:'AC (No HTS)',
    13:'Segregated Account', 14:'Unavailable Cash', 15:'Total Cash',
    16:'TopCo Undrawn Equity',
    17:'April 2022 $110M PPM - Uncommitted Equity',
    18:'June 2022 $140M offering - Uncommitted Equity',
    19:'August 2022 $120M offering - Uncommitted Equity',
    20:'Revolver - DIP',
    21:'Cash Surplus / (Deficit) including Asset Sales',
    22:'Cash Surplus / (Deficit)', 23:'LC Opportunities',
    24:'Cash Surplus / (Deficit) + LC Opportunities',
    25:'Cash Requirement', 26:'Previous Surplus / (Deficit)',
    27:'Variance vs. Previous Forecast',
    28:'Tranche 1 Principal', 29:'Tranche 2 Principal',
    30:'Tranche 3 Principal', 31:'Tranche 4 Principal',
    32:'Cyrus Term Loan', 33:'Tranche 6 Principal',
    34:'Tranche 7 Principal', 35:'Tranche 8 Principal',
    36:'Tranche 9 Principal', 37:'Tranche 10 Principal',
    38:'Tranche 11 Principal', 39:'Cyrus $55M Term Loan',
    40:'ESL $15M Term Loan', 41:'ESL $16.5M Term Loan',
    42:'Cyrus $16.5M Term Loan', 43:'ESL $27M Term Loan',
    44:'ESL $50M Term Loan', 45:'Cyrus $25M Term Loan',
    46:'ESL $70M Term Loan', 47:'Cyrus $8M Term Loan',
    48:'ESL $175M Term Loan (T22)', 49:'Cyrus 7th Term Loan',
    50:'ESL 15th Term Loan', 51:'Sept. 2025 Term Loans',
    52:'Cyrus Eighth Incremental Term Loans',
    58:'Bohemia Loan (Mortgage)', 59:'Lease Opco Loan',
    60:'Lacey, WA Loan (Mortgage) [$11.2M Facility]',
    61:'New Brunswick Loan', 62:'Hackensack, NJ Loan (2022)',
    63:'Hackensack, NJ Loan (2026)', 64:'Wintrust/Manteno Loan',
    65:'Durham, NC Loan',
    66:'UBS Loan Note A', 67:'UBS Loan Note B',
    68:'Proceeds from Asset Sales ($250M Balance)', 69:'Guam Loan',
    70:'Cash Collateralization of BAML LC (5%)',
    71:'Citi LC', 72:'Citi to PNC (Back-to-Back LC 4% cash collateralization)',
    73:'PNC LC', 74:'Cash Collateralization of PNC LC (2%)',
    75:'Total Cash Collateralization of PNC/Citi LC',
    76:'UBS LC', 77:'Cash Collateralization of UBS LC (2%)',
    78:'Balance Sheet Inventory',
    79:'2019 Term Loan', 80:'2020 Term Loan', 81:'2020 Term Loan - LO',
    82:'ESL Note', 83:'HPS RE Loan',
    85:'Inflows', 86:'CITI Reimbursement', 87:'Debt/Financing',
    88:'Supply Chain', 89:'Asset Sales', 90:'Rx', 91:'Misc Inflows',
    92:'Payroll/Bens', 93:'Merch (CIA)', 94:'Merch (On Terms)',
    95:'Rent', 96:'Logistics', 97:'Advertising',
    98:'Other Non-Merch Disbursements', 99:'RiskMgt/Ins', 100:'Taxes',
    101:'RealEstate Transactions', 102:'HW Reserves & Deposits',
    103:'PropCo 2025', 104:'WU', 105:'P-Card', 106:'SHI',
    107:'Collateral', 108:'Interest', 109:'Fees', 110:'PropCo',
    111:'Cstc', 112:'Debt Repayment', 113:'Misc',
    115:'SG&A', 116:'Home Services', 117:'Store Closures',
    119:'Total Operating Receipts',
    120:'Merch/Non-Merch Disbursements',
    121:'Segregated Account Available Cash',
    122:'Inflows Total', 123:'Disbursements Total', 124:'Financing',
    125:'Cash Flow + Financing',
    126:'Unavail Cash Lookup from Model',
    127:'Change in BAML/PNC Cash Collateralization',
    128:'Excess Availability+ Need', 129:'Junior DIP Borrowing Need',
    130:'Total EA', 131:'Change in PNC/ Cash Collateralization',
    137:'Cash Flows no Asset Sales', 138:'Adj Available Cash',
    139:'Adj Cash Requirement',
    143:'Property taxes', 144:'Utilities', 145:'FACILITIES/CAPEX',
    146:'RE MISC', 147:'RE TOAL',
}

# Row 1: col-index reference
for col_idx, label in CF_HDRS.items():
    ws_cf.cell(1, col_idx).value = col_idx
    ws_cf.cell(1, col_idx).font = dfont(sz=8, color=DGREY)
    ws_cf.cell(1, col_idx).fill = fill(LGREY)

# Row 2: header labels
for col_idx, label in CF_HDRS.items():
    paint_header(ws_cf, 2, col_idx, label, bg=NAVY, sz=8, align=ctr(True))

# Row 3: grand-total placeholder
ws_cf.cell(3, 1).value = ' '
ws_cf.cell(3, 1).fill  = fill(AMBER)

# Set reasonable col widths
ws_cf.column_dimensions['A'].width = 12
for c in range(2, 148):
    ws_cf.column_dimensions[get_column_letter(c)].width = 9

# Copy data rows from AI Cash Flow (rows 4 onward → actual starts at row 3 of src)
# AI src: row 1 = DATE header (col A), row 2 = header labels, rows 4+ = data
# We'll copy AI rows 4-1709 into our rows 4+
MAX_ROW = min(ws_src_cf.max_row, 1720)
for src_r in range(4, MAX_ROW + 1):
    dst_r = src_r  # same row numbering
    # Col A = Date
    date_v = ws_src_cf.cell(src_r, 1).value
    if date_v is None:
        continue
    ws_cf.cell(dst_r, 1).value  = date_v
    ws_cf.cell(dst_r, 1).number_format = 'mmm d, yyyy'
    ws_cf.cell(dst_r, 1).font   = dfont()
    ws_cf.cell(dst_r, 1).fill   = fill(LGREY if str(date_v).replace('.','').isalpha() else WHITE)

    # Copy each column from AI source → matching column in our structure
    # AI file col mapping (openpyxl 1-based):
    # A=1(Date), B=2(day), C=3(month), D=4(week), E=5(OpCF), F=6(NonOpCF), G=7(NetCF)
    # J=10(NetChg), K=11(Avail), L=12(ACNoHTS), M=13(Seg), N=14(Unavail), O=15(Total)
    # CG=85(Inflows), CH=86(CITI), CI=87(Debt), CJ=88(SC), CK=89(Assets), CL=90(Rx)
    # CM=91(Misc), CN=92(Payroll), CO=93(MerchCIA), CP=94(MerchTerms)
    # CQ=95(Rent), CR=96(Log), CS=97(Adv), CT=98(OtherNM), CU=99(RiskIns)
    # CV=100(Tax), CW=101(RET), CX=102(HW), CY=103(PropCo25), CZ=104(WU)
    # DA=105(PCard), DB=106(SHI), DC=107(Collat), DD=108(Int), DE=109(Fees)
    # DF=110(PropCo), DG=111(Cstc), DH=112(DebtRep), DI=113(Misc)
    # DV=126(UnavailLookup)

    for ai_col, struct_col in [
        (2,'B'),(3,'C'),(4,'D'),
        (5,'E'),(6,'F'),(7,'G'),
        (10,'J'),(11,'K'),(12,'L'),(13,'M'),(14,'N'),(15,'O'),
        (85,'CG'),(86,'CH'),(87,'CI'),(88,'CJ'),(89,'CK'),(90,'CL'),
        (91,'CM'),(92,'CN'),(93,'CO'),(94,'CP'),(95,'CQ'),(96,'CR'),
        (97,'CS'),(98,'CT'),(99,'CU'),(100,'CV'),(101,'CW'),(102,'CX'),
        (103,'CY'),(104,'CZ'),(105,'DA'),(106,'DB'),(107,'DC'),
        (108,'DD'),(109,'DE'),(110,'DF'),(111,'DG'),(112,'DH'),(113,'DI'),
        (126,'DV'),
    ]:
        v = ws_src_cf.cell(src_r, ai_col).value
        if v is not None:
            col_letter = struct_col if isinstance(struct_col, str) else get_column_letter(struct_col)
            c = ws_cf[f"{col_letter}{dst_r}"]
            c.value  = v
            c.font   = dfont()
            c.border = thin()
            if isinstance(v, (int, float)):
                c.number_format = '#,##0.000;(#,##0.000);"-"'
                c.alignment = rgt()

ws_cf.freeze_panes = 'B3'
ws_cf.row_dimensions[1].height = 14
ws_cf.row_dimensions[2].height = 30
print(f"  Cash Flow: {ws_cf.max_row} rows")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: SOP
# ════════════════════════════════════════════════════════════════════════════
print("Building: SOP …")
ws_sop = wb.create_sheet('SOP')

sop_title_rows = [
    (1, 'Tab', 'Update Instructions', 'Source / Contact'),
    (2, 'Cash Flow', 'Manual changes only to Columns AB – BZ (Cap Table & term loan balances). All other columns formula-driven.', 'Cols AB–BY per Jon Goodin / CAP table. Col BZ per James Coutre.'),
    (3, None, 'All other columns are formula driven.', None),
    (4, 'KES & KCD Cash Flow', 'Copy / paste from Brands finance team.', None),
    (5, 'LC Forecast Changes', 'Per Jon Goodin.', None),
    (6, 'Trapped Cash', 'Cash deposited with PNC for LCs', 'PNC Bank Account 4803431873'),
    (7, None, 'Cash deposited with UBS for LCs', 'BNY Mellon Account 507032'),
    (8, None, 'Cash deposited with PNC for Cash Management', 'PNC Account 4645228406'),
    (9, None, 'Cash deposited with Citibank for Pcard', 'Citibank Account 40784284'),
    (10, None, 'Cash Holdback from Discover', 'Per Brett Bassett'),
    (11, None, 'Cash Holdback from FDC', 'Per Brett Bassett'),
    (12, None, 'Utility Deposits', 'GL Account 14175'),
    (13, None, 'Advances - Landlords & Sub-Tenants', 'GL Account 14180'),
    (14, None, 'Cash in segregated account for RE', 'Per Real Estate finance team'),
    (15, None, 'Holdback from Innovel - environmental', 'Per Luke Valentino'),
    (16, None, 'Real Estate Escrows', 'Per Real Estate finance team'),
    (17, None, 'Safelite rent reserve (Hackensack NJ)', 'Per Real Estate finance team'),
    (18, None, 'Drawn LC restricted cash (insurance/states)', 'Per Jon Goodin'),
    (19, None, 'Hackensack', 'Per Real Estate finance team'),
    (20, None, 'Bohemia', 'Per Real Estate finance team'),
    (21, None, 'Home Warranty reserves and deposits', 'Sum of PNC bank accounts for Warranty reserves'),
    (22, 'UBS reserve', 'Per Zach Straebel.', None),
    (23, '2020 Term Loans', 'Per Jon Goodin.', None),
    (24, 'Inflows Forecasting', 'Each Monday: update date in ABZ1 to prior Friday date; copy/paste values from Cols B:J to Cols ABZ:ACG.', None),
    (25, 'Inflows Actuals', 'From cash position Daily Detail file, named range INFLOWS.', None),
    (26, 'Inflows Detail', 'All formula driven.', None),
    (27, 'Disbursement Actuals', 'All formula driven.', None),
    (28, 'Disbursement Detail', 'All formula driven.', None),
    (29, 'FY File', 'HIDDEN — actual/forecast file for year.', None),
    (30, 'MERCH', 'HIDDEN — copy / paste from James Coutre.', None),
    (31, 'Non-Merch', 'HIDDEN — copy / paste from Adharvana.', None),
    (32, 'Payroll', 'HIDDEN — copy / paste from Payroll finance team.', None),
]

paint_header(ws_sop, 1, 1, 'Tab', bg=NAVY, sz=9)
paint_header(ws_sop, 1, 2, 'Update Instructions', bg=NAVY, sz=9, align=lft(1))
paint_header(ws_sop, 1, 3, 'Source / Contact', bg=NAVY, sz=9, align=lft(1))
ws_sop.column_dimensions['A'].width = 26
ws_sop.column_dimensions['B'].width = 58
ws_sop.column_dimensions['C'].width = 34

for row_num, tab, instr, src_note in sop_title_rows[1:]:
    bg = LGREY if row_num % 2 == 0 else WHITE
    if tab:
        c = ws_sop.cell(row_num, 1, tab)
        c.font = dfont(bold=True, color='1F3864')
        c.fill = fill(LBLUE)
    else:
        ws_sop.cell(row_num, 1).fill = fill(bg)
    for col_i, val in [(2, instr), (3, src_note)]:
        cell = ws_sop.cell(row_num, col_i, val or '')
        cell.font      = dfont()
        cell.fill      = fill(bg)
        cell.border    = thin()
        cell.alignment = lft(1)
    ws_sop.row_dimensions[row_num].height = 16

# Inventory section header
ws_sop.cell(34, 1, 'INVENTORY:').font = dfont(bold=True, color='1F3864')
ws_sop.cell(34, 1).fill = fill(LBLUE)
ws_sop.cell(34, 2, '(Balance sheet inventory schedule — populated by finance team per SOP instructions above)').font = Font(name="Arial", size=8, color=DGREY, italic=True)
print("  SOP: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: KES & KCD Cash Flow
# ════════════════════════════════════════════════════════════════════════════
print("Building: KES & KCD Cash Flow …")
ws_kes = wb.create_sheet('KES & KCD Cash Flow')
ws_src_kes = wb_src['KES & KCD Cash Flow']

KES_HDRS = [
    'Date','Day','Month','Week','Revolver Limit',
    'Starting AR','Gross adds (invoicing & other)',
    'Gross minuses (collections & other)','Ineligible AR',
    'Less AR Advance','Net Eligible AR','Starting Inventory',
    'Inventory Net of Advance Rate','Eligible Inventory',
    'Lesser of elig inven & sublimit','Dilution reserve',
    'Unapplied cash','Available Collateral','Starting Loan Balance',
    'Daily Collections','Collections (paydown from VNB to pay ABL)',
    'KES Wholesale Inflows (VNB)','Int & fees (Actual)',
    'Ending Loan Balance','Int & fees (Reserve)','Rent reserve',
    'Availability Block','Left to Borrow','KCD Royalty Inflows (PNC)',
]

# Row 1: col indices
for i, h in enumerate(KES_HDRS, 1):
    ws_kes.cell(1, i).value = i
    ws_kes.cell(1, i).font  = dfont(sz=8, color=DGREY)
    ws_kes.cell(1, i).fill  = fill(LGREY)
# Row 2: blank (actual has blank row 2)
ws_kes.row_dimensions[2].height = 8
# Row 3: headers
for i, h in enumerate(KES_HDRS, 1):
    paint_header(ws_kes, 3, i, h, bg=NAVY, sz=8, align=ctr(True))
# Row 4: placeholder dot
ws_kes.cell(4, 28).value = '.'

# Copy data from AI source (starting at row 6 in both)
copy_sheet_values(ws_src_kes, ws_kes, max_rows=min(ws_src_kes.max_row, 840))

ws_kes.column_dimensions['A'].width = 11
for c in range(2, 30):
    ws_kes.column_dimensions[get_column_letter(c)].width = 10
ws_kes.freeze_panes = 'E4'
print(f"  KES & KCD: {ws_kes.max_row} rows")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: LC Forecast Changes
# ════════════════════════════════════════════════════════════════════════════
print("Building: LC Forecast Changes …")
ws_lc = wb.create_sheet('LC Forecast Changes')
ws_src_lc = wb_src['LC Forecast Changes']
copy_sheet_values(ws_src_lc, ws_lc)

# Zone A headers (cols 11-28, rows 1-3)
zone_a_row1 = {18:'√', 22:'√'}
zone_a_row2 = {12:'Current', 16:'New'}
zone_a_row3 = {
    11:'Beneficiary', 12:'LC #', 13:'Facility', 14:'Face Amount',
    15:'Collateral Return/Draw Date', 17:'Facility', 18:'Face Amount',
    19:'Collateral Post Date',
    21:'NET CHANGE IN COLLATERAL\nBETTER / (WORSE)',
    22:'PENDING CHANGE IN COLLATERAL\nBETTER / (WORSE)',
    28:'Comments',
}
for col, txt in zone_a_row1.items():
    paint_header(ws_lc, 1, col, txt, bg=GREEN, fg='375623', sz=9)
for col, txt in zone_a_row2.items():
    paint_header(ws_lc, 2, col, txt, bg=LBLUE, fg='1F3864', sz=9)
for col, txt in zone_a_row3.items():
    paint_header(ws_lc, 3, col, txt, bg=NAVY, sz=8, align=ctr(True))

# Section labels
section_labels = {5:'Completed Prior:', 71:'Completed 2022:', 87:'Completed 2023:',
                  123:'Completed 2024:', 155:'Completed 2025:', 171:'Completed 2026:'}
for row, label in section_labels.items():
    c = ws_lc.cell(row, 11, label)
    c.font = dfont(bold=True, color='1F3864'); c.fill = fill(LBLUE)

ws_lc.column_dimensions['A'].width = 8
for c_idx in range(11, 29):
    ws_lc.column_dimensions[get_column_letter(c_idx)].width = 14
ws_lc.freeze_panes = 'K4'
print(f"  LC Forecast Changes: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Trapped Cash
# ════════════════════════════════════════════════════════════════════════════
print("Building: Trapped Cash …")
ws_tc = wb.create_sheet('Trapped Cash')
ws_src_tc = wb_src['Trapped Cash']
copy_sheet_values(ws_src_tc, ws_tc)

# Row 2 group headers
tc_hdrs = {
    3:'Cash — Other than Available Cash',
    6:'Restricted Cash',
    8:'Accounts Receivable / Other Assets',
    10:'Cash and Cash Equivalents',
    12:'Owner', 13:'Collection Date', 14:'Comments',
}
for col, txt in tc_hdrs.items():
    paint_header(ws_tc, 2, col, txt, bg=NAVY, sz=8, align=ctr(True))

ws_tc.column_dimensions['A'].width = 6
ws_tc.column_dimensions['C'].width = 42
for c_idx in [4,6,8,10]:
    ws_tc.column_dimensions[get_column_letter(c_idx)].width = 16
for c_idx in [12,13,14]:
    ws_tc.column_dimensions[get_column_letter(c_idx)].width = 14
ws_tc.freeze_panes = 'C3'
print(f"  Trapped Cash: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: RC
# ════════════════════════════════════════════════════════════════════════════
print("Building: RC …")
ws_rc = wb.create_sheet('RC')
ws_src_rc = wb_src['RC']
copy_sheet_values(ws_src_rc, ws_rc)

ws_rc.cell(1, 2, 'Restricted Cash as of August').font = dfont(bold=True, color='1F3864')
ws_rc.cell(2, 2, '$ USD, except where indicated otherwise').font = Font(name="Arial", size=8, color=DGREY, italic=True)
paint_header(ws_rc, 5, 3, 'Feb Balance', bg=NAVY, sz=9)
ws_rc.cell(28, 2, 'Total Balance').font = dfont(bold=True)
ws_rc.cell(28, 2).fill = fill(LGREY)
ws_rc.cell(28, 3).fill = fill(LGREY)
ws_rc.column_dimensions['B'].width = 46
ws_rc.column_dimensions['C'].width = 18
print(f"  RC: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: UBS reserve
# ════════════════════════════════════════════════════════════════════════════
print("Building: UBS reserve …")
ws_ubs = wb.create_sheet('UBS reserve')
ws_src_ubs = wb_src['UBS reserve']
copy_sheet_values(ws_src_ubs, ws_ubs)

ubs_hdrs_r2 = {
    2:'Maturity', 3:'Types', 4:'Lender', 5:'Property Count',
    6:'Loan Amount', 7:'Lit Value', 8:'Loan/Lit', 9:'Pricing',
    11:'Interest Expense', 12:'Working Capital', 13:'ADA & Life Safety',
    14:'Deferred Maintenance', 15:'General Reserve',
    16:'Insurance Holdback', 17:'Total Reserves',
}
for col, txt in ubs_hdrs_r2.items():
    paint_header(ws_ubs, 2, col, txt, bg=NAVY, sz=8, align=ctr(True))

# Second header block (CMBS)
ubs_hdrs_r9 = {
    2:'Maturity', 3:'Loan Type', 4:'Lender', 5:'Property Count',
    6:'Loan Amount', 9:'Pricing',
    12:'RE Tax Escrow', 13:'Insurance Escrow',
    14:'TI/LC Reserve', 15:'CapEx Reserve', 17:'Total Reserves',
}
for col, txt in ubs_hdrs_r9.items():
    paint_header(ws_ubs, 9, col, txt, bg=LBLUE, fg='1F3864', sz=8, align=ctr(True))

ws_ubs.column_dimensions['A'].width = 8
for c_idx in range(2, 18):
    ws_ubs.column_dimensions[get_column_letter(c_idx)].width = 14
print(f"  UBS reserve: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: 2020 Term Loans
# ════════════════════════════════════════════════════════════════════════════
print("Building: 2020 Term Loans …")
ws_tl = wb.create_sheet('2020 Term Loans')
ws_src_tl = wb_src['2020 Term Loans']
copy_sheet_values(ws_src_tl, ws_tl)

# Row 1 title
ws_tl.cell(1, 3, '2020 Term Loan & 2021 Term Loan').font = dfont(bold=True, color='1F3864', sz=11)

TL_HDRS = [
    'Tranche 1 Principal','Tranche 2 Principal','SKIPPED',
    'Tranche 3 Principal','Tranche 4 Principal','Cyrus Term Loan',
    'Tranche 6 Principal','Tranche 7 Principal','Tranche 8 Principal',
    'Tranche 9 Principal','Tranche 10 Principal','Tranche 11 Principal',
    'Cyrus $55M Term Loan','ESL $15M Term Loan','ESL $16.5M Term Loan',
    'Cyrus $16.5M Term Loan','ESL $27M Term Loan','ESL $50M Term Loan',
    'Cyrus $25M Term Loan','ESL $70M Term Loan','Cyrus $8M Term Loan',
    'ESL $175M Term Loan (T22)','Cyrus 7th Term Loan','ESL 15th Term Loan',
    'Sept. 2025 Term Loans','Cyrus Eighth Incremental Term Loans',
]
for i, h in enumerate(TL_HDRS, 4):
    paint_header(ws_tl, 3, i, h, bg=NAVY, sz=8, align=ctr(True))

ws_tl.column_dimensions['C'].width = 36
for c_idx in range(4, 30):
    ws_tl.column_dimensions[get_column_letter(c_idx)].width = 11
ws_tl.freeze_panes = 'D4'
print(f"  2020 Term Loans: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: RE & Debt init with Manteno
# ════════════════════════════════════════════════════════════════════════════
print("Building: RE & Debt init with Manteno …")
ws_re = wb.create_sheet('RE & Debt init with Manteno')
ws_src_re = wb_src['RE & Debt init with Manteno']
copy_sheet_values(ws_src_re, ws_re)

ws_re.cell(1, 2, 'LIQUIDITY PRO FORMA').font = dfont(bold=True, color='1F3864', sz=12)
paint_header(ws_re, 2, 4, 'Scenario A: Debt exchange closes 12/30/22', bg=LBLUE, fg='1F3864', sz=9)
paint_header(ws_re, 2, 12, 'Scenario B: Debt exchange closes 1/6/23', bg=LBLUE, fg='1F3864', sz=9)

re_hdrs = {
    2:'Date', 3:'Action Due', 4:'UBS Reserve', 5:'UBS Loan Upsize',
    6:'Cash Interest Extension', 7:'Loan Extension',
    8:'Additional Term Loan Cash Interest Expense',
    9:'Total (Exchange 12/30)', 10:'Scenario A Cumulative',
    12:'Debt Exchange Post 12/30',
    13:'Total (Exchange post-12/30)', 14:'Scenario B Cumulative',
    16:'Current Forecast: Cash Surplus/(Deficit)',
    17:'Cash Surplus / Deficit (Exchange 12/30)',
    18:'Cash Surplus / Deficit (Exchange post-12/30)',
    19:'Notes',
}
for col, txt in re_hdrs.items():
    paint_header(ws_re, 4, col, txt, bg=NAVY, sz=8, align=ctr(True))

ws_re.column_dimensions['B'].width = 11
ws_re.column_dimensions['C'].width = 24
for c_idx in range(4, 20):
    ws_re.column_dimensions[get_column_letter(c_idx)].width = 12
ws_re.freeze_panes = 'C5'
print(f"  RE & Debt init: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Inflows Forecasting
# ════════════════════════════════════════════════════════════════════════════
print("Building: Inflows Forecasting …")
ws_if = wb.create_sheet('Inflows Forecasting')
ws_src_if = wb_src['Inflows Forecasting']

# Row 1: version labels; Row 2: sub-column headers
# Actual pattern: col 0=Wk, then every 7 cols = one forecast version
SUB_COLS = ['Sears Stores','Kmart Stores','Home Services',
            'KCD Wholesale','KCD (Royalty)','Online',
            'Service Live Summary','Tenant Income','Total']

ws_if.cell(1, 1, 'Wk').font = dfont(bold=True)
ws_if.cell(1, 2, 'Current Actual/Forecast').font = dfont(bold=True, color='1F3864')
ws_if.cell(1, 2).fill = fill(LBLUE)

# Header row 2: sub-cols for current block
for i, h in enumerate(SUB_COLS, 2):
    paint_header(ws_if, 2, i, h, bg=NAVY, sz=8, align=ctr(True))

ws_if.column_dimensions['A'].width = 6
for c_idx in range(2, 12):
    ws_if.column_dimensions[get_column_letter(c_idx)].width = 14

# Copy data from AI src
copy_sheet_values(ws_src_if, ws_if, max_rows=min(ws_src_if.max_row, 450), max_cols=60)
ws_if.freeze_panes = 'B3'
print(f"  Inflows Forecasting: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Inflows Actuals
# ════════════════════════════════════════════════════════════════════════════
print("Building: Inflows Actuals …")
ws_ia = wb.create_sheet('Inflows Actuals')
ws_src_ia = wb_src['Inflows Actuals']

IA_HDRS_R1 = {
    1:'DATE', 2:'Sears Stores', 3:'Kmart Stores', 4:'Home Services',
    5:'KCD', 6:'Tenant Income', 7:'Auto Centers',
    8:'SHS Assurant Profit Sharing', 9:'Builder Distributors',
    10:'Sears Financial Services', 11:'Service Live Summary',
    12:'HTS', 13:'HTS Prefund', 14:'Costco', 15:'Supply Chain',
    16:'CITI Reimbursement', 17:'CCHS', 18:'Asset Sales', 19:'Rx',
    20:'PropCo/Lease OpCo', 21:'KES', 22:'Sears Mexico (KCD)',
    23:'UBS', 24:'SHI Dividend', 25:'Debt/Financing',
    26:'Misc Inflows', 27:'Daily Total',
    42:'Online', 43:'SEARS B&M', 44:'Sears Protect',
    45:'Assurant', 46:'HS Other',
}
IA_HDRS_R2 = {
    2:'Sears Stores', 3:'Kmart Stores', 4:'Home Services',
    5:'KCD Wholesale', 6:'Tenant Income', 7:'Auto Centers',
    9:'Monark', 10:'Sears Financial Services', 11:'Service Live',
    12:'HTS', 13:'HTS Prefund', 14:'Costco', 15:'Supply Chain',
    16:'CITI Reimbursement', 17:'Cross Country Home Services',
    18:'Asset Sales', 19:'Rx', 20:'PropCo/Lease OpCo',
    21:'KES', 22:'Sears Mexico (KCD)', 23:'UBS Reserve',
    24:'SHI Dividend', 25:'Debt/Financing', 26:'Misc Inflows',
    27:'Daily Total',
}

for col, txt in IA_HDRS_R1.items():
    paint_header(ws_ia, 1, col, txt, bg=NAVY, sz=8, align=ctr(True))
for col, txt in IA_HDRS_R2.items():
    paint_header(ws_ia, 2, col, txt, bg=LBLUE, fg='1F3864', sz=8, align=ctr(True))

copy_sheet_values(ws_src_ia, ws_ia, max_rows=min(ws_src_ia.max_row, 1260), max_cols=50)

ws_ia.column_dimensions['A'].width = 12
for c_idx in range(2, 48):
    ws_ia.column_dimensions[get_column_letter(c_idx)].width = 10
ws_ia.freeze_panes = 'B3'
print(f"  Inflows Actuals: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Inflows Detail
# ════════════════════════════════════════════════════════════════════════════
print("Building: Inflows Detail …")
ws_id2 = wb.create_sheet('Inflows Detail')
ws_src_id2 = wb_src['Inflows Detail']

ID_HDRS_R1 = {
    1:'DATE', 2:'Sears Stores', 3:'Kmart Stores', 4:'Home Services',
    5:'KCD (Royalty)', 6:'Tenant Income', 7:'Online',
    13:'Fidem', 14:'NEXT2', 15:'Supply Chain', 16:'CITI Reimburse',
    17:'CCHS', 18:'Asset Sales', 19:'PH 2.0', 20:'Lease OpCo',
    21:'KES', 22:'Sears Mexico', 23:'UBS Reserve',
    24:'SHI Dividend', 25:'New Debt', 26:'Misc Inflows',
}
ID_HDRS_R4 = {
    2:'Sears Stores', 3:'Kmart Stores', 4:'Home Services',
    5:'KCD', 6:'Tenant Income', 7:'Online',
    8:'SHS Assurant Profit Sharing',
    13:'Fidem/Aress', 14:'Risk/Insurance', 15:'Supply Chain',
    16:'CITI Reimbursement', 17:'Cross Country Home Services',
    18:'Asset Sales', 19:'PropCo 2025', 20:'PropCo/Lease OpCo',
    21:'KES', 22:'Sears Mexico (KCD)', 23:'UBS Reserve',
    24:'SHI Dividend', 25:'Debt/Financing', 26:'Misc Inflows',
    27:'Daily Total',
}

for col, txt in ID_HDRS_R1.items():
    paint_header(ws_id2, 1, col, txt, bg=NAVY, sz=8, align=ctr(True))
for col, txt in ID_HDRS_R4.items():
    paint_header(ws_id2, 4, col, txt, bg=LBLUE, fg='1F3864', sz=8, align=ctr(True))

copy_sheet_values(ws_src_id2, ws_id2, max_rows=min(ws_src_id2.max_row, 270), max_cols=50)

ws_id2.column_dimensions['A'].width = 12
for c_idx in range(2, 30):
    ws_id2.column_dimensions[get_column_letter(c_idx)].width = 11
ws_id2.freeze_panes = 'B5'
print(f"  Inflows Detail: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Disbursement Actuals
# ════════════════════════════════════════════════════════════════════════════
print("Building: Disbursement Actuals …")
ws_da = wb.create_sheet('Disbursement Actuals')
ws_src_da = wb_src['Disbursement Actuals']

DA_HDR_R0 = {
    1:'DATE', 2:'Finance', 3:'FinancialSvcs', 4:'HoldingsCo',
    5:'Home Services', 6:'HR', 7:'KCD', 8:'Legal', 9:'Member Tech',
    10:'Monark', 11:'MSO', 12:'Real Estate', 13:'Retail',
    14:'Retail Online', 15:'Service Live', 16:'SYWR', 17:'Supply Chain',
    18:'PYRL/BENS TOTAL', 19:'Merch CIA (HS)', 20:'Merch Terms (HS)',
    21:'Retail - BM CIA', 22:'Retail - BM Terms', 34:'Grand Total Merch',
    37:'Google', 38:'Other Advertising', 39:'Total Advertising',
    40:'Logistics - UPS', 41:'Logistics - SHS', 43:'Logistics - Other',
    44:'Total Logistics', 51:'Tax', 52:'WU', 54:'Total Rent',
    57:'Outside Serv/Assoc Exp', 63:'Other Services', 64:'Supplies & Postage',
    67:'Raise Mktplc', 70:'Facilities', 73:'SHI', 79:'PNC P-Card Total',
    82:'PropCo', 83:'RE Transactions', 84:'COLLATERAL',
    85:'Misc', 86:'Debt Repayment', 87:'Cash Interest', 88:'Cash Fees',
    90:'Daily Total',
}
DA_HDR_R3 = {
    2:'Finance', 3:'FinancialSvcs', 4:'HoldingsCo', 5:'Home Services',
    6:'HR', 7:'KCD', 8:'Legal', 9:'Member Tech', 10:'Monark',
    11:'MSO', 12:'Real Estate', 13:'Retail', 14:'Retail Online',
    15:'Service Live', 16:'SYWR', 17:'Supply Chain', 18:'PYRL/BENS TOTAL',
    19:'Merch CIA', 20:'Merch On Terms', 21:'Total Home Services',
    22:'Merch CIA', 23:'Merch On Terms', 24:'Total Retail - Stores',
    25:'Merch CIA', 26:'Merch On Terms', 27:'Total Online',
    28:'Merch CIA', 29:'Merch On Terms', 30:'Total KCD',
    31:'Merch CIA', 32:'Merch On Terms', 33:'Total HTS',
    34:'Merch CIA', 35:'Merch On Terms', 36:'MERCH - TOTAL',
    37:'Google', 38:'Other Advertising', 39:'Total Advertising',
    40:'Logistics - UPS', 41:'SHS Logistics', 42:'DC capital',
    43:'Logistics - Other', 44:'Total Logistics', 45:'BYOV Reimburse (HS)',
    46:'Home Services Assurant', 47:'Cinch (Home Warranty)',
    48:'HW Reserves & Deposits', 49:'HelloSuper (Home Warranty)',
    50:'Market place', 51:'RiskMgt/Ins', 52:'Taxes', 53:'Western Union',
    54:'Rent', 55:'Property Tax', 56:'CapEx', 57:'Utilities & Telephone',
    58:'Outside Serv/Assoc Exp', 63:'Other Services',
    64:'Supplies & Postage', 65:'Other Disbursements',
    66:'Legal Services', 67:'Raise Marketplace',
    68:'SHS Customer Refunds', 69:'Facilities',
    70:'1099 contractors', 71:'Enterprise leases', 73:'SHI',
    79:'Total P-Card', 80:'Citi Bad Debt Collection',
    81:'St. Jude Collections', 82:'PropCo 2025', 83:'PropCo',
    84:'RealEstate Transactions', 85:'Collateral: [Sources/(Uses)]',
    86:'Misc', 87:'Debt Repayment', 88:'Cash Interest', 89:'Cash Fees',
    90:'Daily Total',
}

for col, txt in DA_HDR_R0.items():
    paint_header(ws_da, 1, col, txt, bg=NAVY, sz=7, align=ctr(True))
# Group labels row 3
paint_header(ws_da, 3, 2, 'PAYROLL', bg=LBLUE, fg='1F3864', sz=8)
paint_header(ws_da, 3, 19, 'HOME SERVICES', bg=LBLUE, fg='1F3864', sz=8)
paint_header(ws_da, 3, 22, 'Total Retail - Stores', bg=LBLUE, fg='1F3864', sz=8)
paint_header(ws_da, 3, 25, 'Online Business Unit', bg=LBLUE, fg='1F3864', sz=8)
paint_header(ws_da, 3, 28, 'KCD', bg=LBLUE, fg='1F3864', sz=8)
paint_header(ws_da, 3, 31, 'HTS', bg=LBLUE, fg='1F3864', sz=8)
paint_header(ws_da, 3, 34, 'Merch TOTALS', bg=LBLUE, fg='1F3864', sz=8)
for col, txt in DA_HDR_R3.items():
    paint_header(ws_da, 5, col, txt, bg=LBLUE, fg='1F3864', sz=7, align=ctr(True))

copy_sheet_values(ws_src_da, ws_da, max_rows=min(ws_src_da.max_row, 420), max_cols=95)

ws_da.column_dimensions['A'].width = 12
for c_idx in range(2, 96):
    ws_da.column_dimensions[get_column_letter(c_idx)].width = 9
ws_da.freeze_panes = 'B6'
print(f"  Disbursement Actuals: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Disbursement Detail
# ════════════════════════════════════════════════════════════════════════════
print("Building: Disbursement Detail …")
ws_dd = wb.create_sheet('Disbursement Detail')
ws_src_dd = wb_src['Disbursement Detail']

# Same header structure as Disbursement Actuals
for col, txt in DA_HDR_R0.items():
    paint_header(ws_dd, 1, col, txt, bg=NAVY, sz=7, align=ctr(True))
paint_header(ws_dd, 3, 2,  'PAYROLL',                  bg=LBLUE, fg='1F3864', sz=8)
paint_header(ws_dd, 3, 19, 'HOME SERVICES',             bg=LBLUE, fg='1F3864', sz=8)
paint_header(ws_dd, 3, 22, 'Total Retail - Stores',     bg=LBLUE, fg='1F3864', sz=8)
paint_header(ws_dd, 3, 25, 'Online Business Unit',      bg=LBLUE, fg='1F3864', sz=8)
paint_header(ws_dd, 3, 28, 'KCD',                       bg=LBLUE, fg='1F3864', sz=8)
paint_header(ws_dd, 3, 31, 'HTS',                       bg=LBLUE, fg='1F3864', sz=8)
paint_header(ws_dd, 3, 34, 'Merch TOTALS',              bg=LBLUE, fg='1F3864', sz=8)
for col, txt in DA_HDR_R3.items():
    paint_header(ws_dd, 5, col, txt, bg=LBLUE, fg='1F3864', sz=7, align=ctr(True))

copy_sheet_values(ws_src_dd, ws_dd, max_rows=min(ws_src_dd.max_row, 270), max_cols=101)

ws_dd.column_dimensions['A'].width = 12
for c_idx in range(2, 102):
    ws_dd.column_dimensions[get_column_letter(c_idx)].width = 9
ws_dd.freeze_panes = 'B6'
print(f"  Disbursement Detail: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: FY File
# ════════════════════════════════════════════════════════════════════════════
print("Building: FY File …")
ws_fy = wb.create_sheet('FY File')
ws_src_fy = wb_src['FY File']

# Row 1: section labels
ws_fy.cell(1, 1, 46150).number_format = 'mmm d, yyyy'  # as-of date
paint_header(ws_fy, 1, 2, 'INFLOWS', bg=GREEN, fg='375623', sz=9)
paint_header(ws_fy, 1, 3, 'MERCH',   bg=AMBER, fg='7F6000', sz=9)
paint_header(ws_fy, 1, 4, 'NON-MERCH', bg=LBLUE, fg='1F3864', sz=9)
paint_header(ws_fy, 1, 5, 'PAYROLL', bg=RED, fg='C00000', sz=9)

# Row 2: sub-labels
ws_fy.cell(2, 2, 'Inflows').font = Font(name="Arial", size=8, color=DGREY, italic=True)

# Row 3 (primary header row)
FY_HDRS = [
    'Date','Sears Stores','Kmart Stores','Home Services','(gap)',
    'KCD Wholesale','KCD (Royalty)','Online','(gap)',
    'Service Live Summary','Tenant Income','Fidem','New Debt',
    'Supply Chain','CITI Reimburse','CCHS','Asset Sales','Rx',
    'Lease OpCo','Sears Mexico','PH 2.0','KES','UBS Reserve',
    'SHI Dividend','NEXT2','NEXT3','NEXT4','NEXT5','NEXT6',
    'Misc Inflows','Notes','(gap)','AP','Liq.Init',
    'PYRL KmartRetail','PYRL SearsRetail','PYRL HomeServices',
    'COLLATERAL','PH 2.0 out','RE Transactions','PYRL/BENS TOTAL',
    'RiskMgt/Ins','Tax','CstcOut','HW Reserves & Deposits',
    'Western Union','PNC P-Card','SHI','x',
    'Debt Repayment','Cash Interest','Cash Fees','PropCo','Misc',
    'Total','Notes','Total All Inflows/Disbursement','Wk No','Fiscal year',
]
for i, h in enumerate(FY_HDRS, 1):
    paint_header(ws_fy, 3, i, h, bg=NAVY, sz=8, align=ctr(True))

copy_sheet_values(ws_src_fy, ws_fy, max_rows=min(ws_src_fy.max_row, 350), max_cols=60)

ws_fy.column_dimensions['A'].width = 12
for c_idx in range(2, 60):
    ws_fy.column_dimensions[get_column_letter(c_idx)].width = 10
ws_fy.freeze_panes = 'B4'
print(f"  FY File: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: MERCH
# ════════════════════════════════════════════════════════════════════════════
print("Building: MERCH …")
ws_merch = wb.create_sheet('MERCH')
ws_src_merch = wb_src['Merch']

# Row 4 = primary headers (actual file)
MERCH_R4 = [
    'Date','Week','Week Day','Fiscal Month','Fiscal Month Name',
    'CIA','Terms','Total',
    'CIA','Terms','Total HS',
    'CIA','Terms','Total Retail - Stores',
    'CIA','Terms','Online Business Unit',
    'CIA','Terms','Total KCD',
]
MERCH_R3 = {
    6:'Actuals/Forecast', 9:'Home Services', 12:'Total Retail - Stores',
    15:'Online Business Unit', 18:'KCD Support Unit',
}
for col, txt in MERCH_R3.items():
    paint_header(ws_merch, 3, col, txt, bg=LBLUE, fg='1F3864', sz=9)
for i, h in enumerate(MERCH_R4, 1):
    paint_header(ws_merch, 4, i, h, bg=NAVY, sz=8, align=ctr(True))

copy_sheet_values(ws_src_merch, ws_merch, max_rows=min(ws_src_merch.max_row, 400), max_cols=20)

ws_merch.column_dimensions['A'].width = 12
for c_idx in range(2, 21):
    ws_merch.column_dimensions[get_column_letter(c_idx)].width = 12
ws_merch.freeze_panes = 'A5'
print(f"  MERCH: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Non-Merch
# ════════════════════════════════════════════════════════════════════════════
print("Building: Non-Merch …")
ws_nm = wb.create_sheet('Non-Merch')
ws_src_nm = wb_src['Non-Merch']

NM_HDRS = [
    'Date','Payroll/Bens','CIA Merch','Non-CIA Merch','(gap)','Total Merch',
    'Google','Microsoft','Other Advertising','Total Advertising',
    'HelloSuper','Outside Serv/Assoc Exp','Utilities & Telephone',
    'Logistics - UPS','Logistics - Costco','Logistics - SHS',
    'Logistics - Other','Logistics','RE Transactions',
    '1099 contractors','Home Services Assurant','Cinch HW',
    'Raise Mktplc','Replacements/Customer checks','Enterprise leases',
    'Property Tax - Checks','Outside Services - Checks',
    'Occupancy Repairs - Checks','Sales Tax - Checks','Lottery',
    'Utilities & Telephone - Checks','Advertising Expense - Checks',
    'Equipment Expense - Checks','Supplies & Postage - Checks',
    'Merch Checks',
    'State Federal and Foreign Income Tax - Checks',
    'Security & Cleaning Service - Checks',
    'Legal Expense - Checks','Other - Checks','Non-Rent Checks',
    'Equipment - ARI','Equipment - Amazon','Equipment - CA Inc',
    'Equipment - Other','Equipment','CapEx','Facilities',
    'Marketplace','Citi Bad Debt Collection','St. Jude Collections',
    'Legal','Other Services (Store Closure/Security)',
    'Supplies & Postage','Real Estate Taxes','Go Forward Store Rent',
    'GOB/Closed Store Rent','Other Rent (DC/Office)','Total Rent',
    'Cure','Tax','Western Union/CARPACH','BOA P-Card',
    'Global Sourcing','Commercial Paper','Debt Repayment',
    'Cash Interest','Cash Fees','Intercompany (net)','Misc','Total',
    'Misc Notes','Sum','(gap)','Check Point','Difference',
    '(gap)','Corporate','Home Services','Online','Real Estate',
    'Retail B&M','SYW','Sum','Variance',
    '(gap)','NM ACH FCST','NM Checks Fcst',
]

# Row 1: as-of date marker
ws_nm.cell(1, 1, 46150).number_format = 'mmm d, yyyy'
ws_nm.cell(1, 1).font = dfont(bold=True)

# Row 3: headers
for i, h in enumerate(NM_HDRS, 1):
    paint_header(ws_nm, 3, i, h, bg=NAVY, sz=7, align=ctr(True))

copy_sheet_values(ws_src_nm, ws_nm, max_rows=min(ws_src_nm.max_row, 230), max_cols=87)

ws_nm.column_dimensions['A'].width = 12
for c_idx in range(2, 88):
    ws_nm.column_dimensions[get_column_letter(c_idx)].width = 9
ws_nm.freeze_panes = 'B4'
print(f"  Non-Merch: done")

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Payroll
# ════════════════════════════════════════════════════════════════════════════
print("Building: Payroll …")
ws_pay = wb.create_sheet('Payroll')
ws_src_pay = wb_src['Payroll']

# Row 1: as-of date + section labels
ws_pay.cell(1, 1, 46150).number_format = 'mmm d, yyyy'
paint_header(ws_pay, 1, 2,  "Treasury's 16 BU Format: Forecast",
             bg=LBLUE, fg='1F3864', sz=9)
paint_header(ws_pay, 1, 20, "Treasury's 16 BU Format: Actuals",
             bg=GREEN, fg='375623', sz=9)

# Row 2: BU column headers
PAYROLL_BUS_FCST = [
    'Banking days','Finance','FinancialSvcs','HoldingsCo','Home Services',
    'BYOV Reimburse (HS)','HR','KCD','Legal','Member Tech',
    'MSO','Real Estate','Retail','Retail Online','Service Live',
    'SYWR','Supply Chain','Total',
]
PAYROLL_BUS_ACT = [
    '(gap)','Finance','FinancialSvcs','HoldingsCo','Home Services',
    'BYOV Reimburse (HS)','HR','KCD','Legal','Member Tech','Monark',
    'MSO','Real Estate','Retail','Retail Online','Service Live',
    'SYWR','Supply Chain','Total',
]
for i, h in enumerate(PAYROLL_BUS_FCST, 1):
    paint_header(ws_pay, 2, i, h, bg=NAVY, sz=8, align=ctr(True))
for i, h in enumerate(PAYROLL_BUS_ACT, 19):
    paint_header(ws_pay, 2, i, h, bg=NAVY, sz=8, align=ctr(True))

# Divider col 18
ws_pay.column_dimensions['R'].width = 3
ws_pay.cell(2, 18).fill = fill(LGREY)

copy_sheet_values(ws_src_pay, ws_pay, max_rows=min(ws_src_pay.max_row, 470), max_cols=38)

ws_pay.column_dimensions['A'].width = 12
for c_idx in range(2, 18):
    ws_pay.column_dimensions[get_column_letter(c_idx)].width = 11
for c_idx in range(19, 38):
    ws_pay.column_dimensions[get_column_letter(c_idx)].width = 11
ws_pay.freeze_panes = 'B3'
print(f"  Payroll: done")

# ════════════════════════════════════════════════════════════════════════════
# Reorder sheets to match target
# ════════════════════════════════════════════════════════════════════════════
current = wb.sheetnames
ordered = [t for t in TAB_ORDER if t in current]
wb._sheets = [wb[name] for name in ordered]

print(f"\nFinal sheet order ({len(wb.sheetnames)} tabs):")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"  {i:2}. {name}")

wb.save(DST)
print(f"\nSaved → {DST}")
print(f"Size: {os.path.getsize(DST)/1024/1024:.1f} MB")
