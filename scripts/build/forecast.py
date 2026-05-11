"""6-Month Curate Coins Cash Flow & Inventory Projection."""

import json
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----- Inputs -----
SPREADSHEET = '/Users/josh/.openclaw/media/inbound/a49b353f-c14f-44a8-a5e2-f8a583d8e4d9.xlsx'
SOLD = json.load(open('/tmp/sold.json'))
SPOT = 81.88  # Today's silver
EBAY_FEE_PCT = 0.13
SHIPPING_PER_ITEM = 2.0
SUPPLIES_PER_ITEM = 0.50
WEEKLY_BUDGET = 120.0
MONTHS = 6

# ----- Read collection -----
def f(x):
    if x is None: return 0.0
    s = str(x).replace('$','').replace(',','').strip()
    try: return float(s)
    except: return 0.0

wb_in = openpyxl.load_workbook(SPREADSHEET, data_only=True)
ws_in = wb_in.active
header = [c.value for c in ws_in[1]]
inv = []
for row in ws_in.iter_rows(min_row=2, values_only=True):
    if any(c is not None for c in row):
        inv.append(dict(zip(header, row)))

# ----- Realized P&L (last 60d) -----
sold_total_gross = sum(f(s['price']) * f(s['qty_sold']) for s in SOLD)
sold_count = sum(int(f(s['qty_sold']) or 1) for s in SOLD)
fees_total = sold_total_gross * EBAY_FEE_PCT
ship_supply_total = sold_count * (SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM)
sold_net = sold_total_gross - fees_total - ship_supply_total
sold_avg = sold_total_gross / max(sold_count, 1)

# Velocity
velocity_per_month = sold_count / 2.0  # 60 days = 2 months

# ----- Inventory segmentation -----
# Core (don't liquidate): Inherited + premium numismatic
NUMISMATIC_CATS = {'Proof Set', 'Silver Proof Set', 'Mint Set', 'Ancient Coin', 'Indian Head Cent',
                   'Indian Head Nickel', 'Liberty Head Nickel', 'Buffalo Nickel', 'Standing Liberty Quarter',
                   'Barber Coin', 'Confederate Money', 'Silver Certificate', 'Currency',
                   'Shipwreck', 'Medal', 'Diamond', 'Picture'}
core_value, core_cost, core_count = 0, 0, 0
flip_value, flip_cost, flip_count = 0, 0, 0

for r in inv:
    cat = r.get('Category') or 'Unknown'
    cost = f(r.get('Cost Paid'))
    high = f(r.get('High Value'))
    qty = f(r.get('Qty'))
    is_inherited = cost == 0
    is_numismatic_core = cat in NUMISMATIC_CATS
    if is_inherited or is_numismatic_core:
        core_value += high
        core_cost += cost
        core_count += qty
    else:
        flip_value += high
        flip_cost += cost
        flip_count += qty

# ----- Forecast scenarios -----
# Scenario assumptions: monthly sales count and avg price
scenarios = {
    'Conservative': {'sales_per_month': 8, 'avg_price': 100, 'reinvest_pct': 0.30},
    'Base':         {'sales_per_month': 15, 'avg_price': 120, 'reinvest_pct': 0.40},
    'Aggressive':   {'sales_per_month': 25, 'avg_price': 135, 'reinvest_pct': 0.50},
}

# ----- Build workbook -----
wb = openpyxl.Workbook()

# Styles
HDR_FILL = PatternFill('solid', fgColor='2C3E50')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='34495E')
SECTION_FONT = Font(bold=True, color='FFFFFF', size=12)
TOTAL_FILL = PatternFill('solid', fgColor='ECF0F1')
TOTAL_FONT = Font(bold=True)
GOOD_FILL = PatternFill('solid', fgColor='D5F5E3')
BAD_FILL = PatternFill('solid', fgColor='FADBD8')
THIN = Side(border_style='thin', color='BDC3C7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_row(ws, row, vals, font=None, fill=None, fmt=None, border=True):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        if font: c.font = font
        if fill: c.fill = fill
        if fmt: c.number_format = fmt
        if border: c.border = BORDER

# ----- Tab 1: Summary -----
ws = wb.active
ws.title = 'Summary'
ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18

ws['A1'] = 'Curate Coins — 6-Month Outlook'
ws['A1'].font = Font(bold=True, size=16)
ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M MT")}  |  Spot silver: ${SPOT}/oz'
ws['A2'].font = Font(italic=True, size=10, color='7F8C8D')

ws['A4'] = 'CURRENT STATE (60-day actuals)'
ws['A4'].fill = SECTION_FILL; ws['A4'].font = SECTION_FONT
ws.merge_cells('A4:D4')

state = [
    ('Items sold (60d)', sold_count, '0'),
    ('Gross revenue (60d)', sold_total_gross, '$#,##0.00'),
    ('eBay fees (13%)', -fees_total, '$#,##0.00'),
    ('Shipping & supplies', -ship_supply_total, '$#,##0.00'),
    ('NET (60d)', sold_net, '$#,##0.00'),
    ('Avg sale price', sold_avg, '$#,##0.00'),
    ('Sales velocity (per month)', velocity_per_month, '0.0'),
]
for i, (label, val, fmt) in enumerate(state, 5):
    ws.cell(row=i, column=1, value=label).font = Font(bold=(label.startswith('NET')))
    c = ws.cell(row=i, column=2, value=val)
    c.number_format = fmt
    if label.startswith('NET'): c.font = Font(bold=True); c.fill = GOOD_FILL

# Inventory snapshot
ws['A14'] = 'INVENTORY SNAPSHOT'
ws['A14'].fill = SECTION_FILL; ws['A14'].font = SECTION_FONT
ws.merge_cells('A14:D14')

write_row(ws, 15, ['Segment', 'Count (units)', 'Cost basis', 'High value'], font=HDR_FONT, fill=HDR_FILL)
write_row(ws, 16, ['Core / inherited / numismatic (DO NOT LIQUIDATE)', core_count, core_cost, core_value], fmt=None)
ws.cell(row=16, column=3).number_format = '$#,##0'
ws.cell(row=16, column=4).number_format = '$#,##0'
write_row(ws, 17, ['Flippable bullion + common silver', flip_count, flip_cost, flip_value])
ws.cell(row=17, column=3).number_format = '$#,##0'
ws.cell(row=17, column=4).number_format = '$#,##0'
write_row(ws, 18, ['TOTAL', core_count + flip_count, core_cost + flip_cost, core_value + flip_value], font=TOTAL_FONT, fill=TOTAL_FILL)
ws.cell(row=18, column=3).number_format = '$#,##0'
ws.cell(row=18, column=4).number_format = '$#,##0'

# 6-month projection summary
ws['A20'] = '6-MONTH PROJECTION (3 scenarios)'
ws['A20'].fill = SECTION_FILL; ws['A20'].font = SECTION_FONT
ws.merge_cells('A20:D20')

write_row(ws, 21, ['Scenario', 'Conservative', 'Base', 'Aggressive'], font=HDR_FONT, fill=HDR_FILL)
metrics = [
    ('Sales / month',           scenarios['Conservative']['sales_per_month'],
                                 scenarios['Base']['sales_per_month'],
                                 scenarios['Aggressive']['sales_per_month']),
    ('Avg sale price',           scenarios['Conservative']['avg_price'],
                                 scenarios['Base']['avg_price'],
                                 scenarios['Aggressive']['avg_price']),
    ('Gross / month', None, None, None),
    ('Net / month (after 13% + ship)', None, None, None),
    ('Net / 6 months', None, None, None),
    ('Reinvestment % of net',    scenarios['Conservative']['reinvest_pct']*100,
                                 scenarios['Base']['reinvest_pct']*100,
                                 scenarios['Aggressive']['reinvest_pct']*100),
    ('Reinvestment / 6 months', None, None, None),
    ('Cash to Josh / 6 months', None, None, None),
]
for i, m in enumerate(metrics, 22):
    write_row(ws, i, list(m))

# Compute and fill rows 24, 25, 26, 28, 29
for col, (name, sc) in enumerate(scenarios.items(), 2):
    sales = sc['sales_per_month']
    price = sc['avg_price']
    gross = sales * price
    net_unit_fee = price * EBAY_FEE_PCT + SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM
    net_per_month = sales * (price - net_unit_fee)
    net_6m = net_per_month * 6
    reinvest_6m = net_6m * sc['reinvest_pct']
    cash_6m = net_6m - reinvest_6m

    ws.cell(row=24, column=col, value=gross).number_format = '$#,##0'
    ws.cell(row=25, column=col, value=net_per_month).number_format = '$#,##0'
    ws.cell(row=26, column=col, value=net_6m).number_format = '$#,##0'
    ws.cell(row=26, column=col).font = Font(bold=True)
    ws.cell(row=27, column=col).number_format = '0%'
    ws.cell(row=27, column=col, value=sc['reinvest_pct'])
    ws.cell(row=28, column=col, value=reinvest_6m).number_format = '$#,##0'
    ws.cell(row=29, column=col, value=cash_6m).number_format = '$#,##0'
    ws.cell(row=29, column=col).font = Font(bold=True); ws.cell(row=29, column=col).fill = GOOD_FILL

# ----- Tab 2: Realized P&L -----
ws2 = wb.create_sheet('Realized 60d')
ws2.column_dimensions['A'].width = 60
for col in 'BCDEFG':
    ws2.column_dimensions[col].width = 14

ws2['A1'] = 'Realized P&L — Last 60 days'
ws2['A1'].font = Font(bold=True, size=14)
write_row(ws2, 3, ['Item', 'Qty', 'Sale', 'Fees (13%)', 'Ship+Sup', 'Net', 'End Date'], font=HDR_FONT, fill=HDR_FILL)
for i, s in enumerate(SOLD, 4):
    qty = int(f(s['qty_sold']) or 1)
    sale = f(s['price']) * qty
    fees = sale * EBAY_FEE_PCT
    ship = qty * (SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM)
    net = sale - fees - ship
    end = (s.get('end_time') or '')[:10]
    write_row(ws2, i, [s['title'][:80], qty, sale, fees, ship, net, end])
    for col in [3,4,5,6]:
        ws2.cell(row=i, column=col).number_format = '$#,##0.00'

last = len(SOLD) + 4
write_row(ws2, last, ['TOTAL', sold_count, sold_total_gross, fees_total, ship_supply_total, sold_net, ''], font=TOTAL_FONT, fill=TOTAL_FILL)
for col in [3,4,5,6]:
    ws2.cell(row=last, column=col).number_format = '$#,##0.00'

# ----- Tab 3: Cash Flow Timeline -----
ws3 = wb.create_sheet('Cash Flow 6mo')
for col in range(1, 11):
    ws3.column_dimensions[get_column_letter(col)].width = 14
ws3.column_dimensions['A'].width = 24

ws3['A1'] = '6-Month Cash Flow Timeline (Base scenario)'
ws3['A1'].font = Font(bold=True, size=14)

months_labels = []
start = datetime.now().replace(day=1)
for i in range(MONTHS):
    m = start.replace(month=((start.month + i - 1) % 12) + 1, year=start.year + ((start.month + i - 1) // 12))
    months_labels.append(m.strftime('%b %Y'))

write_row(ws3, 3, ['Month'] + months_labels + ['6-Mo Total'], font=HDR_FONT, fill=HDR_FILL)

base = scenarios['Base']
sales = base['sales_per_month']
price = base['avg_price']
gross = sales * price
net = sales * (price - (price * EBAY_FEE_PCT + SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM))
reinvest = net * base['reinvest_pct']
free_cash = net - reinvest
silver_buy_budget = WEEKLY_BUDGET * 4.33

# Realistic ramp: 60% in M1, 75% M2, 90% M3, 100% M4-M6 (account for listing capacity build)
ramp = [0.60, 0.75, 0.90, 1.00, 1.00, 1.00]

rows_data = [
    ('Listings created (cumulative basis)', [int(sales * r * 1.6) for r in ramp]),  # ~60% sell-through, so list more
    ('Sales / month',                        [int(sales * r) for r in ramp]),
    ('Gross revenue',                        [gross * r for r in ramp]),
    ('eBay fees (13%)',                      [-(gross * r) * EBAY_FEE_PCT for r in ramp]),
    ('Shipping + supplies',                  [-(int(sales * r)) * (SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM) for r in ramp]),
    ('NET REVENUE',                          [net * r for r in ramp]),
    ('Inventory reinvestment (40%)',         [-(net * r * base['reinvest_pct']) for r in ramp]),
    ('Silver-buy budget ($120/wk)',          [-silver_buy_budget for _ in ramp]),
    ('FREE CASH TO JOSH',                    None),  # computed below
]

for ridx, (label, vals) in enumerate(rows_data, 4):
    ws3.cell(row=ridx, column=1, value=label)
    if label.startswith('NET') or label.startswith('FREE'):
        ws3.cell(row=ridx, column=1).font = Font(bold=True)
    if vals is None:
        # compute free cash
        for col_i, r in enumerate(ramp, 2):
            free = (net * r) - (net * r * base['reinvest_pct']) - silver_buy_budget
            c = ws3.cell(row=ridx, column=col_i, value=free)
            c.number_format = '$#,##0'
            c.font = Font(bold=True)
            c.fill = GOOD_FILL if free > 0 else BAD_FILL
        total = sum((net * r) - (net * r * base['reinvest_pct']) - silver_buy_budget for r in ramp)
        c = ws3.cell(row=ridx, column=8, value=total)
        c.number_format = '$#,##0'
        c.font = Font(bold=True); c.fill = GOOD_FILL if total > 0 else BAD_FILL
    else:
        for col_i, v in enumerate(vals, 2):
            c = ws3.cell(row=ridx, column=col_i, value=v)
            if isinstance(v, float):
                c.number_format = '$#,##0'
            else:
                c.number_format = '#,##0'
        # Total
        total = sum(vals)
        c = ws3.cell(row=ridx, column=8, value=total)
        if isinstance(vals[0], float):
            c.number_format = '$#,##0'
        else:
            c.number_format = '#,##0'
        c.font = Font(bold=True)

ws3['A14'] = 'Notes:'
ws3['A14'].font = Font(bold=True)
notes = [
    f"• eBay payouts release ~7-10 days after sale (delivery confirmation triggers funds release)",
    f"• Ramp assumes listing throughput builds from current ~5 active to ~50 active by Month 4",
    f"• Reinvestment 40% feeds Tab 4 inventory replenishment math",
    f"• Silver-buy budget ($120/wk = $520/mo) is a Josh-set hard cap, separate from reinvestment",
    f"• Free cash to Josh = net - reinvestment - silver-buy budget",
]
for i, n in enumerate(notes, 15):
    ws3.cell(row=i, column=1, value=n)

# ----- Tab 4: Inventory Replenishment -----
ws4 = wb.create_sheet('Inventory Plan')
for col in 'ABCDEF':
    ws4.column_dimensions[col].width = 22

ws4['A1'] = 'Inventory Replenishment — Avoid Eating the Core'
ws4['A1'].font = Font(bold=True, size=14)

ws4['A3'] = 'CURRENT INVENTORY SEGMENTATION'
ws4['A3'].fill = SECTION_FILL; ws4['A3'].font = SECTION_FONT
ws4.merge_cells('A3:F3')

write_row(ws4, 4, ['Segment', 'Units', 'Cost', 'High value', 'Avg cost/unit', 'Avg value/unit'], font=HDR_FONT, fill=HDR_FILL)
core_avg_cost = core_cost / max(core_count, 1)
core_avg_value = core_value / max(core_count, 1)
flip_avg_cost = flip_cost / max(flip_count, 1)
flip_avg_value = flip_value / max(flip_count, 1)
write_row(ws4, 5, ['Core (DO NOT TOUCH)', core_count, core_cost, core_value, core_avg_cost, core_avg_value])
write_row(ws4, 6, ['Flippable', flip_count, flip_cost, flip_value, flip_avg_cost, flip_avg_value])
for r in [5,6]:
    for col in [3,4,5,6]:
        ws4.cell(row=r, column=col).number_format = '$#,##0.00'

ws4['A8'] = 'BURN-DOWN ANALYSIS (BASE SCENARIO)'
ws4['A8'].fill = SECTION_FILL; ws4['A8'].font = SECTION_FONT
ws4.merge_cells('A8:F8')

# Without replenishment, how long does flippable inventory last at 15 sales/mo?
sales_per_mo = scenarios['Base']['sales_per_month']
months_until_burn = flip_count / sales_per_mo

write_row(ws4, 9, ['Metric', 'Value', '', '', '', ''], font=HDR_FONT, fill=HDR_FILL)
rows = [
    ('Flippable units on hand', flip_count, ''),
    ('Sales / month (Base)', sales_per_mo, ''),
    ('Months until flippable inventory exhausted (no replenishment)', round(months_until_burn, 1), ''),
    ('Avg cost / flippable unit', flip_avg_cost, '$#,##0.00'),
    ('Replenishment needed @ 100% replacement (units/mo)', sales_per_mo, ''),
    ('Replenishment $$/mo @ 100% replacement', sales_per_mo * flip_avg_cost, '$#,##0'),
    ('Replenishment $$/mo from 40% reinvestment', net * 0.40, '$#,##0'),
    ('Silver-buy budget ($120/wk)', silver_buy_budget, '$#,##0'),
    ('Total replenishment capital available / mo', net * 0.40 + silver_buy_budget, '$#,##0'),
    ('SHORTFALL (negative = collection drawdown risk)', (net * 0.40 + silver_buy_budget) - sales_per_mo * flip_avg_cost, '$#,##0'),
]
for i, (lbl, val, fmt) in enumerate(rows, 10):
    ws4.cell(row=i, column=1, value=lbl)
    c = ws4.cell(row=i, column=2, value=val)
    if fmt: c.number_format = fmt
    if 'SHORTFALL' in lbl:
        ws4.cell(row=i, column=1).font = Font(bold=True)
        c.font = Font(bold=True)
        c.fill = BAD_FILL if val < 0 else GOOD_FILL

# Recommendations
ws4['A21'] = 'RECOMMENDATIONS'
ws4['A21'].fill = SECTION_FILL; ws4['A21'].font = SECTION_FONT
ws4.merge_cells('A21:F21')

recs = [
    "1. Target 15 sales/mo (Base) — needs ~25 active listings rotating at any time",
    f"2. Replenish at avg ${flip_avg_cost:.0f}/unit × 15 units/mo = ${sales_per_mo*flip_avg_cost:.0f}/mo capital needed",
    f"3. 40% reinvestment ({net*0.40:.0f}/mo) + $120/wk silver buys ({silver_buy_budget:.0f}/mo) = ${net*0.40+silver_buy_budget:.0f}/mo available",
    f"4. {'SHORTFALL of ${:.0f}/mo'.format(sales_per_mo*flip_avg_cost - (net*0.40+silver_buy_budget)) if (net*0.40+silver_buy_budget) < sales_per_mo*flip_avg_cost else 'Replenishment fully covered'} — adjust by raising reinvestment % or lowering avg sale price",
    "5. Lean on bullion premium (NORFEDs, Libertads, ASE rolls) — high turnover, low cost basis",
    "6. NEVER liquidate Core (inherited + premium numismatic) — that's the appreciation engine",
    "7. Track core_value vs flippable_value monthly — if flippable shrinks below 30% of total, reset listing strategy",
]
for i, r in enumerate(recs, 22):
    ws4.cell(row=i, column=1, value=r)
    ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

# ----- Tab 5: Sensitivity -----
ws5 = wb.create_sheet('Sensitivity')
for col in 'ABCDEFGH':
    ws5.column_dimensions[col].width = 14
ws5.column_dimensions['A'].width = 22

ws5['A1'] = 'Sensitivity Analysis — 6-Month Net to Josh'
ws5['A1'].font = Font(bold=True, size=14)

ws5['A3'] = 'Sales/month (rows) × Avg price (cols)'
ws5['A3'].font = Font(italic=True, size=10)

prices = [80, 100, 120, 140, 160]
sales_range = [5, 10, 15, 20, 25, 30]

write_row(ws5, 4, [''] + [f'${p}' for p in prices], font=HDR_FONT, fill=HDR_FILL)
for i, s in enumerate(sales_range, 5):
    row_vals = [f'{s}/mo']
    for p in prices:
        net_per_month = s * (p - (p * EBAY_FEE_PCT + SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM))
        net_6m = net_per_month * 6
        # Subtract reinvestment at 40%
        cash = net_6m * 0.60
        row_vals.append(cash)
    write_row(ws5, i, row_vals)
    for col in range(2, 2 + len(prices)):
        ws5.cell(row=i, column=col).number_format = '$#,##0'
        v = ws5.cell(row=i, column=col).value
        if v < 5000: ws5.cell(row=i, column=col).fill = BAD_FILL
        elif v > 15000: ws5.cell(row=i, column=col).fill = GOOD_FILL

ws5['A12'] = 'Spot price sensitivity (impacts melt floor + bullion velocity)'
ws5['A12'].font = Font(bold=True)
write_row(ws5, 13, ['Spot $/oz', 'Melt floor: 90% half', 'Melt floor: ASE', 'Notes'], font=HDR_FONT, fill=HDR_FILL)
spots = [70, 75, 80, SPOT, 85, 90, 95]
for i, sp in enumerate(spots, 14):
    half_melt = sp * 0.3617
    ase_melt = sp
    note = '(today)' if abs(sp - SPOT) < 0.01 else ''
    write_row(ws5, i, [sp, half_melt, ase_melt, note])
    ws5.cell(row=i, column=1).number_format = '$#,##0.00'
    ws5.cell(row=i, column=2).number_format = '$#,##0.00'
    ws5.cell(row=i, column=3).number_format = '$#,##0.00'

# ----- Tab 6: Inventory Detail -----
ws6 = wb.create_sheet('Inventory by Category')
for col in 'ABCDEFGHIJ':
    ws6.column_dimensions[col].width = 16
ws6.column_dimensions['A'].width = 28

ws6['A1'] = 'Inventory Detail by Category'
ws6['A1'].font = Font(bold=True, size=14)

write_row(ws6, 3, ['Category', 'Segment', 'Entries', 'Units', 'Silver oz', 'Cost', 'Low value', 'High value', 'Profit ($)', 'ROI %'], font=HDR_FONT, fill=HDR_FILL)

from collections import defaultdict
agg = defaultdict(lambda: {'entries':0,'qty':0,'oz':0,'cost':0,'low':0,'high':0,'profit':0})
for r in inv:
    cat = r.get('Category') or 'Unknown'
    a = agg[cat]
    a['entries'] += 1
    a['qty'] += f(r.get('Qty'))
    a['oz'] += f(r.get('Total Silver Oz'))
    a['cost'] += f(r.get('Cost Paid'))
    a['low'] += f(r.get('Low Value'))
    a['high'] += f(r.get('High Value'))
    a['profit'] += f(r.get('Profit ($)'))

row_i = 4
for cat in sorted(agg.keys()):
    a = agg[cat]
    seg = 'Core' if cat in NUMISMATIC_CATS else 'Flippable'
    if a['cost'] == 0: seg = 'Core (inherited)'
    roi = (a['profit'] / a['cost'] * 100) if a['cost'] > 0 else 0
    write_row(ws6, row_i, [cat, seg, a['entries'], int(a['qty']), round(a['oz'],1), a['cost'], a['low'], a['high'], a['profit'], roi])
    for col in [6,7,8,9]:
        ws6.cell(row=row_i, column=col).number_format = '$#,##0'
    ws6.cell(row=row_i, column=10).number_format = '0.0%'
    if seg == 'Flippable':
        ws6.cell(row=row_i, column=2).fill = GOOD_FILL
    row_i += 1

# ----- Save -----
out = '/Users/josh/.openclaw/workspace/curate-coins-6mo-forecast.xlsx'
wb.save(out)
print(f"\n=== Saved: {out} ===\n")

# ----- Print key takeaways for Josh -----
print("Headline numbers (Base scenario, 6 months):")
base_sc = scenarios['Base']
gross_6m = base_sc['sales_per_month'] * base_sc['avg_price'] * 6
net_per_mo = base_sc['sales_per_month'] * (base_sc['avg_price'] - (base_sc['avg_price']*EBAY_FEE_PCT + SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM))
net_6m = net_per_mo * 6
reinvest_6m = net_6m * base_sc['reinvest_pct']
cash_6m = net_6m - reinvest_6m
print(f"  Gross: ${gross_6m:,.0f}")
print(f"  Net (after fees+ship): ${net_6m:,.0f}")
print(f"  Reinvestment (40%): ${reinvest_6m:,.0f}")
print(f"  Cash to Josh: ${cash_6m:,.0f}")
print(f"\nInventory:")
print(f"  Flippable units: {flip_count:.0f} (~{flip_count/15:.1f} months at 15 sales/mo without replenishment)")
print(f"  Avg flip cost: ${flip_avg_cost:.2f}")
print(f"  Replenishment $/mo needed: ${15*flip_avg_cost:.0f}")
print(f"  Replenishment $/mo available: ${net_per_mo*0.40 + silver_buy_budget:.0f}")
shortfall = (net_per_mo*0.40 + silver_buy_budget) - 15*flip_avg_cost
print(f"  {'SHORTFALL' if shortfall < 0 else 'SURPLUS'}: ${shortfall:,.0f}/mo")
