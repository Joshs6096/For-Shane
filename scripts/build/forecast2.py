"""6-Month Curate Coins Forecast — WITH $1,000 board seed."""

import json
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----- Inputs -----
SPREADSHEET = '/Users/josh/.openclaw/media/inbound/a49b353f-c14f-44a8-a5e2-f8a583d8e4d9.xlsx'
SOLD = json.load(open('/tmp/sold.json'))
SPOT = 81.88
EBAY_FEE_PCT = 0.13
SHIPPING_PER_ITEM = 2.0
SUPPLIES_PER_ITEM = 0.50
WEEKLY_BUDGET = 120.0
MONTHS = 6
BOARD_SEED = 1000.0  # NEW

def f(x):
    if x is None: return 0.0
    s = str(x).replace('$','').replace(',','').strip()
    try: return float(s)
    except: return 0.0

wb_in = openpyxl.load_workbook(SPREADSHEET, data_only=True)
ws_in = wb_in.active
header = [c.value for c in ws_in[1]]
inv = []
for row in ws_in.iter_rows(min_row=2, values_only=True):
    if any(c is not None for c in row):
        inv.append(dict(zip(header, row)))

# Realized
sold_total_gross = sum(f(s['price']) * f(s['qty_sold']) for s in SOLD)
sold_count = sum(int(f(s['qty_sold']) or 1) for s in SOLD)
fees_total = sold_total_gross * EBAY_FEE_PCT
ship_supply_total = sold_count * (SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM)
sold_net = sold_total_gross - fees_total - ship_supply_total
sold_avg = sold_total_gross / max(sold_count, 1)
velocity_per_month = sold_count / 2.0

# Inventory segmentation
NUMISMATIC_CATS = {'Proof Set', 'Silver Proof Set', 'Mint Set', 'Ancient Coin', 'Indian Head Cent',
                   'Indian Head Nickel', 'Liberty Head Nickel', 'Buffalo Nickel', 'Standing Liberty Quarter',
                   'Barber Coin', 'Confederate Money', 'Silver Certificate', 'Currency',
                   'Shipwreck', 'Medal', 'Diamond', 'Picture'}
core_value, core_cost, core_count = 0, 0, 0
flip_value, flip_cost, flip_count = 0, 0, 0
for r in inv:
    cat = r.get('Category') or 'Unknown'
    cost = f(r.get('Cost Paid'))
    high = f(r.get('High Value'))
    qty = f(r.get('Qty'))
    is_inherited = cost == 0
    is_numismatic_core = cat in NUMISMATIC_CATS
    if is_inherited or is_numismatic_core:
        core_value += high; core_cost += cost; core_count += qty
    else:
        flip_value += high; flip_cost += cost; flip_count += qty

flip_avg_cost = flip_cost / max(flip_count, 1)
units_buyable_with_seed = int(BOARD_SEED / flip_avg_cost)

# ----- Scenarios with seed -----
# Seed front-loads M1 inventory, allowing faster ramp + higher steady-state listing cadence
scenarios = {
    'Conservative+Seed': {'sales_per_month': 12, 'avg_price': 100, 'reinvest_pct': 0.30, 'ramp': [0.80, 0.95, 1.00, 1.00, 1.00, 1.00]},
    'Base+Seed':         {'sales_per_month': 20, 'avg_price': 125, 'reinvest_pct': 0.40, 'ramp': [0.85, 1.00, 1.00, 1.00, 1.00, 1.00]},
    'Aggressive+Seed':   {'sales_per_month': 30, 'avg_price': 140, 'reinvest_pct': 0.50, 'ramp': [0.90, 1.00, 1.00, 1.00, 1.00, 1.00]},
}
# Without-seed for comparison
scenarios_no_seed = {
    'Conservative': {'sales_per_month': 8, 'avg_price': 100, 'reinvest_pct': 0.30, 'ramp': [0.60, 0.75, 0.90, 1.00, 1.00, 1.00]},
    'Base':         {'sales_per_month': 15, 'avg_price': 120, 'reinvest_pct': 0.40, 'ramp': [0.60, 0.75, 0.90, 1.00, 1.00, 1.00]},
    'Aggressive':   {'sales_per_month': 25, 'avg_price': 135, 'reinvest_pct': 0.50, 'ramp': [0.60, 0.75, 0.90, 1.00, 1.00, 1.00]},
}

def project(sc, with_seed=False):
    sales = sc['sales_per_month']
    price = sc['avg_price']
    ramp = sc['ramp']
    silver_buy_budget = WEEKLY_BUDGET * 4.33

    monthly = []
    for m, r in enumerate(ramp):
        m_sales = sales * r
        m_gross = m_sales * price
        m_fees = m_gross * EBAY_FEE_PCT
        m_ship = m_sales * (SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM)
        m_net = m_gross - m_fees - m_ship
        m_reinvest = m_net * sc['reinvest_pct']
        # Seed deployed in M1
        m_seed_deployed = BOARD_SEED if (with_seed and m == 0) else 0
        # Cash to Josh = net - reinvestment - silver_buy_budget
        m_cash = m_net - m_reinvest - silver_buy_budget
        monthly.append({
            'sales': m_sales, 'gross': m_gross, 'fees': m_fees, 'ship': m_ship,
            'net': m_net, 'reinvest': m_reinvest, 'silver_buy': silver_buy_budget,
            'seed_deployed': m_seed_deployed, 'cash': m_cash
        })
    return monthly

# ----- Build workbook -----
wb = openpyxl.Workbook()
HDR_FILL = PatternFill('solid', fgColor='2C3E50')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='34495E')
SECTION_FONT = Font(bold=True, color='FFFFFF', size=12)
SEED_FILL = PatternFill('solid', fgColor='F39C12')
SEED_FONT = Font(bold=True, color='FFFFFF')
TOTAL_FILL = PatternFill('solid', fgColor='ECF0F1')
TOTAL_FONT = Font(bold=True)
GOOD_FILL = PatternFill('solid', fgColor='D5F5E3')
BAD_FILL = PatternFill('solid', fgColor='FADBD8')
THIN = Side(border_style='thin', color='BDC3C7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_row(ws, row, vals, font=None, fill=None, border=True):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        if font: c.font = font
        if fill: c.fill = fill
        if border: c.border = BORDER

# ----- Tab 1: Summary -----
ws = wb.active
ws.title = 'Summary'
ws.column_dimensions['A'].width = 38
for c in 'BCDEFGH': ws.column_dimensions[c].width = 16

ws['A1'] = 'Curate Coins — 6-Month Outlook (with $1,000 Board Seed)'
ws['A1'].font = Font(bold=True, size=16)
ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M MT")}  |  Spot silver: ${SPOT}/oz  |  Board seed: ${BOARD_SEED:,.0f}'
ws['A2'].font = Font(italic=True, size=10, color='7F8C8D')

# Seed banner
ws.merge_cells('A4:H4')
ws['A4'] = f'⚡ BOARD SEED: ${BOARD_SEED:,.0f} → buys ~{units_buyable_with_seed} flippable units at ${flip_avg_cost:.0f}/unit avg → ramp accelerates from M4-full → M2-full'
ws['A4'].fill = SEED_FILL; ws['A4'].font = SEED_FONT
ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[4].height = 24

# Current state
ws['A6'] = 'CURRENT STATE (60-day actuals)'
ws['A6'].fill = SECTION_FILL; ws['A6'].font = SECTION_FONT
ws.merge_cells('A6:H6')

state = [
    ('Items sold (60d)', sold_count, '0'),
    ('Gross revenue (60d)', sold_total_gross, '$#,##0.00'),
    ('NET (60d)', sold_net, '$#,##0.00'),
    ('Avg sale price', sold_avg, '$#,##0.00'),
    ('Sales velocity (per month)', velocity_per_month, '0.0'),
]
for i, (label, val, fmt) in enumerate(state, 7):
    ws.cell(row=i, column=1, value=label).font = Font(bold=label.startswith('NET'))
    c = ws.cell(row=i, column=2, value=val)
    c.number_format = fmt
    if label.startswith('NET'): c.font = Font(bold=True); c.fill = GOOD_FILL

# Scenario comparison: with seed vs without
ws['A14'] = '6-MONTH PROJECTION — WITH SEED VS WITHOUT'
ws['A14'].fill = SECTION_FILL; ws['A14'].font = SECTION_FONT
ws.merge_cells('A14:H14')

write_row(ws, 15, ['Metric', 'Conservative', 'Cons+Seed', 'Δ', 'Base', 'Base+Seed', 'Δ', ''], font=HDR_FONT, fill=HDR_FILL)
write_row(ws, 16, ['', '', '', '', '', '', '', ''])

# Calculate totals
def totals(sc, with_seed):
    months = project(sc, with_seed)
    return {
        'sales': sum(m['sales'] for m in months),
        'gross': sum(m['gross'] for m in months),
        'net': sum(m['net'] for m in months),
        'reinvest': sum(m['reinvest'] for m in months),
        'cash': sum(m['cash'] for m in months),
    }

c_no = totals(scenarios_no_seed['Conservative'], False)
c_yes = totals(scenarios['Conservative+Seed'], True)
b_no = totals(scenarios_no_seed['Base'], False)
b_yes = totals(scenarios['Base+Seed'], True)
a_no = totals(scenarios_no_seed['Aggressive'], False)
a_yes = totals(scenarios['Aggressive+Seed'], True)

metrics = [
    ('Total sales (6mo)',      c_no['sales'], c_yes['sales'], c_yes['sales']-c_no['sales'],
                                b_no['sales'], b_yes['sales'], b_yes['sales']-b_no['sales']),
    ('Gross revenue',          c_no['gross'], c_yes['gross'], c_yes['gross']-c_no['gross'],
                                b_no['gross'], b_yes['gross'], b_yes['gross']-b_no['gross']),
    ('Net (after fees+ship)',  c_no['net'], c_yes['net'], c_yes['net']-c_no['net'],
                                b_no['net'], b_yes['net'], b_yes['net']-b_no['net']),
    ('Reinvested',             c_no['reinvest'], c_yes['reinvest'], c_yes['reinvest']-c_no['reinvest'],
                                b_no['reinvest'], b_yes['reinvest'], b_yes['reinvest']-b_no['reinvest']),
    ('CASH TO JOSH',           c_no['cash'], c_yes['cash'], c_yes['cash']-c_no['cash'],
                                b_no['cash'], b_yes['cash'], b_yes['cash']-b_no['cash']),
]
for i, m in enumerate(metrics, 17):
    write_row(ws, i, list(m) + [''])
    if 'sales' not in m[0].lower():
        for col in [2,3,4,5,6,7]:
            ws.cell(row=i, column=col).number_format = '$#,##0'
    if m[0] == 'CASH TO JOSH':
        for col in [1,2,3,4,5,6,7]:
            ws.cell(row=i, column=col).font = Font(bold=True)
        ws.cell(row=i, column=3).fill = GOOD_FILL
        ws.cell(row=i, column=6).fill = GOOD_FILL
        ws.cell(row=i, column=4).fill = GOOD_FILL
        ws.cell(row=i, column=7).fill = GOOD_FILL

# Aggressive on its own row block
ws['A24'] = 'Aggressive'
ws['A24'].fill = SECTION_FILL; ws['A24'].font = SECTION_FONT
ws.merge_cells('A24:H24')
write_row(ws, 25, ['Metric', 'Aggressive', 'Aggr+Seed', 'Δ', '', '', '', ''], font=HDR_FONT, fill=HDR_FILL)
agg_metrics = [
    ('Total sales (6mo)',      a_no['sales'], a_yes['sales'], a_yes['sales']-a_no['sales']),
    ('Gross revenue',          a_no['gross'], a_yes['gross'], a_yes['gross']-a_no['gross']),
    ('Net (after fees+ship)',  a_no['net'], a_yes['net'], a_yes['net']-a_no['net']),
    ('Reinvested',             a_no['reinvest'], a_yes['reinvest'], a_yes['reinvest']-a_no['reinvest']),
    ('CASH TO JOSH',           a_no['cash'], a_yes['cash'], a_yes['cash']-a_no['cash']),
]
for i, m in enumerate(agg_metrics, 26):
    write_row(ws, i, list(m) + ['', '', '', ''])
    if 'sales' not in m[0].lower():
        for col in [2,3,4]:
            ws.cell(row=i, column=col).number_format = '$#,##0'
    if m[0] == 'CASH TO JOSH':
        for col in [1,2,3,4]: ws.cell(row=i, column=col).font = Font(bold=True)
        ws.cell(row=i, column=3).fill = GOOD_FILL
        ws.cell(row=i, column=4).fill = GOOD_FILL

# ----- Tab 2: Cash Flow with Seed (Base) -----
ws3 = wb.create_sheet('Cash Flow with Seed')
for col in range(1, 11): ws3.column_dimensions[get_column_letter(col)].width = 14
ws3.column_dimensions['A'].width = 28

ws3['A1'] = '6-Month Cash Flow with $1,000 Board Seed (Base+Seed scenario)'
ws3['A1'].font = Font(bold=True, size=14)

months_labels = []
start = datetime.now().replace(day=1)
for i in range(MONTHS):
    m = start.replace(month=((start.month + i - 1) % 12) + 1, year=start.year + ((start.month + i - 1) // 12))
    months_labels.append(m.strftime('%b %Y'))

write_row(ws3, 3, ['Line item'] + months_labels + ['6-Mo Total'], font=HDR_FONT, fill=HDR_FILL)

base_seed = project(scenarios['Base+Seed'], with_seed=True)

rows = [
    ('Sales / month', [int(m['sales']) for m in base_seed], '#,##0'),
    ('Gross revenue', [m['gross'] for m in base_seed], '$#,##0'),
    ('eBay fees (13%)', [-m['fees'] for m in base_seed], '$#,##0'),
    ('Shipping + supplies', [-m['ship'] for m in base_seed], '$#,##0'),
    ('NET REVENUE', [m['net'] for m in base_seed], '$#,##0'),
    ('Inventory reinvestment (40%)', [-m['reinvest'] for m in base_seed], '$#,##0'),
    ('Silver-buy budget ($120/wk)', [-m['silver_buy'] for m in base_seed], '$#,##0'),
    ('BOARD SEED INFLOW', [m['seed_deployed'] for m in base_seed], '$#,##0'),
    ('FREE CASH TO JOSH', [m['cash'] for m in base_seed], '$#,##0'),
]
for ridx, (label, vals, fmt) in enumerate(rows, 4):
    ws3.cell(row=ridx, column=1, value=label)
    if 'NET' in label or 'CASH' in label or 'SEED' in label:
        ws3.cell(row=ridx, column=1).font = Font(bold=True)
    if 'SEED' in label:
        ws3.cell(row=ridx, column=1).fill = SEED_FILL
        ws3.cell(row=ridx, column=1).font = SEED_FONT
    for col_i, v in enumerate(vals, 2):
        c = ws3.cell(row=ridx, column=col_i, value=v)
        c.number_format = fmt
        if 'CASH' in label and isinstance(v, (int, float)):
            c.font = Font(bold=True)
            c.fill = GOOD_FILL if v > 0 else BAD_FILL
        if 'SEED' in label and v != 0:
            c.fill = SEED_FILL
            c.font = SEED_FONT
    total = sum(vals)
    c = ws3.cell(row=ridx, column=8, value=total)
    c.number_format = fmt
    c.font = Font(bold=True)
    if 'CASH' in label:
        c.fill = GOOD_FILL if total > 0 else BAD_FILL

# ----- Tab 3: Seed Deployment Plan -----
ws4 = wb.create_sheet('Seed Deployment')
for col in 'ABCDEF': ws4.column_dimensions[col].width = 22

ws4['A1'] = '$1,000 Board Seed — Deployment Plan'
ws4['A1'].font = Font(bold=True, size=14)
ws4['A2'] = 'Strategic split to maximize 6-month return without exposing the company to a single category.'
ws4['A2'].font = Font(italic=True, color='7F8C8D')

ws4['A4'] = 'PROPOSED ALLOCATION'
ws4['A4'].fill = SECTION_FILL; ws4['A4'].font = SECTION_FONT
ws4.merge_cells('A4:F4')

write_row(ws4, 5, ['Bucket', 'Allocation', '$', 'Avg unit cost', 'Units', 'Expected hold time'], font=HDR_FONT, fill=HDR_FILL)

deploy = [
    ('High-velocity bullion premium (NORFEDs, Libertads <$120, ASE rolls)',
     '40%', 400, 100, 4, '30-45 days'),
    ('Common silver upgrades (90% halves, dimes, Walker fills)',
     '30%', 300, 25, 12, '45-60 days'),
    ('Numismatic premium plays (semi-key dates, slabbed semi-keys)',
     '20%', 200, 75, 2.5, '60-90 days'),
    ('Reserve / tactical buys (Whatnot deals, show floor)',
     '10%', 100, 50, 2, 'opportunistic'),
]
total_units = 0
for i, (b, pct, dollars, ucost, units, hold) in enumerate(deploy, 6):
    write_row(ws4, i, [b, pct, dollars, ucost, units, hold])
    ws4.cell(row=i, column=3).number_format = '$#,##0'
    ws4.cell(row=i, column=4).number_format = '$#,##0'
    total_units += units

write_row(ws4, 10, ['TOTAL', '100%', BOARD_SEED, '', round(total_units, 1), ''], font=TOTAL_FONT, fill=TOTAL_FILL)
ws4.cell(row=10, column=3).number_format = '$#,##0'

ws4['A12'] = 'EXPECTED RETURN ON SEED'
ws4['A12'].fill = SECTION_FILL; ws4['A12'].font = SECTION_FONT
ws4.merge_cells('A12:F12')

write_row(ws4, 13, ['Bucket', 'Investment', 'Target margin', 'Expected gross', 'Expected net (87%)', 'ROI'], font=HDR_FONT, fill=HDR_FILL)
returns = [
    ('Bullion premium', 400, '15%', 460, 400, '0%'),  # bullion is tight margin
    ('Common silver upgrades', 300, '40%', 420, 365, '22%'),
    ('Numismatic plays', 200, '60%', 320, 278, '39%'),
    ('Tactical reserve', 100, '50%', 150, 130, '30%'),
]
total_inv = 0; total_gross = 0; total_net = 0
for i, (b, inv_amt, mgn, gross, net, roi) in enumerate(returns, 14):
    write_row(ws4, i, [b, inv_amt, mgn, gross, net, roi])
    ws4.cell(row=i, column=2).number_format = '$#,##0'
    ws4.cell(row=i, column=4).number_format = '$#,##0'
    ws4.cell(row=i, column=5).number_format = '$#,##0'
    total_inv += inv_amt; total_gross += gross; total_net += net

total_roi = (total_net - total_inv) / total_inv * 100
write_row(ws4, 18, ['TOTAL', total_inv, '', total_gross, total_net, f'{total_roi:.0f}%'], font=TOTAL_FONT, fill=TOTAL_FILL)
ws4.cell(row=18, column=2).number_format = '$#,##0'
ws4.cell(row=18, column=4).number_format = '$#,##0'
ws4.cell(row=18, column=5).number_format = '$#,##0'
ws4.cell(row=18, column=5).fill = GOOD_FILL

ws4['A20'] = 'GOVERNANCE NOTES'
ws4['A20'].fill = SECTION_FILL; ws4['A20'].font = SECTION_FONT
ws4.merge_cells('A20:F20')

notes = [
    f"• Seed is board capital — stays in business as equity, not flagged as Josh-cash payback",
    f"• Buy buckets must be Sheldon-vetted per [CUR-42](/CUR/issues/CUR-42) before purchase",
    f"• Reinvestment policy [CUR-73](/CUR/issues/CUR-73) §9 still applies to ongoing earnings (not seed)",
    f"• Bullion premium > $200/unit needs board check-in per pricing policy",
    f"• Track seed deployment separately for clean P&L attribution at Month 6",
    f"• Re-up question for board: at Month 6, do we recycle seed back, deploy further, or extract?",
]
for i, n in enumerate(notes, 21):
    ws4.cell(row=i, column=1, value=n)
    ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

# ----- Tab 4: Inventory Plan (Updated) -----
ws5 = wb.create_sheet('Inventory Plan')
for col in 'ABCDEF': ws5.column_dimensions[col].width = 24

ws5['A1'] = 'Inventory Replenishment with Seed (Base+Seed)'
ws5['A1'].font = Font(bold=True, size=14)

base_seed_total = totals(scenarios['Base+Seed'], True)
net_per_mo = base_seed_total['net'] / 6
silver_buy_budget = WEEKLY_BUDGET * 4.33
sales_per_mo_seed = scenarios['Base+Seed']['sales_per_month']

rows = [
    ('Flippable units on hand (today)', flip_count, ''),
    ('Sales / month (Base+Seed steady-state)', sales_per_mo_seed, ''),
    ('Months runway (no replenishment)', round(flip_count / sales_per_mo_seed, 1), ''),
    ('Avg cost / flippable unit', flip_avg_cost, '$#,##0.00'),
    ('Replenishment $/mo needed @ 100%', sales_per_mo_seed * flip_avg_cost, '$#,##0'),
    ('Replenishment $/mo from 40% reinvest', net_per_mo * 0.40, '$#,##0'),
    ('Silver-buy budget ($120/wk)', silver_buy_budget, '$#,##0'),
    ('Total replenishment $/mo (steady)', net_per_mo * 0.40 + silver_buy_budget, '$#,##0'),
    ('Seed-driven M1 boost', BOARD_SEED, '$#,##0'),
    ('SHORTFALL/SURPLUS (steady-state)', (net_per_mo * 0.40 + silver_buy_budget) - sales_per_mo_seed * flip_avg_cost, '$#,##0'),
]
write_row(ws5, 3, ['Metric', 'Value', 'Fmt'], font=HDR_FONT, fill=HDR_FILL)
for i, (lbl, val, fmt) in enumerate(rows, 4):
    ws5.cell(row=i, column=1, value=lbl)
    c = ws5.cell(row=i, column=2, value=val)
    if fmt: c.number_format = fmt
    if 'SHORTFALL' in lbl or 'SURPLUS' in lbl:
        ws5.cell(row=i, column=1).font = Font(bold=True)
        c.font = Font(bold=True)
        c.fill = BAD_FILL if val < 0 else GOOD_FILL
    if 'Seed' in lbl:
        ws5.cell(row=i, column=1).fill = SEED_FILL
        ws5.cell(row=i, column=1).font = SEED_FONT

# ----- Tab 5: Sensitivity (kept same shape) -----
ws6 = wb.create_sheet('Sensitivity')
for col in 'ABCDEFGH': ws6.column_dimensions[col].width = 14
ws6.column_dimensions['A'].width = 22
ws6['A1'] = 'Sensitivity — 6-Month CASH TO JOSH (with $1k seed)'
ws6['A1'].font = Font(bold=True, size=14)
ws6['A3'] = 'Sales/month (rows) × Avg price (cols), at 40% reinvestment, M1 ramp 85%'
ws6['A3'].font = Font(italic=True, size=10)
prices = [80, 100, 120, 140, 160]
sales_range = [10, 15, 20, 25, 30, 35]
write_row(ws6, 4, [''] + [f'${p}' for p in prices], font=HDR_FONT, fill=HDR_FILL)
for i, s in enumerate(sales_range, 5):
    row_vals = [f'{s}/mo']
    for p in prices:
        net_per_month = s * (p - (p*EBAY_FEE_PCT + SHIPPING_PER_ITEM + SUPPLIES_PER_ITEM))
        # Apply 85% ramp M1, 100% rest
        ramp = [0.85] + [1.0]*5
        net_total = sum(net_per_month * r for r in ramp)
        cash = net_total * 0.60 - silver_buy_budget * 6  # subtract silver budget
        row_vals.append(cash)
    write_row(ws6, i, row_vals)
    for col in range(2, 2 + len(prices)):
        ws6.cell(row=i, column=col).number_format = '$#,##0'
        v = ws6.cell(row=i, column=col).value
        if v < 3000: ws6.cell(row=i, column=col).fill = BAD_FILL
        elif v > 12000: ws6.cell(row=i, column=col).fill = GOOD_FILL

# Save
out = '/Users/josh/.openclaw/workspace/curate-coins-6mo-forecast-with-seed.xlsx'
wb.save(out)
print(f"\nSaved: {out}\n")

# Print headlines
print("=== WITH $1,000 BOARD SEED ===")
print(f"\nBase+Seed scenario, 6 months:")
print(f"  Total sales: {b_yes['sales']:.0f} units (vs {b_no['sales']:.0f} no-seed)")
print(f"  Gross: ${b_yes['gross']:,.0f} (vs ${b_no['gross']:,.0f}, +${b_yes['gross']-b_no['gross']:,.0f})")
print(f"  Net: ${b_yes['net']:,.0f} (vs ${b_no['net']:,.0f}, +${b_yes['net']-b_no['net']:,.0f})")
print(f"  Cash to Josh: ${b_yes['cash']:,.0f} (vs ${b_no['cash']:,.0f}, +${b_yes['cash']-b_no['cash']:,.0f})")
print(f"\nConservative+Seed: cash=${c_yes['cash']:,.0f}")
print(f"Aggressive+Seed: cash=${a_yes['cash']:,.0f}")

print(f"\nSeed deployment buys ~{units_buyable_with_seed} flippable units at ${flip_avg_cost:.2f} avg")
print(f"Expected seed ROI: {total_roi:.0f}% (${total_net-total_inv:,.0f} net gain on $1,000)")
