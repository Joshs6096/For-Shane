"""6-Month Curate Coins Forecast — formula-driven, P&L style cash flow, detailed inventory plan."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ---- Style helpers ----
HDR_FILL = PatternFill('solid', fgColor='2C3E50')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='34495E')
SECTION_FONT = Font(bold=True, color='FFFFFF', size=12)
INPUT_FILL = PatternFill('solid', fgColor='FFF3CD')
INPUT_FONT = Font(bold=True, color='8A6D00')
SEED_FILL = PatternFill('solid', fgColor='F39C12')
SEED_FONT = Font(bold=True, color='FFFFFF')
TOTAL_FILL = PatternFill('solid', fgColor='ECF0F1')
TOTAL_FONT = Font(bold=True)
INCOME_FILL = PatternFill('solid', fgColor='D5F5E3')
EXPENSE_FILL = PatternFill('solid', fgColor='FADBD8')
NET_FILL = PatternFill('solid', fgColor='AED6F1')
THIN = Side(border_style='thin', color='BDC3C7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()

# ============================================================
# TAB 1: SUMMARY (with INPUTS section — single source of truth)
# ============================================================
ws = wb.active
ws.title = 'Summary'
ws.column_dimensions['A'].width = 38
for c in 'BCDEFGH': ws.column_dimensions[c].width = 16

ws['A1'] = 'Curate Coins — 6-Month Outlook'
ws['A1'].font = Font(bold=True, size=16)
ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d")}  |  All values formula-driven from INPUTS block — edit yellow cells to update'
ws['A2'].font = Font(italic=True, size=10, color='7F8C8D')

# ---- INPUTS ----
ws['A4'] = 'INPUTS — edit yellow cells to update everything'
ws['A4'].fill = SECTION_FILL; ws['A4'].font = SECTION_FONT
ws.merge_cells('A4:C4')

inputs = [
    ('Sales / month (steady-state)',  18,    '0',           'B5'),    # B5
    ('Avg sale price ($)',            125,   '$#,##0.00',  'B6'),
    ('eBay fee %',                    0.13,  '0.0%',        'B7'),
    ('Ship + supplies per item ($)',  2.50,  '$#,##0.00',  'B8'),
    ('Reinvestment % of net',         0.40,  '0.0%',        'B9'),
    ('Silver-buy weekly ($)',         120,   '$#,##0.00',  'B10'),
    ('Silver-buy monthly ($)',        '=B10*4.33', '$#,##0.00', 'B11'),
    ('Board seed (Month 1) ($)',      1000,  '$#,##0',      'B12'),
    ('M1 ramp factor',                0.80,  '0%',          'B13'),
    ('M2 ramp factor',                0.95,  '0%',          'B14'),
    ('M3-M6 ramp factor',             1.00,  '0%',          'B15'),
    ('Silver spot ($/oz)',            81.88, '$#,##0.00',  'B16'),
]
for i, (label, val, fmt, _ref) in enumerate(inputs, 5):
    ws.cell(row=i, column=1, value=label)
    c = ws.cell(row=i, column=2, value=val)
    c.number_format = fmt
    c.fill = INPUT_FILL
    c.font = INPUT_FONT
    c.border = BORDER

# Named references for inputs (for clarity in formulas; we'll use cell refs directly)
SALES = 'Summary!$B$5'
PRICE = 'Summary!$B$6'
FEE_PCT = 'Summary!$B$7'
SHIP = 'Summary!$B$8'
REINV_PCT = 'Summary!$B$9'
SILVER_WK = 'Summary!$B$10'
SILVER_MO = 'Summary!$B$11'
SEED = 'Summary!$B$12'
M1_R = 'Summary!$B$13'
M2_R = 'Summary!$B$14'
M3_R = 'Summary!$B$15'
SPOT = 'Summary!$B$16'

# ---- 6-MONTH HEADLINE (formulas) ----
ws['A18'] = '6-MONTH HEADLINE (live from inputs)'
ws['A18'].fill = SECTION_FILL; ws['A18'].font = SECTION_FONT
ws.merge_cells('A18:C18')

# Headline pulls from Cash Flow totals (column H)
headline = [
    ("Total sales (6mo)",            "='Cash Flow (P&L)'!H6",  '0'),
    ("Gross revenue",                "='Cash Flow (P&L)'!H7",  '$#,##0'),
    ("+ Board seed inflow",          "='Cash Flow (P&L)'!H8",  '$#,##0'),
    ("TOTAL INCOME (6mo)",           "='Cash Flow (P&L)'!H9",  '$#,##0'),
    ("- eBay fees (13%)",            "='Cash Flow (P&L)'!H12", '$#,##0'),
    ("- Shipping + supplies",        "='Cash Flow (P&L)'!H13", '$#,##0'),
    ("- Inventory reinvestment",     "='Cash Flow (P&L)'!H14", '$#,##0'),
    ("- Silver-buy spend",           "='Cash Flow (P&L)'!H15", '$#,##0'),
    ("TOTAL EXPENSES (6mo)",         "='Cash Flow (P&L)'!H16", '$#,##0'),
    ("FREE CASH TO JOSH",            "='Cash Flow (P&L)'!H18", '$#,##0'),
]
for i, (label, formula, fmt) in enumerate(headline, 19):
    ws.cell(row=i, column=1, value=label)
    c = ws.cell(row=i, column=2, value=formula)
    c.number_format = fmt
    c.border = BORDER
    if 'INCOME' in label or 'EXPENSES' in label:
        ws.cell(row=i, column=1).font = Font(bold=True)
        c.font = Font(bold=True)
        c.fill = TOTAL_FILL
    if 'CASH TO JOSH' in label:
        ws.cell(row=i, column=1).font = Font(bold=True)
        c.font = Font(bold=True)
        c.fill = NET_FILL

# Key takeaways (still text — these are commentary)
ws['A30'] = 'KEY TAKEAWAYS'
ws['A30'].fill = SECTION_FILL; ws['A30'].font = SECTION_FONT
ws.merge_cells('A30:C30')

takeaways = [
    "• Bottleneck = listing throughput, not capital or inventory.",
    "• Current pace = 3 sales/mo at $152 avg. Target = 18 sales/mo at $125 avg (yellow cells).",
    "• Free cash = Total Income − Total Expenses (see Cash Flow P&L tab).",
    "• Flippable inventory grows during the 6 months — core (188oz collection) never touched.",
    "• To change scenarios, edit the yellow cells in INPUTS — every number recalculates automatically.",
]
for i, t in enumerate(takeaways, 31):
    ws.cell(row=i, column=1, value=t)
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)

# ============================================================
# TAB 2: CASH FLOW (P&L style — Income, Expenses, Net)
# ============================================================
ws2 = wb.create_sheet('Cash Flow (P&L)')
for col in range(1, 11): ws2.column_dimensions[get_column_letter(col)].width = 15
ws2.column_dimensions['A'].width = 32

ws2['A1'] = '6-Month P&L — Income vs Expenses'
ws2['A1'].font = Font(bold=True, size=14)
ws2['A2'] = 'Edit yellow cells in Summary!INPUTS — all figures recalc.'
ws2['A2'].font = Font(italic=True, color='7F8C8D')

months_labels = []
start = datetime.now().replace(day=1)
for i in range(6):
    m = start.replace(month=((start.month + i - 1) % 12) + 1, year=start.year + ((start.month + i - 1) // 12))
    months_labels.append(m.strftime('%b %Y'))

# Header row (row 4)
hdr = ['Line item'] + months_labels + ['6-Mo Total']
for i, v in enumerate(hdr, 1):
    c = ws2.cell(row=4, column=i, value=v)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER

# Helper: ramp formula reference per column. Col B=M1, C=M2, D-G=M3-M6
def ramp_for(col_letter):
    if col_letter == 'B': return M1_R
    if col_letter == 'C': return M2_R
    return M3_R

# ---- INCOME section (row 5 header) ----
ws2['A5'] = 'INCOME'
ws2['A5'].fill = INCOME_FILL; ws2['A5'].font = Font(bold=True, color='1E5F2C')
ws2.merge_cells('A5:H5')

# Row 6: Sales/month = ROUND(SALES * ramp, 0)
ws2.cell(row=6, column=1, value='Sales (units)')
for col_i, col_letter in enumerate(['B','C','D','E','F','G'], 2):
    ws2.cell(row=6, column=col_i, value=f'=ROUND({SALES}*{ramp_for(col_letter)}, 0)')
    ws2.cell(row=6, column=col_i).number_format = '0'
ws2.cell(row=6, column=8, value='=SUM(B6:G6)').number_format = '0'
ws2.cell(row=6, column=8).font = Font(bold=True)

# Row 7: Gross revenue = sales * price
ws2.cell(row=7, column=1, value='Gross revenue')
for col_i, col_letter in enumerate(['B','C','D','E','F','G'], 2):
    ws2.cell(row=7, column=col_i, value=f'={col_letter}6*{PRICE}')
    ws2.cell(row=7, column=col_i).number_format = '$#,##0'
ws2.cell(row=7, column=8, value='=SUM(B7:G7)').number_format = '$#,##0'
ws2.cell(row=7, column=8).font = Font(bold=True)

# Row 8: Board seed inflow (M1 only)
ws2.cell(row=8, column=1, value='Board seed inflow')
ws2.cell(row=8, column=1).fill = SEED_FILL; ws2.cell(row=8, column=1).font = SEED_FONT
ws2.cell(row=8, column=2, value=f'={SEED}')
for col_i in range(3, 8):
    ws2.cell(row=8, column=col_i, value=0)
for col_i in range(2, 8):
    ws2.cell(row=8, column=col_i).number_format = '$#,##0'
    ws2.cell(row=8, column=col_i).fill = SEED_FILL
    ws2.cell(row=8, column=col_i).font = SEED_FONT
ws2.cell(row=8, column=8, value='=SUM(B8:G8)').number_format = '$#,##0'
ws2.cell(row=8, column=8).font = Font(bold=True)

# Row 9: TOTAL INCOME
ws2.cell(row=9, column=1, value='TOTAL INCOME').font = Font(bold=True)
ws2.cell(row=9, column=1).fill = INCOME_FILL
for col_i in range(2, 9):
    ws2.cell(row=9, column=col_i, value=f'={get_column_letter(col_i)}7+{get_column_letter(col_i)}8')
    ws2.cell(row=9, column=col_i).number_format = '$#,##0'
    ws2.cell(row=9, column=col_i).font = Font(bold=True)
    ws2.cell(row=9, column=col_i).fill = INCOME_FILL

# ---- EXPENSES section ----
ws2['A11'] = 'EXPENSES'
ws2['A11'].fill = EXPENSE_FILL; ws2['A11'].font = Font(bold=True, color='8B1A1A')
ws2.merge_cells('A11:H11')

# Row 12: eBay fees = gross * fee_pct
ws2.cell(row=12, column=1, value='eBay fees (13%)')
for col_i, col_letter in enumerate(['B','C','D','E','F','G'], 2):
    ws2.cell(row=12, column=col_i, value=f'={col_letter}7*{FEE_PCT}')
    ws2.cell(row=12, column=col_i).number_format = '$#,##0'
ws2.cell(row=12, column=8, value='=SUM(B12:G12)').number_format = '$#,##0'
ws2.cell(row=12, column=8).font = Font(bold=True)

# Row 13: Shipping + supplies = sales * ship
ws2.cell(row=13, column=1, value='Shipping + supplies')
for col_i, col_letter in enumerate(['B','C','D','E','F','G'], 2):
    ws2.cell(row=13, column=col_i, value=f'={col_letter}6*{SHIP}')
    ws2.cell(row=13, column=col_i).number_format = '$#,##0'
ws2.cell(row=13, column=8, value='=SUM(B13:G13)').number_format = '$#,##0'
ws2.cell(row=13, column=8).font = Font(bold=True)

# Row 14: Inventory reinvestment = (gross - fees - ship) * reinvest_pct
ws2.cell(row=14, column=1, value='Inventory reinvestment')
for col_i, col_letter in enumerate(['B','C','D','E','F','G'], 2):
    ws2.cell(row=14, column=col_i, value=f'=({col_letter}7-{col_letter}12-{col_letter}13)*{REINV_PCT}')
    ws2.cell(row=14, column=col_i).number_format = '$#,##0'
ws2.cell(row=14, column=8, value='=SUM(B14:G14)').number_format = '$#,##0'
ws2.cell(row=14, column=8).font = Font(bold=True)

# Row 15: Silver-buy budget (constant monthly)
ws2.cell(row=15, column=1, value='Silver-buy budget')
for col_i in range(2, 8):
    ws2.cell(row=15, column=col_i, value=f'={SILVER_MO}')
    ws2.cell(row=15, column=col_i).number_format = '$#,##0'
ws2.cell(row=15, column=8, value='=SUM(B15:G15)').number_format = '$#,##0'
ws2.cell(row=15, column=8).font = Font(bold=True)

# Row 16: TOTAL EXPENSES
ws2.cell(row=16, column=1, value='TOTAL EXPENSES').font = Font(bold=True)
ws2.cell(row=16, column=1).fill = EXPENSE_FILL
for col_i in range(2, 9):
    L = get_column_letter(col_i)
    ws2.cell(row=16, column=col_i, value=f'={L}12+{L}13+{L}14+{L}15')
    ws2.cell(row=16, column=col_i).number_format = '$#,##0'
    ws2.cell(row=16, column=col_i).font = Font(bold=True)
    ws2.cell(row=16, column=col_i).fill = EXPENSE_FILL

# ---- NET CASH (Income - Expenses) ----
ws2['A18'] = 'FREE CASH TO JOSH (Income − Expenses)'
ws2.cell(row=18, column=1).font = Font(bold=True)
ws2.cell(row=18, column=1).fill = NET_FILL
for col_i in range(2, 9):
    L = get_column_letter(col_i)
    ws2.cell(row=18, column=col_i, value=f'={L}9-{L}16')
    ws2.cell(row=18, column=col_i).number_format = '$#,##0'
    ws2.cell(row=18, column=col_i).font = Font(bold=True)
    ws2.cell(row=18, column=col_i).fill = NET_FILL

# Apply borders to data block
for r in range(4, 19):
    for col_i in range(1, 9):
        ws2.cell(row=r, column=col_i).border = BORDER

# Notes
ws2['A20'] = 'Notes'
ws2['A20'].font = Font(bold=True)
notes = [
    "• Sales unit count rounds to whole units per month for clean reporting (uses Summary!B5 × ramp factor).",
    "• eBay payouts release 7-10 days after delivery confirmation — cash hits ~10-12 days after sale.",
    "• M1 cash is light because of the ramp factor (B13). M2-M6 = full cadence.",
    "• Board seed is treated as Month 1 income inflow (equity injection, not earnings).",
    "• Inventory reinvestment is an internal transfer — leaves the cash account but stays in company.",
    "• Silver-buy budget is the operating expense for replenishment buys (separate from reinvestment).",
]
for i, n in enumerate(notes, 21):
    ws2.cell(row=i, column=1, value=n)

# ============================================================
# TAB 3: SEED DEPLOYMENT (formula-driven where possible)
# ============================================================
ws3 = wb.create_sheet('Seed Deployment')
for col in 'ABCDEF': ws3.column_dimensions[col].width = 24

ws3['A1'] = '$1,000 Board Seed — Deployment Plan'
ws3['A1'].font = Font(bold=True, size=14)
ws3['A2'] = f'Total seed = Summary!B12 (=${ws["B12"].value:,.0f}). Edit allocation %s in column B; $ amounts recalc.'
ws3['A2'].font = Font(italic=True, color='7F8C8D')

ws3['A4'] = 'PROPOSED ALLOCATION'
ws3['A4'].fill = SECTION_FILL; ws3['A4'].font = SECTION_FONT
ws3.merge_cells('A4:F4')

hdr = ['Bucket', 'Allocation %', '$', 'Avg unit cost', 'Units', 'Expected hold']
for i, h in enumerate(hdr, 1):
    c = ws3.cell(row=5, column=i, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER

deploy = [
    ('High-velocity bullion premium (NORFEDs, Libertads, ASE rolls)', 0.40, 100, '30-45 days'),
    ('Common silver upgrades (90% halves, dimes, Walker fills)',       0.30, 25,  '45-60 days'),
    ('Numismatic plays (semi-key dates, slabbed semi-keys)',           0.20, 75,  '60-90 days'),
    ('Reserve / tactical buys (Whatnot, show floor)',                   0.10, 50,  'opportunistic'),
]
for i, (b, pct, ucost, hold) in enumerate(deploy, 6):
    ws3.cell(row=i, column=1, value=b)
    c = ws3.cell(row=i, column=2, value=pct); c.number_format = '0%'; c.fill = INPUT_FILL; c.font = INPUT_FONT
    ws3.cell(row=i, column=3, value=f'={SEED}*B{i}').number_format = '$#,##0'
    c = ws3.cell(row=i, column=4, value=ucost); c.number_format = '$#,##0'; c.fill = INPUT_FILL; c.font = INPUT_FONT
    ws3.cell(row=i, column=5, value=f'=C{i}/D{i}').number_format = '0.0'
    ws3.cell(row=i, column=6, value=hold)
    for col_i in range(1, 7):
        ws3.cell(row=i, column=col_i).border = BORDER

# Totals row
ws3.cell(row=10, column=1, value='TOTAL').font = Font(bold=True)
ws3.cell(row=10, column=2, value='=SUM(B6:B9)').number_format = '0%'
ws3.cell(row=10, column=3, value='=SUM(C6:C9)').number_format = '$#,##0'
ws3.cell(row=10, column=5, value='=SUM(E6:E9)').number_format = '0.0'
for col_i in range(1, 7):
    ws3.cell(row=10, column=col_i).border = BORDER
    ws3.cell(row=10, column=col_i).fill = TOTAL_FILL
    ws3.cell(row=10, column=col_i).font = Font(bold=True)

# Expected return
ws3['A12'] = 'EXPECTED RETURN ON SEED'
ws3['A12'].fill = SECTION_FILL; ws3['A12'].font = SECTION_FONT
ws3.merge_cells('A12:F12')

ret_hdr = ['Bucket', 'Investment $', 'Margin %', 'Expected gross', 'Expected net (after 13% fees)', 'ROI %']
for i, h in enumerate(ret_hdr, 1):
    c = ws3.cell(row=13, column=i, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER

returns = [
    ('Bullion premium', 6, 0.15),
    ('Common silver upgrades', 7, 0.40),
    ('Numismatic plays', 8, 0.60),
    ('Tactical reserve', 9, 0.50),
]
for i, (lbl, src_row, mgn) in enumerate(returns, 14):
    ws3.cell(row=i, column=1, value=lbl)
    ws3.cell(row=i, column=2, value=f'=C{src_row}').number_format = '$#,##0'
    c = ws3.cell(row=i, column=3, value=mgn); c.number_format = '0%'; c.fill = INPUT_FILL; c.font = INPUT_FONT
    ws3.cell(row=i, column=4, value=f'=B{i}*(1+C{i})').number_format = '$#,##0'
    ws3.cell(row=i, column=5, value=f'=D{i}*(1-{FEE_PCT})').number_format = '$#,##0'
    ws3.cell(row=i, column=6, value=f'=(E{i}-B{i})/B{i}').number_format = '0%'
    for col_i in range(1, 7):
        ws3.cell(row=i, column=col_i).border = BORDER

# Totals
ws3.cell(row=18, column=1, value='TOTAL').font = Font(bold=True)
ws3.cell(row=18, column=2, value='=SUM(B14:B17)').number_format = '$#,##0'
ws3.cell(row=18, column=4, value='=SUM(D14:D17)').number_format = '$#,##0'
ws3.cell(row=18, column=5, value='=SUM(E14:E17)').number_format = '$#,##0'
ws3.cell(row=18, column=6, value='=(E18-B18)/B18').number_format = '0%'
for col_i in range(1, 7):
    ws3.cell(row=18, column=col_i).border = BORDER
    ws3.cell(row=18, column=col_i).fill = TOTAL_FILL
    ws3.cell(row=18, column=col_i).font = Font(bold=True)
ws3.cell(row=18, column=5).fill = INCOME_FILL

# Notes
ws3['A20'] = 'NOTES'
ws3['A20'].fill = SECTION_FILL; ws3['A20'].font = SECTION_FONT
ws3.merge_cells('A20:F20')
notes = [
    "• Yellow cells are inputs — edit allocation %, unit cost, or margin to model variants.",
    "• Seed is company equity, not Josh-payback. Stays on the books.",
    "• Catalytic value > direct ROI: $1k accelerates listing cadence + unlocks steady-state run rate.",
    "• All buys must be Sheldon-vetted per CUR-42 before execution.",
    "• Bullion buys >$200/unit need board check-in.",
]
for i, n in enumerate(notes, 21):
    ws3.cell(row=i, column=1, value=n)
    ws3.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

# ============================================================
# TAB 4: INVENTORY PLAN — DETAILED
# ============================================================
ws4 = wb.create_sheet('Inventory Plan')
for col in 'ABCDEFG': ws4.column_dimensions[col].width = 22

ws4['A1'] = 'Inventory Plan — 6-Month Replenishment & Listing Cadence'
ws4['A1'].font = Font(bold=True, size=14)
ws4['A2'] = 'All values formula-linked. Edit yellow inputs to flex assumptions.'
ws4['A2'].font = Font(italic=True, color='7F8C8D')

# === SECTION A: Current state ===
ws4['A4'] = 'A. CURRENT INVENTORY STATE'
ws4['A4'].fill = SECTION_FILL; ws4['A4'].font = SECTION_FONT
ws4.merge_cells('A4:G4')

state_hdr = ['Segment', 'Description', 'Units', 'Cost basis', 'Low value', 'High value', 'Notes']
for i, h in enumerate(state_hdr, 1):
    c = ws4.cell(row=5, column=i, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER

# Hardcoded current state (computed from collection — these are facts, not inputs)
state = [
    ('Core (inherited)',  '$0 cost basis — heritage holdings', 155, 0,    8000,   31877, 'NEVER liquidate. Appreciation engine.'),
    ('Core (numismatic)', 'Proof sets, key dates, premium graded', 550, 4187, 14000, 13221, 'Hold for grading or premium retail. Liquidate selectively.'),
    ('Flippable bullion', 'Common ASE, Maple Leafs, rounds, bars', 130,  6500, 12000, 18000, 'Primary listing rotation.'),
    ('Flippable silver',  '90% halves/quarters/dimes, Morgans', 212,  6528,  10000, 12293, 'Bullion premium + numismatic blend.'),
]
for i, (seg, desc, units, cost, lo, hi, note) in enumerate(state, 6):
    ws4.cell(row=i, column=1, value=seg)
    ws4.cell(row=i, column=2, value=desc)
    ws4.cell(row=i, column=3, value=units).number_format = '0'
    ws4.cell(row=i, column=4, value=cost).number_format = '$#,##0'
    ws4.cell(row=i, column=5, value=lo).number_format = '$#,##0'
    ws4.cell(row=i, column=6, value=hi).number_format = '$#,##0'
    ws4.cell(row=i, column=7, value=note)
    for col_i in range(1, 8):
        ws4.cell(row=i, column=col_i).border = BORDER

# Totals row
ws4.cell(row=10, column=1, value='TOTAL').font = Font(bold=True)
ws4.cell(row=10, column=3, value='=SUM(C6:C9)').number_format = '0'
ws4.cell(row=10, column=4, value='=SUM(D6:D9)').number_format = '$#,##0'
ws4.cell(row=10, column=5, value='=SUM(E6:E9)').number_format = '$#,##0'
ws4.cell(row=10, column=6, value='=SUM(F6:F9)').number_format = '$#,##0'
for col_i in range(1, 8):
    ws4.cell(row=10, column=col_i).border = BORDER
    ws4.cell(row=10, column=col_i).fill = TOTAL_FILL
    ws4.cell(row=10, column=col_i).font = Font(bold=True)

# Flippable subtotal
ws4.cell(row=11, column=1, value='Flippable subtotal').font = Font(bold=True)
ws4.cell(row=11, column=3, value='=C8+C9').number_format = '0'
ws4.cell(row=11, column=4, value='=D8+D9').number_format = '$#,##0'
ws4.cell(row=11, column=6, value='=F8+F9').number_format = '$#,##0'
for col_i in range(1, 8):
    ws4.cell(row=11, column=col_i).border = BORDER

# === SECTION B: Burn-down math ===
ws4['A13'] = 'B. BURN-DOWN MATH (does the inventory hold up?)'
ws4['A13'].fill = SECTION_FILL; ws4['A13'].font = SECTION_FONT
ws4.merge_cells('A13:G13')

ws4.cell(row=14, column=1, value='Metric').font = HDR_FONT; ws4.cell(row=14, column=1).fill = HDR_FILL
ws4.cell(row=14, column=2, value='Value').font = HDR_FONT; ws4.cell(row=14, column=2).fill = HDR_FILL
for col_i in [1,2]: ws4.cell(row=14, column=col_i).border = BORDER

burn = [
    ('Flippable units on hand',                    '=C11',                        '0'),
    ('Sales / month (steady-state)',               f'={SALES}',                   '0'),
    ('Months runway with zero replenishment',      f'=C11/{SALES}',                '0.0'),
    ('Avg cost per flippable unit',                '=D11/C11',                    '$#,##0.00'),
    ('Replenishment $/mo needed (100% replace)',   f'={SALES}*B18',                '$#,##0'),
    ('Replenishment $/mo from 40% reinvest',       f"='Cash Flow (P&L)'!H14/6",   '$#,##0'),
    ('Silver-buy budget ($/mo)',                   f'={SILVER_MO}',                '$#,##0'),
    ('Total replenishment $/mo available',         '=B20+B21',                    '$#,##0'),
    ('SURPLUS / SHORTFALL',                        '=B22-B19',                    '$#,##0'),
]
for i, (lbl, formula, fmt) in enumerate(burn, 15):
    ws4.cell(row=i, column=1, value=lbl)
    c = ws4.cell(row=i, column=2, value=formula)
    c.number_format = fmt
    c.border = BORDER
    ws4.cell(row=i, column=1).border = BORDER
    if 'SURPLUS' in lbl:
        ws4.cell(row=i, column=1).font = Font(bold=True)
        c.font = Font(bold=True)
        c.fill = INCOME_FILL  # color logic in Excel; default green; user sees red when neg

# === SECTION C: Listing cadence target ===
ws4['A25'] = 'C. LISTING CADENCE TARGET (what we have to do operationally)'
ws4['A25'].fill = SECTION_FILL; ws4['A25'].font = SECTION_FONT
ws4.merge_cells('A25:G25')

ws4.cell(row=26, column=1, value='Metric').font = HDR_FONT; ws4.cell(row=26, column=1).fill = HDR_FILL
ws4.cell(row=26, column=2, value='Value').font = HDR_FONT; ws4.cell(row=26, column=2).fill = HDR_FILL
ws4.cell(row=26, column=3, value='Notes').font = HDR_FONT; ws4.cell(row=26, column=3).fill = HDR_FILL
for col_i in [1,2,3]: ws4.cell(row=26, column=col_i).border = BORDER

# Listing assumptions
ws4.cell(row=27, column=1, value='Sell-through rate (target % within 30d)')
c = ws4.cell(row=27, column=2, value=0.40); c.number_format = '0%'; c.fill = INPUT_FILL; c.font = INPUT_FONT
ws4.cell(row=27, column=3, value='Editable. Higher = fewer listings needed.')

ws4.cell(row=28, column=1, value='Active listings needed')
ws4.cell(row=28, column=2, value=f'={SALES}/B27').number_format = '0'
ws4.cell(row=28, column=3, value='= sales/mo ÷ sell-through rate')

ws4.cell(row=29, column=1, value='Current active listings')
c = ws4.cell(row=29, column=2, value=5); c.number_format = '0'; c.fill = INPUT_FILL; c.font = INPUT_FONT
ws4.cell(row=29, column=3, value='Current state — 5 active.')

ws4.cell(row=30, column=1, value='Listing gap (need to add)')
ws4.cell(row=30, column=2, value='=B28-B29').number_format = '0'
ws4.cell(row=30, column=3, value='Listings to draft + publish to hit target.')

ws4.cell(row=31, column=1, value='Listings to draft per week (8-week ramp)')
ws4.cell(row=31, column=2, value='=B30/8').number_format = '0.0'
ws4.cell(row=31, column=3, value='Spread the gap over 2 months.')

ws4.cell(row=32, column=1, value='Listings to draft per week (steady-state, post-ramp)')
ws4.cell(row=32, column=2, value=f'={SALES}/4.33').number_format = '0.0'
ws4.cell(row=32, column=3, value='Replace what sells each week.')

for r in range(27, 33):
    for col_i in [1,2,3]:
        ws4.cell(row=r, column=col_i).border = BORDER

# === SECTION D: Monthly buying plan ===
ws4['A34'] = 'D. MONTHLY BUYING PLAN ($ allocation per month)'
ws4['A34'].fill = SECTION_FILL; ws4['A34'].font = SECTION_FONT
ws4.merge_cells('A34:G34')

buy_hdr = ['Bucket', '% of monthly buy', '$/mo', 'Units/mo', 'Channels', 'Vetting']
for i, h in enumerate(buy_hdr, 1):
    c = ws4.cell(row=35, column=i, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER

buy_buckets = [
    ('Bullion premium plays', 0.45, 100, 'eBay, Whatnot live shows, dealers', 'Sheldon-vet (CUR-42)'),
    ('Common silver upgrades', 0.30, 25, 'eBay BIN searches, Whatnot bulk lots', 'Self-vet via melt+10% rule'),
    ('Numismatic semi-keys', 0.20, 75, 'eBay auctions, GreatCollections', 'Sheldon-vet mandatory'),
    ('Tactical / show floor', 0.05, 50, 'Coin shows, local dealers', 'On-the-spot judgment'),
]
for i, (b, pct, ucost, ch, vet) in enumerate(buy_buckets, 36):
    ws4.cell(row=i, column=1, value=b)
    c = ws4.cell(row=i, column=2, value=pct); c.number_format = '0%'; c.fill = INPUT_FILL; c.font = INPUT_FONT
    ws4.cell(row=i, column=3, value=f'=B22*B{i}').number_format = '$#,##0'  # B22 = total replenishment available
    ws4.cell(row=i, column=4, value=f'=C{i}/{ucost}').number_format = '0.0'
    ws4.cell(row=i, column=5, value=ch)
    ws4.cell(row=i, column=6, value=vet)
    for col_i in range(1, 7):
        ws4.cell(row=i, column=col_i).border = BORDER

# Totals
ws4.cell(row=40, column=1, value='TOTAL').font = Font(bold=True)
ws4.cell(row=40, column=2, value='=SUM(B36:B39)').number_format = '0%'
ws4.cell(row=40, column=3, value='=SUM(C36:C39)').number_format = '$#,##0'
ws4.cell(row=40, column=4, value='=SUM(D36:D39)').number_format = '0.0'
for col_i in range(1, 7):
    ws4.cell(row=40, column=col_i).border = BORDER
    ws4.cell(row=40, column=col_i).fill = TOTAL_FILL
    ws4.cell(row=40, column=col_i).font = Font(bold=True)

# === SECTION E: Inventory turnover & projections ===
ws4['A42'] = 'E. INVENTORY TRAJECTORY (units on hand by month, base case)'
ws4['A42'].fill = SECTION_FILL; ws4['A42'].font = SECTION_FONT
ws4.merge_cells('A42:G42')

trj_hdr = ['Metric', 'Start'] + months_labels[:6]
for i, h in enumerate(trj_hdr, 1):
    c = ws4.cell(row=43, column=i, value=h)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER

# Row 44: Flippable units start of month
ws4.cell(row=44, column=1, value='Flippable units (start)')
ws4.cell(row=44, column=2, value='=C11').number_format = '0'  # start
# col 3 = Jun (M1 end). Each month: prev_end - sold + bought
# units bought from total monthly buy / avg unit cost
ws4.cell(row=45, column=1, value='Sales (units sold)')
ws4.cell(row=46, column=1, value='Units bought (replenish)')
ws4.cell(row=47, column=1, value='Flippable units (end)')

for col_i in range(3, 9):  # cols C-H
    L = get_column_letter(col_i)
    PL = get_column_letter(col_i - 1)
    cf_col = get_column_letter(col_i - 1)  # B->B, C->B... actually map
    # Cash Flow row 6 = sales for that month; col B in CF = May (col 2). Our col_i 3 = Jun.
    # Map: ws4 col 3 (Jun) -> CF col B (May, M1) ... not aligned. Let me realign.
    pass

# Actually, simpler approach: align ws4 row 44+ columns with months_labels (B = Start, C = May, D = Jun, ...)
# Re-do this section properly
ws4.cell(row=44, column=2, value='=C11').number_format = '0'
# Subsequent flow:
for i, col_i in enumerate(range(3, 9)):  # i: 0..5, col 3..8 = May..Oct
    L = get_column_letter(col_i)
    PL = get_column_letter(col_i - 1)
    cf_col = get_column_letter(i + 2)  # CF col B for M1 = i=0 -> B
    # Sales sold this month
    ws4.cell(row=45, column=col_i, value=f"='Cash Flow (P&L)'!{cf_col}6").number_format = '0'
    # Units bought = monthly replenishment / avg unit cost
    ws4.cell(row=46, column=col_i, value=f'=B22/B18').number_format = '0.0'
    # Flippable end = prev start - sales + bought
    ws4.cell(row=47, column=col_i, value=f'={PL}44-{L}45+{L}46').number_format = '0'
    # Next start = previous end
    if col_i < 8:
        next_col = get_column_letter(col_i + 1)
        ws4.cell(row=44, column=col_i + 1, value=f'={L}47').number_format = '0'

# Borders
for r in range(43, 48):
    for col_i in range(1, 9):
        ws4.cell(row=r, column=col_i).border = BORDER

# === SECTION F: Guardrails ===
ws4['A49'] = 'F. GUARDRAILS'
ws4['A49'].fill = SECTION_FILL; ws4['A49'].font = SECTION_FONT
ws4.merge_cells('A49:G49')

guards = [
    "• Core inventory (rows 6-7) is NEVER liquidated — that's the appreciation engine. Treat as off-limits.",
    "• If flippable units drop below 30% of total inventory, pause sales and reset listing strategy.",
    "• If SURPLUS/SHORTFALL (B23) goes negative, raise reinvestment % or lower sales velocity.",
    "• All buys must be Sheldon-vetted per CUR-42 before execution.",
    "• Bullion buys >$200/unit need board check-in.",
    "• Track monthly: flippable count, avg listing age, sell-through rate, gross-margin %.",
]
for i, g in enumerate(guards, 50):
    ws4.cell(row=i, column=1, value=g)
    ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)

# ============================================================
# Save
# ============================================================
out = '/Users/josh/.openclaw/workspace/curate-coins-6mo-forecast.xlsx'
wb.save(out)
print(f"Saved: {out}")
