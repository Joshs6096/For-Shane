#!/usr/bin/env python3
"""Build CUR-28 grading-recommendations.xlsx for board review."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ----- Styles -----
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

B1_FILL = PatternFill("solid", fgColor="D1FAE5")  # green
B2_FILL = PatternFill("solid", fgColor="FEF3C7")  # amber
B3_FILL = PatternFill("solid", fgColor="FEE2E2")  # red

TITLE_FONT = Font(name="Calibri", size=14, bold=True)
SUB_FONT = Font(name="Calibri", size=10, italic=True, color="6B7280")
BOLD = Font(name="Calibri", size=11, bold=True)

THIN = Side(border_style="thin", color="9CA3AF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, row_idx, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def apply_borders_and_wrap(ws, start_row, end_row, n_cols, fill=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = WRAP
            if fill is not None:
                cell.fill = fill


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Sheet 1: Summary
# ============================================================
ws = wb.active
ws.title = "Summary"

ws["A1"] = "CUR-28 — Grade-vs-Fee Triage on the 188oz Holdings"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:F1")

ws["A2"] = "Author: Sheldon (Numismatic Grader & Researcher)  ·  Date: 2026-05-07  ·  Spot baseline: $74/oz silver  ·  Source: coin-collector MCP (live)"
ws["A2"].font = SUB_FONT
ws.merge_cells("A2:F2")

ws["A4"] = "Recommendation (one line)"
ws["A4"].font = BOLD
ws["A4"].fill = HEADER_FILL
ws["A4"].font = Font(bold=True, color="FFFFFF")
ws.merge_cells("A4:F4")
ws["A5"] = ("Submit ONE batch — 13 American Silver Eagle Proofs (mixed years 2006–2022) to PCGS Bulk Modern. "
            "~$285 fees, ~$135 expected net profit (photo-gated). Everything else falls into Bucket 2 (photo-gated) "
            "or Bucket 3 (sell raw / hold raw). At $74 spot, modern raw bullion grading economics are broken — "
            "MS70 premium ($25–40/coin) is too thin to pay back grading friction.")
ws["A5"].alignment = WRAP
ws.merge_cells("A5:F5")
ws.row_dimensions[5].height = 60

ws["A7"] = "Headline finding"
ws["A7"].font = Font(bold=True, color="FFFFFF")
ws["A7"].fill = HEADER_FILL
ws.merge_cells("A7:F7")
ws["A8"] = ("THE SPOT-COLLAPSE TRAP: at $74/oz silver, the MS70 premium ($25–40 over melt for common-date ASE) is too small "
            "a fraction of total realized value to justify grading friction. Walked example: 2014 ASE Bullion (qty 15, $26 cost) "
            "nets only ~$3/coin lift from grading vs. selling raw. This kills the v1 hypothesis of submitting modern bullion. "
            "Modern PROOFS still pencil (PR70 vs PR69 spread is $50–90/coin); CLASSIC US in MS condition still pencils (rarity premium).")
ws["A8"].alignment = WRAP
ws.merge_cells("A8:F8")
ws.row_dimensions[8].height = 75

# Bucket totals table
ws["A10"] = "Portfolio scope — coin-collector MCP, 2026-05-07 21:38 UTC"
ws["A10"].font = BOLD
ws.merge_cells("A10:F10")

scope_headers = ["Metric", "Value"]
scope_rows = [
    ("Catalog entries", "416"),
    ("Total coins (set-expanded)", "2,574"),
    ("Total cost paid", "$17,877.57"),
    ("Total high-value (catalog)", "$77,866.07"),
    ("Total low-value (catalog)", "$45,230.91"),
    ("Unrealized profit", "$59,988.50  (+335.55%)"),
    ("Silver oz on hand", "188.033 oz / 339 entries / 2,365 coins"),
    ("Silver spot baseline (CUR-29 W19)", "$74/oz  (MCP spot field is null; using dashboard baseline)"),
]

ws.cell(row=11, column=1, value=scope_headers[0]).font = HEADER_FONT
ws.cell(row=11, column=1).fill = HEADER_FILL
ws.cell(row=11, column=1).alignment = CENTER
ws.cell(row=11, column=2, value=scope_headers[1]).font = HEADER_FONT
ws.cell(row=11, column=2).fill = HEADER_FILL
ws.cell(row=11, column=2).alignment = CENTER

for i, (k, v) in enumerate(scope_rows, start=12):
    ws.cell(row=i, column=1, value=k).font = BOLD
    ws.cell(row=i, column=2, value=v)
    for c in range(1, 3):
        ws.cell(row=i, column=c).border = BORDER
        ws.cell(row=i, column=c).alignment = LEFT

# Bucket roll-up
roll_start = 22
ws.cell(row=roll_start, column=1, value="Bucket roll-up").font = BOLD
ws.merge_cells(start_row=roll_start, start_column=1, end_row=roll_start, end_column=6)

bucket_hdr = ["Bucket", "Action", "Coins / Clusters", "Est. fees", "Expected net", "Confidence"]
for c, h in enumerate(bucket_hdr, start=1):
    cell = ws.cell(row=roll_start + 1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

bucket_rows = [
    ("Bucket 1 — Submit Now", "Submit one batch (Batch A: 13 ASE Proofs → PCGS Bulk Modern)",
     "13 coins / 1 batch", "~$285", "~$135 (base case)", "Medium (photo-gated)", B1_FILL),
    ("Bucket 2 — Photos required", "Macro-photo audit by Elvis; Sheldon writes v3 go/no-go after",
     "8 candidate clusters (~30+ coins)", "Variable",
     "Variable; high stakes on 1976-S Ike crackout", "Medium (photo-dependent)", B2_FILL),
    ("Bucket 3 — Do not submit", "Sell raw (modern bullion) / hold raw (junk silver) / sell intact (sets) / hold graded (existing slabs)",
     "~80–100 modern bullion + class-wide junk silver / sets / slabs", "$0", "$0 grading lift; raw sale captures spot premium",
     "High", B3_FILL),
]
for i, (b, act, coins, fees, net, conf, fill) in enumerate(bucket_rows, start=roll_start + 2):
    vals = [b, act, coins, fees, net, conf]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.alignment = WRAP
        cell.border = BORDER
        cell.fill = fill
    ws.row_dimensions[i].height = 60

# Decision asks
da_start = roll_start + 6
ws.cell(row=da_start, column=1, value="Decisions the board needs to make").font = BOLD
ws.merge_cells(start_row=da_start, start_column=1, end_row=da_start, end_column=6)

da_hdr = ["#", "Decision", "Default if no answer", "Owner"]
for c, h in enumerate(da_hdr, start=1):
    cell = ws.cell(row=da_start + 1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
ws.merge_cells(start_row=da_start + 1, start_column=2, end_row=da_start + 1, end_column=3)

decisions = [
    ("1", "Approve Batch A (13 ASE Proofs → PCGS Bulk Modern, ~$285 fees, ~$135 net) pending photo audit?",
     "Defer to v3", "Board → Elvis pulls photos"),
    ("2", "Crack one Bicentennial Silver Proof Set (1 of 2 sets, hedged) for the 1976-S Ike DCAM PR70 chase?",
     "Hold both intact", "Board → Sheldon executes if greenlit"),
    ("3", "Disclose contents of 2024 Mint Limited Edition Silver Proof Set so we can run crackout math?",
     "Hold intact", "Board / Elvis"),
    ("4", "Reconcile inventory `high_value_each` against eBay 30-day sold medians (catalog overstates real market on common-date bullion)?",
     "Sheldon flags but no action", "CEO"),
]
for i, (num, dec, default, owner) in enumerate(decisions, start=da_start + 2):
    ws.cell(row=i, column=1, value=num).alignment = CENTER
    ws.cell(row=i, column=2, value=dec).alignment = WRAP
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
    ws.cell(row=i, column=4, value=default).alignment = WRAP
    ws.cell(row=i, column=5, value=owner).alignment = WRAP
    for c in range(1, 6):
        ws.cell(row=i, column=c).border = BORDER
    ws.row_dimensions[i].height = 45

# Method and gates
gates_start = da_start + 7
ws.cell(row=gates_start, column=1, value="Economic gates (must clear all 3 to submit)").font = BOLD
ws.merge_cells(start_row=gates_start, start_column=1, end_row=gates_start, end_column=6)

gate_rows = [
    ("Gate 1 — Raw cost", "≤ $65/coin for modern bullion; ≤ 50% of MS65 retail for classic US"),
    ("Gate 2 — Service routing", "PCGS for modern bullion + modern proofs; NGC for classic US"),
    ("Gate 3 — Batch size", "n ≥ 5 of same series/window on a single submission form"),
    ("Gate 4 (new this rev)", "Modern bullion gate effectively closed at $74 spot — MS70 premium too thin; PROOFS only"),
    ("Express gate", "Never NGC Express on modern bullion (CUR-27 v2 lesson)"),
]
for i, (g, t) in enumerate(gate_rows, start=gates_start + 1):
    ws.cell(row=i, column=1, value=g).font = BOLD
    ws.cell(row=i, column=2, value=t)
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
    for c in range(1, 7):
        ws.cell(row=i, column=c).border = BORDER
        ws.cell(row=i, column=c).alignment = LEFT

set_widths(ws, [22, 30, 26, 22, 22, 22])

# ============================================================
# Sheet 2: Bucket 1 — Submit Now
# ============================================================
ws = wb.create_sheet("Bucket 1 - Submit Now")

ws["A1"] = "Bucket 1 — Submit Now (Batch A)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:K1")

ws["A2"] = ("Single batch: 13 American Silver Eagle Proofs (mixed years 2006–2022) → PCGS Bulk Modern, single shipment. "
            "Skip First Strike / Advance Release labels (window long-closed for these years). "
            "Photo audit required before sub form is finalized.")
ws["A2"].font = SUB_FONT
ws["A2"].alignment = WRAP
ws.merge_cells("A2:K2")
ws.row_dimensions[2].height = 45

# Per-coin table
b1_hdr = [
    "Inv ID", "Year", "Coin", "Qty", "Cost / each",
    "Catalog high $", "Comp source",
    "Predicted PR70 retail", "PR69 fallback", "Predicted lift / coin (if PR70)",
    "Notes / justification"
]
hdr_row = 4
for c, h in enumerate(b1_hdr, start=1):
    cell = ws.cell(row=hdr_row, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

b1_rows = [
    (162, "2021", "American Eagle Silver Proof (Type 2)", 4, 86.00, 219.99,
     "eBay sold 2021-W PR70 DCAM, 2026 Q1–Q2",
     "$130–180", "$80–95", "+$30–55",
     "Type 2 reverse die premium real; PR69 close to raw cost"),
    (116, "2006", "American Eagle Silver Proof Bullion Coin", 3, 86.00, 194.99,
     "PCGS CoinFacts + eBay sold 2026",
     "$150–200", "$90–110", "+$40–70",
     "20th-anniversary year — premium label window if any FS/Advance Release strikes are present"),
    (148, "2014", "American Eagle Silver Proof Coin", 3, 86.00, 165.00,
     "eBay sold 2026 30-day median",
     "$130–160", "$90–110", "+$25–45",
     "Mid-vintage; 70-rate ~70% historically"),
    (155, "2020", "American Eagle Silver Proof", 1, 86.10, 399.00,
     "eBay sold 2020-W PR70 DCAM 2026",
     "$200–300 (W variants higher)", "$100–130", "+$60–120",
     "$399 catalog suggests 2020-W or Congratulations Set example. PHOTO AUDIT CRITICAL — large lift on FS variant; if standard issue, lift is in the base-case band"),
    (167, "2022", "American Eagle Silver Proof", 1, 86.10, 126.00,
     "eBay sold 2026 30-day median",
     "$130–150", "$90–100", "+$15–25",
     "Recent year, modest lift; batches well"),
    (123, "2007", "American Eagle Silver Proof Coin", 1, 86.10, 140.00,
     "eBay sold 2026 30-day median",
     "$130–160", "$85–110", "+$20–35",
     "Mid-vintage; routine candidate; batches well"),
]

start = hdr_row + 1
for i, row in enumerate(b1_rows, start=start):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.alignment = WRAP
        cell.border = BORDER
        cell.fill = B1_FILL
    ws.row_dimensions[i].height = 60

# Aggregate economics
agg_start = start + len(b1_rows) + 1
ws.cell(row=agg_start, column=1, value="Aggregate economics (base case, 75% PR70 conversion)").font = BOLD
ws.merge_cells(start_row=agg_start, start_column=1, end_row=agg_start, end_column=11)

econ = [
    ("Coins in batch", "13"),
    ("Total raw cost (already paid)", "$1,118.40"),
    ("PCGS Bulk Modern fee", "13 × $20 = $260"),
    ("Return shipping", "~$25"),
    ("Total fees", "~$285"),
    ("Expected gross lift @ 75% PR70 rate", "~9.75 coins × $50 avg lift = ~$485"),
    ("Less: 3.25 coins at PR69 (no upgrade)", "~$65 absorbed fees"),
    ("Net before/after fees", "~$420 net before fees / ~$135 net AFTER fees"),
    ("Realized portfolio value (post-grade)", "~$1,600–2,000 vs ~$1,100 raw"),
    ("Swing factor (upside)", "ID 155 (2020 ASE Proof) — if FS or 2020-W variant, batch lift jumps materially"),
]
for i, (k, v) in enumerate(econ, start=agg_start + 1):
    ws.cell(row=i, column=1, value=k).font = BOLD
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
    ws.cell(row=i, column=5, value=v)
    ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=11)
    for c in range(1, 12):
        ws.cell(row=i, column=c).border = BORDER
        ws.cell(row=i, column=c).alignment = LEFT

# Pre-flight + service notes
notes_start = agg_start + len(econ) + 2
ws.cell(row=notes_start, column=1, value="Pre-flight and service notes").font = BOLD
ws.merge_cells(start_row=notes_start, start_column=1, end_row=notes_start, end_column=11)

notes_rows = [
    ("Recommended grading service", "PCGS, modern bulk submission tier, single shipment"),
    ("Do NOT split", "Across NGC and PCGS — defeats batch economics (Gate 3 violation)"),
    ("Skip", "First Strike / Advance Release label add-ons (window long-closed for these years)"),
    ("Elvis pre-flight", "Confirm physical possession of all 13 coins; pull macro-photos of every obverse + reverse for Sheldon's per-coin call before sub form is filled out"),
    ("Slab status", "All 13 = `confirmed_raw` (held in OGP from Mint)"),
    ("Mint mark detection", "ID 155 requires explicit mint-mark photo (W vs S) — drives variant call"),
]
for i, (k, v) in enumerate(notes_rows, start=notes_start + 1):
    ws.cell(row=i, column=1, value=k).font = BOLD
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
    ws.cell(row=i, column=5, value=v)
    ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=11)
    for c in range(1, 12):
        ws.cell(row=i, column=c).border = BORDER
        ws.cell(row=i, column=c).alignment = LEFT

set_widths(ws, [8, 8, 28, 6, 12, 12, 26, 18, 18, 16, 38])

# ============================================================
# Sheet 3: Bucket 2 — Photos required
# ============================================================
ws = wb.create_sheet("Bucket 2 - Photos required")

ws["A1"] = "Bucket 2 — Submit Only If Photos Confirm"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:I1")

ws["A2"] = "Ranked by potential portfolio impact (highest stakes first). Each cluster gates on Elvis macro-photo audit before Sheldon writes v3 go/no-go."
ws["A2"].font = SUB_FONT
ws["A2"].alignment = WRAP
ws.merge_cells("A2:I2")
ws.row_dimensions[2].height = 30

b2_hdr = [
    "Rank", "Inv ID(s)", "Coin / Set", "Qty",
    "Cost / each", "Catalog high $", "Stakes",
    "Photo audit ask", "Recommendation if photos clear (service)"
]
hdr_row = 4
for c, h in enumerate(b2_hdr, start=1):
    cell = ws.cell(row=hdr_row, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

b2_rows = [
    ("B1", "172", "1776–1976 United States Bicentennial Silver Proof Set", "2 sets / 10 coins",
     "$45/set", "$190/set",
     "Very high — 1976-S Silver Ike DCAM PR70 = $649–730 recent eBay; Heritage record $6,900. Set intact = $190.",
     "Macro photos of the Ike's reverse: cameo contrast and any field marks. PR69→PR70 cliff is the most binary call in the entire portfolio.",
     "If photos clear: submit ONE Ike (1 of 2 sets) to PCGS as single-coin Bulk Modern; keep second set sealed as hedge. PCGS (PR70 demand stronger on PCGS holders for the Ike series)."),

    ("B2", "169", "2024 Mint Limited Edition Silver Proof Set", "1 set / 5 coins",
     "$180", "$399.99",
     "Medium-high — set value $399.99; crackout depends on contents (board contents disclosure pending). ASE PR70 = $130–180; LE Set premium partially carried by packaging.",
     "Photo + contents disclosure required (which 5 coins). Same protocol as v1 ask.",
     "Default if no info: keep set intact, sell as set. If contents include high-PR70-spread coins, submit individually to PCGS."),

    ("B3", "5", "1859 Indian Head Cent (Type 1 Copper-Nickel)", "17 coins",
     "$20", "$159.95",
     "Medium-high — first-year-of-issue, semi-key. 1859/1858 DDO variety = $1,000+ in MS. PCGS retail VG $35–65, F $80–160, XF $200–300, AU $400–500, MS60+ $700–2,000+. Risk: classic copper has ~40–50% body-bag rate on raw lots.",
     "All 17 coins, both sides, focused on (a) cleaning indicators (hairlines, color anomalies), (b) variety check on obverse for 1859/1858 DDO.",
     "If photos clear: submit only the 8–10 cleanest examples as NGC Standard tier (not Bulk; classic copper needs higher tier for legitimate strike attribution). Hold remainder raw. Predicted lift if 8 of 17 grade XF+: ~$320 gross / ~$40 net + variety lottery upside. Borderline. NGC."),

    ("B4", "228", "1982 Libertad", "3 coins",
     "$0 (gifted)", "$226.10",
     "Medium — first-year-of-issue Mexican Libertad. NGC MS65 ~$300; MS68 ~$2,500; MS69 ~$8,000+. GATE 3 CONFLICT (n<5).",
     "Luster, contact marks, edge (1982 known for poor strike on eagle's wing). MS65+ is the threshold for grading to pay back.",
     "If any photo clearly shows MS66+ surfaces: submit that single coin under NGC Standard tier ($40); keep rest raw. Don't blanket-submit. NGC."),

    ("B5", "232", "1996 American Silver Eagle PCS PR69 DCAM (already slabbed)", "1 coin",
     "$0 (gifted)", "$345",
     "Low-medium — PCS is a low-tier slab; PCGS or NGC PR70 = $400–700 for 1996 (a tougher year). Crossover at PCGS = $25.",
     "Existing slab photo to assess upgrade likelihood (any haze, milk spots, hairline?).",
     "If photos show pristine surfaces: PCGS crossover as one-off, add-on to Batch A submission. Risk: PCGS may downgrade to PR69; net ~$0 lift. PCGS."),

    ("B6", "416", "2026 Australian Swan Proof", "2 coins",
     "$97.50", "$275",
     "Medium — recently acquired (2026-05-05); likely follow-on to CUR-2's NGC bullion-Swan submission. Catalog $275 high suggests PR70 retail expectation.",
     "DEFER until CUR-2 NGC submission #1 returns 2026-06-16 → 06-30. Use realized grades to calibrate Perth-mint 70-rate before adding more Swans.",
     "If Perth-Swan 70-rate ≥ 70%: batch with any other 2026 RCM/Perth modern proof bullion at PCGS. If ≤ 70%: sell raw at premium retail. PCGS (per CUR-27 policy)."),

    ("B7", "—", "WLH (Walking Liberty Half) common-date short-set", "11 coins (1936–1947 dates)",
     "$75–140", "varies",
     "Low-medium — common dates raw at $75–140 most likely XF/AU. PCGS AU58 = $50–90; MS62 = $100–150; MS63 = $150–300; MS64 = $300–650; MS65 = $700–2,000.",
     "Liberty's head-detail, skirt-line completeness, luster cartwheels. Only MS-grade candidates pay back.",
     "Likely 0–2 of 11 grade MS64+. Submit at most 2–3 individually if photos justify. Don't pre-commit. NGC."),

    ("B8", "243, 249, 301", "1891 Morgan trio", "3 coins",
     "varies", "varies",
     "Medium IF a 1891-CC variety is among them; otherwise low. 1891-P/O/S common; 1891-CC = NGC MS62 ~$700; MS63 ~$1,500.",
     "Mintmark check on reverse below eagle (CC, O, S, none).",
     "If any 1891-CC: submit that one to NGC Standard tier. Otherwise: hold raw or sell raw at $200–275. NGC."),
]

start = hdr_row + 1
for i, row in enumerate(b2_rows, start=start):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.alignment = WRAP
        cell.border = BORDER
        cell.fill = B2_FILL
    ws.row_dimensions[i].height = 110

set_widths(ws, [6, 12, 26, 14, 12, 14, 38, 38, 50])

# ============================================================
# Sheet 4: Bucket 3 — Don't Submit
# ============================================================
ws = wb.create_sheet("Bucket 3 - Don't submit")

ws["A1"] = "Bucket 3 — Do Not Submit (gate-fail or worth more raw)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:E1")

ws["A2"] = "Class-level standing rules from v1, accepted as policy by CEO 2026-05-06, applied to live MCP inventory."
ws["A2"].font = SUB_FONT
ws.merge_cells("A2:E2")

b3_hdr = ["Class", "Examples / Inv IDs", "Why no grading", "Action", "Trigger to revisit"]
hdr_row = 4
for c, h in enumerate(b3_hdr, start=1):
    cell = ws.cell(row=hdr_row, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

b3_rows = [
    ("Modern silver bullion 2024–2026 acquired at $78–110/ea",
     "Koalas qty 6; Britannias qty 7; Maples qty 7+; Pandas qty 5; Queens Beasts qty 2; Libertads 2022–23 qty 6; raw ASE bullion outside 2014 cluster; many others (~80–100 coins)",
     "SPOT-COLLAPSE TRAP. At $74 spot, MS70 premium ($25–40/coin) too thin to pay back grading friction. Cost basis $78–110 already at/above raw market.",
     "SELL RAW at premium retail ($110–220/ea by series) while spot is at $74.",
     "Silver retraces below $50/oz → re-evaluate; or class-specific shortage premium spike."),

    ("2014 ASE Bullion (representative case for class above)",
     "ID 146, qty 15, cost $26/ea",
     "Most-promising case in entire inventory; even at sub-melt cost, lift from grading = $3/coin vs $38 raw profit. Headline data point for the spot-collapse trap.",
     "SELL RAW. Add to CUR-10 SELL bucket at $80–90/coin (premium-bullion pricing, sells fast at current spot).",
     "Spot moves >±10% from $74."),

    ("Circulated common-date silver dollars",
     "Morgans 1889–1926 + Peace dollars at $66 cost basis VF/EF, qty ~15 across IDs 6, 7, 8, 10, 12, 13, 14, 17, 19, 22",
     "Not MS-grade; circulated common-date Morgans at VF/EF have flat slab demand. Slabbing fee exceeds grade lift.",
     "Sell raw at $115–250/ea per CUR-10 hold/sell.",
     "If any 1893-S, 1895, or 1889-CC surfaces (none seen)."),

    ("Junk silver bags (sub-MS65 raw)",
     "Roosevelt dimes ID 382 qty 99; Mercury dimes IDs 381, 414 qty 103; Silver Quarters IDs 383, 415 qty 85; Franklin Halves ID 385 qty 17; Kennedy Halves 1965–1970 ID 174 qty 40; 1964 Kennedys IDs 28, 380, 384 qty 10",
     "Sells at/near spot; no rarity premium for circulated common-date silver. Per CUR-11 spot-trigger ≥$95/oz.",
     "HOLD. Per CUR-11 spot-trigger ≥$95/oz to liquidate.",
     "Spot ≥ $95/oz triggers CUR-11 sell."),

    ("Already-graded coins (PCGS/NGC/CAC MS70/PR70 + PCS PR69 except crossover candidate)",
     "IDs 333, 335, 231, 238, 268, 280, 279, 343, 354, 366, 413, 418",
     "Already at top retail tier. Re-grading risk > upside.",
     "HOLD or SELL graded; no re-grade exercise.",
     "Specific tier-jump opportunities (e.g., CAC sticker on PCGS coin that lacks one) — case-by-case, not class-wide."),

    ("Pre-2024 standard-issue Silver Proof Sets / Mint Sets",
     "IDs 45, 60, 72, 82, 85, 87, 89, 103, 106, 107, 111, 112, 115, 120, 124, 143, 149, 150, 152, 154, 156, 160, 161, 164, 166, 364, others",
     "Crackout math fails on standard issues at current MS70 premiums; intact set holds packaging premium.",
     "HOLD INTACT, sell as sets.",
     "Specific high-stakes year identified (e.g., 1976-S Bicentennial in B1)."),

    ("Foreign single-coin oddities",
     "USSR Olympic Basketball; lone Filipina; Liberian Platinum; single Isle of Man; etc. (n<5 each)",
     "Gate 3 fail (batch size < 5). No grading lift.",
     "Sell raw to specialist buyers.",
     "Bundling opportunity into a curated foreign lot for batch grading (unlikely)."),

    ("NORFEDs / private mint silver rounds and bars",
     "Silver Round/Bar category, 19 entries / 40 coins",
     "No graded market — privately minted rounds/bars don't slab competitively.",
     "Sell raw at melt + small premium.",
     "Spot moves materially."),

    ("Cleaned, damaged, or low-grade cents",
     "ID 419 (Indian Head 1904); ID 420 (1905) at $2.50 cost; ID 4 (1848 Braided Hair 3-coin); ID 387 (1968 IGC slab MS67 RD)",
     "Low cost-basis raw cents fail Gate 1; IGC slab is third-tier and crossover is only worthwhile if confirmed superior surfaces.",
     "Sell raw / hold IGC slabs intact unless photos suggest crossover.",
     "Photo audit on ID 387 confirms surfaces support PCGS upgrade."),

    ("Active SELL listings — already on the market raw",
     "CUR-3 through CUR-8 (6 active listings)",
     "Already in motion — no grading injection at this stage.",
     "Continue listed.",
     "Listing expires unsold + relisting strategy review."),
]

start = hdr_row + 1
for i, row in enumerate(b3_rows, start=start):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.alignment = WRAP
        cell.border = BORDER
        cell.fill = B3_FILL
    ws.row_dimensions[i].height = 110

set_widths(ws, [22, 38, 38, 28, 30])

# ============================================================
# Sheet 5: Comp sources
# ============================================================
ws = wb.create_sheet("Comp sources")

ws["A1"] = "Comp Sources Cited"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:C1")

ws["A2"] = "Pricing decays — every comp carries a date and source. Refresh before the next submission greenlight."
ws["A2"].font = SUB_FONT
ws.merge_cells("A2:C2")

cs_hdr = ["Source", "Used for", "Notes"]
for c, h in enumerate(cs_hdr, start=1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

cs_rows = [
    ("PCGS CoinFacts (pcgs.com/coinfacts)",
     "Series retail bands and 70-rate priors",
     "1976-S Silver Ike, 1859 IHC, 1982 Libertad, modern ASE Proof series. Authoritative for PCGS-graded comparables."),
    ("eBay sold listings (ebay.com, 30-day sold filter)",
     "Realized-price ground truth for modern bullion + ASE Proof",
     "2024–2026 modern ASE bullion ($44–55 MS70); ASE Proof ($130–180 PR70 DCAM); Bicentennial Silver Ike DCAM PR70 ($649–730 recent). Use 30-day median, not list price."),
    ("Heritage Auctions realized (ha.com)",
     "High-end / record comps",
     "1976-S Silver Ike DCAM PR70 record $6,900 anchor."),
    ("GreatCollections (greatcollections.com)",
     "Mid-market US auction comps",
     "Cross-check on classic US series."),
    ("NGC and PCGS fee schedules",
     "Submission cost calc",
     "NGC Bulk Modern $17; PCGS Bulk Modern $20. Always include return shipping ~$25 in batch math."),
    ("Internal Sheldon resources",
     "Durable lessons applied this pass",
     "grading-cliff-1976S-silver-ike; ngc-vs-pcgs-modern-bullion; modern-bullion-grading-economics; silver-spot-regime; spot-collapse-bullion-grading (new this revision)."),
    ("CUR-29 W19 dashboard",
     "Spot baseline",
     "$74/oz silver spot used throughout. MCP `spot_prices` field is null this run."),
    ("coin-collector MCP",
     "Inventory ground truth",
     "All cost basis, qty, catalog high-value, and inventory tallies in this xlsx pulled live from MCP 2026-05-07 21:38 UTC. Reconciles to W19 dashboard."),
]
for i, (s, u, n) in enumerate(cs_rows, start=5):
    ws.cell(row=i, column=1, value=s).alignment = WRAP
    ws.cell(row=i, column=2, value=u).alignment = WRAP
    ws.cell(row=i, column=3, value=n).alignment = WRAP
    for c in range(1, 4):
        ws.cell(row=i, column=c).border = BORDER
    ws.row_dimensions[i].height = 50

set_widths(ws, [32, 30, 60])

# ============================================================
# Save
# ============================================================
out = "/Users/josh/.paperclip/instances/default/workspaces/bf3d8eaa-282f-494b-bfab-b370b8561669/agents/sheldon/CUR-28-grading-recommendations.xlsx"
wb.save(out)
print(f"WROTE {out}")
