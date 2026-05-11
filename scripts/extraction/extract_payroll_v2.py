"""
Payroll trajectory extraction from SP_Analysis daily cash forecast files.
Handles both Format A (2022) and Format B (2023+) monthly total structures.

Format A (Sep/Dec 2022):
  - Month labels like 'Feb 22', 'July 22' appear BEFORE the daily rows for that month
  - Monthly total is a None-date row appearing BEFORE the next month label

Format B (Apr 2023+):
  - Month labels like 'Feb 23', 'July 23' appear BEFORE the daily rows
  - Monthly total is a string row like 'Feb', 'Jul' appearing AFTER the daily rows
"""

import openpyxl
import datetime
import json
import os
import re

BASE_DIR = "/Users/josh/Downloads/SP_Analysis"

TARGET_FILES = [
    ("9.19.22_Sep BOD",         "2022-09-19",  "Daily Cash Fcst - 9.19.22_BU view_Sep BOD.xlsx",         "A"),
    ("12.7.22_DecBOD",          "2022-12-07",  "Daily Cash Fcst - 12.7.22_BU view_DecBOD Treasury_liquidity update final.xlsx", "A"),
    ("4.25.23_AprBOD",          "2023-04-25",  "Daily Cash Fcst - 4.25.23_BU view_April 27 BOD.xlsx",     "B"),
    ("8.22.23_AugBOD",          "2023-08-22",  "Daily Cash Fcst - 8.22.23_BU view_AugBOD.xlsx",           "B"),
    ("1.23.24_BOD",             "2024-01-23",  "Daily Cash Fcst - 1.23.24_BU view_BOD.xlsx",              "B"),
    ("7.23.24_JulBOD",          "2024-07-23",  "Daily Cash Fcst - 7.23.24_deck_JulBOD_FY25.xlsx",         "B"),
    ("12.31.24",                "2024-12-31",  "Daily Cash Fcst - 12.31.24_BU view.xlsx",                 "B"),
    ("1.21.25_JanBOD",          "2025-01-21",  "Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx", "B"),
    ("5.20.25_MayBOD",          "2025-05-20",  "Daily Cash Fcst - 5.20.25_BU view_deck version_MayBOD.xlsx", "B"),
    ("10.21.25_OctBOD",         "2025-10-21",  "Daily Cash Fcst - 10.21.25_BU view_deck version_OctBOD.xlsx", "B"),
    ("12.30.25_year_end",       "2025-12-30",  "Daily Cash Fcst - 12.30.25_BU view_deck version final.xlsx", "B"),
    ("02.17.26_deck",           "2026-02-17",  "Daily Cash Fcst - 02.17.26_BU view_deck version.xlsx",    "B"),
    ("03.03.26_post-reductions","2026-03-03",  "Daily Cash Fcst - 03.03.26_post-reductions_deck version_MarBOD.xlsx", "B"),
    ("04.21.26_AprBOD",         "2026-04-21",  "Daily Cash Fcst - 04.21.26_ BU version_deck version_AprBOD.xlsx", "B"),
    ("05.08.26",                "2026-05-08",  "Daily Cash Fcst - 05.08.26_BU view.xlsx",                 "B"),
]

# Abbreviated month name -> full month name
MONTH_ABBREVS = {
    'jan': 'January', 'feb': 'February', 'mar': 'March', 'apr': 'April',
    'may': 'May', 'jun': 'June', 'jul': 'July', 'aug': 'August',
    'sep': 'September', 'oct': 'October', 'nov': 'November', 'dec': 'December',
}


def is_month_abbrev(val):
    """Check if a string is just a 3-letter month abbreviation (Format B total row)."""
    if not isinstance(val, str):
        return False
    v = val.strip().lower()
    return v in MONTH_ABBREVS


def is_month_label(val):
    """Check if a string is a month+year label like 'Aug 22', 'Sept 22', 'Feb 23'."""
    if not isinstance(val, str):
        return False
    v = val.strip()
    # Matches: 'Feb 22', 'March 22', 'July 22', 'Sept 22', 'June 23', etc.
    return bool(re.match(r'^[A-Za-z]+ \d{2,4}$', v)) and v != 'Banking days'


def extract_payroll(label, file_date_str, fname, fmt):
    filepath = os.path.join(BASE_DIR, fname)
    result = {
        "label": label,
        "file_date": file_date_str,
        "filename": fname,
        "format": fmt,
        "errors": [],
        "monthly_data": [],
        "weekly_data": {},
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        result["errors"].append(f"Cannot open file: {e}")
        return result

    # Find 'Payroll' sheet
    payroll_sheet = None
    for sname in wb.sheetnames:
        if sname.lower() == "payroll":
            payroll_sheet = sname
            break
    if payroll_sheet is None:
        result["errors"].append("No 'Payroll' sheet found")
        wb.close()
        return result

    try:
        ws = wb[payroll_sheet]
        all_rows = []
        for row in ws.iter_rows(max_row=600, max_col=40, values_only=True):
            all_rows.append(row)

        # Determine column indices from header row (row index 1 = row 2)
        header = all_rows[1] if len(all_rows) > 1 else []

        # Find Total columns
        fcst_total_idx = None
        act_total_idx = None
        occurrence = 0
        for j, v in enumerate(header):
            if v == 'Total':
                if occurrence == 0:
                    fcst_total_idx = j
                elif occurrence == 1:
                    act_total_idx = j
                occurrence += 1

        if fcst_total_idx is None:
            result["errors"].append("Cannot find 'Total' column in header")
            wb.close()
            return result

        # Build BU name -> column index for forecast and actuals
        fcst_bu_cols = {}  # bu_name -> col_idx (0-based)
        act_bu_cols = {}

        for j in range(1, (fcst_total_idx or 0) + 1):
            v = header[j] if j < len(header) else None
            if v and v != 'Total':
                fcst_bu_cols[v] = j

        if act_total_idx is not None:
            for j in range(fcst_total_idx + 2, act_total_idx + 1):
                v = header[j] if j < len(header) else None
                if v and v != 'Total':
                    act_bu_cols[v] = j

        result["fcst_total_col"] = fcst_total_idx + 1  # 1-based for reporting
        result["act_total_col"] = act_total_idx + 1 if act_total_idx else None
        result["fcst_bus"] = list(fcst_bu_cols.keys())
        result["act_bus"] = list(act_bu_cols.keys())

        def get_bu_breakdown(row, bu_cols):
            out = {}
            for bu, col_idx in bu_cols.items():
                v = row[col_idx] if col_idx < len(row) else None
                if isinstance(v, (int, float)) and v != 0:
                    out[bu] = round(v, 2)
            return out

        # === EXTRACT MONTHLY TOTALS ===
        monthly_data = []

        if fmt == "A":
            # Format A: None-date rows with numeric data are monthly totals
            # They appear between month label rows
            # Map: find month label rows, then the None-date total row just before the NEXT label
            label_indices = []  # (row_idx_0based, label_str)
            for i, row in enumerate(all_rows):
                col_a = row[0] if len(row) > 0 else None
                if is_month_label(col_a):
                    label_indices.append((i, col_a))

            for k, (lbl_idx, lbl_str) in enumerate(label_indices):
                # The total for month k is the None-date row between label k and label k+1
                end_idx = label_indices[k + 1][0] if k + 1 < len(label_indices) else len(all_rows)
                for i in range(lbl_idx + 1, end_idx):
                    row = all_rows[i]
                    col_a = row[0]
                    fcst_val = row[fcst_total_idx] if fcst_total_idx < len(row) else None
                    act_val = row[act_total_idx] if act_total_idx and act_total_idx < len(row) else None
                    if col_a is None and isinstance(fcst_val, (int, float)) and fcst_val != 0:
                        fcst_bu = get_bu_breakdown(row, fcst_bu_cols)
                        act_bu = get_bu_breakdown(row, act_bu_cols) if act_bu_cols else {}
                        monthly_data.append({
                            "month_label": lbl_str,
                            "row": i + 1,
                            "forecast_total": round(fcst_val, 2),
                            "actuals_total": round(act_val, 2) if isinstance(act_val, (int, float)) else None,
                            "forecast_bu": fcst_bu,
                            "actuals_bu": act_bu,
                        })
                        break

        elif fmt == "B":
            # Format B: monthly total rows have an abbreviated month string in col A
            # like 'Feb', 'Jul', 'Oct' - these appear AFTER the daily rows
            for i, row in enumerate(all_rows):
                col_a = row[0] if len(row) > 0 else None
                if is_month_abbrev(col_a):
                    fcst_val = row[fcst_total_idx] if fcst_total_idx < len(row) else None
                    act_val = row[act_total_idx] if act_total_idx and act_total_idx < len(row) else None
                    if isinstance(fcst_val, (int, float)) and fcst_val != 0:
                        # Find the corresponding full label (the month label row before this)
                        month_full_label = col_a.strip()
                        # Look backwards for the month label
                        for j in range(i - 1, max(i - 30, 0), -1):
                            prev_a = all_rows[j][0] if len(all_rows[j]) > 0 else None
                            if is_month_label(prev_a):
                                month_full_label = prev_a
                                break
                        fcst_bu = get_bu_breakdown(row, fcst_bu_cols)
                        act_bu = get_bu_breakdown(row, act_bu_cols) if act_bu_cols else {}
                        monthly_data.append({
                            "month_label": month_full_label,
                            "month_abbrev": col_a.strip(),
                            "row": i + 1,
                            "forecast_total": round(fcst_val, 2),
                            "actuals_total": round(act_val, 2) if isinstance(act_val, (int, float)) else None,
                            "forecast_bu": fcst_bu,
                            "actuals_bu": act_bu,
                        })

        result["monthly_data"] = monthly_data

        # === FIND MOST RECENT WEEK WITH ACTUAL DATA ===
        # Collect all daily rows (datetime in col A, non-zero total)
        daily_rows = []
        for i, row in enumerate(all_rows):
            col_a = row[0] if len(row) > 0 else None
            if isinstance(col_a, datetime.datetime):
                fcst_val = row[fcst_total_idx] if fcst_total_idx < len(row) else None
                act_val = row[act_total_idx] if act_total_idx and act_total_idx < len(row) else None
                if isinstance(fcst_val, (int, float)) and fcst_val != 0:
                    daily_rows.append((col_a, i, row))

        # Find most recent week where actuals > 0 (not forecast-only)
        # Actuals are non-zero for dates that have been paid
        actual_daily = [(d, i, r) for d, i, r in daily_rows
                        if act_total_idx and r[act_total_idx] and isinstance(r[act_total_idx], (int, float)) and r[act_total_idx] != 0]

        if actual_daily:
            last_act_date = actual_daily[-1][0]
            week_start = last_act_date - datetime.timedelta(days=6)
            week_rows = [(d, i, r) for d, i, r in actual_daily if d >= week_start]

            week_fcst_total = sum(r[fcst_total_idx] for d, i, r in week_rows
                                  if isinstance(r[fcst_total_idx], (int, float)))
            week_act_total = sum(r[act_total_idx] for d, i, r in week_rows
                                 if act_total_idx and act_total_idx < len(r) and isinstance(r[act_total_idx], (int, float)))

            week_fcst_bu = {}
            week_act_bu = {}
            for bu, col_idx in fcst_bu_cols.items():
                t = sum(r[col_idx] for d, i, r in week_rows
                        if col_idx < len(r) and isinstance(r[col_idx], (int, float)))
                if t != 0:
                    week_fcst_bu[bu] = round(t, 2)
            for bu, col_idx in act_bu_cols.items():
                t = sum(r[col_idx] for d, i, r in week_rows
                        if col_idx < len(r) and isinstance(r[col_idx], (int, float)))
                if t != 0:
                    week_act_bu[bu] = round(t, 2)

            result["weekly_data"] = {
                "week_end_date": last_act_date.strftime("%Y-%m-%d"),
                "week_start_date": week_start.strftime("%Y-%m-%d"),
                "dates": [d.strftime("%Y-%m-%d") for d, i, r in week_rows],
                "weekly_total_forecast": round(week_fcst_total, 2),
                "weekly_total_actuals": round(week_act_total, 2),
                "bu_breakdown_forecast": week_fcst_bu,
                "bu_breakdown_actuals": week_act_bu,
                "data_type": "actuals",
            }
        elif daily_rows:
            # Fall back to most recent forecast week
            last_date = daily_rows[-1][0]
            week_start = last_date - datetime.timedelta(days=6)
            week_rows = [(d, i, r) for d, i, r in daily_rows if d >= week_start]

            week_fcst_total = sum(r[fcst_total_idx] for d, i, r in week_rows
                                  if isinstance(r[fcst_total_idx], (int, float)))
            week_fcst_bu = {}
            for bu, col_idx in fcst_bu_cols.items():
                t = sum(r[col_idx] for d, i, r in week_rows
                        if col_idx < len(r) and isinstance(r[col_idx], (int, float)))
                if t != 0:
                    week_fcst_bu[bu] = round(t, 2)

            result["weekly_data"] = {
                "week_end_date": last_date.strftime("%Y-%m-%d"),
                "week_start_date": week_start.strftime("%Y-%m-%d"),
                "dates": [d.strftime("%Y-%m-%d") for d, i, r in week_rows],
                "weekly_total_forecast": round(week_fcst_total, 2),
                "weekly_total_actuals": None,
                "bu_breakdown_forecast": week_fcst_bu,
                "bu_breakdown_actuals": {},
                "data_type": "forecast",
            }

    except Exception as e:
        import traceback
        result["errors"].append(f"Extraction error: {e}\n{traceback.format_exc()}")

    wb.close()
    return result


def parse_month_label(label):
    """Parse a month label like 'Aug 22', 'Sept 22', 'Feb 23' into a sortable date string."""
    if not label:
        return None
    # Handle abbreviated labels like 'Oct' (Format B - look at context)
    parts = label.strip().split()
    if len(parts) == 1:
        return None
    month_str = parts[0].lower()[:3]
    year_str = parts[-1]
    # Month mapping
    months = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
               'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
    # Handle 'sept' -> sep
    if month_str == 'sep' or month_str == 'sep':
        month_str = 'sep'
    if month_str == 'jun':
        month_str = 'jun'
    m = months.get(month_str[:3])
    if m is None:
        return None
    try:
        y = int(year_str)
        if y < 100:
            y += 2000
        return f"{y:04d}-{m:02d}"
    except:
        return None


def main():
    all_results = []
    for label, file_date, fname, fmt in TARGET_FILES:
        print(f"Processing: {label}...", end=" ", flush=True)
        data = extract_payroll(label, file_date, fname, fmt)
        all_results.append(data)
        if data["errors"]:
            print(f"ERRORS: {data['errors']}")
        else:
            n_months = len(data["monthly_data"])
            n_with_actuals = sum(1 for m in data["monthly_data"] if m["actuals_total"] is not None)
            print(f"OK ({n_months} months, {n_with_actuals} with actuals)")

    # =============================================
    # BUILD PAYROLL TIMELINE
    # =============================================
    print("\n" + "="*80)
    print("PAYROLL TIMELINE - Monthly Actuals (dollars in thousands)")
    print("="*80)

    timeline_rows = []
    for r in all_results:
        label = r["label"]
        file_date = r["file_date"]
        monthly = r.get("monthly_data", [])
        weekly = r.get("weekly_data", {})

        # Find last month with actuals
        months_with_actuals = [m for m in monthly if m.get("actuals_total") is not None and m["actuals_total"] != 0]
        last_actual = months_with_actuals[-1] if months_with_actuals else None

        # Also find the most recent forecast month
        last_forecast = monthly[-1] if monthly else None

        # For weekly: use actuals if available, else forecast
        weekly_total = (weekly.get("weekly_total_actuals") or weekly.get("weekly_total_forecast"))
        weekly_type = weekly.get("data_type", "n/a")

        row = {
            "label": label,
            "file_date": file_date,
            # Best monthly data point (last full month with actuals)
            "last_actual_month_label": last_actual["month_label"] if last_actual else None,
            "last_actual_monthly_total": last_actual["actuals_total"] if last_actual else None,
            "last_actual_bu_breakdown": last_actual.get("actuals_bu") or last_actual.get("forecast_bu") if last_actual else None,
            # Most recent forecast month (for context)
            "latest_forecast_month_label": last_forecast["month_label"] if last_forecast else None,
            "latest_forecast_monthly_total": last_forecast["forecast_total"] if last_forecast else None,
            # Weekly data
            "weekly_date_range": f"{weekly.get('week_start_date')} to {weekly.get('week_end_date')}" if weekly else None,
            "weekly_total": weekly_total,
            "weekly_type": weekly_type,
            "weekly_bu_breakdown": weekly.get("bu_breakdown_actuals") or weekly.get("bu_breakdown_forecast"),
            "errors": r.get("errors", []),
        }
        timeline_rows.append(row)

        # Print summary
        am = row["last_actual_monthly_total"]
        wt = row["weekly_total"]
        am_str = f"${abs(am)/1e6:.2f}M" if am else "N/A"
        wt_str = f"${abs(wt)/1e6:.2f}M [{weekly_type}]" if wt else "N/A"
        print(f"\n{label} ({file_date})")
        print(f"  Last full month w/actuals: {row['last_actual_month_label']} -> {am_str}/mo")
        if row["last_actual_bu_breakdown"]:
            bu = row["last_actual_bu_breakdown"]
            sorted_bu = sorted(bu.items(), key=lambda x: abs(x[1]), reverse=True)
            print(f"  BU breakdown (monthly actuals):")
            for bu_name, bu_amt in sorted_bu:
                pct = abs(bu_amt)/abs(am)*100 if am else 0
                print(f"    {bu_name:20s}: ${abs(bu_amt)/1e6:.2f}M ({pct:.0f}%)")
        print(f"  Most recent week ({row['weekly_date_range']}): {wt_str}/wk | implied monthly: ${abs(wt)*52/12/1e6:.2f}M" if wt else "  Most recent week: N/A")

    # =============================================
    # KEY QUESTIONS
    # =============================================
    print("\n" + "="*80)
    print("KEY QUESTION 1: Sep 2022 vs May 2026")
    print("="*80)

    sep22 = next((r for r in timeline_rows if r["label"] == "9.19.22_Sep BOD"), None)
    may26 = next((r for r in timeline_rows if r["label"] == "05.08.26"), None)

    # Sep 2022: best data = Aug 2022 actuals (Sep is partial in Sep file; Dec file has Sep full)
    # May 2026: best data = Apr 2026 actuals (most recent full month before May 8 file date)

    sep22_monthly = sep22["last_actual_monthly_total"] if sep22 else None
    sep22_month_lbl = sep22["last_actual_month_label"] if sep22 else None

    # For May 2026, we need Apr 2026 actuals from the May 8 file
    may26_data = next((r for r in all_results if r["label"] == "05.08.26"), None)
    may26_monthly = None
    may26_month_lbl = None
    may26_bu = None
    if may26_data:
        months_with_act = [m for m in may26_data["monthly_data"] if m.get("actuals_total") and abs(m["actuals_total"]) > 1e6]
        if months_with_act:
            last = months_with_act[-1]
            may26_monthly = last["actuals_total"]
            may26_month_lbl = last["month_label"]
            may26_bu = last.get("actuals_bu") or last.get("forecast_bu")

    print(f"\nSep 2022 best monthly payroll ({sep22_month_lbl}): ${abs(sep22_monthly)/1e6:.2f}M/mo" if sep22_monthly else "Sep 2022: N/A")
    print(f"May 2026 best monthly payroll ({may26_month_lbl}): ${abs(may26_monthly)/1e6:.2f}M/mo" if may26_monthly else "May 2026: N/A")

    if sep22_monthly and may26_monthly:
        reduction = abs(sep22_monthly) - abs(may26_monthly)
        pct = reduction / abs(sep22_monthly) * 100
        print(f"\nAbsolute reduction: ${reduction/1e6:.2f}M/mo ({pct:.1f}% decline)")
        print(f"Annualized reduction: ${reduction*12/1e6:.1f}M")

        # Implied headcount
        print(f"\nImplied headcount analysis:")
        for avg_comp in [70000, 85000, 100000]:
            monthly_cost = avg_comp / 12
            sep_hc = abs(sep22_monthly) / monthly_cost
            may_hc = abs(may26_monthly) / monthly_cost
            print(f"  @${avg_comp:,} avg annual total comp: Sep 22 ~{sep_hc:,.0f} -> May 26 ~{may_hc:,.0f} (reduction of ~{sep_hc-may_hc:,.0f})")

    # =============================================
    # KEY QUESTION 2: Apr 21, 2026 vs May 8, 2026
    # =============================================
    print("\n" + "="*80)
    print("KEY QUESTION 2: Pre vs Post Payroll Change (Apr 21 vs May 8, 2026)")
    print("="*80)

    apr26_data = next((r for r in all_results if r["label"] == "04.21.26_AprBOD"), None)
    may26_data_full = next((r for r in all_results if r["label"] == "05.08.26"), None)

    # Look at the actual payroll runs close to each file date
    # Apr 21 file: most recent actuals week ending around Apr 18-21
    # May 8 file: most recent actuals week ending around May 5-8

    for data, lbl in [(apr26_data, "Apr 21, 2026 (PRE)"), (may26_data_full, "May 8, 2026 (POST)")]:
        if data:
            wd = data.get("weekly_data", {})
            wt = wd.get("weekly_total_actuals") or wd.get("weekly_total_forecast")
            dr = f"{wd.get('week_start_date')} to {wd.get('week_end_date')}"
            wtype = wd.get("data_type","")
            print(f"\n{lbl}:")
            print(f"  Most recent week ({dr}): ${abs(wt)/1e6:.3f}M/wk [{wtype}]" if wt else "  N/A")
            print(f"  Implied monthly: ${abs(wt)*52/12/1e6:.2f}M" if wt else "")
            if wd.get("bu_breakdown_actuals"):
                bu = wd["bu_breakdown_actuals"]
                sorted_bu = sorted(bu.items(), key=lambda x: abs(x[1]), reverse=True)
                print(f"  BU breakdown:")
                for bu_name, bu_amt in sorted_bu[:10]:
                    print(f"    {bu_name:20s}: ${abs(bu_amt)/1e6:.3f}M")

    # Compare the monthly actuals close to each date
    print("\n  --- Monthly actuals comparison (last full month before each file date) ---")
    for data, lbl in [(apr26_data, "Apr 21, 2026 (PRE)"), (may26_data_full, "May 8, 2026 (POST)")]:
        if data:
            months_with_act = [m for m in data["monthly_data"]
                               if m.get("actuals_total") and abs(m["actuals_total"]) > 1e6]
            if months_with_act:
                last = months_with_act[-1]
                print(f"  {lbl}: {last['month_label']} actual = ${abs(last['actuals_total'])/1e6:.2f}M")
                bu = last.get("actuals_bu") or {}
                if bu:
                    for bu_name, bu_amt in sorted(bu.items(), key=lambda x: abs(x[1]), reverse=True):
                        print(f"    {bu_name:20s}: ${abs(bu_amt)/1e6:.3f}M")

    # =============================================
    # COMPLETE MONTHLY PAYROLL TABLE
    # =============================================
    print("\n" + "="*80)
    print("COMPLETE MONTHLY PAYROLL ACTUALS TABLE")
    print("="*80)
    print(f"{'Source File':<30} {'Month':<12} {'Monthly Actual':>16} {'Monthly Fcst':>14} {'Source'}")
    print("-"*85)

    for r in all_results:
        label = r["label"]
        for m in r["monthly_data"]:
            act = m.get("actuals_total")
            fcst = m.get("forecast_total")
            month = m.get("month_label", "?")
            src = "ACTUAL" if act is not None else "fcst"
            preferred = act if act is not None else fcst
            act_str = f"${abs(act)/1e6:.2f}M" if act is not None else "—"
            fcst_str = f"${abs(fcst)/1e6:.2f}M" if fcst is not None else "—"
            print(f"{label:<30} {month:<12} {act_str:>16} {fcst_str:>14} {src}")

    # =============================================
    # SAVE JSON
    # =============================================
    # Build the consolidated timeline for the JSON
    consolidated_timeline = []
    for r in all_results:
        entry = {
            "label": r["label"],
            "file_date": r["file_date"],
            "format_type": r.get("format"),
            "monthly_data": r.get("monthly_data", []),
            "weekly_data": r.get("weekly_data", {}),
            "errors": r.get("errors", []),
        }
        consolidated_timeline.append(entry)

    # Key findings
    key_findings = {
        "sep_2022_best_monthly_payroll": {
            "amount": sep22_monthly,
            "month": sep22_month_lbl,
            "source": "actuals",
            "note": "Best available full-month actual from Sep 19, 2022 BOD file"
        },
        "may_2026_best_monthly_payroll": {
            "amount": may26_monthly,
            "month": may26_month_lbl,
            "source": "actuals",
            "note": "Most recent full-month actual from May 8, 2026 file"
        },
    }

    if sep22_monthly and may26_monthly:
        reduction = abs(sep22_monthly) - abs(may26_monthly)
        key_findings["absolute_monthly_reduction"] = round(reduction, 2)
        key_findings["pct_reduction"] = round(reduction / abs(sep22_monthly) * 100, 1)
        key_findings["annualized_reduction"] = round(reduction * 12, 2)
        key_findings["implied_headcount_at_85k_avg"] = {
            "sep_2022": round(abs(sep22_monthly) / (85000/12)),
            "may_2026": round(abs(may26_monthly) / (85000/12)),
            "reduction": round((abs(sep22_monthly) - abs(may26_monthly)) / (85000/12)),
        }

    output = {
        "analysis_date": "2026-05-11",
        "description": "Payroll cost trajectory 2022-2026 extracted from daily cash forecast files. All amounts in dollars (negative = outflow).",
        "methodology": {
            "format_A": "Sep/Dec 2022 files: monthly totals in None-date rows before next month label",
            "format_B": "Apr 2023+ files: monthly totals in abbreviated month-name rows (e.g., 'Feb', 'Oct')",
            "actuals_vs_forecast": "Actuals are populated for completed dates; zeros/None indicate forecast-only",
            "weekly_data": "Most recent week of actuals before file date; days summed individually",
        },
        "payroll_timeline": consolidated_timeline,
        "key_findings": key_findings,
        "summary_table": [],
    }

    # Build a clean summary table
    for r in all_results:
        monthly = r.get("monthly_data", [])
        weeks = r.get("weekly_data", {})
        months_with_act = [m for m in monthly if m.get("actuals_total") and abs(m["actuals_total"]) > 1e6]
        last_act = months_with_act[-1] if months_with_act else None
        wt = weeks.get("weekly_total_actuals") or weeks.get("weekly_total_forecast")

        output["summary_table"].append({
            "label": r["label"],
            "file_date": r["file_date"],
            "last_full_month_with_actuals": last_act["month_label"] if last_act else None,
            "last_full_month_actuals_amount": last_act["actuals_total"] if last_act else None,
            "last_full_month_actuals_amount_millions": round(abs(last_act["actuals_total"])/1e6, 2) if last_act else None,
            "last_full_month_bu_breakdown": (last_act.get("actuals_bu") or last_act.get("forecast_bu")) if last_act else None,
            "most_recent_week_end": weeks.get("week_end_date"),
            "most_recent_week_total": wt,
            "most_recent_week_millions": round(abs(wt)/1e6, 3) if wt else None,
            "weekly_implied_monthly_millions": round(abs(wt)*52/12/1e6, 2) if wt else None,
            "most_recent_week_bu": weeks.get("bu_breakdown_actuals") or weeks.get("bu_breakdown_forecast"),
            "weekly_data_type": weeks.get("data_type"),
        })

    out_path = os.path.join(BASE_DIR, "analysis_payroll_headcount.json")
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\n\nOutput saved to: {out_path}")
    return output


if __name__ == "__main__":
    main()
