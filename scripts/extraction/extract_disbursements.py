"""
Extract monthly disbursement totals from 17 cash forecast files.
Uses BU beta sheet, finds rows by label, columns by year-month header.
"""

import openpyxl
import json
import os
import re
from datetime import datetime

SP_DIR = "/Users/josh/Downloads/SP_Analysis"

# The 17 target files mapped to short IDs
FILES = [
    ("1.23.24_BOD",       "Daily Cash Fcst - 1.23.24_BU view_BOD.xlsx"),
    ("2.20.24_BOD",       "Daily Cash Fcst - 2.20.24_BU view_BOD.xlsx"),
    ("3.19.24_MarBOD",    "Daily Cash Fcst - 3.19.24_BU view_MarBOD.xlsx"),
    ("4.30.24_online_adj","Daily Cash Fcst - 4.30.24_BU view_final_online adj.xlsx"),
    ("5.7.24_MayBOD",     "Daily Cash Fcst - 5.7.24_BU view_MayBOD_5.9.24.xlsx"),
    ("6.18.24_JunBOD",    "Daily Cash Fcst - 6.18.24_BU view_JunBOD.xlsx"),
    ("7.23.24_JulBOD",    "Daily Cash Fcst - 7.23.24_deck_JulBOD_FY25.xlsx"),
    ("8.19.24_AugBOD",    "Daily Cash Fcst - 8.19.24_BU view_AugBOD.xlsx"),
    ("9.24.24_BOD",       "Daily Cash Fcst - 9.24.24_BU view_deck version_BOD.xlsx"),
    ("10.15.24_OctBOD",   "Daily Cash Fcst - 10.15.24_BU view_deck version_OctBOD.xlsx"),
    ("11.12.24_NovBOD",   "Daily Cash Fcst - 11.12.24_BU view_deck version_NovBOD.xlsx"),
    ("12.10.24_BOD",      "Daily Cash Fcst - 12.10.24_BU view_deck version_BOD_SC wo 3 RE Interco IF_$9M.xlsx"),
    ("12.31.24",          "Daily Cash Fcst - 12.31.24_BU view.xlsx"),
    ("1.21.25_JanBOD",    "Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx"),
    ("3.18.25_MarBOD",    "Daily Cash Fcst - 3.18.25_BU view_deck version_MarBOD.xlsx"),
    ("03.03.26_post-reductions_MarBOD", "Daily Cash Fcst - 03.03.26_post-reductions_deck version_MarBOD.xlsx"),
    ("05.08.26_BU_view",  "Daily Cash Fcst - 05.08.26_BU view.xlsx"),
]

# Row label searches: (category_key, [search_terms_in_priority_order])
# Anchored with ^ $ for exact match; plain for substring
ROW_SEARCHES = [
    ("payroll_total",      ["1. Payroll", "Payroll/Bens"]),
    ("payroll_direct",     ["1a. Direct Payroll"]),
    ("payroll_backoffice", ["1b. Back-office Payroll"]),
    ("merch_total",        ["3. Merch Disbursements"]),
    ("merch_subtotal",     ["^Merch$"]),
    ("logistics_total",    ["6. Logistics Vendors"]),
    ("rent_facilities",    ["Rent & facililities", "Rent & Facilities Disb"]),
    ("rent_only",          ["^Rent$"]),
    ("property_tax",       ["Property Tax"]),
    ("utilities",          ["Utilities & Telephone"]),
    ("capex",              ["^CapEx$"]),
    ("riskmgt",            ["RiskMgt/Ins"]),
    ("advertising",        ["^Advertising$"]),
    ("operating_disb",     ["Operating Disbursements w/", "Operating Disbursements "]),
    ("total_disb",         ["Disbursements Total"]),
    ("interest",           ["^Interest$"]),
    ("debt_repayment",     ["Debt\nRepayment", "Debt Repayment"]),
    ("total_inflows",      ["Total Inflows"]),
    ("net_op_cf",          ["Net Operating Cash Flow w/", "Net Operating Cash Flow"]),
]

CATS_OF_INTEREST = [
    "payroll_total", "merch_total", "logistics_total", "rent_facilities", "rent_only",
    "property_tax", "riskmgt", "advertising", "interest", "debt_repayment",
    "operating_disb", "total_disb", "total_inflows"
]


def normalize(s):
    if s is None:
        return ""
    return str(s).strip()


def find_label_col(ws, max_col=60):
    """Find column containing '$ in millions' label."""
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=max_col, values_only=True):
        for col_idx, val in enumerate(row):
            if normalize(val) == "$ in millions":
                return col_idx + 1
    return None


def build_col_month_map(ws, max_col=200):
    """
    Build {YYYY-MM: col_1based} from any row containing 'Mon YYYY' patterns.
    """
    month_map = {}
    month_abbr = {
        "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
        "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
        "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12"
    }
    for row in ws.iter_rows(min_row=1, max_row=8, max_col=max_col, values_only=True):
        for col_idx, val in enumerate(row):
            v = normalize(val)
            m = re.match(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})$', v)
            if m:
                key = f"{m.group(2)}-{month_abbr[m.group(1)]}"
                if key not in month_map:
                    month_map[key] = col_idx + 1
    return month_map


def find_rows_by_label(ws, label_col, searches, max_row=320):
    """Find row numbers by label text."""
    labels = {}
    for row_idx, row in enumerate(ws.iter_rows(
            min_row=1, max_row=max_row, min_col=label_col, max_col=label_col, values_only=True)):
        val = normalize(row[0])
        if val:
            labels[row_idx + 1] = val

    results = {}
    for cat_key, search_terms in searches:
        found_row = None
        for term in search_terms:
            is_anchored = term.startswith("^") or term.endswith("$")
            if is_anchored:
                clean = term.strip("^$")
                for rn, lbl in sorted(labels.items()):
                    if lbl.strip() == clean:
                        found_row = rn
                        break
            else:
                for rn, lbl in sorted(labels.items()):
                    if term.lower() in lbl.lower():
                        found_row = rn
                        break
            if found_row:
                break
        results[cat_key] = found_row
    return results


def get_cell_value(ws, row, col):
    """Read a single cell from a read-only worksheet."""
    for r in ws.iter_rows(min_row=row, max_row=row, min_col=col, max_col=col, values_only=True):
        val = r[0]
        if isinstance(val, (int, float)):
            return round(float(val), 4)
        return None
    return None


def extract_file(file_id, filename):
    path = os.path.join(SP_DIR, filename)
    result = {"file_id": file_id, "filename": filename, "months": {}, "errors": []}

    if not os.path.exists(path):
        result["errors"].append(f"File not found: {filename}")
        return result

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        result["errors"].append(f"Cannot open: {e}")
        return result

    if "BU beta" not in wb.sheetnames:
        result["errors"].append("No 'BU beta' sheet")
        wb.close()
        return result

    ws = wb["BU beta"]

    label_col = find_label_col(ws)
    if label_col is None:
        result["errors"].append("Could not find '$ in millions' label column")
        wb.close()
        return result

    result["label_col"] = label_col
    month_col_map = build_col_month_map(ws)
    result["months_available"] = sorted(month_col_map.keys())
    row_map = find_rows_by_label(ws, label_col, ROW_SEARCHES)
    result["row_map"] = {k: v for k, v in row_map.items() if v is not None}

    # Extract Jan 2024 - May 2026
    TARGET_MONTHS = []
    for year in [2024, 2025, 2026]:
        for month in range(1, 13):
            if year == 2026 and month > 5:
                break
            TARGET_MONTHS.append(f"{year}-{month:02d}")

    for month_key in TARGET_MONTHS:
        col = month_col_map.get(month_key)
        if col is None:
            continue
        month_data = {}
        for cat_key, _ in ROW_SEARCHES:
            row_num = row_map.get(cat_key)
            if row_num is not None:
                month_data[cat_key] = get_cell_value(ws, row_num, col)
            else:
                month_data[cat_key] = None
        result["months"][month_key] = month_data

    wb.close()
    return result


def main():
    print("Starting disbursement extraction...")
    all_results = {}

    for file_id, filename in FILES:
        print(f"  {file_id} ...", end=" ", flush=True)
        res = extract_file(file_id, filename)
        all_results[file_id] = res
        months_found = len(res.get("months", {}))
        errors = res.get("errors", [])
        print(f"{months_found} months | errors: {errors if errors else 'none'}")

    # Authoritative source per month
    authoritative_sources = {
        "2024-01": "1.23.24_BOD",
        "2024-02": "2.20.24_BOD",
        "2024-03": "3.19.24_MarBOD",
        "2024-04": "4.30.24_online_adj",
        "2024-05": "5.7.24_MayBOD",
        "2024-06": "6.18.24_JunBOD",
        "2024-07": "7.23.24_JulBOD",
        "2024-08": "8.19.24_AugBOD",
        "2024-09": "9.24.24_BOD",
        "2024-10": "10.15.24_OctBOD",
        "2024-11": "11.12.24_NovBOD",
        "2024-12": "12.31.24",
        "2025-01": "1.21.25_JanBOD",
        "2025-02": "3.18.25_MarBOD",
        "2025-03": "3.18.25_MarBOD",
        "2025-04": "03.03.26_post-reductions_MarBOD",
        "2025-05": "03.03.26_post-reductions_MarBOD",
        "2025-06": "03.03.26_post-reductions_MarBOD",
        "2025-07": "03.03.26_post-reductions_MarBOD",
        "2025-08": "03.03.26_post-reductions_MarBOD",
        "2025-09": "03.03.26_post-reductions_MarBOD",
        "2025-10": "03.03.26_post-reductions_MarBOD",
        "2025-11": "03.03.26_post-reductions_MarBOD",
        "2025-12": "03.03.26_post-reductions_MarBOD",
        "2026-01": "03.03.26_post-reductions_MarBOD",
        "2026-02": "03.03.26_post-reductions_MarBOD",
        "2026-03": "03.03.26_post-reductions_MarBOD",
        "2026-04": "05.08.26_BU_view",
        "2026-05": "05.08.26_BU_view",
    }

    # Build authoritative monthly table
    auth_table = {}
    for month_key, source_id in authoritative_sources.items():
        res = all_results.get(source_id, {})
        mdata = res.get("months", {}).get(month_key)
        if mdata:
            auth_table[month_key] = {
                "source": source_id,
                "values": {cat: mdata.get(cat) for cat in CATS_OF_INTEREST}
            }
        else:
            auth_table[month_key] = {
                "source": source_id,
                "values": {},
                "note": "month not found in this file"
            }

    def get_auth(month_key, cat):
        return auth_table.get(month_key, {}).get("values", {}).get(cat)

    def avg_months(months_list, cat):
        vals = [get_auth(m, cat) for m in months_list if get_auth(m, cat) is not None]
        if not vals:
            return None
        return round(sum(vals) / len(vals), 4)

    # Dec 2025 vs Mar 2026 comparison
    dec25_vs_mar26 = {}
    for cat in CATS_OF_INTEREST:
        d25 = get_auth("2025-12", cat)
        m26 = get_auth("2026-03", cat)
        delta = None
        if d25 is not None and m26 is not None:
            delta = round(m26 - d25, 4)
        dec25_vs_mar26[cat] = {
            "dec_2025_$M": d25,
            "mar_2026_$M": m26,
            "change_$M": delta,
        }

    dec25_payroll = get_auth("2025-12", "payroll_total")
    mar26_payroll = get_auth("2026-03", "payroll_total")
    dec25_total = get_auth("2025-12", "total_disb")
    mar26_total = get_auth("2026-03", "total_disb")

    payroll_annualized = None
    total_annualized = None
    if dec25_payroll is not None and mar26_payroll is not None:
        payroll_delta = mar26_payroll - dec25_payroll
        payroll_annualized = {
            "monthly_payroll_reduction_$M": round(payroll_delta, 4),
            "annualized_run_rate_savings_$M": round(payroll_delta * 12, 4),
            "note": "Negative = cost reduction (disbursements went down)"
        }
    if dec25_total is not None and mar26_total is not None:
        total_delta = mar26_total - dec25_total
        total_annualized = {
            "monthly_total_disb_change_$M": round(total_delta, 4),
            "annualized_run_rate_change_$M": round(total_delta * 12, 4),
        }

    dec_mar_comparison = {
        "dec_2025_source": auth_table.get("2025-12", {}).get("source"),
        "mar_2026_source": auth_table.get("2026-03", {}).get("source"),
        "categories": dec25_vs_mar26,
        "payroll_annualized": payroll_annualized,
        "total_disb_annualized": total_annualized,
    }

    # Oct-Dec 2024 pre-reduction benchmark vs Feb-Mar 2026
    pre_months = ["2024-10", "2024-11", "2024-12"]
    post_months = ["2026-02", "2026-03"]
    pre_post_comparison = {}
    for cat in CATS_OF_INTEREST:
        pre = avg_months(pre_months, cat)
        post = avg_months(post_months, cat)
        delta = round(post - pre, 4) if (pre is not None and post is not None) else None
        pre_post_comparison[cat] = {
            "pre_avg_oct_dec_2024_$M": pre,
            "post_avg_feb_mar_2026_$M": post,
            "change_$M": delta,
            "annualized_change_$M": round(delta * 12, 4) if delta is not None else None,
        }

    # Rent trajectory by quarter
    quarters = [
        ("Q1-2024", ["2024-01", "2024-02", "2024-03"]),
        ("Q2-2024", ["2024-04", "2024-05", "2024-06"]),
        ("Q3-2024", ["2024-07", "2024-08", "2024-09"]),
        ("Q4-2024", ["2024-10", "2024-11", "2024-12"]),
        ("Q1-2025", ["2025-01", "2025-02", "2025-03"]),
        ("Q2-2025", ["2025-04", "2025-05", "2025-06"]),
        ("Q3-2025", ["2025-07", "2025-08", "2025-09"]),
        ("Q4-2025", ["2025-10", "2025-11", "2025-12"]),
        ("Q1-2026", ["2026-01", "2026-02", "2026-03"]),
    ]

    rent_trajectory = {}
    for q_label, months in quarters:
        rent_vals = [get_auth(m, "rent_facilities") for m in months if get_auth(m, "rent_facilities") is not None]
        rent_trajectory[q_label] = {
            "months": months,
            "avg_monthly_rent_facilities_$M": round(sum(rent_vals) / len(rent_vals), 4) if rent_vals else None,
            "data_points": len(rent_vals),
        }

    interest_trajectory = {}
    for q_label, months in quarters:
        int_vals = [get_auth(m, "interest") for m in months if get_auth(m, "interest") is not None]
        interest_trajectory[q_label] = {
            "avg_monthly_interest_$M": round(sum(int_vals) / len(int_vals), 4) if int_vals else None,
            "data_points": len(int_vals),
        }

    output = {
        "extraction_date": datetime.now().isoformat(),
        "note": "All values in $M. Negative values = disbursements/outflows. Positive = inflows.",
        "authoritative_monthly_table": auth_table,
        "analysis": {
            "dec_2025_vs_mar_2026": dec_mar_comparison,
            "oct_dec_2024_vs_feb_mar_2026_pre_post": pre_post_comparison,
            "rent_trajectory_by_quarter": rent_trajectory,
            "interest_trajectory_by_quarter": interest_trajectory,
        },
        "raw_by_file": {
            fid: {
                "errors": res.get("errors", []),
                "months_extracted": list(res.get("months", {}).keys()),
                "row_map": res.get("row_map", {}),
                "label_col": res.get("label_col"),
                "months_available_in_file": res.get("months_available", []),
            }
            for fid, res in all_results.items()
        },
    }

    out_path = os.path.join(SP_DIR, "analysis_disbursements_2024_2026.json")
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)
    print(f"\nSaved to: {out_path}")

    # ---- PRINT SUMMARY ----
    print("\n" + "=" * 110)
    print("MONTHLY DISBURSEMENTS TABLE (authoritative source per month, $M)")
    print("=" * 110)
    header = f"{'Month':<10} {'Source':<38} {'Payroll':>8} {'Merch':>8} {'Logistics':>10} {'Rent+Fac':>9} {'RiskMgt':>8} {'Interest':>9} {'TotalDisb':>10}"
    print(header)
    print("-" * 110)

    for month_key in sorted(auth_table.keys()):
        entry = auth_table[month_key]
        vals = entry.get("values", {})
        src = entry.get("source", "")[:36]
        py = vals.get("payroll_total")
        me = vals.get("merch_total")
        lo = vals.get("logistics_total")
        re = vals.get("rent_facilities")
        rm = vals.get("riskmgt")
        it = vals.get("interest")
        td = vals.get("total_disb")

        def fmt(v):
            return f"{v:>8.2f}" if v is not None else f"{'N/A':>8}"

        print(f"{month_key:<10} {src:<38} {fmt(py)} {fmt(me)} {fmt(lo):>10} {fmt(re):>9} {fmt(rm):>8} {fmt(it):>9} {fmt(td):>10}")

    print("\n" + "=" * 80)
    print("DEC 2025 vs MAR 2026 COMPARISON (cost reduction evidence)")
    print("=" * 80)
    comp = dec_mar_comparison
    print(f"  Dec 2025 source: {comp.get('dec_2025_source')}")
    print(f"  Mar 2026 source: {comp.get('mar_2026_source')}")
    print(f"\n  {'Category':<25} {'Dec-2025':>10} {'Mar-2026':>10} {'Delta $M':>10} {'Ann. Delta':>12}")
    print("  " + "-" * 70)
    for cat, vals in comp.get("categories", {}).items():
        d = vals.get("dec_2025_$M")
        m = vals.get("mar_2026_$M")
        c = vals.get("change_$M")
        ann = round(c * 12, 2) if c is not None else None
        d_str = f"{d:>10.2f}" if d is not None else f"{'N/A':>10}"
        m_str = f"{m:>10.2f}" if m is not None else f"{'N/A':>10}"
        c_str = f"{c:>10.2f}" if c is not None else f"{'N/A':>10}"
        a_str = f"{ann:>12.2f}" if ann is not None else f"{'N/A':>12}"
        print(f"  {cat:<25} {d_str} {m_str} {c_str} {a_str}")

    if payroll_annualized:
        print(f"\n  Payroll monthly reduction: ${payroll_annualized['monthly_payroll_reduction_$M']:.2f}M")
        print(f"  Payroll annualized run-rate: ${payroll_annualized['annualized_run_rate_savings_$M']:.2f}M/yr")
    if total_annualized:
        print(f"  Total disb monthly change:  ${total_annualized['monthly_total_disb_change_$M']:.2f}M")
        print(f"  Total disb annualized:      ${total_annualized['annualized_run_rate_change_$M']:.2f}M/yr")

    print("\n" + "=" * 60)
    print("RENT & FACILITIES TRAJECTORY (avg monthly $M)")
    print("=" * 60)
    for q, data in rent_trajectory.items():
        r = data.get("avg_monthly_rent_facilities_$M")
        pts = data.get("data_points", 0)
        r_str = f"${r:.2f}M" if r is not None else "N/A"
        print(f"  {q}: {r_str}/month  ({pts} data pts)")

    print("\n" + "=" * 60)
    print("INTEREST TRAJECTORY (avg monthly $M)")
    print("=" * 60)
    for q, data in interest_trajectory.items():
        it = data.get("avg_monthly_interest_$M")
        pts = data.get("data_points", 0)
        it_str = f"${it:.2f}M" if it is not None else "N/A"
        print(f"  {q}: {it_str}/month  ({pts} data pts)")

    return output


if __name__ == "__main__":
    main()
