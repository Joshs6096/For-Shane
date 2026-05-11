"""
DEFINITIVE Inflows Extractor for Transform SR Holding Management LLC
Sep 2022 – May 2026

Data Sources:
- PRIMARY: "Inflows Actuals" sheet in each file — carries complete confirmed monthly totals
  going back to Feb 2022. Uses the most-recent file's version for final actuals.
  "Daily Total" column = total operating inflows (Sears + Kmart + HS + KCD + HTS + SC +
  CITI + Asset Sales + CCHS + Misc + etc.)
  "Equity" and "Debt/Financing" columns = financing inflows (tracked separately).

- May 2026 SPECIAL CASE: Inflows Actuals only has partial actuals (through May 8).
  Use Inflows Detail from the most-recent file for the projected full-month total.

Deduplication rule: For each calendar month, use the value from the MOST RECENT file
(i.e., latest BOD date). Rationale: more recent files contain revised/confirmed actuals.
"""

import openpyxl
import os
import json
import re
from datetime import datetime
from collections import defaultdict

BASE_DIR = "/Users/josh/Downloads/SP_Analysis"

MONTH_ABBREVS = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}


def parse_month_year_label(label):
    """Parse 'Month YYYY' -> (year, month) or None."""
    if not label or not isinstance(label, str):
        return None
    s = label.strip()
    m = re.match(r'^([A-Za-z]+)\s+(\d{4})$', s)
    if m:
        mon_str = m.group(1).lower()[:3]
        yr = int(m.group(2))
        mon = MONTH_ABBREVS.get(mon_str)
        if mon:
            return (yr, mon)
    return None


def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(',', ''))
    except:
        return 0.0


def get_file_date(fname):
    """Extract file date from filename pattern like '9.19.22' or '05.08.26'."""
    m = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})', fname)
    if m:
        mo, da, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yr < 100:
            yr += 2000
        try:
            return datetime(yr, mo, da)
        except:
            pass
    return datetime(2099, 1, 1)


def build_col_map(headers):
    """Build column map from row of header values (0-indexed)."""
    col_map = {}
    for i, h in enumerate(headers):
        if not h or not isinstance(h, str):
            continue
        hl = ' '.join(h.lower().strip().replace('\n', ' ').split())

        if 'sears stores' in hl or hl == 'sears b&m':
            col_map.setdefault('sears', i)
        elif 'kmart' in hl:
            col_map.setdefault('kmart', i)
        elif 'home services' in hl:
            col_map.setdefault('home_svcs', i)
        elif hl in ('kcd', 'kcd wholesale', 'kcd (royalty)'):
            col_map.setdefault('kcd', i)
        elif 'supply chain' in hl:
            col_map['supply_chain'] = i
        elif 'citi reimbursement' in hl or 'citi reimburse' in hl:
            col_map['citi_reimb'] = i
        elif 'cross country' in hl or hl == 'cchs':
            col_map['cchs'] = i
        elif 'asset sales' in hl:
            col_map['asset_sales'] = i
        elif ('debt' in hl and 'financ' in hl) or hl == 'new debt':
            col_map.setdefault('debt_financ', i)
        elif 'daily total' in hl:
            col_map['daily_total'] = i
        elif 'hts prefund' in hl:
            col_map['hts_prefund'] = i
        elif hl == 'hts':
            col_map.setdefault('hts', i)
        elif 'costco' in hl:
            col_map['costco'] = i
        elif 'misc inflows' in hl:
            col_map['misc'] = i
        elif 'tenant income' in hl:
            col_map['tenant'] = i
        elif 'propco' in hl or 'lease opco' in hl:
            col_map.setdefault('propco', i)
        elif hl == 'equity':
            col_map.setdefault('equity', i)
        elif hl == 'kes' or 'kes prefund' in hl:
            col_map.setdefault('kes', i)
        elif 'online' in hl:
            col_map.setdefault('online', i)
        elif hl == 'rx':
            col_map['rx'] = i
        elif 'sears mexico' in hl:
            col_map['sears_mexico'] = i
        elif 'monark' in hl:
            col_map.setdefault('monark', i)
        elif 'service live' in hl:
            col_map.setdefault('service_live', i)

    return col_map


def extract_inflows_actuals(ws, source_file, file_date):
    """
    Extract monthly totals from the Inflows Actuals sheet.
    Returns list of (file_date, record_dict).
    """
    results = []
    all_rows = list(ws.iter_rows(min_row=1, max_row=2000, values_only=True))
    if len(all_rows) < 2:
        return results

    # Use row 2 as canonical header (row 1 is sometimes abbreviated)
    h2 = all_rows[1]
    col_map = build_col_map(h2)

    # Some files have slightly different header structure - try row 1 as fallback
    if 'daily_total' not in col_map:
        h1 = all_rows[0]
        col_map2 = build_col_map(h1)
        col_map.update({k: v for k, v in col_map2.items() if k not in col_map})

    for row in all_rows:
        v = row[0]
        if not v or not isinstance(v, str):
            continue
        parsed = parse_month_year_label(v.strip())
        if not parsed:
            continue
        yr, mon = parsed

        def g(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return 0.0
            return safe_float(row[idx])

        daily_total = g('daily_total')
        sears = g('sears')
        hs = g('home_svcs')

        # Skip empty rows (forecast placeholders)
        if daily_total == 0.0 and sears == 0.0 and hs == 0.0:
            continue

        rec = {
            'year': yr,
            'month': mon,
            'date_label': f"{yr}-{mon:02d}",
            'source_file': source_file,
            'file_date': file_date.isoformat()[:10],
            'source_sheet': 'Inflows Actuals',
            'is_forecast': False,
            # Operating inflows (Daily Total)
            'total_inflows': round(daily_total, 4),
            # Financing inflows (separate)
            'equity_inflows': round(g('equity'), 4),
            'debt_financing': round(g('debt_financ'), 4),
            # Total including financing
            'total_with_financing': round(
                daily_total + g('equity') + g('debt_financ'), 4
            ),
            # Category breakdown
            'sears_stores': round(sears, 4),
            'kmart_stores': round(g('kmart'), 4),
            'home_services': round(hs, 4),
            'kcd_wholesale': round(g('kcd'), 4),
            'supply_chain': round(g('supply_chain'), 4),
            'citi_reimb': round(g('citi_reimb'), 4),
            'cchs': round(g('cchs'), 4),
            'asset_sales': round(g('asset_sales'), 4),
            'hts': round(g('hts') + g('hts_prefund'), 4),
            'costco': round(g('costco'), 4),
            'tenant_income': round(g('tenant'), 4),
            'propco': round(g('propco'), 4),
            'online': round(g('online'), 4),
            'rx': round(g('rx'), 4),
            'misc_inflows': round(g('misc'), 4),
            'kes': round(g('kes'), 4),
            'sears_mexico': round(g('sears_mexico'), 4),
        }
        results.append((file_date, rec))

    return results


def extract_may_2026_from_detail(ws, source_file, file_date):
    """
    Extract May 2026 projected full-month total from Inflows Detail sheet.
    Used as fallback when Inflows Actuals only has partial data.
    """
    all_rows = list(ws.iter_rows(min_row=1, max_row=100, values_only=True))
    if not all_rows:
        return None

    # Find header row
    header_row = None
    for row in all_rows[:10]:
        vals = [str(v).strip() if v else '' for v in row]
        if 'Sears Stores' in vals or 'Daily Total' in vals:
            header_row = row
            break

    if header_row is None:
        # Try row 5 (some files have header repeated there)
        if len(all_rows) >= 5:
            header_row = all_rows[4]

    if header_row is None:
        return None

    col_map = build_col_map(header_row)

    for row in all_rows:
        v = row[0]
        if not v or not isinstance(v, str):
            continue
        parsed = parse_month_year_label(v.strip())
        if not parsed:
            continue
        yr, mon = parsed
        if yr != 2026 or mon != 5:
            continue

        def g(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return 0.0
            return safe_float(row[idx])

        daily_total = g('daily_total')
        if daily_total == 0.0:
            # Compute from parts
            daily_total = sum([
                g('sears'), g('kmart'), g('home_svcs'), g('kcd'),
                g('supply_chain'), g('citi_reimb'), g('cchs'),
                g('asset_sales'), g('debt_financ'), g('misc'),
                g('hts'), g('hts_prefund'), g('costco'), g('tenant'),
                g('propco'), g('online'), g('rx'),
            ])

        if daily_total < 1.0:
            continue

        rec = {
            'year': 2026,
            'month': 5,
            'date_label': '2026-05',
            'source_file': source_file,
            'file_date': file_date.isoformat()[:10],
            'source_sheet': 'Inflows Detail (May 2026 forecast)',
            'is_forecast': True,
            'total_inflows': round(daily_total, 4),
            'equity_inflows': 0.0,
            'debt_financing': round(g('debt_financ'), 4),
            'total_with_financing': round(daily_total + g('debt_financ'), 4),
            'sears_stores': round(g('sears'), 4),
            'kmart_stores': round(g('kmart'), 4),
            'home_services': round(g('home_svcs'), 4),
            'kcd_wholesale': round(g('kcd'), 4),
            'supply_chain': round(g('supply_chain'), 4),
            'citi_reimb': round(g('citi_reimb'), 4),
            'cchs': round(g('cchs'), 4),
            'asset_sales': round(g('asset_sales'), 4),
            'hts': round(g('hts') + g('hts_prefund'), 4),
            'costco': round(g('costco'), 4),
            'tenant_income': round(g('tenant'), 4),
            'propco': round(g('propco'), 4),
            'online': round(g('online'), 4),
            'rx': round(g('rx'), 4),
            'misc_inflows': round(g('misc'), 4),
            'kes': round(g('kes'), 4),
            'sears_mexico': round(g('sears_mexico'), 4),
        }
        return rec

    return None


def main():
    files = sorted([
        f for f in os.listdir(BASE_DIR)
        if f.endswith('.xlsx') and f.startswith('Daily Cash')
    ])

    print(f"Processing {len(files)} files...\n")

    # Collect all records from Inflows Actuals across all files
    month_data = defaultdict(list)  # (yr, mon) -> [(file_date, record)]
    may_2026_candidates = []  # (file_date, record) from Inflows Detail

    for fname in files:
        fpath = os.path.join(BASE_DIR, fname)
        fdate = get_file_date(fname)
        print(f"  {fname[-50:]}")

        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        except Exception as e:
            print(f"    ERROR opening: {e}")
            continue

        # Extract from Inflows Actuals
        if 'Inflows Actuals' in wb.sheetnames:
            ws = wb['Inflows Actuals']
            try:
                recs = extract_inflows_actuals(ws, fname, fdate)
                for fd, rec in recs:
                    key = (rec['year'], rec['month'])
                    month_data[key].append((fd, rec))
                print(f"    Inflows Actuals: {len(recs)} month rows")
            except Exception as e:
                print(f"    ERROR in Inflows Actuals: {e}")

        # Extract May 2026 from Inflows Detail (for files dated after May 1, 2026)
        if fdate >= datetime(2026, 5, 1) and 'Inflows Detail' in wb.sheetnames:
            ws2 = wb['Inflows Detail']
            try:
                rec = extract_may_2026_from_detail(ws2, fname, fdate)
                if rec:
                    may_2026_candidates.append((fdate, rec))
                    print(f"    Inflows Detail: May 2026 forecast found (${rec['total_inflows']:.1f}M)")
            except Exception as e:
                print(f"    ERROR in Inflows Detail: {e}")

        wb.close()

    print(f"\nTotal unique months found: {len(month_data)}")

    # For each month, pick the MOST RECENT file's data
    final_data = {}
    for key, records in month_data.items():
        sorted_recs = sorted(records, key=lambda x: x[0], reverse=True)
        final_data[key] = sorted_recs[0][1]

    # For May 2026: the Inflows Actuals partial value ($8.56M) is not the full month.
    # Override with the most recent Inflows Detail forecast.
    if may_2026_candidates:
        may_2026_candidates.sort(key=lambda x: x[0], reverse=True)
        best_may = may_2026_candidates[0][1]
        final_data[(2026, 5)] = best_may
        print(f"\nMay 2026: Using Inflows Detail forecast = ${best_may['total_inflows']:.1f}M "
              f"(from {best_may['source_file'][-35:]})")

    # Filter to Sep 2022 - May 2026 window
    in_window = {
        k: v for k, v in final_data.items()
        if (k[0] > 2022 or (k[0] == 2022 and k[1] >= 9))
        and (k[0] < 2026 or (k[0] == 2026 and k[1] <= 5))
    }

    # Build chronological list
    monthly_data = [in_window[k] for k in sorted(in_window.keys())]
    print(f"Months in Sep 2022 – May 2026 window: {len(monthly_data)}")

    # ─── INSIGHT ANALYSIS ───

    total_series = [r['total_inflows'] for r in monthly_data]
    total_w_fin_series = [r['total_with_financing'] for r in monthly_data]

    sep_2022 = in_window.get((2022, 9))
    may_2026 = in_window.get((2026, 5))

    # Peak months (by total operating inflows)
    peak_months = sorted(monthly_data, key=lambda r: r['total_inflows'], reverse=True)[:10]

    # Asset sales - first month over $5M
    asset_milestones = [r for r in monthly_data if r.get('asset_sales', 0) >= 5.0]
    first_big_asset_sales = asset_milestones[0] if asset_milestones else None

    # Debt/financing - when did it become meaningful (>$10M)?
    debt_milestones = [r for r in monthly_data if r.get('debt_financing', 0) >= 10.0]
    first_big_debt = debt_milestones[0] if debt_milestones else None

    # Category averages - first half vs second half of period
    mid = len(monthly_data) // 2
    first_half = monthly_data[:mid]
    second_half = monthly_data[mid:]

    def avg_cat(recs, cat):
        vals = [r.get(cat, 0) for r in recs]
        return round(sum(vals) / len(vals), 2) if vals else 0.0

    categories = [
        'sears_stores', 'kmart_stores', 'home_services', 'kcd_wholesale',
        'supply_chain', 'citi_reimb', 'cchs', 'asset_sales', 'debt_financing',
        'equity_inflows', 'hts', 'costco', 'tenant_income', 'propco',
        'online', 'misc_inflows', 'kes', 'sears_mexico'
    ]

    category_trends = {}
    for cat in categories:
        early = avg_cat(first_half, cat)
        late = avg_cat(second_half, cat)
        if early > 0.01 or late > 0.01:
            pct = round((late - early) / early * 100, 1) if early > 0 else None
            category_trends[cat] = {
                'first_half_avg_monthly': early,
                'second_half_avg_monthly': late,
                'direction': 'grew' if late > early else ('declined' if late < early else 'flat'),
                'pct_change': pct
            }

    # Year-by-year summary
    by_year = defaultdict(list)
    for r in monthly_data:
        by_year[r['year']].append(r)

    annual_summary = {}
    for yr, recs in sorted(by_year.items()):
        annual_summary[str(yr)] = {
            'months_of_data': len(recs),
            'total_operating_inflows': round(sum(r['total_inflows'] for r in recs), 2),
            'total_with_financing': round(sum(r['total_with_financing'] for r in recs), 2),
            'avg_monthly_operating': round(sum(r['total_inflows'] for r in recs) / len(recs), 2),
            'avg_monthly_home_services': round(sum(r['home_services'] for r in recs) / len(recs), 2),
            'avg_monthly_asset_sales': round(sum(r['asset_sales'] for r in recs) / len(recs), 2),
            'avg_monthly_debt_financing': round(sum(r['debt_financing'] for r in recs) / len(recs), 2),
        }

    output = {
        'metadata': {
            'generated': datetime.now().isoformat()[:19],
            'period': 'Sep 2022 – May 2026 (45 months)',
            'entity': 'Transform SR Holding Management LLC',
            'total_months': len(monthly_data),
            'files_processed': len(files),
            'primary_source': 'Inflows Actuals sheet (confirmed monthly receipts)',
            'may_2026_note': 'May 2026 uses Inflows Detail forecast (Inflows Actuals only ~1 week of data as of file date 5/8/26)',
            'units': '$M (millions)',
            'columns': {
                'total_inflows': 'Operating cash inflows (Daily Total in Inflows Actuals)',
                'equity_inflows': 'Equity capital contributions',
                'debt_financing': 'Net debt/financing proceeds',
                'total_with_financing': 'total_inflows + equity + debt_financing',
                'sears_stores': 'Sears B&M store cash receipts',
                'kmart_stores': 'Kmart store cash receipts',
                'home_services': 'Home Services (warranty, repairs, installation) receipts',
                'kcd_wholesale': 'KCD brand licensing/wholesale',
                'supply_chain': 'Supply chain / logistics billings',
                'citi_reimb': 'CITI card reimbursements',
                'cchs': 'Cross Country Home Services receipts',
                'asset_sales': 'Proceeds from asset dispositions (RE, equipment, other)',
                'hts': 'HTS (Home & Tool Solutions) + HTS Prefund',
                'costco': 'Costco Sears connection program receipts',
                'tenant_income': 'Tenant / sublease income',
                'propco': 'PropCo / Lease OpCo distributions',
                'online': 'Online / e-commerce receipts',
                'misc_inflows': 'Miscellaneous inflows',
            }
        },
        'insights': {
            'sep_2022_vs_may_2026': {
                'sep_2022': {
                    'total_inflows': sep_2022['total_inflows'] if sep_2022 else None,
                    'total_with_financing': sep_2022['total_with_financing'] if sep_2022 else None,
                    'sears_stores': sep_2022['sears_stores'] if sep_2022 else None,
                    'kmart_stores': sep_2022['kmart_stores'] if sep_2022 else None,
                    'home_services': sep_2022['home_services'] if sep_2022 else None,
                    'asset_sales': sep_2022['asset_sales'] if sep_2022 else None,
                    'supply_chain': sep_2022['supply_chain'] if sep_2022 else None,
                    'citi_reimb': sep_2022['citi_reimb'] if sep_2022 else None,
                    'debt_financing': sep_2022['debt_financing'] if sep_2022 else None,
                },
                'may_2026': {
                    'total_inflows': may_2026['total_inflows'] if may_2026 else None,
                    'total_with_financing': may_2026['total_with_financing'] if may_2026 else None,
                    'sears_stores': may_2026['sears_stores'] if may_2026 else None,
                    'kmart_stores': may_2026['kmart_stores'] if may_2026 else None,
                    'home_services': may_2026['home_services'] if may_2026 else None,
                    'asset_sales': may_2026['asset_sales'] if may_2026 else None,
                    'supply_chain': may_2026['supply_chain'] if may_2026 else None,
                    'citi_reimb': may_2026['citi_reimb'] if may_2026 else None,
                    'debt_financing': may_2026['debt_financing'] if may_2026 else None,
                    'is_forecast': may_2026['is_forecast'] if may_2026 else None,
                },
                'change_total_inflows': round(
                    (may_2026['total_inflows'] - sep_2022['total_inflows'])
                    / sep_2022['total_inflows'] * 100, 1
                ) if sep_2022 and may_2026 else None,
            },
            'overall_stats': {
                'all_time_max_monthly_operating': round(max(total_series), 2) if total_series else None,
                'all_time_max_monthly_w_financing': round(max(total_w_fin_series), 2) if total_w_fin_series else None,
                'all_time_min_monthly_operating': round(min(total_series), 2) if total_series else None,
                'all_time_avg_monthly_operating': round(sum(total_series) / len(total_series), 2) if total_series else None,
                'total_operating_inflows_full_period': round(sum(total_series), 2),
                'total_with_financing_full_period': round(sum(total_w_fin_series), 2),
            },
            'peak_inflow_months': [
                {
                    'date': r['date_label'],
                    'total_inflows': r['total_inflows'],
                    'total_with_financing': r['total_with_financing'],
                    'primary_driver': max(
                        [('sears', r['sears_stores']), ('kmart', r['kmart_stores']),
                         ('home_svcs', r['home_services']), ('asset_sales', r['asset_sales']),
                         ('citi', r['citi_reimb']), ('debt', r['debt_financing'])],
                        key=lambda x: x[1]
                    )[0],
                    'source_file': r['source_file'][-45:]
                }
                for r in peak_months
            ],
            'asset_sales': {
                'first_month_over_5M': {
                    'date': first_big_asset_sales['date_label'] if first_big_asset_sales else None,
                    'amount': first_big_asset_sales['asset_sales'] if first_big_asset_sales else None,
                },
                'months_with_meaningful_asset_sales': len([r for r in monthly_data if r['asset_sales'] >= 5.0]),
                'peak_asset_sales_month': max(monthly_data, key=lambda r: r['asset_sales'])['date_label'] if monthly_data else None,
                'peak_asset_sales_amount': max(r['asset_sales'] for r in monthly_data) if monthly_data else None,
            },
            'debt_financing_emergence': {
                'first_month_over_10M': {
                    'date': first_big_debt['date_label'] if first_big_debt else None,
                    'amount': first_big_debt['debt_financing'] if first_big_debt else None,
                },
                'months_with_financing': len([r for r in monthly_data if r['debt_financing'] >= 1.0]),
                'total_financing_raised': round(sum(r['debt_financing'] + r['equity_inflows'] for r in monthly_data), 2),
            },
            'category_trends': category_trends,
            'annual_summary': annual_summary,
        },
        'monthly_data': monthly_data,
    }

    out_path = os.path.join(BASE_DIR, 'analysis_inflows_trend.json')
    with open(out_path, 'w') as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\n{'='*70}")
    print(f"SAVED: {out_path}")
    print(f"{'='*70}")

    # Print summary
    print(f"\n=== KEY FINDINGS ===")
    print(f"Period: Sep 2022 – May 2026 ({len(monthly_data)} months)")

    if sep_2022:
        print(f"\nSep 2022 (first month):")
        print(f"  Total Operating Inflows: ${sep_2022['total_inflows']:.1f}M")
        print(f"  + Equity/Debt: ${sep_2022['equity_inflows']:.1f}M + ${sep_2022['debt_financing']:.1f}M")
        print(f"  = Total w/Financing: ${sep_2022['total_with_financing']:.1f}M")
        print(f"  Sears: ${sep_2022['sears_stores']:.1f}M | Kmart: ${sep_2022['kmart_stores']:.1f}M | HS: ${sep_2022['home_services']:.1f}M")
        print(f"  CITI: ${sep_2022['citi_reimb']:.1f}M | SC: ${sep_2022['supply_chain']:.1f}M | AS: ${sep_2022['asset_sales']:.1f}M")

    if may_2026:
        print(f"\nMay 2026 (last month, {'' if not may_2026['is_forecast'] else 'FORECAST'}):")
        print(f"  Total Operating Inflows: ${may_2026['total_inflows']:.1f}M")
        print(f"  + Debt: ${may_2026['debt_financing']:.1f}M")
        print(f"  = Total w/Financing: ${may_2026['total_with_financing']:.1f}M")
        print(f"  Sears: ${may_2026['sears_stores']:.1f}M | Kmart: ${may_2026['kmart_stores']:.1f}M | HS: ${may_2026['home_services']:.1f}M")
        if sep_2022:
            pct = (may_2026['total_inflows'] - sep_2022['total_inflows']) / sep_2022['total_inflows'] * 100
            print(f"  Change vs Sep 2022: {pct:+.0f}%")

    print(f"\nAll-time stats (operating inflows, no financing):")
    print(f"  Max: ${max(total_series):.1f}M ({peak_months[0]['date_label']})")
    print(f"  Min: ${min(total_series):.1f}M")
    print(f"  Avg: ${sum(total_series)/len(total_series):.1f}M/month")
    print(f"  Period Total: ${sum(total_series):.0f}M")

    print(f"\nTop 5 peak months:")
    for r in peak_months[:5]:
        print(f"  {r['date_label']}: ${r['total_inflows']:.1f}M operating | ${r['total_with_financing']:.1f}M w/financing")

    if first_big_asset_sales:
        print(f"\nAsset sales first >$5M: {first_big_asset_sales['date_label']} (${first_big_asset_sales['asset_sales']:.1f}M)")

    if first_big_debt:
        print(f"Debt/financing first >$10M: {first_big_debt['date_label']} (${first_big_debt['debt_financing']:.1f}M)")

    print(f"\nCategory trends (first half avg -> second half avg):")
    for cat, trend in sorted(category_trends.items(), key=lambda x: -x[1]['first_half_avg_monthly']):
        pct = f"{trend['pct_change']:+.0f}%" if trend['pct_change'] is not None else 'N/A'
        print(f"  {cat:25s}: ${trend['first_half_avg_monthly']:.1f}M -> ${trend['second_half_avg_monthly']:.1f}M  ({pct} {trend['direction']})")

    print(f"\nAnnual summary:")
    for yr, summ in output['insights']['annual_summary'].items():
        print(f"  {yr}: {summ['months_of_data']}mo | Ops=${summ['total_operating_inflows']:.0f}M | w/Fin=${summ['total_with_financing']:.0f}M | AvgMo=${summ['avg_monthly_operating']:.0f}M")

    return output


if __name__ == '__main__':
    main()
