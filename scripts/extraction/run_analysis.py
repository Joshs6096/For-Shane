"""
Cash Forecast Model Evolution & Architecture Analysis
Covers Part A (Model Evolution), Part B (Data Flow), Part C (Automation Architecture)
"""

import json
import os
import re
from collections import defaultdict
import openpyxl
from openpyxl import load_workbook

BASE = "/Users/josh/Downloads/SP_Analysis"

TARGET_FILES = [
    ("9.19.22",  "Daily Cash Fcst - 9.19.22_BU view_Sep BOD.xlsx"),
    ("1.17.23",  "Daily Cash Fcst - 1.17.23_BU view_JanBOD.xlsx"),
    ("8.22.23",  "Daily Cash Fcst - 8.22.23_BU view_AugBOD.xlsx"),
    ("1.23.24",  "Daily Cash Fcst - 1.23.24_BU view_BOD.xlsx"),
    ("7.23.24",  "Daily Cash Fcst - 7.23.24_deck_JulBOD_FY25.xlsx"),
    ("1.21.25",  "Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx"),
    ("05.08.26", "Daily Cash Fcst - 05.08.26_BU view.xlsx"),
]

FORMULA_RE = re.compile(r'^=')

def analyze_sheet(ws):
    """Return (rows, cols, formula_count, hardcoded_count, sample_headers)."""
    rows = ws.max_row or 0
    cols = ws.max_column or 0
    formula_count = 0
    hardcoded_count = 0
    sample_headers = []
    got_headers = False

    # Sample first 200 rows for formula ratio (speed); grab headers from row 1
    sample_rows = min(rows, 200)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=sample_rows, values_only=False), start=1):
        for cell in row:
            if row_idx == 1 and not got_headers:
                if cell.value is not None:
                    sample_headers.append(str(cell.value)[:40])
            if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                formula_count += 1
            elif cell.value is not None:
                hardcoded_count += 1
        if row_idx == 1:
            got_headers = True

    return rows, cols, formula_count, hardcoded_count, sample_headers[:10]


def get_file_size_mb(path):
    return round(os.path.getsize(path) / (1024 * 1024), 2)


def categorize_sheet(name):
    """Assign a category to a sheet by name."""
    n = name.lower()
    if any(x in n for x in ['fy file', 'fyfile', 'fy  file']):
        return 'backend_fy_file'
    if 'payroll' in n:
        return 'backend_payroll'
    if 'merch' in n and 'non' not in n:
        return 'backend_merch'
    if 'non-merch' in n or 'non merch' in n:
        return 'backend_non_merch'
    if 'inflow' in n:
        return 'inflows_detail'
    if 'disburs' in n or 'disb' in n:
        return 'disbursement_detail'
    if 'cash flow' in n or 'cashflow' in n or 'cf ' in n:
        return 'cash_flow'
    if 'summary' in n or 'summ' in n:
        return 'summary'
    if 'bank' in n or 'balan' in n:
        return 'bank_balance'
    if any(x in n for x in ['bu ', 'bus', 'segment', 'retail', 'wholesale', 'online', 'direct']):
        return 'bu_view'
    if 'cover' in n or 'index' in n or 'toc' in n:
        return 'cover'
    if 'chart' in n or 'graph' in n:
        return 'chart'
    return 'other'


def analyze_file(label, filename):
    path = os.path.join(BASE, filename)
    size_mb = get_file_size_mb(path)
    print(f"\n{'='*60}")
    print(f"  {label}  |  {filename[:60]}  |  {size_mb} MB")
    print('='*60)

    # data_only=False to see formulas; read_only for speed on large files
    wb = load_workbook(path, read_only=True, data_only=False)
    sheet_names = wb.sheetnames

    sheets = []
    categories = defaultdict(list)

    for sname in sheet_names:
        ws = wb[sname]
        rows, cols, f_count, h_count, headers = analyze_sheet(ws)
        cat = categorize_sheet(sname)
        categories[cat].append(sname)
        sheet_info = {
            "name": sname,
            "rows": rows,
            "cols": cols,
            "formula_cells_sampled": f_count,
            "hardcoded_cells_sampled": h_count,
            "formula_pct": round(f_count / max(f_count + h_count, 1) * 100, 1),
            "category": cat,
            "sample_headers": headers,
        }
        sheets.append(sheet_info)
        print(f"  {sname:<40} {rows:>6}r x {cols:>4}c  | fml={f_count:>5} hrd={h_count:>5}  cat={cat}")

    wb.close()
    return {
        "label": label,
        "filename": filename,
        "size_mb": size_mb,
        "sheet_count": len(sheet_names),
        "sheets": sheets,
        "category_map": {k: v for k, v in categories.items()},
    }


# ─── PART A: Analyze all target files ────────────────────────────────────────
print("\n\n========= PART A: MODEL EVOLUTION ANALYSIS =========")
all_results = []
for label, fname in TARGET_FILES:
    result = analyze_file(label, fname)
    all_results.append(result)

# Summarize sheet name evolution
print("\n\n========= SHEET NAME EVOLUTION =========")
all_sheet_sets = []
for r in all_results:
    s = set(sh['name'] for sh in r['sheets'])
    all_sheet_sets.append((r['label'], s))

# New sheets each version
prev_sheets = set()
new_sheets_by_version = {}
for label, sheets in all_sheet_sets:
    new = sheets - prev_sheets
    new_sheets_by_version[label] = sorted(new)
    prev_sheets = prev_sheets | sheets
    print(f"\n{label}: {len(sheets)} sheets total, {len(new)} new")
    for s in sorted(new):
        print(f"   + {s}")


# ─── PART B: Deep-dive on 05.08.26 backend sheets ────────────────────────────
print("\n\n========= PART B: DATA FLOW – 05.08.26 BACKEND SHEETS =========")

LATEST = os.path.join(BASE, "Daily Cash Fcst - 05.08.26_BU view.xlsx")
wb26 = load_workbook(LATEST, read_only=True, data_only=False)

BACKEND_KEYWORDS = ['fy file', 'payroll', 'merch', 'non-merch', 'non merch',
                    'inflow', 'disburs', 'cash flow']

backend_details = {}

for sname in wb26.sheetnames:
    n = sname.lower()
    is_backend = any(kw in n for kw in BACKEND_KEYWORDS)
    if not is_backend:
        continue

    ws = wb26[sname]
    rows = ws.max_row or 0
    cols = ws.max_column or 0

    # Collect first 3 rows to understand structure
    sample_rows_data = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        sample_rows_data.append([str(v)[:50] if v is not None else None for v in row[:20]])

    # Find date range: scan row 1 for date-like values
    date_cells = []
    for cell in ws[1]:
        if cell.value and hasattr(cell.value, 'strftime'):
            date_cells.append(str(cell.value.date()))
        elif isinstance(cell.value, str) and any(c.isdigit() for c in cell.value):
            date_cells.append(str(cell.value)[:20])

    date_range = None
    if len(date_cells) >= 2:
        date_range = f"{date_cells[0]} → {date_cells[-1]}"
    elif date_cells:
        date_range = date_cells[0]

    # Count formulas in first 50 rows
    f_cnt = h_cnt = 0
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=False):
        for cell in row:
            if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                f_cnt += 1
            elif cell.value is not None:
                h_cnt += 1

    # Scan for cross-sheet references (=SheetName! patterns)
    cross_refs = set()
    for row in ws.iter_rows(min_row=1, max_row=30, values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                # Extract sheet references
                matches = re.findall(r"'([^']+)'!", cell.value)
                matches += re.findall(r"([A-Za-z][A-Za-z0-9_ ]+)!", cell.value)
                cross_refs.update(matches)

    backend_details[sname] = {
        "rows": rows,
        "cols": cols,
        "date_range": date_range,
        "formula_cells_50rows": f_cnt,
        "hardcoded_cells_50rows": h_cnt,
        "cross_sheet_refs": sorted(cross_refs)[:15],
        "sample_row1": sample_rows_data[0] if sample_rows_data else [],
        "sample_row2": sample_rows_data[1] if len(sample_rows_data) > 1 else [],
        "sample_row3": sample_rows_data[2] if len(sample_rows_data) > 2 else [],
    }

    print(f"\n  [{sname}]  {rows}r x {cols}c  | dates: {date_range}")
    print(f"    formulas(50r)={f_cnt}  hardcoded={h_cnt}")
    if cross_refs:
        print(f"    cross-refs → {sorted(cross_refs)[:8]}")
    print(f"    row1 sample: {sample_rows_data[0][:8] if sample_rows_data else 'N/A'}")

wb26.close()

# ─── PART C: Automation Architecture ─────────────────────────────────────────
print("\n\n========= PART C: AUTOMATION ARCHITECTURE DESIGN =========")

automation_arch = {
    "pipeline_name": "Daily Cash Forecast Auto-Generator",
    "target_file": "Daily Cash Fcst - {DATE}_BU view.xlsx",
    "run_schedule": "Weekdays 07:00 ET (before treasury opens)",
    "layers": {
        "1_data_collection": {
            "description": "Pull all raw data from source systems",
            "sources": {
                "netsuite_mcp": {
                    "what_it_provides": [
                        "AP disbursements (vendor payments cleared by date)",
                        "AR receipts (customer cash received)",
                        "Bank account balances (GL cash accounts)",
                        "Intercompany settlements",
                        "Accrued liabilities / scheduled payments",
                        "Purchase orders (future cash commitments)"
                    ],
                    "query_approach": "SuiteQL: SELECT date, account, amount, entity FROM Transaction WHERE type IN ('VendPymt','CustPymt','Check','Deposit') AND trandate = SYSDATE-1",
                    "unit": "USD thousands (divide by 1000 to match model)",
                    "latency": "T+0, available by 06:30 ET",
                    "gaps": [
                        "Does not contain payroll detail (ADP feeds NS summary only)",
                        "MERCH vendor terms may lag actual bank posting by 1-2 days",
                        "JPM bank position is authoritative; NS GL may have timing differences"
                    ]
                },
                "adp_payroll": {
                    "what_it_provides": [
                        "Payroll funding date and gross amount by pay group",
                        "Employee count by BU",
                        "Employer taxes and benefits (FICA, FUTA, SUTA)",
                        "Net payroll wire amount"
                    ],
                    "delivery": "SFTP drop: /payroll/daily/payroll_register_{YYYYMMDD}.csv",
                    "frequency": "Bi-weekly per pay cycle; preview file available T-3 days",
                    "latency": "File available by 06:00 ET on pay dates, otherwise prior file"
                },
                "jpm_bank_position": {
                    "what_it_provides": [
                        "Prior-day ledger balance by account",
                        "Current-day collected balance",
                        "Intraday transaction detail (wires, ACH, checks)",
                        "Controlled disbursement (positive pay) details"
                    ],
                    "delivery": "BAI2 file via SFTP or JPM ACCESS API",
                    "latency": "Prior-day position available by 06:30 ET",
                    "format": "BAI2 standard; parse with bai2 Python library"
                },
                "merch_feed": {
                    "what_it_provides": [
                        "Merchandise AP aging by vendor and due date",
                        "Open POs approved for payment",
                        "Freight and import duty payments",
                        "Return-to-vendor credits"
                    ],
                    "delivery": "Internal ERP extract or shared drive: /merch/ap_aging_{YYYYMMDD}.xlsx",
                    "latency": "Available by 07:00 ET",
                    "note": "May require manual upload if no API; flagged for future automation"
                },
                "non_merch_feed": {
                    "what_it_provides": [
                        "Facilities / rent payments",
                        "IT and SaaS vendor payments",
                        "Insurance premiums",
                        "Professional services",
                        "Capital expenditures (non-merch)"
                    ],
                    "delivery": "NetSuite MCP (AP subledger) supplemented by internal categorization map",
                    "latency": "Available same time as NS pull"
                },
                "xlsb_team_template": {
                    "description": "Structural shell file (no data, all formulas/formatting intact)",
                    "path": "/templates/cash_fcst_template.xlsb",
                    "role": "File generation target — data is injected into named ranges",
                    "note": "openpyxl cannot read .xlsb; template must be maintained as .xlsx clone or use xlwings for xlsb"
                }
            }
        },
        "2_data_transformation": {
            "description": "Normalize, categorize, and reconcile raw data before injection",
            "steps": [
                {
                    "step": "Unit conversion",
                    "detail": "All sources → USD thousands (÷1000). JPM BAI2 is in cents (÷100000)"
                },
                {
                    "step": "BU categorization",
                    "detail": "Map vendor/entity IDs to BU using NS custom segment or internal mapping table (CSV lookup)"
                },
                {
                    "step": "Category tagging",
                    "detail": "Tag each disbursement as MERCH, Non-Merch, Payroll, Interco, Tax, CapEx"
                },
                {
                    "step": "Timing adjustment",
                    "detail": "Apply float calendar: check clearance +1 day, ACH +1 day, wire same day"
                },
                {
                    "step": "Sign convention",
                    "detail": "Inflows positive, outflows negative. Verify NS AP = negative, AR = positive"
                },
                {
                    "step": "Consolidation roll-up",
                    "detail": "Sum BU-level to consolidated; eliminate interco entries"
                },
                {
                    "step": "Prior-day actual vs. forecast reconciliation",
                    "detail": "Compare yesterday's forecast cells to JPM actuals; compute variance; flag >$500K misses"
                }
            ]
        },
        "3_file_generation": {
            "description": "Write transformed data into the Excel template",
            "approach": "openpyxl write-only mode for data sheets; preserve formula sheets untouched",
            "steps": [
                {
                    "step": "Clone template",
                    "detail": "shutil.copy(template_path, output_path). Never modify master template."
                },
                {
                    "step": "Open with openpyxl (data_only=False)",
                    "detail": "Load clone, keep all formulas intact"
                },
                {
                    "step": "Inject FY File backend sheet",
                    "detail": "Write NS daily actuals to FY File sheet: date column + BU rows. This is the data spine."
                },
                {
                    "step": "Inject Payroll backend sheet",
                    "detail": "Write ADP pay dates and amounts by BU/pay group"
                },
                {
                    "step": "Inject MERCH backend sheet",
                    "detail": "Write MERCH AP aging detail (vendor, amount, due date)"
                },
                {
                    "step": "Inject Non-Merch backend sheet",
                    "detail": "Write Non-Merch AP detail from NS subledger"
                },
                {
                    "step": "Inject bank balance row",
                    "detail": "Write JPM prior-day balance to bank balance summary row (feeds opening balance)"
                },
                {
                    "step": "Set date header",
                    "detail": "Write today's date to cover/title cell"
                },
                {
                    "step": "Save and close",
                    "detail": "wb.save(output_path); log file hash for audit trail"
                }
            ],
            "named_ranges_to_inject": [
                "fy_file_data_start",
                "payroll_data_start",
                "merch_data_start",
                "non_merch_data_start",
                "bank_balance_row",
                "as_of_date_cell"
            ]
        },
        "4_quality_checks": {
            "description": "Automated QC before file is distributed",
            "checks": [
                {
                    "check": "BS_BALANCE_TIES",
                    "rule": "Total ending cash (Cash Flow sheet) == JPM bank balance",
                    "tolerance_k": 100,
                    "action_on_fail": "FLAG – do not distribute; alert treasury"
                },
                {
                    "check": "INFLOWS_COMPLETENESS",
                    "rule": "Inflows Detail total > 0 for each weekday (no silent zero days)",
                    "tolerance_k": 0,
                    "action_on_fail": "FLAG – possible NS data feed failure"
                },
                {
                    "check": "MERCH_DISBURSEMENT_RANGE",
                    "rule": "MERCH disbursements within ±30% of rolling 5-day average",
                    "tolerance_pct": 30,
                    "action_on_fail": "WARN – requires analyst review before distribution"
                },
                {
                    "check": "PAYROLL_DATE_MATCH",
                    "rule": "If today is a payroll date (ADP calendar), payroll disbursement > $0",
                    "tolerance_k": 0,
                    "action_on_fail": "FLAG – ADP file may be missing"
                },
                {
                    "check": "PRIOR_DAY_VARIANCE",
                    "rule": "Actual vs. prior-day forecast variance < $1M per BU",
                    "tolerance_k": 1000,
                    "action_on_fail": "WARN – log variance for monthly attribution report"
                },
                {
                    "check": "FORMULA_INTEGRITY",
                    "rule": "No #REF!, #VALUE!, #N/A errors in key output cells",
                    "tolerance_k": 0,
                    "action_on_fail": "FLAG – template may be corrupted; roll back to prior version"
                },
                {
                    "check": "FILE_SIZE_SANITY",
                    "rule": "Output file size within 20% of prior day file size",
                    "tolerance_pct": 20,
                    "action_on_fail": "FLAG – data truncation or corruption suspected"
                }
            ]
        }
    },
    "distribution": {
        "output_path": "/Users/josh/Downloads/SP_Analysis/daily_output/Daily Cash Fcst - {DATE}_BU view.xlsx",
        "email_recipients": ["treasury@company.com", "cfo@company.com"],
        "sharepoint_path": "/Treasury/Cash Forecasts/{YEAR}/{MONTH}/",
        "delivery_time_target": "08:00 ET"
    },
    "monitoring": {
        "run_log": "/logs/cash_fcst_run_{DATE}.json",
        "alert_channel": "Slack #treasury-ops",
        "sla": "File delivered by 08:00 ET; QC report attached"
    }
}

print(json.dumps(automation_arch, indent=2)[:2000])
print("... [truncated for console]")

# ─── ASSEMBLE FINAL OUTPUT ────────────────────────────────────────────────────
final_output = {
    "generated_at": "2026-05-11",
    "analyst": "Cash Forecast Architecture Analysis",
    "part_a_model_evolution": {
        "files_analyzed": [
            {
                "label": r["label"],
                "filename": r["filename"],
                "size_mb": r["size_mb"],
                "sheet_count": r["sheet_count"],
                "categories": r["category_map"],
                "sheets": r["sheets"],
            }
            for r in all_results
        ],
        "new_sheets_by_version": new_sheets_by_version,
        "size_growth_summary": [
            {"label": r["label"], "size_mb": r["size_mb"], "sheet_count": r["sheet_count"]}
            for r in all_results
        ],
    },
    "part_b_data_flow": {
        "source_file": "Daily Cash Fcst - 05.08.26_BU view.xlsx",
        "backend_sheets": backend_details,
    },
    "part_c_automation_architecture": automation_arch,
}

out_path = os.path.join(BASE, "analysis_automation_architecture.json")
with open(out_path, "w") as f:
    json.dump(final_output, f, indent=2)

print(f"\n\nSaved to: {out_path}")
print(f"File size: {os.path.getsize(out_path) / 1024:.1f} KB")
