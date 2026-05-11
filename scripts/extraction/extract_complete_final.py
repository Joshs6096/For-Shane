"""
Complete final extraction:
- Liquidity sheet: monthly Available Cash, Inflows, Outflows, Net CF (most current forecast months)
- Cash Flow sheet: daily columns include Available Cash, Unavailable Cash, Total Cash
  We extract monthly summary rows (the month-total rows where col A = month name like "Sep", "Oct")
- Payroll: from "Payroll" sheet or payroll backup sheets
- SC Interco $9M: full detail from Dec 2024
- Jan 2025 April append: what April 2025 shows
"""
import openpyxl
import json
from pathlib import Path

BASE = Path("/Users/josh/Downloads/SP_Analysis")

FILES = [
    ("2024-09-24", "Daily Cash Fcst - 9.24.24_BU view_deck version_BOD.xlsx"),
    ("2024-10-15", "Daily Cash Fcst - 10.15.24_BU view_deck version_OctBOD.xlsx"),
    ("2024-11-12", "Daily Cash Fcst - 11.12.24_BU view_deck version_NovBOD.xlsx"),
    ("2024-12-10", "Daily Cash Fcst - 12.10.24_BU view_deck version_BOD_SC wo 3 RE Interco IF_$9M.xlsx"),
    ("2024-12-31", "Daily Cash Fcst - 12.31.24_BU view.xlsx"),
    ("2025-01-21", "Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx"),
    ("2025-03-18", "Daily Cash Fcst - 3.18.25_BU view_deck version_MarBOD.xlsx"),
    ("2025-04-18", "Daily Cash Fcst - 4.18.25_BU view_prelim AprBOD.xlsx"),
    ("2026-05-08", "Daily Cash Fcst - 05.08.26_BU view.xlsx"),
]

def sf(v):
    if v is None: return None
    try: return float(v)
    except: return None

def r3(v):
    if v is None: return None
    return round(float(v), 3)

MONTH_NAMES = ["january","february","march","april","may","june",
               "july","august","september","october","november","december"]
MONTH_SHORT = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]

def is_month_label(s):
    sl = str(s).lower().strip()
    return any(sl == m or sl.startswith(m) for m in MONTH_NAMES + MONTH_SHORT)

def extract_cash_flow_monthly(ws):
    """
    Cash Flow sheet has columns:
    Col indices (0-based): Date, Week, OpCF, NonOpCF, NetCF, NetChg, AvailCash, AC(NoHTS),
                           SegAcct, UnavailCash, TotalCash, ...
    Monthly subtotal rows have col[0]=month name, col[1]=None or blank
    These are the rollup rows for each month.
    """
    # From header row: col 2 = Op CF, col 4 = Net CF, col 6 = Available Cash, col 9 = Unavailable, col 10 = Total Cash
    # We confirmed col indices from R2 output:
    # 0=Date, 1=Week, 2=Op CF, 3=Non-Op CF, 4=Net CF, 5=Net Chg in Cash, 6=Avail Cash,
    # 7=AC(NoHTS), 8=Seg Acct, 9=Unavail Cash, 10=Total Cash, 11=TopCo Undrawn Equity

    monthly_rows = []
    header_found = False

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=3000, values_only=True), 1):
        if row is None or all(v is None for v in row): continue

        c0 = row[0]
        c1 = row[1] if len(row) > 1 else None
        c2 = row[2] if len(row) > 2 else None

        # Monthly subtotal rows: col A = month name (e.g., "Sep", "Oct"), col B = blank/None
        # AND the row has numeric data in cols 2-11
        if c0 is not None and isinstance(c0, str) and is_month_label(c0):
            # Check if it's a monthly summary (col B null/blank) vs a daily row (col B = day name)
            if c1 is None or (isinstance(c1, str) and c1.strip() == ""):
                # Get the key values
                row_list = list(row)
                # pad to 15 cols
                while len(row_list) < 15:
                    row_list.append(None)

                avail_cash = sf(row_list[6])
                unavail_cash = sf(row_list[9])
                total_cash = sf(row_list[10])
                net_cf = sf(row_list[4])
                op_cf = sf(row_list[2])

                monthly_rows.append({
                    "month": str(c0).strip(),
                    "row": row_idx,
                    "net_cf_M": r3(net_cf),
                    "available_cash_M": r3(avail_cash),
                    "unavailable_cash_M": r3(unavail_cash),
                    "total_cash_M": r3(total_cash),
                    "operating_cf_M": r3(op_cf),
                })

    return monthly_rows

def extract_payroll_monthly(wb, date_str):
    """Extract monthly payroll from payroll-specific sheets."""
    payroll_data = {}

    # Try the "Payroll" sheet (detail sheet with payroll cycle data)
    for sname in wb.sheetnames:
        sl = sname.lower()
        if sl == "payroll" or ("fy" in sl and "payroll" in sl and ("24" in sl or "25" in sl or "26" in sl or "27" in sl)):
            ws = wb[sname]
            # Look for monthly totals
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True), 1):
                if row is None: continue
                texts = [str(v).lower().strip() for v in row[:10] if v is not None and isinstance(v, str)]
                row_text = " ".join(texts)
                nums = [sf(v) for v in row if sf(v) is not None and abs(sf(v)) > 0.5]

                # Monthly summary rows in payroll tabs
                if ("total" in row_text or "grand" in row_text or "monthly" in row_text) and nums:
                    payroll_data[f"{sname}_R{i}"] = {
                        "label": row_text[:80],
                        "values_M": [r3(v) for v in nums[:15]]
                    }
                    break  # take first matching row

    # Also check B&M Anita sheets for payroll/bens
    for sname in wb.sheetnames:
        if "anita" in sname.lower() or "shs cash" in sname.lower():
            ws = wb[sname]
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), 1):
                if row is None: continue
                texts = [str(v).lower().strip() for v in row[:10] if v is not None and isinstance(v, str)]
                row_text = " ".join(texts)
                if "payroll/bens" in row_text or "pyrl/bens" in row_text:
                    nums = [sf(v) for v in row if sf(v) is not None and abs(sf(v)) > 0.1]
                    if nums:
                        payroll_data[f"{sname}_R{i}"] = {
                            "label": row_text[:80],
                            "values_M": [r3(v) for v in nums[:15]]
                        }

    return payroll_data

def extract_liquidity_full(ws):
    """Extract all monthly rows from Liquidity sheet."""
    monthly = []
    for row in ws.iter_rows(min_row=1, max_row=500, values_only=True):
        if row is None or all(v is None for v in row): continue

        c0 = row[0]
        c1 = row[1] if len(row) > 1 else None

        # Skip headers / cumulative rows
        if c0 is not None and isinstance(c0, str):
            if "cumulative" in str(c0).lower(): continue
            if "cash flow" in str(c0).lower(): continue
            if "operating" in str(c0).lower(): continue
            if "available" in str(c0).lower() and str(c0).strip() == "Available": continue
            if "cash" in str(c0).lower() and str(c0).strip() in ["Cash", "Cash Flow"]: continue

        # Try 2022-2024 format: [month_name, avail_cash_start, inflows, outflows, ...]
        if c0 is not None and isinstance(c0, str) and any(m in str(c0).lower() for m in MONTH_NAMES):
            month_str = str(c0).strip()
            avail_start = sf(c1)
            inflows = sf(row[2]) if len(row) > 2 else None
            outflows = sf(row[3]) if len(row) > 3 else None
            if inflows is not None or avail_start is not None:
                monthly.append({
                    "month": month_str,
                    "available_cash_start_M": r3(avail_start),
                    "inflows_M": r3(inflows),
                    "outflows_M": r3(outflows),
                    "net_cf_M": r3(inflows + outflows) if (inflows is not None and outflows is not None) else None,
                })

        # Try 2026 format: [float, month_name, avail_cash_start, inflows, outflows, ...]
        elif c1 is not None and isinstance(c1, str) and any(m in str(c1).lower() for m in MONTH_NAMES):
            if "cumulative" in str(c1).lower(): continue
            month_str = str(c1).strip()
            avail_start = sf(row[2]) if len(row) > 2 else None
            inflows = sf(row[3]) if len(row) > 3 else None
            outflows = sf(row[4]) if len(row) > 4 else None
            if inflows is not None or avail_start is not None:
                monthly.append({
                    "month": month_str,
                    "available_cash_start_M": r3(avail_start),
                    "inflows_M": r3(inflows),
                    "outflows_M": r3(outflows),
                    "net_cf_M": r3(inflows + outflows) if (inflows is not None and outflows is not None) else None,
                })

    return monthly

def process_file(date_str, filename):
    filepath = BASE / filename
    print(f"\n{'='*70}")
    print(f"FILE: {filename}")

    result = {
        "file": filename,
        "date": date_str,
        "sheets": [],
        "liquidity_monthly": [],
        "cashflow_monthly_summary": [],
        "payroll_bens": {},
        "sc_interco_detail": None,
        "jan25_april_append": None,
        "notes": []
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        result["sheets"] = list(wb.sheetnames)

        # 1. Liquidity sheet — all monthly rows
        liq_names = [s for s in wb.sheetnames if "liquidity" in s.lower()]
        if liq_names:
            ws = wb[liq_names[0]]
            liq_data = extract_liquidity_full(ws)
            result["liquidity_monthly"] = liq_data
            print(f"  Liquidity: {len(liq_data)} month rows")

        # 2. Cash Flow sheet — monthly summary rows
        cf_names = [s for s in wb.sheetnames if s.lower() == "cash flow"]
        if cf_names:
            ws = wb[cf_names[0]]
            cf_monthly = extract_cash_flow_monthly(ws)
            result["cashflow_monthly_summary"] = cf_monthly
            print(f"  Cash Flow monthly rows: {len(cf_monthly)}")
            # Print last 12 months
            for m in cf_monthly[-12:]:
                print(f"    {m['month']}: avail={m['available_cash_M']}, unavail={m['unavailable_cash_M']}, total={m['total_cash_M']}, net_cf={m['net_cf_M']}")

        # 3. Payroll
        payroll = extract_payroll_monthly(wb, date_str)
        result["payroll_bens"] = payroll
        if payroll:
            print(f"  Payroll data found: {list(payroll.keys())}")

        # 4. Dec 2024: SC Interco $9M detail
        if "12.10.24" in filename:
            sc_sheet = wb.get("SC Interco") if hasattr(wb, 'get') else None
            if "SC Interco" in wb.sheetnames:
                ws_sc = wb["SC Interco"]
                sc_rows = []
                for i, row in enumerate(ws_sc.iter_rows(min_row=1, max_row=40, values_only=True), 1):
                    non_none = [v for v in row if v is not None]
                    if non_none:
                        sc_rows.append({"row": i, "values": [str(v)[:40] for v in non_none[:10]]})
                result["sc_interco_detail"] = sc_rows

        # 5. Jan 2025: April append
        if "1.21.25" in filename:
            # The Liquidity sheet contains historical months plus forecast months
            # April rows in the liquidity data (there can be Apr 2022, Apr 2023, Apr 2024, Apr 2025)
            liq = result["liquidity_monthly"]
            april_rows = [m for m in liq if "april" in m["month"].lower() or
                         (m["month"].lower().startswith("apr") and not any(yr in m["month"] for yr in ["2022","2023","2024"]))]
            result["jan25_april_append"] = {
                "note": "Liquidity sheet April rows (all years shown)",
                "april_rows": [m for m in liq if "apr" in m["month"].lower()][:10],
                "explanation": ("File name 'Apr append_deck2' suggests an April 2025 forecast horizon "
                               "was appended to the Jan 2025 BOD deck. The Liquidity sheet covers "
                               f"Feb 2022 through Jan 2026 horizon ({len(liq)} total months).")
            }

        wb.close()

    except Exception as e:
        result["notes"].append(f"ERROR: {e}")
        import traceback; traceback.print_exc()

    return result

def main():
    all_results = []
    for date_str, filename in FILES:
        r = process_file(date_str, filename)
        all_results.append(r)

    # Print concise comparison table
    print("\n\n" + "="*80)
    print("KEY METRICS COMPARISON — Most Recent Forecast Month Per File")
    print("="*80)
    print(f"{'File Date':<12} {'Last Month (Liq)':<18} {'Avail Start':<14} {'Inflows':<12} {'Outflows':<12} {'Net CF':<10}")
    print("-"*80)

    comparison_rows = []
    for r in all_results:
        liq = r["liquidity_monthly"]
        last_liq = liq[-1] if liq else None

        cf_monthly = r["cashflow_monthly_summary"]
        # Get the most recent months (those that are forecast months, not actuals from 2020)
        # Filter to months that appear to be in the forecast period
        recent_cf = [m for m in cf_monthly if m["available_cash_M"] is not None][-3:] if cf_monthly else []

        row = {
            "date": r["date"],
            "liquidity_last_month": last_liq["month"] if last_liq else None,
            "available_cash_start_M": last_liq["available_cash_start_M"] if last_liq else None,
            "inflows_M": last_liq["inflows_M"] if last_liq else None,
            "outflows_M": last_liq["outflows_M"] if last_liq else None,
            "net_cf_M": last_liq["net_cf_M"] if last_liq else None,
            "cashflow_sheet_recent_months": recent_cf,
        }
        comparison_rows.append(row)

        print(f"{r['date']:<12} {str(last_liq['month'] if last_liq else 'N/A'):<18} "
              f"{str(last_liq['available_cash_start_M'] if last_liq else 'N/A'):<14} "
              f"{str(last_liq['inflows_M'] if last_liq else 'N/A'):<12} "
              f"{str(last_liq['outflows_M'] if last_liq else 'N/A'):<12} "
              f"{str(last_liq['net_cf_M'] if last_liq else 'N/A'):<10}")

    # Compare Dec 2024 vs May 2026
    print("\n\nCOMPARISON: Available Cash trend (Dec 2024 → May 2026)")
    print("="*60)

    dec_r = next((r for r in all_results if "12.10.24" in r["file"]), None)
    may_r = next((r for r in all_results if "05.08.26" in r["file"]), None)

    if dec_r and may_r:
        # December 2024 (from the Dec 2024 file liquidity): January row is the last
        dec_jan = next((m for m in reversed(dec_r["liquidity_monthly"]) if "jan" in m["month"].lower()), None)
        dec_dec = next((m for m in reversed(dec_r["liquidity_monthly"]) if "dec" in m["month"].lower()), None)
        may_last = may_r["liquidity_monthly"][-1] if may_r["liquidity_monthly"] else None
        may_first = may_r["liquidity_monthly"][0] if may_r["liquidity_monthly"] else None

        print(f"\nDec 2024 file - December row: avail={dec_dec['available_cash_start_M'] if dec_dec else 'N/A'}")
        print(f"Dec 2024 file - January (last) row: avail={dec_jan['available_cash_start_M'] if dec_jan else 'N/A'}")
        print(f"May 2026 file - First row (Feb 2025): avail={may_first['available_cash_start_M'] if may_first else 'N/A'}")
        print(f"May 2026 file - Last row (Jan 2026): avail={may_last['available_cash_start_M'] if may_last else 'N/A'}")

    # SC Interco $9M detail
    if dec_r and dec_r.get("sc_interco_detail"):
        print("\n\nDec 2024 SC INTERCO SHEET - $9M Adjustment Detail")
        print("="*60)
        for r in dec_r["sc_interco_detail"]:
            print(f"  R{r['row']}: {r['values'][:5]}")

    # Build output JSON
    output = {
        "extraction_date": "2026-05-11",
        "methodology": {
            "liquidity_sheet": ("Monthly summary: Available Cash (beginning of month), "
                               "Inflows, Outflows, Net CF. Covers entire forecast horizon per file."),
            "cash_flow_sheet": ("Daily time-series with columns: Date, Week, Op CF, Non-Op CF, "
                               "Net CF, Net Change, Available Cash, AC(No HTS), Seg Acct, "
                               "Unavailable Cash, Total Cash. Monthly subtotals extracted."),
            "payroll": "From B&M Anita / SHS Cash Changes sheets (Payroll/Bens row)",
            "units": "All values in $Millions"
        },
        "comparison_table": comparison_rows,
        "files": []
    }

    for r in all_results:
        file_entry = {
            "file": r["file"],
            "date": r["date"],
            "sheets_count": len(r["sheets"]),
            "sheet_names": r["sheets"],
            "liquidity_monthly": r["liquidity_monthly"],
            "cashflow_monthly_summary": r["cashflow_monthly_summary"],
            "payroll_bens": r["payroll_bens"],
            "sc_interco_detail_dec2024": r.get("sc_interco_detail"),
            "jan25_april_append": r.get("jan25_april_append"),
            "notes": r["notes"]
        }
        output["files"].append(file_entry)

    out_path = BASE / "analysis_2025_h2_2026.json"
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\n\nJSON saved to: {out_path}")
    print(f"File size: {out_path.stat().st_size / 1024:.1f} KB")

if __name__ == "__main__":
    main()
