"""
Targeted extraction: reads the 'Net Cash Flow' and 'Cash Flow' summary sheets
for each of the 9 files to pull Available Cash, Total Cash, Unavailable Cash,
Net CF, Inflows Total, Disbursements Total, and Payroll/Bens.
"""
import openpyxl
import json
from pathlib import Path

BASE = Path("/Users/josh/Downloads/SP_Analysis")

FILES = [
    ("2024-09-24", "Daily Cash Fcst - 9.24.24_BU view_deck version_BOD.xlsx"),
    ("2024-10-15", "Daily Cash Fcst - 10.15.24_BU view_deck version_OctBOD.xlsx"),
    ("2024-11-12", "Daily Cash Fcst - 11.12.24_BU view_deck version_NovBOD.xlsx"),
    ("2024-12-10", "Daily Cash Fcst - 12.10.24_BU view_deck version_BOD_SC wo 3 RE Interco IF_$9M.xlsx"),
    ("2024-12-31", "Daily Cash Fcst - 12.31.24_BU view.xlsx"),
    ("2025-01-21", "Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx"),
    ("2025-03-18", "Daily Cash Fcst - 3.18.25_BU view_deck version_MarBOD.xlsx"),
    ("2025-04-18", "Daily Cash Fcst - 4.18.25_BU view_prelim AprBOD.xlsx"),
    ("2026-05-08", "Daily Cash Fcst - 05.08.26_BU view.xlsx"),
]

def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except:
        return None

def to_millions(v, raw_val=None):
    """Smart conversion to millions based on magnitude."""
    if v is None:
        return None
    abs_v = abs(v)
    if abs_v == 0:
        return 0.0
    if abs_v >= 100_000:   # dollars -> millions
        return round(v / 1_000_000, 3)
    elif abs_v >= 1_000:   # thousands -> millions
        return round(v / 1_000, 3)
    else:
        return round(v, 3)

def dump_sheet_first_rows(ws, n_rows=30, n_cols=150, label=""):
    """Print first n_rows x n_cols as TSV for inspection."""
    print(f"\n--- {label} ---")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=n_rows, values_only=True), 1):
        row_data = [str(v) if v is not None else "" for v in row[:n_cols]]
        # Only print rows that have something
        non_empty = [x for x in row_data if x.strip()]
        if non_empty:
            print(f"  R{i}: {' | '.join(row_data[:30])}")

def extract_ncf_sheet(ws, filename):
    """Extract metrics from the Net Cash Flow summary sheet."""
    results = {}
    print(f"\n  [Net Cash Flow sheet] for {filename[:40]}")

    # Collect all rows with labels and values
    rows_data = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
        if row[0] is not None or row[1] is not None:
            rows_data.append((i, list(row)))

    # Print first 40 rows to understand structure
    for i, row in rows_data[:40]:
        text_cells = [(j, v) for j, v in enumerate(row[:5]) if v is not None]
        num_cells = [(j, v) for j, v in enumerate(row) if isinstance(v, (int, float)) and j > 3]
        print(f"    R{i}: {text_cells} | nums[4+]: {[(j, v) for j, v in num_cells[:10]]}")

    return results

def scan_for_metrics(ws, label=""):
    """Scan all rows for key metric labels and extract the entire row of values."""
    TARGET_KEYWORDS = {
        "available_cash": ["available cash"],
        "total_cash": ["total cash"],
        "unavailable_cash": ["unavailable", "not available"],
        "net_cf": ["net cf", "net cash flow"],
        "inflows_total": ["total inflows", "inflows total"],
        "disbursements_total": ["total disbursements", "disbursements total"],
        "payroll_bens": ["payroll/bens", "pyrl/bens total", "payroll & bens", "payroll total", "total payroll"],
    }

    found = {k: [] for k in TARGET_KEYWORDS}

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=2000, values_only=True), 1):
        # Get all string values in first 5 cols
        row_text = " ".join(
            str(v).lower().strip() for v in row[:5] if v is not None and isinstance(v, str)
        )
        if not row_text:
            continue

        # Get numeric values from entire row
        num_vals = []
        for v in row:
            f = safe_float(v)
            if f is not None and f != 0:
                num_vals.append(f)

        for metric, keywords in TARGET_KEYWORDS.items():
            if any(kw in row_text for kw in keywords):
                # Get first 30 values from the row (skip first 5 label cols)
                row_nums = []
                for j, v in enumerate(row):
                    if j < 5:
                        continue
                    f = safe_float(v)
                    if f is not None:
                        row_nums.append(f)

                entry = {
                    "row": i,
                    "label": " | ".join(str(v) for v in row[:5] if v is not None)[:150],
                    "values_raw": row_nums[:30],
                }
                found[metric].append(entry)
                print(f"    [{metric}] R{i}: {entry['label'][:80]} | raw_vals[5+]: {row_nums[:8]}")

    return found

def get_sheet_date_range(ws):
    """Find min/max date-like values in first 5 rows."""
    import datetime
    dates = []
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for v in row:
            if isinstance(v, (datetime.datetime, datetime.date)):
                dates.append(v)
    if dates:
        dates.sort()
        return str(dates[0])[:10], str(dates[-1])[:10]
    return None, None

def process_file(date_str, filename):
    filepath = BASE / filename
    print(f"\n{'='*70}")
    print(f"FILE: {filename}")
    print(f"DATE: {date_str}")

    result = {
        "file": filename,
        "date": date_str,
        "sheets": [],
        "key_metrics_by_sheet": {},
        "forecast_horizon": {},
        "special_findings": []
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        result["sheets"] = list(wb.sheetnames)
        print(f"  Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

        # Focus on summary/key sheets
        KEY_SHEET_KEYWORDS = ["net cash flow", "cash flow", "liquidity", "summary", "treasury"]
        sheets_to_scan = []
        for s in wb.sheetnames:
            sl = s.lower()
            if any(kw in sl for kw in KEY_SHEET_KEYWORDS):
                sheets_to_scan.append(s)

        # Always scan first 2 sheets
        for s in wb.sheetnames[:2]:
            if s not in sheets_to_scan:
                sheets_to_scan.insert(0, s)

        print(f"  Scanning sheets: {sheets_to_scan[:6]}")

        for sheet_name in sheets_to_scan[:6]:
            ws = wb[sheet_name]
            d_start, d_end = get_sheet_date_range(ws)
            result["forecast_horizon"][sheet_name] = {
                "start": d_start, "end": d_end
            }
            print(f"\n  Sheet: '{sheet_name}' | Date range: {d_start} to {d_end}")

            metrics = scan_for_metrics(ws, sheet_name)
            if any(len(v) > 0 for v in metrics.values()):
                result["key_metrics_by_sheet"][sheet_name] = metrics

        # SPECIAL: Dec 2024 — look for SC / RE / Interco / $9M sheets
        if "12.10.24" in filename:
            print(f"\n  [SPECIAL] Dec 2024: searching for SC/RE/Interco data")
            for sname in wb.sheetnames:
                sl = sname.lower()
                if any(kw in sl for kw in ["sc", "re", "interco", "supply", "real estate", "9m", "adjust"]):
                    ws = wb[sname]
                    print(f"    Scanning special sheet: {sname}")
                    metrics = scan_for_metrics(ws, sname)
                    if any(len(v) > 0 for v in metrics.values()):
                        result["key_metrics_by_sheet"][f"SPECIAL_{sname}"] = metrics
                    result["special_findings"].append(f"Sheet '{sname}' found and scanned")

        # SPECIAL: Jan 2025 — look for April 2025 append data
        if "1.21.25" in filename:
            print(f"\n  [SPECIAL] Jan 2025: searching for April 2025 append data")
            # Scan all sheets for April 2025 references
            for sname in wb.sheetnames:
                sl = sname.lower()
                if "apr" in sl or "april" in sl or "2025" in sl or "append" in sl or "deck2" in sl:
                    ws = wb[sname]
                    print(f"    Scanning April sheet: {sname}")
                    metrics = scan_for_metrics(ws, sname)
                    result["key_metrics_by_sheet"][f"APRIL_{sname}"] = metrics
                    result["special_findings"].append(f"April-related sheet '{sname}' scanned")

        wb.close()

    except Exception as e:
        result["special_findings"].append(f"ERROR: {e}")
        print(f"  ERROR: {e}")
        import traceback
        traceback.print_exc()

    return result

def main():
    all_results = []
    for date_str, filename in FILES:
        r = process_file(date_str, filename)
        all_results.append(r)

    # Build concise summary
    summary = []
    for r in all_results:
        row = {"date": r["date"], "file": r["file"], "sheets": r["sheets"]}

        # Aggregate metrics across all sheets
        for sheet_name, metrics in r.get("key_metrics_by_sheet", {}).items():
            for metric_key, entries in metrics.items():
                if entries:
                    key = f"{metric_key}_examples"
                    if key not in row:
                        row[key] = []
                    for e in entries[:3]:
                        row[key].append({
                            "sheet": sheet_name,
                            "label": e["label"][:100],
                            "values_raw": e["values_raw"][:10]
                        })

        summary.append(row)

    output = {
        "extraction_date": "2026-05-11",
        "note": "Targeted extraction from Net Cash Flow / Cash Flow / Liquidity / Summary sheets",
        "summary": summary,
        "detail": all_results
    }

    out_path = BASE / "analysis_2025_h2_2026.json"
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\n\nJSON written to: {out_path}")

if __name__ == "__main__":
    main()
