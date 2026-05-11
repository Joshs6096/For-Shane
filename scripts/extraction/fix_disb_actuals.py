"""
Patch Daily Cash Fcst - 5.8.26_10tab.xlsx
→ Replace blank Disbursement Actuals with real data read from the xlsb team file.
Values in the xlsb are stored in millions.  Date col 0 is an Excel serial.
"""
import warnings; warnings.filterwarnings('ignore')
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pyxlsb import open_workbook as open_xlsb
from datetime import date, timedelta, datetime

XLSB = '/Users/josh/Desktop/Daily Cash Fcst - 05.08.26.xlsb'
DST  = '/Users/josh/Downloads/Daily Cash Fcst - 5.8.26_10tab.xlsx'

# ── Style helpers (same as build script) ──────────────────────────────────
NAVY='1F3864'; WHITE='FFFFFF'; LBLUE='D6E4F0'; LGREY='F2F2F2'
DGREY='595959'; AMBER='FFF2CC'

def hdr(sz=9,bold=True,color=WHITE): return Font(name='Arial',size=sz,bold=bold,color=color)
def body(sz=9,bold=False,color="000000"): return Font(name='Arial',size=sz,bold=bold,color=color)
def fill(h): return PatternFill("solid",fgColor=h)
def ctr(wrap=False): return Alignment(horizontal='center',vertical='center',wrap_text=wrap)
def lft(): return Alignment(horizontal='left',vertical='center')
def rgt(): return Alignment(horizontal='right',vertical='center')
def thin():
    s=Side(style='thin',color='BBBBBB')
    return Border(left=s,right=s,top=s,bottom=s)
def med_bottom():
    m=Side(style='medium',color=NAVY); t=Side(style='thin',color='BBBBBB')
    return Border(left=t,right=t,top=t,bottom=m)

def ph(ws,row,col,text,bg=NAVY,fg=WHITE,sz=9,wrap=True):
    c=ws.cell(row,col,text)
    c.font=hdr(sz,True,fg); c.fill=fill(bg)
    c.alignment=ctr(wrap); c.border=med_bottom()
    return c

def xl_date(serial):
    """Excel serial → Python date."""
    if serial and isinstance(serial,(int,float)) and serial > 1000:
        return date(1899,12,30) + timedelta(days=int(serial))
    return None

# ── Read all data from xlsb ────────────────────────────────────────────────
print("Reading Disbursement Actuals from xlsb …")
with open_xlsb(XLSB) as wb_x:
    with wb_x.get_sheet('Disbursement Actuals') as ws_x:
        raw_rows = list(ws_x.rows())

# Header rows 1-5 (0-indexed 0-4)
hdr_rows = raw_rows[:5]

# Data rows: only rows where col 0 is a date serial
data_rows = []
for row in raw_rows[5:]:
    cell0 = next((c for c in row if c.c == 0), None)
    if cell0 and isinstance(cell0.v,(int,float)) and cell0.v > 1000:
        vals = {c.c: c.v for c in row if c.v is not None}
        data_rows.append(vals)

print(f"  Header rows: 5")
print(f"  Data rows: {len(data_rows)}")
print(f"  Date range: {xl_date(data_rows[0][0])} → {xl_date(data_rows[-1][0])}")

# ── Column header map from xlsb header rows ────────────────────────────────
# Build a 3-level header: from rows 1, 3, 5 (0-indexed 0, 2, 4)
# Row 1 (0): top-level group names
# Row 3 (2): mid-level section names
# Row 5 (4): bottom-level detail names
row1_map = {c.c: c.v for c in hdr_rows[0] if c.v is not None}
row3_map = {c.c: c.v for c in hdr_rows[2] if c.v is not None}
row5_map = {c.c: c.v for c in hdr_rows[4] if c.v is not None}

# Determine which xlsb columns to include in output
# Include: 0 (date) + all non-helper columns (skip cols > 89 which are audit/check cols)
MAX_DATA_COL = 89   # col 89 = Daily Total; cols 90+ are audit/check columns

# ── Open output xlsx and rebuild Disbursement Actuals ─────────────────────
print("Opening output xlsx …")
wb = load_workbook(DST)

# Delete old sheet and create fresh
if 'Disbursement Actuals' in wb.sheetnames:
    old_pos = wb.sheetnames.index('Disbursement Actuals')
    del wb['Disbursement Actuals']
    ws = wb.create_sheet('Disbursement Actuals', old_pos)
else:
    ws = wb.create_sheet('Disbursement Actuals')

ws.sheet_state = 'visible'

# ── Build header rows ──────────────────────────────────────────────────────
# Map xlsb col index → xlsx col index (1-based)
# We include cols 0..MAX_DATA_COL → xlsx cols 1..(MAX_DATA_COL+1)
def xl_col(xlsb_c): return xlsb_c + 1   # 0-indexed → 1-indexed

# Row 1: top-level group headers (col 0 = DATE, rest from row1_map)
ws.cell(1, 1, 'DATE').font = hdr(sz=9, color=WHITE)
ws.cell(1, 1).fill = fill(NAVY)
ws.cell(1, 1).alignment = ctr()
ws.cell(1, 1).border = med_bottom()

for xlsb_c, label in row1_map.items():
    if xlsb_c == 0 or xlsb_c > MAX_DATA_COL:
        continue
    xc = xl_col(xlsb_c)
    c = ws.cell(1, xc, str(label) if label else '')
    c.font = hdr(sz=7, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = ctr(wrap=True)
    c.border = med_bottom()

# Row 2: HIDE row markers → use as subtle sub-group label row (light blue)
# Skip entirely — not useful for display; row 2 stays empty for spacing

# Row 3: mid-level section headers
for xlsb_c, label in row3_map.items():
    if xlsb_c > MAX_DATA_COL:
        continue
    xc = xl_col(xlsb_c)
    c = ws.cell(3, xc, str(label))
    c.font = hdr(sz=8, color=NAVY)
    c.fill = fill(LBLUE)
    c.alignment = ctr(wrap=True)
    c.border = med_bottom()

# Row 4: blank (HIDE row in xlsb — skip)

# Row 5: detail column headers
for xlsb_c, label in row5_map.items():
    if xlsb_c > MAX_DATA_COL:
        continue
    xc = xl_col(xlsb_c)
    c = ws.cell(5, xc, str(label))
    c.font = hdr(sz=7, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = ctr(wrap=True)
    c.border = med_bottom()

# Row heights for header rows
ws.row_dimensions[1].height = 28
ws.row_dimensions[3].height = 28
ws.row_dimensions[5].height = 36

# ── Write data rows ────────────────────────────────────────────────────────
print("Writing data rows …")
FMT_DATE = 'mmm d, yyyy'
FMT_NUM  = '#,##0.000;(#,##0.000);"-"'

OUT_START = 6   # first data row in xlsx

for ri, vals in enumerate(data_rows):
    xlsx_row = OUT_START + ri
    d = xl_date(vals.get(0))
    if d is None:
        continue

    # Date cell
    dc = ws.cell(xlsx_row, 1, datetime(d.year, d.month, d.day))
    dc.number_format = FMT_DATE
    dc.font = body(sz=8)
    dc.fill = fill(WHITE)
    dc.border = thin()
    dc.alignment = lft()

    # Data cols 1..MAX_DATA_COL
    for xlsb_c in range(1, MAX_DATA_COL + 1):
        v = vals.get(xlsb_c)
        xc = xl_col(xlsb_c)
        cell = ws.cell(xlsx_row, xc)
        if v is not None:
            cell.value = v
            cell.number_format = FMT_NUM
        cell.font = body(sz=8)
        cell.fill = fill(WHITE)
        cell.border = thin()
        cell.alignment = rgt()

print(f"  Wrote {len(data_rows)} data rows (rows 6–{OUT_START + len(data_rows) - 1})")

# ── Column widths ──────────────────────────────────────────────────────────
ws.column_dimensions['A'].width = 13          # date
for xlsb_c in range(1, MAX_DATA_COL + 1):
    xc = xl_col(xlsb_c)
    ws.column_dimensions[get_column_letter(xc)].width = 9

# Wider for PYRL/BENS TOTAL (col 18 → xlsx 19) and Daily Total (col 90 → xlsx 91... wait no)
ws.column_dimensions[get_column_letter(xl_col(17))].width = 12  # PYRL/BENS TOTAL
ws.column_dimensions[get_column_letter(xl_col(35))].width = 12  # MERCH TOTAL
ws.column_dimensions[get_column_letter(xl_col(89))].width = 12  # Daily Total

# Freeze at row 6, col 2 (date frozen)
ws.freeze_panes = 'B6'

# ── Re-apply sheet order (must stay in position 9, 0-indexed 8) ──────────
# openpyxl appended the new sheet at the end; move it back to position 9
TAB_ORDER = [
    'Cash Flow','LC Forecast Changes','Trapped Cash','UBS reserve',
    '2020 Term Loans','Inflows Forecasting','Inflows Actuals',
    'Inflows Detail','Disbursement Actuals','Disbursement Detail',
    'RE & Debt init with Manteno','KES & KCD Cash Flow','SOP','RC',
    'FY File','MERCH','Non-Merch','Payroll',
]
present = {s: wb[s] for s in wb.sheetnames}
wb._sheets = [present[t] for t in TAB_ORDER if t in present]

# ── Save ───────────────────────────────────────────────────────────────────
wb.save(DST)
import os
sz = os.path.getsize(DST)
print(f"\nSaved → {DST}")
print(f"Size: {sz:,} bytes ({sz/1024/1024:.1f} MB)")

# Quick verification
print("\nVerification — Disbursement Actuals:")
print(f"  max_row={ws.max_row}, max_col={ws.max_column}")
may8 = None
for r in range(6, ws.max_row+1):
    v = ws.cell(r, 1).value
    if isinstance(v, datetime) and v.month==5 and v.day==8 and v.year==2026:
        may8 = r
        break
if may8:
    total = ws.cell(may8, xl_col(89)).value
    print(f"  May 8 2026 (row {may8}): Daily Total = {total:.4f}M = ${total*1e6:,.0f}")
else:
    print("  May 8 2026 row not found")
