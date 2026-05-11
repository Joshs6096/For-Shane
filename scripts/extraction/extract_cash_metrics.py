"""
Cash Forecast Extractor
Analyzes 9 historical cash forecast files and outputs structured JSON.
"""
import openpyxl
import json
import re
from pathlib import Path

BASE = Path("/Users/josh/Downloads/SP_Analysis")

FILES = [
    ("2024-09-24", "Daily Cash Fcst - 9.24.24_BU view_deck version_BOD.xlsx"),
    ("2024-10-15", "Daily Cash Fcst - 10.15.24_BU view_deck version_OctBOD.xlsx"),
    ("2024-11-12", "Daily Cash Fcst - 11.12.24_BU view_deck version_NovBOD.xlsx"),
    ("2024-12-10", "Daily Cash Fcst - 12.10.24_BU view_deck version_BOD_SC wo 3 RE Interco IF_$9M.xlsx"),
    ("2024-12-31", "Daily Cash Fcst - 12.31.24_BU view.xlsx"),
    ("2025-01-21", "Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx"),
    ("2025-03-18", "Daily Cash Fcst - 3.18.25_BU view_deck version_MarBOD.xlsx"),
    ("2025-04-18", "Daily Cash Fcst - 4.18.25_BU view_prelim AprBOD.xlsx"),
    ("2026-05-08", "Daily Cash Fcst - 05.08.26_BU view.xlsx"),
]

# Keywords to search for (lowercased patterns)
KEYWORD_PATTERNS = {
    "available_cash": ["available cash"],
    "total_cash": ["total cash"],
    "unavailable_cash": ["unavailable", "not available"],
    "net_cf": ["net cf", "net cash flow", "net cash"],
    "inflows_total": ["total inflows", "inflows total", "total receipts", "total collections"],
    "disbursements_total": ["total disbursements", "disbursements total", "total payments", "total outflows"],
    "payroll_bens": ["payroll", "payroll & bens", "payroll/bens", "wages", "salaries"],
    "supply_chain_interco": ["supply chain", "intercompany", "interco", "sc wo", "re interco"],
    "april_data": ["april", "apr"],
}

def cell_to_millions(val):
    """Convert a cell value to millions. If already small (< 1000), assume already in millions."""
    if val is None:
        return None
    try:
        v = float(val)
        if abs(v) == 0:
            return 0.0
        # If value is >= 100,000 it's likely in dollars -> convert to millions
        if abs(v) >= 100000:
            return round(v / 1_000_000, 3)
        # If value is between 1000 and 100000, might be in thousands
        elif abs(v) >= 1000:
            return round(v / 1_000, 3)
        else:
            # Assume already in millions
            return round(v, 3)
    except (TypeError, ValueError):
        return None

def get_row_numbers(row):
    """Extract all numeric values from a row."""
    nums = []
    for cell in row:
        if cell.value is not None:
            v = cell_to_millions(cell.value)
            if v is not None:
                nums.append(v)
    return nums

def row_label(row):
    """Get text from first few cells of a row."""
    parts = []
    for cell in row[:5]:
        if cell.value is not None and isinstance(cell.value, str):
            parts.append(str(cell.value).strip())
    return " | ".join(parts)

def matches_keyword(text, keywords):
    text_lower = text.lower()
    return any(kw in text_lower for kw in keywords)

def find_date_range(ws):
    """Scan top rows for date-like headers to determine forecast horizon."""
    import datetime
    dates = []
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            v = cell.value
            if isinstance(v, datetime.datetime):
                dates.append(v)
            elif isinstance(v, str):
                # look for month/year patterns
                m = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[\s\-]*(20\d\d)', v, re.I)
                if m:
                    dates.append(m.group(0))
    if dates:
        if all(isinstance(d, str) for d in dates):
            return f"{dates[0]} to {dates[-1]}" if len(dates) > 1 else dates[0]
        dt_dates = [d for d in dates if hasattr(d, 'year')]
        if dt_dates:
            dt_dates.sort()
            return f"{dt_dates[0].strftime('%Y-%m-%d')} to {dt_dates[-1].strftime('%Y-%m-%d')}"
    return "Not detected in first 10 rows"

def extract_file(date_str, filename):
    filepath = BASE / filename
    result = {
        "file": filename,
        "date": date_str,
        "sheets": [],
        "metrics": {k: [] for k in KEYWORD_PATTERNS},
        "forecast_horizon": {},
        "special_notes": [],
        "raw_rows_sample": []
    }

    print(f"\n{'='*60}")
    print(f"Processing: {filename}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        result["sheets"] = wb.sheetnames
        print(f"  Sheets: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Get date range for this sheet
            horizon = find_date_range(ws)
            result["forecast_horizon"][sheet_name] = horizon

            print(f"  Sheet: {sheet_name} | Horizon: {horizon}")

            # Scan rows
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                # Build full row text from string cells
                row_text = " ".join(
                    str(c.value).strip() for c in row
                    if c.value is not None and isinstance(c.value, str)
                )
                if not row_text.strip():
                    continue

                row_nums = get_row_numbers(row)
                label = row_label(row)

                matched_any = False
                for metric_key, keywords in KEYWORD_PATTERNS.items():
                    if matches_keyword(row_text, keywords):
                        entry = {
                            "sheet": sheet_name,
                            "row": row_idx,
                            "label": label,
                            "row_text_snippet": row_text[:200],
                            "values_millions": row_nums[:20]  # first 20 numeric values
                        }
                        result["metrics"][metric_key].append(entry)
                        matched_any = True
                        print(f"    [{metric_key}] Row {row_idx}: {label[:80]} | nums: {row_nums[:8]}")

        wb.close()

    except Exception as e:
        result["special_notes"].append(f"ERROR: {e}")
        print(f"  ERROR: {e}")

    return result

def main():
    all_results = []

    for date_str, filename in FILES:
        r = extract_file(date_str, filename)
        all_results.append(r)

    # Build comparison
    def get_best_value(metrics_list, key):
        """Get the first non-empty numeric value for a metric across all matches."""
        for entry in metrics_list.get(key, []):
            vals = entry.get("values_millions", [])
            non_zero = [v for v in vals if v is not None and abs(v) > 0.001]
            if non_zero:
                return non_zero[0]
        return None

    # Summary table
    summary = []
    for r in all_results:
        row = {
            "file_date": r["date"],
            "file": r["file"],
            "available_cash_first": get_best_value(r["metrics"], "available_cash"),
            "total_cash_first": get_best_value(r["metrics"], "total_cash"),
            "unavailable_cash_first": get_best_value(r["metrics"], "unavailable_cash"),
            "net_cf_first": get_best_value(r["metrics"], "net_cf"),
            "inflows_total_first": get_best_value(r["metrics"], "inflows_total"),
            "disbursements_total_first": get_best_value(r["metrics"], "disbursements_total"),
            "payroll_bens_first": get_best_value(r["metrics"], "payroll_bens"),
        }
        summary.append(row)

    # Trend comparison
    dec10 = next((r for r in all_results if r["date"] == "2024-12-10"), None)
    jan21 = next((r for r in all_results if r["date"] == "2025-01-21"), None)
    may08 = next((r for r in all_results if r["date"] == "2026-05-08"), None)

    comparisons = {}

    if dec10 and may08:
        dec_avail = get_best_value(dec10["metrics"], "available_cash")
        may_avail = get_best_value(may08["metrics"], "available_cash")
        dec_total = get_best_value(dec10["metrics"], "total_cash")
        may_total = get_best_value(may08["metrics"], "total_cash")
        comparisons["dec10_vs_may08_available_cash"] = {
            "dec_2024_12_10": dec_avail,
            "may_2026_05_08": may_avail,
            "change": round(may_avail - dec_avail, 3) if (dec_avail and may_avail) else None
        }
        comparisons["dec10_vs_may08_total_cash"] = {
            "dec_2024_12_10": dec_total,
            "may_2026_05_08": may_total,
            "change": round(may_total - dec_total, 3) if (dec_total and may_total) else None
        }

    if jan21 and may08:
        jan_avail = get_best_value(jan21["metrics"], "available_cash")
        may_avail = get_best_value(may08["metrics"], "available_cash")
        jan_total = get_best_value(jan21["metrics"], "total_cash")
        may_total = get_best_value(may08["metrics"], "total_cash")
        comparisons["jan21_vs_may08_available_cash"] = {
            "jan_2025_01_21": jan_avail,
            "may_2026_05_08": may_avail,
            "change": round(may_avail - jan_avail, 3) if (jan_avail and may_avail) else None
        }
        comparisons["jan21_vs_may08_total_cash"] = {
            "jan_2025_01_21": jan_total,
            "may_2026_05_08": may_total,
            "change": round(may_total - jan_total, 3) if (jan_total and may_total) else None
        }

    # Special: Dec 2024 SC interco notes
    if dec10:
        sc_entries = dec10["metrics"].get("supply_chain_interco", [])
        comparisons["dec10_sc_interco_9M_entries"] = sc_entries[:10]

    # Special: Jan 2025 April append
    if jan21:
        apr_entries = jan21["metrics"].get("april_data", [])
        comparisons["jan21_april_append_entries"] = apr_entries[:10]

    output = {
        "extraction_date": "2026-05-11",
        "files_analyzed": len(all_results),
        "summary_table": summary,
        "comparisons": comparisons,
        "files": all_results
    }

    out_path = BASE / "analysis_2025_h2_2026.json"
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\n{'='*60}")
    print(f"Output saved to: {out_path}")
    print(f"\nSUMMARY TABLE:")
    for row in summary:
        print(f"  {row['file_date']} | avail={row['available_cash_first']} | total={row['total_cash_first']} | net_cf={row['net_cf_first']} | inflows={row['inflows_total_first']} | disb={row['disbursements_total_first']} | payroll={row['payroll_bens_first']}")

if __name__ == "__main__":
    main()
