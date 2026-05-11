"""Extract BU beta annual total data from all files."""
import openpyxl, os, sys

BASE_DIR = "/Users/josh/Downloads/SP_Analysis"

files_to_check = [
    ('9.19.22_Sep BOD',  '9.19.22_BU view_Sep BOD'),
    ('12.7.22_DecBOD',   '12.7.22_BU view_DecBOD'),
    ('1.17.23_JanBOD',   '1.17.23_BU view_JanBOD'),
    ('4.25.23_AprBOD',   '4.25.23_BU view_April 27 BOD'),
    ('7.25.23_JulBOD',   '7.25.23_BU view_JulBOD'),
    ('11.14.23_NovBOD',  '11.14.23_BU view_adUPDATE'),
    ('2.20.24_BOD',      '2.20.24_BU view_BOD'),
    ('6.18.24_JunBOD',   '6.18.24_BU view_JunBOD'),
    ('10.15.24_OctBOD',  '10.15.24_BU view_deck version_OctBOD'),
    ('12.31.24_BU view', '12.31.24_BU view'),
    ('3.18.25_MarBOD',   '3.18.25_BU view_deck version_MarBOD'),
    ('7.22.25_JulBOD',   '7.22.25_BU view_deck version_condensed_JulBOD'),
    ('12.15.25_DecBOD',  '12.15.25_BU view_DecBOD'),
    ('01.16.26_JanBOD',  '01.16.26_BU view new format_asset sales_JanBOD'),
    ('04.21.26_AprBOD',  '04.21.26_ BU version_deck version_AprBOD'),
    ('05.08.26_BU view', '05.08.26_BU view'),
]

print('Scanning BU beta sheets...', flush=True)

for label, partial in files_to_check:
    filepath = None
    for fn in os.listdir(BASE_DIR):
        if partial.lower() in fn.lower() and fn.endswith('.xlsx'):
            filepath = os.path.join(BASE_DIR, fn)
            break
    if not filepath:
        print(f'{label}: FILE NOT FOUND', flush=True)
        continue

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        bu_sheets = [s for s in wb.sheetnames if 'bu beta' in s.lower()]
        hist_sheets = [s for s in wb.sheetnames if 'historica' in s.lower() or 'historicals' in s.lower()]

        sheets_to_try = bu_sheets + hist_sheets
        if not sheets_to_try:
            print(f'{label}: no BU beta or Historicals sheet', flush=True)
            wb.close()
            continue

        for sn in sheets_to_try[:2]:
            ws = wb[sn]
            print(f'\n{label} -> {sn}', flush=True)
            # Scan all rows for Sears/Kmart
            sears_row = kmart_row = None
            store_count_rows = []
            for r in range(1, 50):
                found = False
                for c in range(1, 20):
                    v = ws.cell(row=r, column=c).value
                    if v is None: continue
                    vn = str(v).lower().strip()
                    if ('sears store' in vn) or (vn == 'sears stores'):
                        sears_row = r
                        sears_label_col = c
                        found = True
                    if ('kmart store' in vn) or (vn == 'kmart stores'):
                        kmart_row = r
                        found = True
                    if any(k in vn for k in ['store count', '# stores', 'open store']):
                        store_count_rows.append((r, c, str(v)))

            if sears_row:
                print(f'  Sears row={sears_row}, Kmart row={kmart_row}', flush=True)
                # Print header row just above and both data rows
                for hr in range(max(1, sears_row-4), sears_row):
                    row_vals = []
                    for c in range(1, min((ws.max_column or 70)+1, 80)):
                        v = ws.cell(row=hr, column=c).value
                        if v is not None:
                            row_vals.append(f'[{c}]{v}')
                    if row_vals:
                        print(f'  HDR R{hr}: {" | ".join(str(x) for x in row_vals[:30])[:300]}', flush=True)
                for dr in [sears_row, kmart_row]:
                    if dr:
                        row_vals = []
                        for c in range(1, min((ws.max_column or 70)+1, 80)):
                            v = ws.cell(row=dr, column=c).value
                            if v is not None:
                                row_vals.append(f'[{c}]{v}')
                        print(f'  DATA R{dr}: {" | ".join(str(x) for x in row_vals[:30])[:400]}', flush=True)
            else:
                print(f'  No Sears/Kmart rows found in {sn}', flush=True)
                # Print first 10 rows to debug
                for r in range(1, 12):
                    row_vals = []
                    for c in range(1, 30):
                        v = ws.cell(row=r, column=c).value
                        if v is not None:
                            row_vals.append(f'[{c}]{v}')
                    if row_vals:
                        print(f'  R{r}: {" | ".join(str(x) for x in row_vals[:15])[:250]}', flush=True)

            if store_count_rows:
                print(f'  Store count rows: {store_count_rows}', flush=True)

        wb.close()
    except Exception as e:
        print(f'{label}: ERROR {e}', flush=True)

print('\n\nDONE', flush=True)
