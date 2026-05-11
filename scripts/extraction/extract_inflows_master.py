"""
Master Inflows Extractor for Transform SR Holding Management LLC
Sep 2022 – May 2026

Strategy:
  Primary source: "Inflows Detail" sheet – has granular monthly subtotals by category
  Fallback: "Cash Flow" sheet – uses dynamic column mapping by header name

Output: /Users/josh/Downloads/SP_Analysis/analysis_inflows_trend.json
"""

import openpyxl
import os
import json
import re
from datetime import datetime

BASE_DIR = "/Users/josh/Downloads/SP_Analysis"

MONTH_ABBREVS = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}

FULL_MONTHS = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12
}


def parse_month_label(label):
    """Parse a month label like 'SEP', 'Sep 2022', 'Jan 2026' etc.
    Returns (year, month) tuple or None."""
    if not label or not isinstance(label, str):
        return None
    s = label.strip()

    # Pattern: "Month Year" e.g. "Jan 2026", "SEP 2022", "January 2023"
    m = re.match(r'^([A-Za-z]+)\s+(\d{4})$', s)
    if m:
        mon_str = m.group(1).lower()[:3]
        yr = int(m.group(2))
        mon = MONTH_ABBREVS.get(mon_str)
        if mon:
            return (yr, mon)

    # Pattern: full month name with year e.g. "September 2022"
    m = re.match(r'^([A-Za-z]+)\s+(\d{4})$', s)
    if m:
        mon_str = m.group(1).lower()
        yr = int(m.group(2))
        mon = FULL_MONTHS.get(mon_str) or MONTH_ABBREVS.get(mon_str[:3])
        if mon:
            return (yr, mon)

    # Just 3-letter abbrev with no year (ambiguous – need context)
    m = re.match(r'^([A-Za-z]{3,})$', s)
    if m:
        mon_str = m.group(1).lower()[:3]
        mon = MONTH_ABBREVS.get(mon_str)
        if mon:
            return (None, mon)  # year unknown

    return None


def safe_val(v):
    """Convert cell value to float, treating None/'#N/A'/errors as 0."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        v_clean = v.strip().replace(',', '')
        try:
            return float(v_clean)
        except:
            return 0.0
    return 0.0


def build_col_map_by_header(headers):
    """Build a dict mapping category names to 0-based column indices by reading header labels."""
    col_map = {}
    for i, h in enumerate(headers):
        if not h or not isinstance(h, str):
            continue
        hl = h.lower().strip().replace('\n', ' ').replace('\r', ' ')
        # Consolidate whitespace
        hl = ' '.join(hl.split())

        # Total inflows
        if 'inflows total' in hl:
            col_map.setdefault('inflows_total', i)
        # Inflows (Sears Stores composite or retail composite in Cash Flow sheet)
        elif hl == 'inflows':
            col_map.setdefault('cf_inflows', i)
        # CITI Reimbursement
        elif ('citi' in hl and 'reimb' in hl) or 'citi reimbursement' in hl or 'citi reimburse' in hl:
            col_map.setdefault('citi_reimb', i)
        # Debt / Financing
        elif ('debt' in hl and 'financ' in hl):
            col_map.setdefault('debt_financ', i)
        # Supply Chain
        elif 'supply chain' in hl:
            col_map.setdefault('supply_chain', i)
        # Asset Sales (not "proceeds from asset sales" balance tracker)
        elif 'asset sales' in hl and 'proceeds' not in hl:
            col_map.setdefault('asset_sales', i)
        # Home Services
        elif 'home services' in hl:
            col_map.setdefault('home_svcs', i)
        # Misc Inflows
        elif hl == 'misc inflows':
            col_map.setdefault('misc', i)
        # HTS Reimbursement (older files)
        elif 'hts' in hl and 'reimb' in hl:
            col_map.setdefault('hts_reimb', i)

    return col_map


def extract_from_inflows_detail(ws, file_year_hint=None):
    """
    Extract monthly totals from the Inflows Detail sheet.
    Returns list of dicts, one per month-total row found.
    """
    results = []

    # Read header row (row 1, 0-indexed)
    all_rows = list(ws.iter_rows(min_row=1, max_row=2000, values_only=True))
    if not all_rows:
        return results

    # The header may be in row 1 or row 5 (some files repeat headers)
    # Find the canonical header row - look for 'Sears Stores' indicator
    header_row_idx = 0
    for idx in range(min(10, len(all_rows))):
        row = all_rows[idx]
        row_vals = [str(v).strip() if v else '' for v in row]
        if 'Sears Stores' in row_vals:
            header_row_idx = idx
            break

    headers = all_rows[header_row_idx]

    # Build column map for inflows detail categories
    # Standard positions (1-indexed in description, 0-indexed here):
    # Col 0: DATE label
    # Col 1: Sears Stores
    # Col 2: Kmart Stores
    # Col 3: Home Services
    # Then varies by era - use dynamic mapping

    id_col_map = {}
    for i, h in enumerate(headers):
        if not h or not isinstance(h, str):
            continue
        hl = h.lower().strip().replace('\n', ' ')
        hl = ' '.join(hl.split())

        if 'sears stores' in hl:
            id_col_map['sears'] = i
        elif 'kmart' in hl:
            id_col_map['kmart'] = i
        elif 'home services' in hl:
            id_col_map['home_svcs'] = i
        elif ('kcd wholesale' in hl) or ('kcd (royalty)' in hl) or (hl == 'kcd'):
            id_col_map.setdefault('kcd', i)
        elif 'supply chain' in hl:
            id_col_map['supply_chain'] = i
        elif ('citi' in hl and 'reimb' in hl) or 'citi reimbursement' in hl:
            id_col_map['citi_reimb'] = i
        elif 'asset sales' in hl:
            id_col_map['asset_sales'] = i
        elif 'hts' in hl and 'hts prefund' not in hl:
            id_col_map.setdefault('hts', i)
        elif 'hts prefund' in hl:
            id_col_map['hts_prefund'] = i
        elif 'costco' in hl:
            id_col_map['costco'] = i
        elif 'new debt' in hl or ('debt' in hl and 'financ' in hl):
            id_col_map['debt_financ'] = i
        elif 'misc inflows' in hl:
            id_col_map['misc'] = i
        elif 'daily total' in hl:
            id_col_map['daily_total'] = i
        elif 'tenant income' in hl:
            id_col_map['tenant'] = i
        elif 'cchs' in hl or 'cross country' in hl:
            id_col_map['cchs'] = i
        elif 'lease opco' in hl or 'propco' in hl:
            id_col_map.setdefault('propco', i)
        elif 'online' in hl and hl != 'online monetization':
            id_col_map.setdefault('online', i)
        elif 'rx' in hl:
            id_col_map['rx'] = i

    # Track year context for disambiguation
    current_year = file_year_hint
    last_month = None

    for row_idx, row in enumerate(all_rows):
        col0 = row[0]
        if col0 is None:
            continue

        # Check if this is a month total row
        parsed = None
        if isinstance(col0, str) and col0.strip():
            parsed = parse_month_label(col0.strip())

        if parsed is None:
            continue

        yr, mon = parsed

        # Resolve year if missing from label
        if yr is None:
            # Try to infer from context
            if current_year is not None:
                # If month jumped backward, it's likely next year
                if last_month is not None and mon < last_month and last_month >= 10:
                    current_year += 1
                yr = current_year
            else:
                yr = file_year_hint or 2022

        current_year = yr
        last_month = mon

        # Build the record
        def g(key):
            idx = id_col_map.get(key)
            if idx is None or idx >= len(row):
                return 0.0
            return safe_val(row[idx])

        # Total inflows - prefer daily_total column, else sum key categories
        total = g('daily_total')
        sears = g('sears')
        kmart = g('kmart')
        home_svcs = g('home_svcs')
        kcd = g('kcd')
        supply_chain = g('supply_chain')
        citi_reimb = g('citi_reimb')
        asset_sales = g('asset_sales')
        debt_financ = g('debt_financ')
        misc = g('misc')
        hts = g('hts')
        hts_prefund = g('hts_prefund')
        costco = g('costco')
        tenant = g('tenant')
        cchs = g('cchs')
        propco = g('propco')
        online = g('online')
        rx = g('rx')

        # If no daily_total column, compute total from parts
        if total == 0.0:
            total = (sears + kmart + home_svcs + kcd + supply_chain +
                     citi_reimb + asset_sales + debt_financ + misc +
                     hts + hts_prefund + costco + tenant + cchs +
                     propco + online + rx)

        if total == 0.0 and sears == 0.0:
            # Likely empty row
            continue

        record = {
            'year': yr,
            'month': mon,
            'date_label': f"{yr}-{mon:02d}",
            'source': 'Inflows Detail',
            'total_inflows': round(total, 4),
            'sears_stores': round(sears, 4),
            'kmart_stores': round(kmart, 4),
            'home_services': round(home_svcs, 4),
            'kcd_wholesale': round(kcd, 4),
            'supply_chain': round(supply_chain, 4),
            'citi_reimb': round(citi_reimb, 4),
            'asset_sales': round(asset_sales, 4),
            'debt_financing': round(debt_financ, 4),
            'misc_inflows': round(misc, 4),
            'hts': round(hts + hts_prefund, 4),
            'costco': round(costco, 4),
            'tenant_income': round(tenant, 4),
            'cchs': round(cchs, 4),
            'propco': round(propco, 4),
            'online': round(online, 4),
            'rx': round(rx, 4),
        }
        results.append(record)

    return results


def extract_from_cash_flow(ws, file_year_hint=None):
    """
    Fallback: extract monthly totals from the Cash Flow sheet.
    Uses dynamic column mapping by header label.
    """
    results = []

    all_rows = list(ws.iter_rows(min_row=1, max_row=2000, values_only=True))
    if len(all_rows) < 2:
        return results

    headers = all_rows[1]  # Row 2 = header labels (row 1 has numeric indices)
    col_map = build_col_map_by_header(headers)

    current_year = file_year_hint
    last_month = None

    for row_idx, row in enumerate(all_rows):
        col0 = row[0]
        if col0 is None:
            continue

        parsed = None
        if isinstance(col0, str) and col0.strip():
            parsed = parse_month_label(col0.strip())

        if parsed is None:
            continue

        yr, mon = parsed
        if yr is None:
            if current_year is not None:
                if last_month is not None and mon < last_month and last_month >= 10:
                    current_year += 1
                yr = current_year
            else:
                yr = file_year_hint or 2022

        current_year = yr
        last_month = mon

        def g(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return 0.0
            return safe_val(row[idx])

        total = g('inflows_total')
        # "cf_inflows" in Cash Flow is the main retail inflow bucket (Sears+Kmart+HTS+Costco combo)
        cf_inflows = g('cf_inflows')

        if total == 0.0 and cf_inflows == 0.0:
            continue

        record = {
            'year': yr,
            'month': mon,
            'date_label': f"{yr}-{mon:02d}",
            'source': 'Cash Flow',
            'total_inflows': round(total, 4),
            'sears_stores': 0.0,
            'kmart_stores': 0.0,
            'home_services': round(g('home_svcs'), 4),
            'kcd_wholesale': 0.0,
            'supply_chain': round(g('supply_chain'), 4),
            'citi_reimb': round(g('citi_reimb'), 4),
            'asset_sales': round(g('asset_sales'), 4),
            'debt_financing': round(g('debt_financ'), 4),
            'misc_inflows': round(g('misc'), 4),
            'hts': round(g('hts_reimb'), 4),
            'costco': 0.0,
            'tenant_income': 0.0,
            'cchs': 0.0,
            'propco': 0.0,
            'online': 0.0,
            'rx': 0.0,
        }
        results.append(record)

    return results


def get_file_year_hint(filename):
    """Extract year from filename like '9.19.22' -> 2022, '01.16.26' -> 2026."""
    # Look for patterns like 9.19.22 or 01.16.26
    m = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})', filename)
    if m:
        yr = int(m.group(3))
        if yr < 100:
            yr += 2000
        return yr
    # Fallback: look for 4-digit year
    m = re.search(r'20(\d{2})', filename)
    if m:
        return int('20' + m.group(1))
    return None


def extract_file(filepath):
    """Extract monthly inflows from a single file."""
    fname = os.path.basename(filepath)
    year_hint = get_file_year_hint(fname)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        print(f"  ERROR opening {fname}: {e}")
        return []

    results = []
    try:
        if 'Inflows Detail' in wb.sheetnames:
            ws = wb['Inflows Detail']
            results = extract_from_inflows_detail(ws, file_year_hint=year_hint)
            if not results and 'Cash Flow' in wb.sheetnames:
                ws = wb['Cash Flow']
                results = extract_from_cash_flow(ws, file_year_hint=year_hint)
        elif 'Cash Flow' in wb.sheetnames:
            ws = wb['Cash Flow']
            results = extract_from_cash_flow(ws, file_year_hint=year_hint)
        else:
            print(f"  SKIP {fname}: no Cash Flow or Inflows Detail sheet")
    except Exception as e:
        print(f"  ERROR extracting {fname}: {e}")
        import traceback
        traceback.print_exc()
    finally:
        wb.close()

    return results


def deduplicate_months(all_records):
    """
    For each (year, month) combination, keep only the record from the
    most-recent file (files are named by date, so sort by file date).
    Also track which file provided each month's actuals.
    """
    from collections import defaultdict
    month_records = defaultdict(list)
    for rec in all_records:
        key = (rec['year'], rec['month'])
        month_records[key].append(rec)

    final = {}
    for key, recs in month_records.items():
        # For the analysis, we prefer actuals from the earliest file that would
        # have had that month as actual (actuals are locked, forecasts vary).
        # Simple heuristic: prefer records with higher total (actuals tend to be
        # more complete). If totals are similar, use the one from the file whose
        # date is closest after month end.
        # Since we track source_file, sort by recency then pick first.
        # Actually just keep all, we'll average later – but for now, prefer max total.
        best = max(recs, key=lambda r: abs(r.get('total_inflows', 0)))
        final[key] = best

    return final


def main():
    files = sorted([
        f for f in os.listdir(BASE_DIR)
        if f.endswith('.xlsx') and f.startswith('Daily Cash')
    ])

    print(f"Processing {len(files)} files...\n")

    all_records = []
    file_contributions = {}

    for fname in files:
        fpath = os.path.join(BASE_DIR, fname)
        print(f"Processing: {fname}")
        recs = extract_file(fpath)
        print(f"  -> {len(recs)} month records extracted")

        for rec in recs:
            rec['source_file'] = fname
        all_records.extend(recs)
        file_contributions[fname] = len(recs)

    print(f"\nTotal raw records: {len(all_records)}")

    # Filter to Sep 2022 - May 2026 range
    filtered = [
        r for r in all_records
        if (r['year'] > 2022 or (r['year'] == 2022 and r['month'] >= 9))
        and (r['year'] < 2026 or (r['year'] == 2026 and r['month'] <= 5))
    ]
    print(f"Records in Sep 2022 – May 2026 window: {len(filtered)}")

    # Deduplicate - for each month, pick the best record
    # Strategy: group by (year, month), prefer actuals (file date close to or after month end)
    from collections import defaultdict
    month_groups = defaultdict(list)
    for rec in filtered:
        key = (rec['year'], rec['month'])
        month_groups[key].append(rec)

    # For deduplication:
    # - If we have multiple records for the same month, prefer the one from the
    #   BOD file closest to that month (i.e., actuals most likely locked in)
    # - Use the record with the lowest file date that is >= month end
    def file_date_score(rec):
        fname = rec.get('source_file', '')
        m = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})', fname)
        if m:
            mo, da, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if yr < 100:
                yr += 2000
            try:
                return datetime(yr, mo, da)
            except:
                pass
        return datetime(2099, 1, 1)

    deduplicated = {}
    for key, recs in month_groups.items():
        yr, mon = key
        # Find the file with the earliest date that is AFTER or AT month end
        # (i.e., the first BOD meeting that could have confirmed actuals for this month)
        month_end = datetime(yr, mon, 28)  # approximate month end

        # Sort by file date
        sorted_recs = sorted(recs, key=lambda r: file_date_score(r))

        # Find first file dated at or after month end
        after_month = [r for r in sorted_recs if file_date_score(r) >= month_end]
        if after_month:
            # Earliest file after month end = most likely to have actuals
            deduplicated[key] = after_month[0]
        else:
            # All files predate month end = these are forecasts - take the last one (most recent)
            deduplicated[key] = sorted_recs[-1]

    print(f"Deduplicated months: {len(deduplicated)}")

    # Build chronological list
    sorted_months = sorted(deduplicated.keys())
    monthly_data = []
    for key in sorted_months:
        monthly_data.append(deduplicated[key])

    # Compute summary statistics
    total_inflows_series = [r['total_inflows'] for r in monthly_data]

    peak_months = sorted(monthly_data, key=lambda r: r['total_inflows'], reverse=True)[:10]

    # Category trends
    categories = ['sears_stores', 'kmart_stores', 'home_services', 'kcd_wholesale',
                  'supply_chain', 'citi_reimb', 'asset_sales', 'debt_financing',
                  'misc_inflows', 'hts', 'costco', 'tenant_income', 'cchs',
                  'propco', 'online', 'rx']

    # Sep 2022 vs May 2026
    sep_2022 = deduplicated.get((2022, 9))
    may_2026 = deduplicated.get((2026, 5))

    # Asset sales first meaningful month (>$5M)
    asset_sales_significant = [
        r for r in monthly_data
        if r.get('asset_sales', 0) > 5.0
    ]
    asset_sales_first = asset_sales_significant[0] if asset_sales_significant else None

    # Category averages - compare first 12 months vs last 12 months
    first_12 = monthly_data[:12] if len(monthly_data) >= 12 else monthly_data
    last_12 = monthly_data[-12:] if len(monthly_data) >= 12 else monthly_data

    def avg_cat(records, cat):
        vals = [r.get(cat, 0) for r in records]
        return round(sum(vals) / len(vals), 4) if vals else 0.0

    category_trends = {}
    for cat in categories:
        early_avg = avg_cat(first_12, cat)
        late_avg = avg_cat(last_12, cat)
        direction = 'grew' if late_avg > early_avg else ('declined' if late_avg < early_avg else 'flat')
        category_trends[cat] = {
            'first_12mo_avg': early_avg,
            'last_12mo_avg': late_avg,
            'change_direction': direction,
            'change_pct': round((late_avg - early_avg) / early_avg * 100, 1) if early_avg != 0 else None
        }

    # Build output
    output = {
        'metadata': {
            'generated': datetime.now().isoformat(),
            'period': 'Sep 2022 – May 2026',
            'entity': 'Transform SR Holding Management LLC',
            'total_months': len(monthly_data),
            'files_processed': len(files),
            'note': 'All figures in $M. Monthly totals from Inflows Detail sheet (primary) or Cash Flow sheet (fallback). Deduplication: earliest file dated at/after month-end used as actuals; else most-recent forecast.'
        },
        'insights': {
            'sep_2022_vs_may_2026': {
                'sep_2022': {
                    'total_inflows': sep_2022.get('total_inflows') if sep_2022 else None,
                    'source_file': sep_2022.get('source_file') if sep_2022 else None,
                },
                'may_2026': {
                    'total_inflows': may_2026.get('total_inflows') if may_2026 else None,
                    'source_file': may_2026.get('source_file') if may_2026 else None,
                }
            },
            'peak_inflow_months': [
                {
                    'date': r['date_label'],
                    'total_inflows': r['total_inflows'],
                    'source_file': r.get('source_file', '')[-45:]
                }
                for r in peak_months
            ],
            'asset_sales_first_significant_month': {
                'date': asset_sales_first['date_label'] if asset_sales_first else None,
                'amount': asset_sales_first.get('asset_sales') if asset_sales_first else None,
                'threshold_used': '$5M'
            },
            'category_trends': category_trends,
            'overall_trend': {
                'first_12mo_avg_total': round(sum(r['total_inflows'] for r in first_12) / len(first_12), 2) if first_12 else None,
                'last_12mo_avg_total': round(sum(r['total_inflows'] for r in last_12) / len(last_12), 2) if last_12 else None,
                'all_time_max': round(max(total_inflows_series), 2) if total_inflows_series else None,
                'all_time_min': round(min(total_inflows_series), 2) if total_inflows_series else None,
                'all_time_avg': round(sum(total_inflows_series) / len(total_inflows_series), 2) if total_inflows_series else None,
            }
        },
        'monthly_data': monthly_data
    }

    out_path = os.path.join(BASE_DIR, 'analysis_inflows_trend.json')
    with open(out_path, 'w') as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\nSaved to: {out_path}")
    print(f"\n=== KEY FINDINGS ===")
    print(f"Total months with data: {len(monthly_data)}")

    if sep_2022:
        print(f"Sep 2022 Total Inflows: ${sep_2022['total_inflows']:.1f}M")
        print(f"  Sears: ${sep_2022['sears_stores']:.1f}M | Kmart: ${sep_2022['kmart_stores']:.1f}M | HomeSvcs: ${sep_2022['home_services']:.1f}M")
        print(f"  AssetSales: ${sep_2022['asset_sales']:.1f}M | SC: ${sep_2022['supply_chain']:.1f}M")
    else:
        print("Sep 2022: NO DATA")

    if may_2026:
        print(f"\nMay 2026 Total Inflows: ${may_2026['total_inflows']:.1f}M")
        print(f"  Sears: ${may_2026['sears_stores']:.1f}M | Kmart: ${may_2026['kmart_stores']:.1f}M | HomeSvcs: ${may_2026['home_services']:.1f}M")
        print(f"  AssetSales: ${may_2026['asset_sales']:.1f}M | SC: ${may_2026['supply_chain']:.1f}M")
    else:
        print("\nMay 2026: NO DATA")

    if asset_sales_first:
        print(f"\nAsset Sales first >$5M month: {asset_sales_first['date_label']} (${asset_sales_first['asset_sales']:.1f}M)")

    print(f"\nPeak inflow months:")
    for r in peak_months[:5]:
        print(f"  {r['date_label']}: ${r['total_inflows']:.1f}M")

    print(f"\nCategory trends (first 12mo avg -> last 12mo avg):")
    for cat, trend in category_trends.items():
        if trend['first_12mo_avg'] > 0.1 or trend['last_12mo_avg'] > 0.1:
            pct = f"{trend['change_pct']:+.0f}%" if trend['change_pct'] is not None else 'N/A'
            print(f"  {cat}: ${trend['first_12mo_avg']:.1f}M -> ${trend['last_12mo_avg']:.1f}M ({pct})")

    return output


if __name__ == '__main__':
    main()
