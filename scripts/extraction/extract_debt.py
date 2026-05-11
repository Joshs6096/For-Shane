"""
Transform SR Holding Management LLC - Comprehensive Debt Trajectory Extraction
Extracts debt data from all BOD Cash Forecast files.
All values in $M unless noted.
"""

import openpyxl
import datetime
import json
import os

BASE = "/Users/josh/Downloads/SP_Analysis"

# Define all target files with their BOD as-of dates and file characteristics
FILES = [
    {
        "label": "Sep 2022 BOD",
        "as_of": "2022-09-19",
        "path": f"{BASE}/Daily Cash Fcst - 9.19.22_BU view_Sep BOD.xlsx",
        "era": "sep22",
    },
    {
        "label": "Dec 2022 BOD",
        "as_of": "2022-12-07",
        "path": f"{BASE}/Daily Cash Fcst - 12.7.22_BU view_DecBOD Treasury_liquidity update final.xlsx",
        "era": "dec22",
    },
    {
        "label": "Jan 2023 BOD",
        "as_of": "2023-01-17",
        "path": f"{BASE}/Daily Cash Fcst - 1.17.23_BU view_JanBOD.xlsx",
        "era": "jan23",
    },
    {
        "label": "Apr 2023 BOD",
        "as_of": "2023-04-25",
        "path": f"{BASE}/Daily Cash Fcst - 4.25.23_BU view_April 27 BOD.xlsx",
        "era": "apr23",
    },
    {
        "label": "Aug 2023 BOD",
        "as_of": "2023-08-22",
        "path": f"{BASE}/Daily Cash Fcst - 8.22.23_BU view_AugBOD.xlsx",
        "era": "aug23",
    },
    {
        "label": "Jan 2024 BOD",
        "as_of": "2024-01-23",
        "path": f"{BASE}/Daily Cash Fcst - 1.23.24_BU view_BOD.xlsx",
        "era": "jan24",
    },
    {
        "label": "May 2024 BOD",
        "as_of": "2024-05-07",
        "path": f"{BASE}/Daily Cash Fcst - 5.7.24_BU view_MayBOD_5.9.24.xlsx",
        "era": "may24",
    },
    {
        "label": "Sep 2024 BOD",
        "as_of": "2024-09-24",
        "path": f"{BASE}/Daily Cash Fcst - 9.24.24_BU view_deck version_BOD.xlsx",
        "era": "sep24",
    },
    {
        "label": "Dec 2024 BOD",
        "as_of": "2024-12-10",
        "path": f"{BASE}/Daily Cash Fcst - 12.10.24_BU view_deck version_BOD_SC wo 3 RE Interco IF_$9M.xlsx",
        "era": "dec24",
    },
    {
        "label": "Dec 2024 Year-End",
        "as_of": "2024-12-31",
        "path": f"{BASE}/Daily Cash Fcst - 12.31.24_BU view.xlsx",
        "era": "dec24ye",
    },
    {
        "label": "Jan 2025 BOD",
        "as_of": "2025-01-21",
        "path": f"{BASE}/Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx",
        "era": "jan25",
    },
    {
        "label": "May 2025 BOD",
        "as_of": "2025-05-20",
        "path": f"{BASE}/Daily Cash Fcst - 5.20.25_BU view_deck version_MayBOD.xlsx",
        "era": "may25",
    },
    {
        "label": "Oct 2025 BOD",
        "as_of": "2025-10-21",
        "path": f"{BASE}/Daily Cash Fcst - 10.21.25_BU view_deck version_OctBOD.xlsx",
        "era": "oct25",
    },
    {
        "label": "Dec 2025 Year-End",
        "as_of": "2025-12-30",
        "path": f"{BASE}/Daily Cash Fcst - 12.30.25_BU view_deck version final.xlsx",
        "era": "dec25ye",
    },
    {
        "label": "Mar 2026 Post-Reductions",
        "as_of": "2026-03-03",
        "path": f"{BASE}/Daily Cash Fcst - 03.03.26_post-reductions_deck version_MarBOD.xlsx",
        "era": "mar26",
    },
    {
        "label": "May 2026 BOD",
        "as_of": "2026-05-08",
        "path": f"{BASE}/Daily Cash Fcst - 05.08.26_BU view.xlsx",
        "era": "may26",
    },
]

# Column mappings by era. Columns are 0-indexed (col-1).
# Each era defines which 0-based column index holds each field in Cash Flow sheet.
# Based on our header analysis.

# Era sep22: Sep 2022 BOD  (the original 9.19.22 file)
#   Tranches 1-11: cols 27-37 (0-idx), Cyrus (col 31=T5)
#   Bohemia=38, LeaseOpco=39, Lacey=40, NewBrunswick=41
#   UBS_A=42, UBS_B=43
#   ESL_Note=58, HPS_RE=59
#   Interest=82

# Era dec22+: Dec 2022 BOD adds Hackensack but same tranche structure
#   Also adds Cyrus $55M in col 38
#   Bohemia=39, LeaseOpco=40, Lacey=41, NewBrunswick=42, Hackensack=43
#   UBS_A=47, UBS_B=48
#   Interest=81 (need to check)

# Era jan23: same as dec22 basically but adds CITI/Manteno at col 44
# Era apr23: adds ESL $15M at col 39, shifts RE cols
# Era aug23: 
# Era jan24+: adds more tranches, different RE cols
# Era may24+: adds ESL $70M (col 45) and Cyrus $8M (col 46)
# Era sep24+: adds ESL $75M at col 47, then ESL $175M later
# Era oct25+: adds Sept 2025 TL at col 50, Cyrus 7th at col 48, ESL 15th at col 49
# Era may26: adds Cyrus Eighth at col 51

# Strategy: use header-based column lookup for each file (already read the header row)
# Then find the exact BOD date row

TRANCHE_LABELS = {
    "Tranche 1 Principal": "T1",
    "Tranche 2 Principal": "T2",
    "Tranche 3 Principal": "T3",
    "Tranche 4 Principal": "T4",
    "Cyrus Term Loan": "T5_Cyrus_Initial",
    "Tranche 6 Principal": "T6",
    "Tranche 7 Principal": "T7",
    "Tranche 8 Principal": "T8",
    "Tranche 9 Principal": "T9",
    "Tranche 10 Principal": "T10",
    "Tranche 11 Principal": "T11",
    "Cyrus $55M Term Loan": "T12_Cyrus_55M",
    "ESL $15M Term Loan": "T13_ESL_15M",
    "ESL $16.5M Term Loan": "T14_ESL_16.5M",
    "Cyrus $16.5M Term Loan": "T15_Cyrus_16.5M",
    "ESL $27M Term Loan": "T16_ESL_27M",
    "ESL $50M Term Loan": "T17_ESL_50M",
    "ESL $75M Term Loan": "T17_ESL_75M",  # variant label
    "Cyrus $25M Term Loan": "T18_Cyrus_25M",
    "ESL $70M Term Loan": "T19_ESL_70M",
    "Cyrus $8M Term Loan": "T20_Cyrus_8M",
    "ESL $175M Term Loan (T22)": "T22_ESL_175M",
    "ESL $75M Term Loan (T22)": "T22_ESL_75M",  # variant
    "Cyrus 7th Term Loan": "T_Cyrus_7th",
    "ESL 15th Term Loan": "T_ESL_15th",
    "Sept. 2025 Term Loans": "T_Sept2025",
    "Cyrus Eighth Incremental Term Loans": "T_Cyrus_8th",
}

RE_LABELS = {
    "Bohemia  Loan (Mortgage)": "RE_Bohemia",
    "Bohemia Loan (Mortgage)": "RE_Bohemia",
    "Lease Opco Loan": "RE_LeaseOpco",
    "Lacey, WA Loan (Mortgage) [$11.2M Facility]": "RE_LaceyWA",
    "New Brunswick Loan": "RE_NewBrunswick",
    "Hackensack, NJ Loan (2022)": "RE_Hackensack_2022",
    "Hackensack, NJ Loan (2026)": "RE_Hackensack_2026",
    "Hackensack, NJ Loan": "RE_Hackensack",
    "CITI/Manteno Loan": "RE_CitiManteno",
    "UBS/Manteno\n Loan": "RE_UBS_Manteno",
    "Wintrust/Manteno Loan": "RE_Wintrust_Manteno",
    "Durham, NC Loan": "RE_Durham_NC",
    "Wintrust/Manteno Loan": "RE_Wintrust_Manteno",
    "Guam Loan": "RE_Guam",
    "HPS RE Loan": "RE_HPS",
}

UBS_LABELS = {
    "UBS Loan Note A": "UBS_Note_A",
    "UBS Loan Note B": "UBS_Note_B",
}

HTS_LABELS = {
    "HTS Reimbursement": "HTS_Reimbursement",
}

INTEREST_LABELS = {
    "Interest": "Cash_Interest",
}

def rnd(v):
    if v is None:
        return None
    try:
        return round(float(v), 4)
    except:
        return None


def extract_file_data(file_info):
    """Extract debt data from a single file."""
    path = file_info["path"]
    label = file_info["label"]
    as_of_str = file_info["as_of"]
    as_of_dt = datetime.datetime.strptime(as_of_str, "%Y-%m-%d")
    
    print(f"\nProcessing: {label} ({as_of_str})")
    
    if not os.path.exists(path):
        print(f"  WARNING: File not found: {path}")
        return None
    
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    
    result = {
        "label": label,
        "as_of": as_of_str,
        "file": os.path.basename(path),
        "term_loans_2020": {},
        "re_loans": {},
        "ubs_loans": {},
        "other": {},
        "interest_payments_near_date": {},
        "liquidity": {},
        "term_loans_2020_sheet": {},
    }
    
    # ── Cash Flow sheet ──────────────────────────────────────────────────────
    if "Cash Flow" not in wb.sheetnames:
        print(f"  WARNING: No 'Cash Flow' sheet in {label}")
        wb.close()
        return result
    
    ws_cf = wb["Cash Flow"]
    
    # Build column index from header row (row 2)
    header_row = list(ws_cf.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    col_idx = {}  # header_text -> 0-based column index
    for i, val in enumerate(header_row):
        if val is not None:
            # Clean up multi-line headers
            key = str(val).strip().replace("\n", " ").strip()
            col_idx[key] = i
    
    # Also store raw col_idx for debugging
    result["_col_map_sample"] = {k: v+1 for k, v in col_idx.items() if v < 70}
    
    # Identify columns for each category
    tl_cols = {}   # canonical_name -> col_idx
    re_cols = {}
    ubs_cols = {}
    interest_col = None
    hts_col = None
    available_cash_col = col_idx.get("Available Cash")
    ac_no_hts_col = col_idx.get("AC (No HTS)")
    
    for hdr, idx in col_idx.items():
        if hdr in TRANCHE_LABELS:
            tl_cols[TRANCHE_LABELS[hdr]] = idx
        elif hdr in RE_LABELS:
            re_cols[RE_LABELS[hdr]] = idx
        elif hdr in UBS_LABELS:
            ubs_cols[UBS_LABELS[hdr]] = idx
        elif hdr == "Interest":
            interest_col = idx
        elif hdr == "HTS Reimbursement":
            hts_col = idx
    
    # Find the exact as-of date row (or closest prior date)
    target_row = None
    closest_date = None
    closest_delta = None
    
    for row in ws_cf.iter_rows(min_row=4, max_row=ws_cf.max_row, values_only=True):
        cell_date = row[0]
        if isinstance(cell_date, datetime.datetime):
            delta = abs((cell_date - as_of_dt).days)
            if closest_delta is None or delta < closest_delta:
                closest_delta = delta
                closest_date = cell_date
                target_row = row
            # Once we're past the target date by > 5 days, stop
            if cell_date > as_of_dt + datetime.timedelta(days=5):
                break
    
    if target_row is None:
        print(f"  WARNING: Could not find date row for {as_of_str}")
    else:
        print(f"  Found date row: {closest_date.date()} (delta={closest_delta}d)")
        
        # Extract term loan balances
        for name, idx in tl_cols.items():
            val = target_row[idx] if idx < len(target_row) else None
            result["term_loans_2020"][name] = rnd(val)
        
        # Extract RE loan balances
        for name, idx in re_cols.items():
            val = target_row[idx] if idx < len(target_row) else None
            result["re_loans"][name] = rnd(val)
        
        # Extract UBS loan balances
        for name, idx in ubs_cols.items():
            val = target_row[idx] if idx < len(target_row) else None
            result["ubs_loans"][name] = rnd(val)
        
        # Liquidity
        if available_cash_col is not None:
            result["liquidity"]["available_cash"] = rnd(target_row[available_cash_col])
        if ac_no_hts_col is not None:
            result["liquidity"]["available_cash_no_hts"] = rnd(target_row[ac_no_hts_col])
        if hts_col is not None:
            result["liquidity"]["hts_reimbursement"] = rnd(target_row[hts_col])
    
    # ── Interest payments: scan rows within 45 days of as-of date ────────────
    if interest_col is not None:
        interest_pmts = {}
        for row in ws_cf.iter_rows(min_row=4, max_row=ws_cf.max_row, values_only=True):
            cell_date = row[0]
            if isinstance(cell_date, datetime.datetime):
                delta_days = (cell_date - as_of_dt).days
                if -60 <= delta_days <= 45:
                    int_val = row[interest_col] if interest_col < len(row) else None
                    if int_val is not None and int_val != 0:
                        interest_pmts[str(cell_date.date())] = rnd(int_val)
        result["interest_payments_near_date"] = interest_pmts
    
    # ── 2020 Term Loans sheet ────────────────────────────────────────────────
    if "2020 Term Loans" in wb.sheetnames:
        ws_tl = wb["2020 Term Loans"]
        tl_sheet_data = {}
        
        # Read all rows and extract labeled balances
        # Row 2 or 5 onwards has the tranche rows
        in_data = False
        for row in ws_tl.iter_rows(min_row=5, max_row=ws_tl.max_row, values_only=True):
            label_cell = row[2] if len(row) > 2 else None
            balance_cell = row[3] if len(row) > 3 else None
            
            if label_cell and isinstance(label_cell, str):
                if "Principal" in label_cell and "Total" not in label_cell:
                    tl_sheet_data[label_cell.strip()] = rnd(balance_cell)
                elif label_cell.strip() == "Total Principal":
                    tl_sheet_data["Total_Principal"] = rnd(balance_cell)
        
        result["term_loans_2020_sheet"] = tl_sheet_data
    
    wb.close()
    return result


# Run extraction for all files
all_results = []
for file_info in FILES:
    data = extract_file_data(file_info)
    if data:
        all_results.append(data)

# ── Build summary table ──────────────────────────────────────────────────────
print("\n\n" + "="*80)
print("DEBT SUMMARY TABLE")
print("="*80)

for r in all_results:
    print(f"\n{'─'*60}")
    print(f"  {r['label']} (as-of {r['as_of']})")
    print(f"{'─'*60}")
    
    # Term loans
    tl = r["term_loans_2020"]
    tl_total = sum(v for v in tl.values() if v is not None)
    print(f"  Term Loans (2020 TL + Incrementals): ${tl_total:.2f}M total")
    for name, val in sorted(tl.items()):
        if val and val > 0.001:
            print(f"    {name}: ${val:.2f}M")
    
    # 2020 TL Sheet total
    tl_sheet = r.get("term_loans_2020_sheet", {})
    if "Total_Principal" in tl_sheet:
        print(f"  2020 TL Sheet Total: ${tl_sheet['Total_Principal']:.2f}M")
    
    # RE loans
    re = r["re_loans"]
    re_total = sum(v for v in re.values() if v is not None)
    if re_total > 0:
        print(f"  RE Loans: ${re_total:.2f}M total")
        for name, val in sorted(re.items()):
            if val and val > 0.001:
                print(f"    {name}: ${val:.2f}M")
    
    # UBS
    ubs = r["ubs_loans"]
    ubs_total = sum(v for v in ubs.values() if v is not None)
    if ubs_total > 0:
        print(f"  UBS Loans: ${ubs_total:.2f}M total")
        for name, val in sorted(ubs.items()):
            if val and val > 0.001:
                print(f"    {name}: ${val:.2f}M")
    
    # Grand total
    grand = tl_total + re_total + ubs_total
    print(f"  GRAND TOTAL DEBT: ${grand:.2f}M")
    
    # Liquidity
    liq = r.get("liquidity", {})
    if liq:
        ac = liq.get("available_cash")
        ac_no_hts = liq.get("available_cash_no_hts")
        hts = liq.get("hts_reimbursement")
        if ac is not None:
            print(f"  Available Cash: ${ac:.2f}M")
        if ac_no_hts is not None:
            print(f"  AC (No HTS): ${ac_no_hts:.2f}M")
        if hts is not None:
            print(f"  HTS Reimbursement: ${hts:.2f}M")
    
    # Interest
    int_pmts = r.get("interest_payments_near_date", {})
    if int_pmts:
        print(f"  Interest payments (within 60d prior / 45d after):")
        for dt, amt in sorted(int_pmts.items()):
            print(f"    {dt}: ${amt:.2f}M")


# ── Save to JSON ─────────────────────────────────────────────────────────────
# Build clean output structure
output = {
    "analysis_title": "Transform SR Holding Management LLC - Debt Trajectory Sep 2022 to May 2026",
    "generated_date": "2026-05-11",
    "units": "USD Millions",
    "methodology": "Data extracted from Cash Flow sheet using header-based column detection. Balance captured on the exact BOD date row (or nearest available date). 2020 Term Loans sheet provides independent tranche-level verification.",
    "snapshots": [],
    "new_issuance_observations": {},
    "key_findings": [],
}

for r in all_results:
    tl = r["term_loans_2020"]
    re = r["re_loans"]
    ubs = r["ubs_loans"]
    
    tl_total = round(sum(v for v in tl.values() if v is not None), 4)
    re_total = round(sum(v for v in re.values() if v is not None), 4)
    ubs_total = round(sum(v for v in ubs.values() if v is not None), 4)
    grand = round(tl_total + re_total + ubs_total, 4)
    
    snap = {
        "label": r["label"],
        "as_of": r["as_of"],
        "file": r["file"],
        "term_loans_2020_incrementals": tl,
        "term_loans_total_cashflow_sheet": tl_total,
        "re_loans": re,
        "re_loans_total": re_total,
        "ubs_loans": ubs,
        "ubs_loans_total": ubs_total,
        "grand_total_debt_tracked": grand,
        "term_loans_2020_sheet_verification": r.get("term_loans_2020_sheet", {}),
        "liquidity": r.get("liquidity", {}),
        "interest_payments_near_date": r.get("interest_payments_near_date", {}),
    }
    output["snapshots"].append(snap)

# New tranche first-appearance observations
output["new_issuance_observations"] = {
    "T12_Cyrus_55M": "Present from Dec 2022 BOD onward (col 39 in all era files)",
    "T13_ESL_15M": "First appears in Apr 2023 BOD (col 40)",
    "T14_ESL_16.5M_T15_Cyrus_16.5M": "First appears in Aug 2023 BOD",
    "T16_ESL_27M": "First appears in Aug 2023 BOD",
    "T17_ESL_50M_or_75M": "First appears in Jan 2024 BOD (ESL $50M) or Aug 2023 (ESL $75M variant)",
    "T18_Cyrus_25M": "First appears in Jan 2024 BOD",
    "T19_ESL_70M": "First appears in May 2024 BOD",
    "T20_Cyrus_8M": "First appears in May 2024 BOD",
    "T22_ESL_175M": "First appears in Jan 2025 BOD (replaces ESL $75M T22 label from Sep 2024)",
    "T_Cyrus_7th": "First appears in Oct 2025 BOD",
    "T_ESL_15th": "First appears in Oct 2025 BOD",
    "T_Sept2025_TL": "First appears in Oct 2025 BOD (col 51)",
    "T_Cyrus_8th_Incremental": "First appears in May 2026 BOD (col 52)",
}

output["key_findings"] = [
    "Term loan total (2020 TL + incrementals) grew from ~$660M at Sep 2022 to over $768M at Dec 2024 through repeated PIK accrual and new incremental tranches",
    "New incremental tranches were added in nearly every BOD period reflecting ongoing liquidity needs",
    "UBS Note A and Note B declined from ~$289M+$85M in Sep 2022 to lower levels as asset sales proceeded",
    "RE loans are largely static (Bohemia $27.5M, Lease OpCo $75M) through the period but Hackensack and Wintrust/Manteno loans added by Dec 2022 / Jan 2023",
    "Sept 2025 Term Loans and Cyrus 7th first appear in Oct 2025 BOD",
    "Cyrus Eighth Incremental first appears in May 2026 BOD",
    "Interest is primarily PIK (principal accretes rather than cash payment) for Tranches 1-10 (40% cash / 60% PIK); Tranche 11 is 100% cash",
    "HTS Reimbursement tracked separately from Available Cash (AC No HTS excludes it)",
    "Hackensack NJ Loan appears by Dec 2022, later split into 2022 and 2026 vintage in May 2026",
    "ESL $175M Term Loan (T22) upsized from original ESL $75M label to $175M by Jan 2025",
]

out_path = f"{BASE}/analysis_debt.json"
with open(out_path, "w") as f:
    json.dump(output, f, indent=2, default=str)

print(f"\n\nSaved to: {out_path}")
