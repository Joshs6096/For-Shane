"""
Full SC Interco sheet from Dec 2024 file + payroll/bens from Cash Flow sheets.
"""
import openpyxl
from pathlib import Path

BASE = Path("/Users/josh/Downloads/SP_Analysis")

# 1. Full SC Interco sheet
print("="*60)
print("Dec 2024 - SC Interco sheet (full)")
wb = openpyxl.load_workbook(BASE / "Daily Cash Fcst - 12.10.24_BU view_deck version_BOD_SC wo 3 RE Interco IF_$9M.xlsx", data_only=True, read_only=True)
ws = wb["SC Interco"]
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), 1):
    non_none = [v for v in row if v is not None]
    if non_none:
        print(f"  R{i}: {[str(v)[:30] for v in non_none[:10]]}")
wb.close()

# 2. Payroll/Bens from Cash Flow sheets - all 9 files
FILES = [
    ("2024-09-24", "Daily Cash Fcst - 9.24.24_BU view_deck version_BOD.xlsx"),
    ("2024-10-15", "Daily Cash Fcst - 10.15.24_BU view_deck version_OctBOD.xlsx"),
    ("2024-11-12", "Daily Cash Fcst - 11.12.24_BU view_deck version_NovBOD.xlsx"),
    ("2024-12-10", "Daily Cash Fcst - 12.10.24_BU view_deck version_BOD_SC wo 3 RE Interco IF_$9M.xlsx"),
    ("2024-12-31", "Daily Cash Fcst - 12.31.24_BU view.xlsx"),
    ("2025-01-21", "Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx"),
    ("2025-03-18", "Daily Cash Fcst - 3.18.25_BU view_deck version_MarBOD.xlsx"),
    ("2025-04-18", "Daily Cash Fcst - 4.18.25_BU view_prelim AprBOD.xlsx"),
    ("2026-05-08", "Daily Cash Fcst - 05.08.26_BU view.xlsx"),
]

print("\n\n" + "="*60)
print("Payroll/Bens from Cash Flow sheets (monthly summary rows)")
print("="*60)

for date_str, filename in FILES:
    filepath = BASE / filename
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    payroll_data = {}

    # Look in Cash Flow sheet
    for sheet_name in ["Cash Flow", "Net Cash Flow", "SHS Cash Changes 5.1.26", "SHS Cash Changes 8.28",
                       "B&M Anita 8.2.24 IF & merch", "B&M Anita 10.1.24"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), 1):
            if row is None: continue
            texts = [str(v).lower().strip() for v in row[:10] if v is not None and isinstance(v, str)]
            row_text = " ".join(texts)

            # Monthly payroll totals (look for monthly summary rows)
            if any(kw in row_text for kw in ["pyrl/bens", "payroll/bens", "payroll & bens total", "payroll total", "total payroll"]):
                # Get numeric values - these should be monthly totals
                nums = [v for v in row if v is not None and isinstance(v, (int, float))]
                if nums and any(abs(n) > 1 for n in nums):
                    key = f"{sheet_name}_R{i}"
                    payroll_data[key] = {
                        "label": " | ".join(str(v) for v in row[:5] if v is not None)[:100],
                        "values": [round(float(n), 3) for n in nums[:20]]
                    }

            # Also look for PYRL/BENS TOTAL monthly row
            if "pyrl/bens total" in row_text or "payroll/bens total" in row_text:
                nums = [v for v in row if v is not None and isinstance(v, (int, float))]
                if nums:
                    payroll_data[f"PYRL_TOTAL_{sheet_name}_R{i}"] = {
                        "label": row_text[:80],
                        "values": [round(float(n), 3) for n in nums[:20]]
                    }

    print(f"\n{date_str} | {filename[:45]}")
    if payroll_data:
        for k, v in payroll_data.items():
            print(f"  [{k}]: {v['label'][:60]}")
            print(f"    vals: {v['values'][:12]}")
    else:
        print("  No payroll row found in standard sheets")

    wb.close()

# 3. Available Cash / Total Cash from Cash Flow monthly summary - for all files
print("\n\n" + "="*60)
print("Available Cash / Total Cash monthly rows from Cash Flow sheet")
print("="*60)
for date_str, filename in FILES:
    filepath = BASE / filename
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    for sheet_name in ["Cash Flow", "Net Cash Flow"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        found = False
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), 1):
            if row is None: continue
            texts = [str(v).lower().strip() for v in row[:10] if v is not None and isinstance(v, str)]
            row_text = " ".join(texts)
            if any(kw in row_text for kw in ["available cash", "total cash", "unavailable"]):
                nums = [v for v in row if v is not None and isinstance(v, (int, float))]
                print(f"{date_str} | {sheet_name} R{i}: {row_text[:60]}")
                print(f"  vals: {[round(float(n),3) for n in nums[:15]]}")
                found = True

    wb.close()
