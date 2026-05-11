"""
Retail Store Analysis: Sears and Kmart inflows across all BOD files
Extracts longitudinal data from /Users/josh/Downloads/SP_Analysis/
"""

import os
import json
import openpyxl
from collections import defaultdict

BASE_DIR = "/Users/josh/Downloads/SP_Analysis"

# Map of BOD period label -> filename (partial match)
BOD_FILES = [
    ("9.19.22_Sep BOD",   "9.19.22_BU view_Sep BOD"),
    ("12.7.22_DecBOD",    "12.7.22_BU view_DecBOD"),
    ("1.17.23_JanBOD",    "1.17.23_BU view_JanBOD"),
    ("4.25.23_AprBOD",    "4.25.23_BU view_April 27 BOD"),
    ("7.25.23_JulBOD",    "7.25.23_BU view_JulBOD"),
    ("11.14.23_NovBOD",   "11.14.23_BU view_adUPDATE"),
    ("2.20.24_BOD",       "2.20.24_BU view_BOD"),
    ("6.18.24_JunBOD",    "6.18.24_BU view_JunBOD"),
    ("10.15.24_OctBOD",   "10.15.24_BU view_deck version_OctBOD"),
    ("12.31.24_BU view",  "12.31.24_BU view"),
    ("3.18.25_MarBOD",    "3.18.25_BU view_deck version_MarBOD"),
    ("7.22.25_JulBOD",    "7.22.25_BU view_deck version_condensed_JulBOD"),
    ("12.15.25_DecBOD",   "12.15.25_BU view_DecBOD"),
    ("01.16.26_JanBOD",   "01.16.26_BU view new format_asset sales_JanBOD"),
    ("04.21.26_AprBOD",   "04.21.26_ BU version_deck version_AprBOD"),
    ("05.08.26_BU view",  "05.08.26_BU view"),
]

# Keywords to look for in headers
SEARS_KEYWORDS = ["sears store", "sears retail", "sears", "s-retail", "s retail"]
KMART_KEYWORDS = ["kmart", "k-mart", "k mart"]
TERRITORY_KEYWORDS = ["guam", "usvi", "u.s.v.i", "puerto rico", "virgin island", "international", "territory"]
INFLOW_SECTION_KEYWORDS = ["inflow", "receipt", "revenue", "cash in"]
STORE_COUNT_KEYWORDS = ["store count", "# stores", "number of store", "operating store", "open store"]


def find_file(partial_name):
    """Find a file in BASE_DIR whose name contains partial_name."""
    for fn in os.listdir(BASE_DIR):
        if partial_name.lower() in fn.lower() and fn.endswith(".xlsx"):
            return os.path.join(BASE_DIR, fn)
    return None


def get_cell_value(ws, row, col):
    """Safely get a cell value."""
    try:
        cell = ws.cell(row=row, column=col)
        return cell.value
    except Exception:
        return None


def normalize(text):
    """Lowercase and strip for comparison."""
    if text is None:
        return ""
    return str(text).lower().strip()


def find_col_by_keyword(ws, header_row, keywords):
    """Find columns in a specific row that match any keyword."""
    matches = {}
    max_col = ws.max_column or 100
    for col in range(1, min(max_col + 1, 300)):
        val = get_cell_value(ws, header_row, col)
        norm = normalize(val)
        for kw in keywords:
            if kw in norm:
                matches[col] = str(val)
                break
    return matches


def scan_sheet_for_retail_data(ws, sheet_name, label):
    """
    Scan a worksheet for Sears/Kmart/Territory columns and extract values.
    Returns a dict of findings.
    """
    result = {
        "sheet": sheet_name,
        "sears_columns": {},
        "kmart_columns": {},
        "territory_columns": {},
        "store_count_rows": {},
        "inflow_section_found": False,
        "raw_sears_values": [],
        "raw_kmart_values": [],
        "raw_territory_values": [],
        "raw_store_counts": [],
        "header_scan": {},
    }

    max_row = min(ws.max_row or 200, 500)
    max_col = min(ws.max_column or 100, 400)

    # Scan first 30 rows for headers
    for row in range(1, min(31, max_row + 1)):
        row_data = {}
        for col in range(1, min(max_col + 1, 400)):
            val = get_cell_value(ws, row, col)
            norm = normalize(val)
            if any(kw in norm for kw in SEARS_KEYWORDS):
                result["sears_columns"][col] = str(val)
                row_data[col] = str(val)
            if any(kw in norm for kw in KMART_KEYWORDS):
                result["kmart_columns"][col] = str(val)
                row_data[col] = str(val)
            if any(kw in norm for kw in TERRITORY_KEYWORDS):
                result["territory_columns"][col] = str(val)
                row_data[col] = str(val)
            if any(kw in norm for kw in INFLOW_SECTION_KEYWORDS):
                result["inflow_section_found"] = True
        if row_data:
            result["header_scan"][row] = row_data

    # Also scan for store count keywords in row labels (col 1-5)
    for row in range(1, max_row + 1):
        for col in range(1, 6):
            val = get_cell_value(ws, row, col)
            norm = normalize(val)
            if any(kw in norm for kw in STORE_COUNT_KEYWORDS):
                # Capture the whole row
                row_vals = {}
                for c in range(1, min(max_col + 1, 100)):
                    v = get_cell_value(ws, row, c)
                    if v is not None:
                        row_vals[c] = v
                result["store_count_rows"][row] = {"label": str(val), "values": row_vals}

    # If we found Sears/Kmart columns, extract the values from subsequent rows
    all_retail_cols = {}
    for col, name in result["sears_columns"].items():
        all_retail_cols[col] = ("sears", name)
    for col, name in result["kmart_columns"].items():
        all_retail_cols[col] = ("kmart", name)
    for col, name in result["territory_columns"].items():
        all_retail_cols[col] = ("territory", name)

    if all_retail_cols:
        # Find what row these headers are in and extract values below
        # Look for numeric values in those columns
        for row in range(1, max_row + 1):
            row_label = get_cell_value(ws, row, 1) or get_cell_value(ws, row, 2) or ""
            for col, (biz_type, col_name) in all_retail_cols.items():
                val = get_cell_value(ws, row, col)
                if isinstance(val, (int, float)) and val != 0:
                    entry = {
                        "row": row,
                        "col": col,
                        "col_name": col_name,
                        "row_label": str(row_label),
                        "value": val
                    }
                    if biz_type == "sears":
                        result["raw_sears_values"].append(entry)
                    elif biz_type == "kmart":
                        result["raw_kmart_values"].append(entry)
                    elif biz_type == "territory":
                        result["raw_territory_values"].append(entry)

    return result


def analyze_file(label, partial_name):
    """Analyze a single BOD file."""
    filepath = find_file(partial_name)
    if not filepath:
        return {"label": label, "error": f"File not found for pattern: {partial_name}"}

    filename = os.path.basename(filepath)
    print(f"\n{'='*60}")
    print(f"Processing: {label}")
    print(f"File: {filename}")

    result = {
        "label": label,
        "filename": filename,
        "sheets_analyzed": [],
        "sears_retail_inflows": {},
        "kmart_inflows": {},
        "territory_inflows": {},
        "store_counts": {},
        "notable_sheets": [],
        "column_headers_found": {},
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        sheet_names = wb.sheetnames
        print(f"  Sheets: {sheet_names}")

        # Flag interesting sheet names
        for sn in sheet_names:
            norm_sn = sn.lower()
            if any(kw in norm_sn for kw in ["sears", "kmart", "retail", "store", "territory", "guam", "usvi"]):
                result["notable_sheets"].append(sn)

        # Analyze each sheet
        for sn in sheet_names:
            try:
                ws = wb[sn]
                sheet_result = scan_sheet_for_retail_data(ws, sn, label)
                result["sheets_analyzed"].append(sheet_result)

                # Aggregate findings
                if sheet_result["sears_columns"]:
                    result["column_headers_found"][f"{sn}_sears"] = sheet_result["sears_columns"]
                if sheet_result["kmart_columns"]:
                    result["column_headers_found"][f"{sn}_kmart"] = sheet_result["kmart_columns"]
                if sheet_result["territory_columns"]:
                    result["column_headers_found"][f"{sn}_territory"] = sheet_result["territory_columns"]
                if sheet_result["store_count_rows"]:
                    result["store_counts"][sn] = sheet_result["store_count_rows"]

                # Collect numeric values
                if sheet_result["raw_sears_values"]:
                    result["sears_retail_inflows"][sn] = sheet_result["raw_sears_values"]
                if sheet_result["raw_kmart_values"]:
                    result["kmart_inflows"][sn] = sheet_result["raw_kmart_values"]
                if sheet_result["raw_territory_values"]:
                    result["territory_inflows"][sn] = sheet_result["raw_territory_values"]

            except Exception as e:
                result["sheets_analyzed"].append({"sheet": sn, "error": str(e)})

        wb.close()

    except Exception as e:
        result["error"] = str(e)
        print(f"  ERROR: {e}")

    return result


def main():
    all_results = []

    for label, partial_name in BOD_FILES:
        r = analyze_file(label, partial_name)
        all_results.append(r)

        # Print summary
        if "error" not in r:
            sears_sheets = list(r["sears_retail_inflows"].keys())
            kmart_sheets = list(r["kmart_inflows"].keys())
            terr_sheets = list(r["territory_inflows"].keys())
            print(f"  Notable sheets: {r['notable_sheets']}")
            print(f"  Sears data in sheets: {sears_sheets}")
            print(f"  Kmart data in sheets: {kmart_sheets}")
            print(f"  Territory data in sheets: {terr_sheets}")
            print(f"  Store count rows: {list(r['store_counts'].keys())}")
            if r["column_headers_found"]:
                print(f"  Column headers found: {r['column_headers_found']}")

    # Save to JSON
    output_path = os.path.join(BASE_DIR, "analysis_retail_stores.json")
    with open(output_path, "w") as f:
        json.dump(all_results, f, indent=2, default=str)

    print(f"\n\nSaved to: {output_path}")
    print(f"Total files processed: {len(all_results)}")

    # Print longitudinal summary
    print("\n" + "="*80)
    print("LONGITUDINAL SUMMARY: Sears & Kmart Inflows")
    print("="*80)
    for r in all_results:
        label = r.get("label", "unknown")
        has_sears = bool(r.get("sears_retail_inflows"))
        has_kmart = bool(r.get("kmart_inflows"))
        has_terr = bool(r.get("territory_inflows"))
        has_stores = bool(r.get("store_counts"))
        print(f"{label:30s} | Sears: {'YES' if has_sears else 'no':3s} | Kmart: {'YES' if has_kmart else 'no':3s} | Territory: {'YES' if has_terr else 'no':3s} | StoreCounts: {'YES' if has_stores else 'no':3s}")


if __name__ == "__main__":
    main()
