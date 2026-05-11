"""
Final precise extraction:
- 'Liquidity' sheet: monthly Available Cash, Inflows, Outflows, Net CF
- 'Cash Flow' (or 'Net Cash Flow') sheet: Inflows Total, Disbursements Total, Payroll/Bens
- Dec 2024 'SC Interco' sheet: $9M adjustment detail
- Jan 2025 April append detail
"""
import openpyxl
import json
from pathlib import Path
from datetime import datetime, date

BASE = Path("/Users/josh/Downloads/SP_Analysis")

FILES = [
    ("2024-09-24", "Daily Cash Fcst - 9.24.24_BU view_deck version_BOD.xlsx"),
    ("2024-10-15", "Daily Cash Fcst - 10.15.24_BU view_deck version_OctBOD.xlsx"),
    ("2024-11-12", "Daily Cash Fcst - 11.12.24_BU view_deck version_NovBOD.xlsx"),
    ("2024-12-10", "Daily Cash Fcst - 12.10.24_BU view_deck version_BOD_SC wo 3 RE Interco IF_$9M.xlsx"),
    ("2024-12-31", "Daily Cash Fcst - 12.31.24_BU view.xlsx"),
    ("2025-01-21", "Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx"),
    ("2025-03-18", "Daily Cash Fcst - 3.18.25_BU view_deck version_MarBOD.xlsx"),
    ("2025-04-18", "Daily Cash Fcst - 4.18.25_BU view_prelim AprBOD.xlsx"),
    ("2026-05-08", "Daily Cash Fcst - 05.08.26_BU view.xlsx"),
]

def sf(v):
    """Safe float conversion."""
    if v is None: return None
    try: return float(v)
    except: return None

def r3(v):
    if v is None: return None
    return round(v, 3)

def extract_liquidity_sheet(ws, file_date):
    """
    The Liquidity sheet has multiple year-sections.
    Structure per row: [month_label | available_cash_start | inflows | outflows | ... | net_cf | ending_avail_cash | ...]
    Based on observation:
      col[0] = month label (or starting available cash for 2026+ format)
      col[1] = starting Available Cash (or month label in 2026 format)
      col[2] = Inflows
      col[3] = Outflows (negative)
      col[4..] = other items
      The penultimate numeric cols = ending available cash and total cash
    """
    monthly = []
    current_year = None
    cumulative_rows = []

    MONTH_NAMES = ["january","february","march","april","may","june",
                   "july","august","september","october","november","december"]

    for row in ws.iter_rows(min_row=1, max_row=500, values_only=True):
        if row is None or all(v is None for v in row): continue

        # Check col[0] for month name or year label
        c0 = str(row[0]).strip() if row[0] is not None else ""
        c1 = str(row[1]).strip() if row[1] is not None else ""

        # Detect year header rows like "Feb 24", "Mar 24" etc
        if any(m in c0.lower() for m in ["feb 2","mar 2","jan 2","apr 2","sep 2","oct 2","nov 2","dec 2"]):
            continue  # skip sub-headers

        # Detect cumulative row
        if "cumulative" in c0.lower() or "cumulative" in c1.lower():
            cumulative_rows.append(row)
            continue

        # Try to parse month from c0 or c1
        month_label = None
        avail_cash_start = None
        inflows = None
        outflows = None

        # 2022-2024 format: [month_name | avail_cash_start | inflows | outflows | ...]
        # 2026 format: [avail_cash_float | month_name | avail_cash_2 | inflows | outflows | ...]
        c0_lower = c0.lower()
        c1_lower = c1.lower()

        if any(m in c0_lower for m in MONTH_NAMES):
            month_label = c0
            avail_cash_start = sf(row[1])
            inflows = sf(row[2])
            outflows = sf(row[3])
        elif any(m in c1_lower for m in MONTH_NAMES):
            # 2026 format - col0 is a float (ending cash from prev month)
            month_label = c1
            avail_cash_start = sf(row[2])
            inflows = sf(row[3])
            outflows = sf(row[4])
        else:
            continue

        if month_label is None: continue

        # Get all numeric values in row
        all_nums = [sf(v) for v in row if sf(v) is not None]

        # Net CF is typically near the end of the middle-column section
        # From observation: Net CF appears as last before ending cash
        # Get non-zero numerics after the first 5 columns
        tail_nums = [sf(v) for v in row[5:] if sf(v) is not None]

        monthly.append({
            "month": month_label,
            "available_cash_start_M": r3(avail_cash_start),
            "inflows_M": r3(inflows),
            "outflows_M": r3(outflows),
            "net_cf_M": r3(inflows + outflows) if (inflows is not None and outflows is not None) else None,
            "all_row_nums": [r3(v) for v in all_nums[:20]],
            "tail_nums": [r3(v) for v in tail_nums[:10]],
        })

    return monthly

def extract_ncf_sheet(ws):
    """Extract from Net Cash Flow summary sheet (older files).
    Row 20: Total Inflows
    Row 40: Total Disbursements
    """
    data = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=70, values_only=True), 1):
        if row is None: continue
        text_vals = [str(v).lower().strip() for v in row[:5] if v is not None and isinstance(v, str)]
        row_text = " ".join(text_vals)
        nums = [sf(v) for v in row[2:] if sf(v) is not None]

        if "total inflows" in row_text or ("inflows total" in row_text):
            data["inflows_total_raw"] = nums[:15]
        elif "total disbursements" in row_text or "disbursements total" in row_text:
            data["disbursements_total_raw"] = nums[:15]
        elif "pyrl/bens" in row_text or "payroll/bens" in row_text or "payroll & bens" in row_text:
            if "total" not in row_text:
                data["payroll_bens_raw"] = nums[:15]
        elif "starting cash" in row_text:
            data["starting_cash"] = nums[0] if nums else None
        elif "ending cash" in row_text and "/" not in row_text and "requirement" not in row_text:
            data["ending_cash_samples"] = nums[:15]
    return data

def extract_cashflow_sheet(ws):
    """Extract from 'Cash Flow' sheet."""
    data = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True), 1):
        if row is None: continue
        text_vals = [str(v).lower().strip() for v in row[:10] if v is not None and isinstance(v, str)]
        row_text = " ".join(text_vals)
        nums = [sf(v) for v in row if sf(v) is not None and abs(sf(v)) > 0.001]

        if "available cash" in row_text:
            data.setdefault("available_cash_rows", []).append({"row": i, "text": row_text[:80], "nums": nums[:15]})
        if "total cash" in row_text:
            data.setdefault("total_cash_rows", []).append({"row": i, "text": row_text[:80], "nums": nums[:15]})
        if "unavailable" in row_text:
            data.setdefault("unavailable_cash_rows", []).append({"row": i, "text": row_text[:80], "nums": nums[:15]})
        if "inflows total" in row_text or "total inflows" in row_text:
            data.setdefault("inflows_total_rows", []).append({"row": i, "text": row_text[:80], "nums": nums[:15]})
        if "disbursements total" in row_text or "total disbursements" in row_text:
            data.setdefault("disbursements_total_rows", []).append({"row": i, "text": row_text[:80], "nums": nums[:15]})
        if "payroll/bens" in row_text or "pyrl/bens" in row_text or "payroll & bens" in row_text:
            data.setdefault("payroll_bens_rows", []).append({"row": i, "text": row_text[:80], "nums": nums[:15]})
        if "net cash flow" in row_text or "net cf" in row_text:
            data.setdefault("net_cf_rows", []).append({"row": i, "text": row_text[:80], "nums": nums[:15]})

    return data

def extract_sc_interco_sheet(ws, sheet_name):
    """Extract the SC Interco / $9M adjustment sheet from Dec 2024 file."""
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True), 1):
        if row is None: continue
        text_vals = [str(v).strip() for v in row[:10] if v is not None]
        nums = [sf(v) for v in row if sf(v) is not None]
        if text_vals or nums:
            rows.append({
                "row": i,
                "labels": text_vals[:5],
                "nums_M": [r3(v) for v in nums[:10]]
            })
    return {"sheet_name": sheet_name, "rows": rows[:60]}

def process_file(date_str, filename):
    filepath = BASE / filename
    print(f"\n{'='*70}")
    print(f"FILE: {filename}")
    print(f"DATE: {date_str}")

    result = {
        "file": filename,
        "date": date_str,
        "sheets": [],
        "liquidity_monthly_data": [],
        "ncf_sheet_data": {},
        "cashflow_sheet_data": {},
        "forecast_horizon": {"start": None, "end": None},
        "special_sc_interco": None,
        "special_apr_append": None,
        "notes": []
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        result["sheets"] = list(wb.sheetnames)

        # 1. Liquidity sheet
        liq_names = [s for s in wb.sheetnames if "liquidity" in s.lower()]
        if liq_names:
            ws_liq = wb[liq_names[0]]
            monthly = extract_liquidity_sheet(ws_liq, date_str)
            result["liquidity_monthly_data"] = monthly
            if monthly:
                result["forecast_horizon"]["start"] = monthly[0]["month"]
                result["forecast_horizon"]["end"] = monthly[-1]["month"]
            print(f"  Liquidity: {len(monthly)} months extracted")
            for m in monthly[-5:]:
                print(f"    {m['month']}: avail={m['available_cash_start_M']}, inflows={m['inflows_M']}, outflows={m['outflows_M']}, net_cf={m['net_cf_M']}")
        else:
            result["notes"].append("No Liquidity sheet found")
            print("  No Liquidity sheet found")

        # 2. Net Cash Flow sheet (older format)
        ncf_names = [s for s in wb.sheetnames if s.lower() == "net cash flow"]
        if ncf_names:
            ws_ncf = wb[ncf_names[0]]
            result["ncf_sheet_data"] = extract_ncf_sheet(ws_ncf)
            print(f"  Net Cash Flow sheet extracted")

        # 3. Cash Flow sheet
        cf_names = [s for s in wb.sheetnames if s.lower() == "cash flow"]
        if cf_names:
            ws_cf = wb[cf_names[0]]
            result["cashflow_sheet_data"] = extract_cashflow_sheet(ws_cf)
            print(f"  Cash Flow sheet extracted")

        # 4. SPECIAL: Dec 2024 - SC Interco sheet
        if "12.10.24" in filename:
            sc_names = [s for s in wb.sheetnames if "sc interco" in s.lower() or "interco" in s.lower() or
                        ("sc" in s.lower() and "if" not in s.lower() and "forecast" not in s.lower())]
            print(f"  SC/Interco candidate sheets: {sc_names}")
            for sname in sc_names[:3]:
                ws_sc = wb[sname]
                sc_data = extract_sc_interco_sheet(ws_sc, sname)
                result["special_sc_interco"] = sc_data
                print(f"  SC Interco sheet '{sname}': {len(sc_data['rows'])} rows")
                for r in sc_data["rows"][:15]:
                    print(f"    R{r['row']}: {r['labels'][:3]} | {r['nums_M'][:8]}")
                break  # take first match

        # 5. SPECIAL: Jan 2025 - April append
        if "1.21.25" in filename:
            # Look for sheets with "Apr" or "2025" in name
            apr_names = [s for s in wb.sheetnames if "apr" in s.lower() or "deck2" in s.lower() or "append" in s.lower()]
            print(f"  April candidate sheets: {apr_names}")
            # Also scan the Liquidity sheet for April 2025 data
            if liq_names:
                apr_rows = [m for m in result["liquidity_monthly_data"] if "april" in m["month"].lower() or "apr" in m["month"].lower()]
                result["special_apr_append"] = {
                    "apr_months_in_liquidity": apr_rows,
                    "note": "April rows from Liquidity sheet; file name suggests Apr 2025 extension was appended"
                }
                print(f"  April rows in Liquidity: {apr_rows}")

        wb.close()

    except Exception as e:
        result["notes"].append(f"ERROR: {e}")
        print(f"  ERROR: {e}")
        import traceback; traceback.print_exc()

    return result

def main():
    all_results = []
    for date_str, filename in FILES:
        r = process_file(date_str, filename)
        all_results.append(r)

    # Build clean summary table
    print("\n\n" + "="*70)
    print("CLEAN SUMMARY TABLE")
    print("="*70)
    print(f"{'Date':<12} {'Horizon Start':<20} {'Horizon End':<20} {'Last Month Avail Cash':>22} {'Last Month Inflows':>18} {'Last Month Outflows':>19}")

    summary_table = []
    for r in all_results:
        monthly = r.get("liquidity_monthly_data", [])

        # Find the most recent forecast month (not a historical month)
        last = monthly[-1] if monthly else None
        first = monthly[0] if monthly else None

        row = {
            "date": r["date"],
            "file": r["file"],
            "sheets_count": len(r["sheets"]),
            "forecast_start": r["forecast_horizon"]["start"],
            "forecast_end": r["forecast_horizon"]["end"],
            "last_month": last["month"] if last else None,
            "last_available_cash_M": last["available_cash_start_M"] if last else None,
            "last_inflows_M": last["inflows_M"] if last else None,
            "last_outflows_M": last["outflows_M"] if last else None,
            "last_net_cf_M": last["net_cf_M"] if last else None,
            "all_months": [{"month": m["month"], "avail": m["available_cash_start_M"],
                            "inflows": m["inflows_M"], "outflows": m["outflows_M"],
                            "net_cf": m["net_cf_M"]} for m in monthly],
        }
        summary_table.append(row)

        # Print summary
        print(f"{r['date']:<12} {str(r['forecast_horizon']['start']):<20} {str(r['forecast_horizon']['end']):<20} "
              f"{str(last['available_cash_start_M'] if last else 'N/A'):>22} "
              f"{str(last['inflows_M'] if last else 'N/A'):>18} "
              f"{str(last['outflows_M'] if last else 'N/A'):>19}")

    # Comparison: Dec 2024 vs May 2026
    print("\n\nCOMPARISON: Dec 10, 2024 vs May 8, 2026")
    print("-"*60)
    dec_data = next((r for r in all_results if "12.10.24" in r["file"]), None)
    jan_data = next((r for r in all_results if "1.21.25" in r["file"]), None)
    may_data = next((r for r in all_results if "05.08.26" in r["file"]), None)

    def get_month_data(file_result, month_kw):
        if not file_result: return None
        for m in file_result.get("liquidity_monthly_data", []):
            if month_kw.lower() in m["month"].lower():
                return m
        return None

    # Dec 2024 last forecast month
    if dec_data:
        dec_monthly = dec_data.get("liquidity_monthly_data", [])
        print(f"\nDec 10, 2024 file — Liquidity months:")
        for m in dec_monthly[-8:]:
            print(f"  {m['month']}: avail={m['available_cash_start_M']}, inflows={m['inflows_M']}, outflows={m['outflows_M']}, net_cf={m['net_cf_M']}")

    if may_data:
        may_monthly = may_data.get("liquidity_monthly_data", [])
        print(f"\nMay 8, 2026 file — Liquidity months (all):")
        for m in may_monthly:
            print(f"  {m['month']}: avail={m['available_cash_start_M']}, inflows={m['inflows_M']}, outflows={m['outflows_M']}, net_cf={m['net_cf_M']}")

    # SC Interco special
    if dec_data and dec_data.get("special_sc_interco"):
        print(f"\nDec 2024 SC Interco sheet: {dec_data['special_sc_interco']['sheet_name']}")
        for r in dec_data["special_sc_interco"]["rows"][:20]:
            print(f"  R{r['row']}: {r['labels'][:3]} | {r['nums_M'][:6]}")

    # Save full JSON
    output = {
        "extraction_date": "2026-05-11",
        "summary_table": summary_table,
        "full_data": all_results
    }

    out_path = BASE / "analysis_2025_h2_2026.json"
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\n\nJSON saved to: {out_path}")

if __name__ == "__main__":
    main()
