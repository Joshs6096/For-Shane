"""
Cash Forecast Inflow Extractor
Extracts monthly inflow totals from 17 cash forecast files (Jan 2024 – May 2026).
Sources:
  - Inflows Actuals sheet: per-category monthly actuals (past data)
  - Liquidity sheet: monthly aggregate inflows/outflows + available cash
Output: /Users/josh/Downloads/SP_Analysis/analysis_inflows_2024_2026.json
"""

import openpyxl
import json
import re
from pathlib import Path
from datetime import datetime

BASE = Path('/Users/josh/Downloads/SP_Analysis')

FILES = [
    ('2024-01-23', 'Daily Cash Fcst - 1.23.24_BU view_BOD.xlsx'),
    ('2024-02-20', 'Daily Cash Fcst - 2.20.24_BU view_BOD.xlsx'),
    ('2024-03-19', 'Daily Cash Fcst - 3.19.24_BU view_MarBOD.xlsx'),
    ('2024-04-30', 'Daily Cash Fcst - 4.30.24_BU view_final_online adj.xlsx'),
    ('2024-05-07', 'Daily Cash Fcst - 5.7.24_BU view_MayBOD_5.9.24.xlsx'),
    ('2024-06-18', 'Daily Cash Fcst - 6.18.24_BU view_JunBOD.xlsx'),
    ('2024-07-23', 'Daily Cash Fcst - 7.23.24_deck_JulBOD_FY25.xlsx'),
    ('2024-08-19', 'Daily Cash Fcst - 8.19.24_BU view_AugBOD.xlsx'),
    ('2024-09-24', 'Daily Cash Fcst - 9.24.24_BU view_deck version_BOD.xlsx'),
    ('2024-10-15', 'Daily Cash Fcst - 10.15.24_BU view_deck version_OctBOD.xlsx'),
    ('2024-11-12', 'Daily Cash Fcst - 11.12.24_BU view_deck version_NovBOD.xlsx'),
    ('2024-12-10', 'Daily Cash Fcst - 12.10.24_BU view_deck version_BOD_SC wo 3 RE Interco IF_$9M.xlsx'),
    ('2024-12-31', 'Daily Cash Fcst - 12.31.24_BU view.xlsx'),
    ('2025-01-21', 'Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx'),
    ('2025-03-18', 'Daily Cash Fcst - 3.18.25_BU view_deck version_MarBOD.xlsx'),
    ('2026-03-03', 'Daily Cash Fcst - 03.03.26_post-reductions_deck version_MarBOD.xlsx'),
    ('2026-05-08', 'Daily Cash Fcst - 05.08.26_BU view.xlsx'),
]

MONTH_ABBREVS = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}

def clean(v):
    """Return float or None; skip strings like '#N/A', '#VALUE!'."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str) and v.startswith('#'):
        return None
    return None

def parse_month_label(label):
    """Parse 'Jan 2024' -> (2024, 1). Returns None if not parseable."""
    if not label or not isinstance(label, str):
        return None
    m = re.match(r'^([A-Za-z]+)\s+(\d{4})$', label.strip())
    if not m:
        return None
    mon_str = m.group(1).lower()[:3]
    year = int(m.group(2))
    if mon_str not in MONTH_ABBREVS:
        return None
    return (year, MONTH_ABBREVS[mon_str])

def extract_inflows_actuals(ws):
    """
    Extract monthly category-level inflow rows from the Inflows Actuals sheet.
    Returns dict: {(year, month): {category: value, ...}}
    """
    # Read all rows into memory
    rows = list(ws.iter_rows(min_row=1, max_row=1200, max_col=35, values_only=True))

    # Build column map from header rows (rows 0 and 1, 0-indexed)
    # Header row 1 (index 0) has the primary categories; row 2 (index 1) has sub-labels
    # Col 0 = DATE/label, cols 1+ are categories
    # We use row index 1 (the second header row) as canonical names
    col_names = {}
    if len(rows) >= 2:
        for j, v in enumerate(rows[1]):
            if v and isinstance(v, str) and v.strip() and not v.startswith('#'):
                col_names[j] = v.strip().replace('\n', ' ')
    # Fallback to row index 0 for any missing
    if len(rows) >= 1:
        for j, v in enumerate(rows[0]):
            if j not in col_names and v and isinstance(v, str) and v.strip():
                col_names[j] = v.strip().replace('\n', ' ')

    result = {}
    for row in rows:
        label = row[0] if row else None
        parsed = parse_month_label(label)
        if parsed is None:
            continue
        year, month = parsed
        # Skip YTD rows
        if isinstance(label, str) and 'YTD' in label:
            continue

        entry = {}
        for j, v in enumerate(row):
            if j == 0:
                continue
            cv = clean(v)
            if cv is not None and cv != 0.0:
                col_label = col_names.get(j, f'col_{j}')
                if col_label:
                    entry[col_label] = cv

        result[(year, month)] = entry

    return result


def extract_liquidity(ws):
    """
    Extract monthly aggregate rows from the Liquidity sheet.
    Returns dict: {(year, month): {avail_cash, inflows, outflows, net_cash,
                                    equity, other_financing, asset_sales, misc_other,
                                    interest_fees, debt_repayment, all_other, total}}

    NOTE: Liquidity blocks run Feb–Jan (fiscal year). The "January" at the end
    of each block belongs to the NEXT calendar year (e.g., in a block starting
    "February 2022", the closing "January" is January 2023).
    We track the last seen month number to detect year rollover.
    """
    rows = list(ws.iter_rows(min_row=1, max_row=200, max_col=20, values_only=True))

    result = {}

    FULL_MONTHS = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12
    }

    current_year = None
    last_month_num = None  # track for year rollover

    for row in rows:
        label = row[1] if len(row) > 1 else None
        if not label or not isinstance(label, str):
            continue
        label = label.strip()

        # Check for cumulative row (skip)
        if 'cumulative' in label.lower():
            last_month_num = None  # reset on new block
            continue

        # Parse: "February 2022" style (block header — anchors the year)
        m_full = re.match(r'^([A-Za-z]+)\s+(\d{4})$', label)
        if m_full:
            mon_name = m_full.group(1).lower()
            year = int(m_full.group(2))
            if mon_name in FULL_MONTHS:
                current_year = year
                month = FULL_MONTHS[mon_name]
                last_month_num = month
            else:
                continue
        elif label.lower() in FULL_MONTHS and current_year:
            # Just a month name — use current_year but watch for Jan rollover
            month = FULL_MONTHS[label.lower()]
            year = current_year
            # If month number went from high to low (e.g. Dec→Jan), we've crossed
            # a year boundary. This happens in Feb–Jan fiscal blocks.
            if last_month_num is not None and month < last_month_num:
                year = current_year + 1
            last_month_num = month
        else:
            continue

        # Extract values - column layout varies slightly between versions
        # Standard layout (newer files, 2025+):
        #   col2=Available Cash, col3=Inflows, col4=Outflows, col5=Net Cash
        #   col6=Equity, col7=Other Financing (TL/etc), col8=Asset Sales
        #   col9=Lease Opco & Misc Other, col12=Interest & Fees, col13=Debt Repayment
        #   col14=All Other, col15=Net Cash Total
        # Older layout (2022-2024):
        #   col2=Available Cash, col3=Inflows, col4=Outflows, col5=Net Cash
        #   col6=Equity, col7=Other Financing, col8=Asset Sales
        #   col9=Lease Opco & Misc Other, col10=Interest & Fees, col11=Debt Repayment
        #   col12=All Other, col13=Total

        def g(idx):
            if idx < len(row):
                return clean(row[idx])
            return None

        entry = {
            'available_cash': g(2),
            'total_inflows': g(3),
            'total_outflows': g(4),
            'net_operating_cash': g(5),
            'equity': g(6),
            'other_financing': g(7),
            'asset_sales': g(8),
            'lease_opco_misc': g(9),
        }

        # Detect column layout: older files have interest at col10, newer at col12
        # We check which column has a negative number consistent with fees
        # Simple approach: try to read both and pick whichever has data
        c10, c11, c12, c13, c14 = g(10), g(11), g(12), g(13), g(14)

        if c12 is not None or c13 is not None:
            # Newer layout
            entry['interest_fees'] = c12
            entry['debt_repayment'] = c13
            entry['all_other'] = c14
            entry['net_cash_total'] = g(15)
        elif c10 is not None or c11 is not None:
            # Older layout
            entry['interest_fees'] = c10
            entry['debt_repayment'] = c11
            entry['all_other'] = c12
            entry['net_cash_total'] = c13
        else:
            entry['interest_fees'] = None
            entry['debt_repayment'] = None
            entry['all_other'] = None
            entry['net_cash_total'] = None

        result[(year, month)] = entry

    return result


def extract_cash_flow_avail(ws, as_of_date_str):
    """
    From the Cash Flow sheet, extract Available Cash and Total Cash on the as-of date.
    Returns (avail_cash, total_cash) or (None, None).
    Col 10 = Available Cash, Col 14 = Total Cash
    """
    target_year = int(as_of_date_str[:4])
    target_month = int(as_of_date_str[5:7])
    target_day = int(as_of_date_str[8:10])

    for row in ws.iter_rows(min_row=3, max_row=5000, max_col=16, values_only=True):
        v0 = row[0]
        if isinstance(v0, datetime):
            if v0.year == target_year and v0.month == target_month and v0.day == target_day:
                avail = clean(row[10]) if len(row) > 10 else None
                total = clean(row[14]) if len(row) > 14 else None
                return avail, total
    return None, None


def process_file(as_of_date, fname):
    fpath = BASE / fname
    print(f'\nProcessing {fname} (as-of: {as_of_date})')

    wb = openpyxl.load_workbook(str(fpath), data_only=True, read_only=True)
    sheets = wb.sheetnames

    result = {
        'as_of_date': as_of_date,
        'filename': fname,
        'inflows_actuals_by_month': {},
        'liquidity_by_month': {},
        'cash_on_as_of_date': {},
    }

    # 1. Extract from Inflows Actuals
    if 'Inflows Actuals' in sheets:
        ws = wb['Inflows Actuals']
        actuals = extract_inflows_actuals(ws)
        for (yr, mo), data in sorted(actuals.items()):
            key = f'{yr}-{mo:02d}'
            result['inflows_actuals_by_month'][key] = data
        print(f'  Inflows Actuals: {len(actuals)} monthly rows')
    else:
        print(f'  WARNING: No Inflows Actuals sheet')

    # 2. Extract from Liquidity
    if 'Liquidity' in sheets:
        ws = wb['Liquidity']
        liq = extract_liquidity(ws)
        for (yr, mo), data in sorted(liq.items()):
            key = f'{yr}-{mo:02d}'
            result['liquidity_by_month'][key] = data
        print(f'  Liquidity: {len(liq)} monthly rows')
    else:
        print(f'  WARNING: No Liquidity sheet')

    # 3. Extract Available Cash and Total Cash from Cash Flow sheet on as-of date
    if 'Cash Flow' in sheets:
        ws = wb['Cash Flow']
        avail, total = extract_cash_flow_avail(ws, as_of_date)
        result['cash_on_as_of_date'] = {
            'available_cash': avail,
            'total_cash': total,
        }
        print(f'  Cash on {as_of_date}: available={avail}, total={total}')

    wb.close()
    return result


def main():
    all_results = []

    for as_of_date, fname in FILES:
        fpath = BASE / fname
        if not fpath.exists():
            print(f'FILE NOT FOUND: {fname}')
            continue
        try:
            r = process_file(as_of_date, fname)
            all_results.append(r)
        except Exception as e:
            print(f'  ERROR processing {fname}: {e}')
            import traceback
            traceback.print_exc()

    # Build a consolidated monthly inflows table
    # For each file, collect inflows_actuals (past) and liquidity (both past and forward)
    # Key question: what months does each file cover?

    output = {
        'generated': datetime.now().isoformat(),
        'description': 'Monthly inflow extraction from 17 SP cash forecast files, Jan 2024 - May 2026',
        'files': [],
        'consolidated_inflows_actuals': {},
        'consolidated_liquidity': {},
    }

    for r in all_results:
        file_entry = {
            'as_of_date': r['as_of_date'],
            'filename': r['filename'],
            'cash_on_as_of_date': r['cash_on_as_of_date'],
            'inflows_actuals_months_available': sorted(r['inflows_actuals_by_month'].keys()),
            'liquidity_months_available': sorted(r['liquidity_by_month'].keys()),
            'inflows_actuals_by_month': r['inflows_actuals_by_month'],
            'liquidity_by_month': r['liquidity_by_month'],
        }
        output['files'].append(file_entry)

        # Consolidate: last-file-wins for each month (most recent file has best data)
        for k, v in r['inflows_actuals_by_month'].items():
            output['consolidated_inflows_actuals'][k] = v
        for k, v in r['liquidity_by_month'].items():
            output['consolidated_liquidity'][k] = v

    # Write output
    out_path = BASE / 'analysis_inflows_2024_2026.json'
    with open(out_path, 'w') as f:
        json.dump(output, f, indent=2, default=str)

    print(f'\n\nWrote {out_path}')
    print(f'Files processed: {len(output["files"])}')
    print(f'Consolidated inflows actuals months: {sorted(output["consolidated_inflows_actuals"].keys())}')
    print(f'Consolidated liquidity months: {len(output["consolidated_liquidity"])} months')


if __name__ == '__main__':
    main()
