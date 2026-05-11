"""
Comprehensive retail inflow extractor for Sears, Kmart, and Territories.
Strategy:
1. 'Inflows Actuals' sheet: col 1=date, col 2=Sears Stores, col 3=Kmart Stores
   -> sum all rows by month
2. 'Inflows Forecasting' sheet: weekly row data
   -> extract future weekly forecasts
3. 'BU beta' or 'Historicals view KK': monthly summary aggregates
4. 'B&M Anita' or 'SC IF' sheets: look for store count data
5. Territory data in Inflows Actuals or Disbursements

All values assumed in $M unless noted.
"""

import os, json
from collections import defaultdict
import openpyxl
from datetime import datetime

BASE_DIR = "/Users/josh/Downloads/SP_Analysis"

BOD_FILES = [
    ("9.19.22_Sep BOD",   "9.19.22_BU view_Sep BOD"),
    ("12.7.22_DecBOD",    "12.7.22_BU view_DecBOD"),
    ("1.17.23_JanBOD",    "1.17.23_BU view_JanBOD"),
    ("4.25.23_AprBOD",    "4.25.23_BU view_April 27 BOD"),
    ("7.25.23_JulBOD",    "7.25.23_BU view_JulBOD"),
    ("11.14.23_NovBOD",   "11.14.23_BU view_adUPDATE"),
    ("2.20.24_BOD",       "2.20.24_BU view_BOD"),
    ("6.18.24_JunBOD",    "6.18.24_BU view_JunBOD"),
    ("10.15.24_OctBOD",   "10.15.24_BU view_deck version_OctBOD"),
    ("12.31.24_BU view",  "12.31.24_BU view"),
    ("3.18.25_MarBOD",    "3.18.25_BU view_deck version_MarBOD"),
    ("7.22.25_JulBOD",    "7.22.25_BU view_deck version_condensed_JulBOD"),
    ("12.15.25_DecBOD",   "12.15.25_BU view_DecBOD"),
    ("01.16.26_JanBOD",   "01.16.26_BU view new format_asset sales_JanBOD"),
    ("04.21.26_AprBOD",   "04.21.26_ BU version_deck version_AprBOD"),
    ("05.08.26_BU view",  "05.08.26_BU view"),
]

def find_file(partial_name):
    for fn in os.listdir(BASE_DIR):
        if partial_name.lower() in fn.lower() and fn.endswith(".xlsx"):
            return os.path.join(BASE_DIR, fn)
    return None

def safe_val(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str) and v.strip() == "":
        return None
    return None

def normalize(v):
    if v is None: return ""
    return str(v).lower().strip()

def get_col_header_map(ws, header_row):
    """Return dict of col -> normalized header text for a given row."""
    m = {}
    for col in range(1, min((ws.max_column or 100)+1, 300)):
        v = ws.cell(row=header_row, column=col).value
        if v is not None:
            m[col] = normalize(v)
    return m

def find_header_row(ws, keywords, max_scan=20):
    """Find first row where any keyword appears in any column."""
    for row in range(1, min(max_scan+1, (ws.max_row or 20)+1)):
        for col in range(1, min((ws.max_column or 50)+1, 100)):
            v = normalize(ws.cell(row=row, column=col).value)
            if any(kw in v for kw in keywords):
                return row
    return None

# ── Inflows Actuals (daily actuals) ──────────────────────────────────────────
def extract_inflows_actuals(ws, label):
    """
    Format: Row 1 = header labels, col1=date, col2=Sears Stores, col3=Kmart Stores
    Some files have row 1 = col indices, row 2 = header names.
    Returns monthly_data: {YYYY-MM: {sears: x, kmart: x, guam: x, usvi: x}}
    """
    # Scan first 5 rows for headers
    header_row = None
    sears_col = kmart_col = guam_col = usvi_col = date_col = None

    for r in range(1, 8):
        for c in range(1, 60):
            v = normalize(ws.cell(row=r, column=c).value)
            if "sears store" in v or (v == "sears stores") or ("sears" in v and "store" in v):
                sears_col = c
                header_row = r
            if "kmart store" in v or (v == "kmart stores") or ("kmart" in v and "store" in v):
                kmart_col = c
            if "guam" in v:
                guam_col = c
            if "usvi" in v or "u.s.v.i" in v or "virgin island" in v:
                usvi_col = c
        if sears_col and kmart_col:
            break

    # Find date column: col 1 with datetime or date-like values
    date_col = 1

    if not (sears_col and kmart_col):
        return None, None, None, None

    # Collect monthly data
    monthly = defaultdict(lambda: {"sears": 0.0, "kmart": 0.0, "guam": 0.0, "usvi": 0.0, "days": 0})
    data_start = (header_row or 1) + 1

    for r in range(data_start, min((ws.max_row or 500)+1, 2000)):
        date_val = ws.cell(row=r, column=date_col).value
        if date_val is None:
            continue
        # Parse date
        if isinstance(date_val, datetime):
            ym = date_val.strftime("%Y-%m")
        elif hasattr(date_val, 'year'):
            ym = f"{date_val.year}-{date_val.month:02d}"
        else:
            continue

        s_val = safe_val(ws.cell(row=r, column=sears_col).value)
        k_val = safe_val(ws.cell(row=r, column=kmart_col).value)

        if s_val is not None:
            monthly[ym]["sears"] += s_val
        if k_val is not None:
            monthly[ym]["kmart"] += k_val
        if guam_col:
            g = safe_val(ws.cell(row=r, column=guam_col).value)
            if g: monthly[ym]["guam"] += g
        if usvi_col:
            u = safe_val(ws.cell(row=r, column=usvi_col).value)
            if u: monthly[ym]["usvi"] += u
        monthly[ym]["days"] += 1

    return dict(monthly), sears_col, kmart_col, header_row

# ── Inflows Forecasting (weekly forecasts) ────────────────────────────────────
def extract_inflows_forecasting(ws, label):
    """
    Format: Row 2 = header row with Sears Stores, Kmart Stores etc.
    Col 1 = week number or month label
    """
    sears_col = kmart_col = None
    header_row = None

    for r in range(1, 10):
        for c in range(1, 60):
            v = normalize(ws.cell(row=r, column=c).value)
            if "sears store" in v:
                sears_col = c
                header_row = r
            if "kmart store" in v:
                kmart_col = c
        if sears_col:
            break

    if not (sears_col and kmart_col):
        return None

    weekly = []
    data_start = header_row + 1
    for r in range(data_start, min((ws.max_row or 200)+1, 500)):
        wk_label = ws.cell(row=r, column=1).value
        if wk_label is None:
            continue
        s = safe_val(ws.cell(row=r, column=sears_col).value)
        k = safe_val(ws.cell(row=r, column=kmart_col).value)
        if s is not None or k is not None:
            weekly.append({"week": str(wk_label), "sears": s, "kmart": k})

    return weekly

# ── Historical view / BU beta ────────────────────────────────────────────────
def extract_historicals(ws, label):
    """
    'Historicals view KK' or 'BU beta' sheets have monthly aggregates:
    Row 3/4 = month headers, Row 4/6 = Sears Stores, Row 5/7 = Kmart Stores
    """
    sears_row = kmart_row = month_header_row = None
    guam_row = usvi_row = None

    for r in range(1, 20):
        for c in range(1, 5):
            v = normalize(ws.cell(row=r, column=c).value)
            if "sears store" in v:
                sears_row = r
            if "kmart store" in v:
                kmart_row = r
            if "guam" in v:
                guam_row = r
            if "usvi" in v or "virgin island" in v:
                usvi_row = r

    if not (sears_row and kmart_row):
        return None

    # Find the column with month headers - usually row just above sears_row
    # or within rows 1-5
    # Find columns that have month labels
    month_cols = {}
    for scan_row in range(1, max(sears_row, 2)+1):
        for c in range(3, 50):
            v = ws.cell(row=scan_row, column=c).value
            if v is not None and isinstance(v, str):
                vn = v.lower()
                if any(m in vn for m in ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]):
                    month_cols[c] = v
                    month_header_row = scan_row

    if not month_cols:
        # Try numeric columns
        for c in range(3, 50):
            v = ws.cell(row=sears_row, column=c).value
            if isinstance(v, (int, float)) and v != 0:
                month_cols[c] = f"col_{c}"

    result = []
    for c, month_label in sorted(month_cols.items()):
        s = safe_val(ws.cell(row=sears_row, column=c).value)
        k = safe_val(ws.cell(row=kmart_row, column=c).value)
        g = safe_val(ws.cell(row=guam_row, column=c).value) if guam_row else None
        u = safe_val(ws.cell(row=usvi_row, column=c).value) if usvi_row else None
        if s is not None or k is not None:
            result.append({"month": month_label, "col": c, "sears": s, "kmart": k, "guam": g, "usvi": u})

    return result

# ── Store count scanner ───────────────────────────────────────────────────────
def extract_store_counts(ws):
    """Look for store count rows anywhere in the sheet."""
    results = []
    STORE_KW = ["store count", "# stores", "# of stores", "number of stores", "open stores",
                "operating stores", "sears stores open", "kmart stores open", "b&m stores"]
    for r in range(1, min((ws.max_row or 300)+1, 500)):
        for c in range(1, 6):
            v = normalize(ws.cell(row=r, column=c).value)
            if any(kw in v for kw in STORE_KW):
                row_data = {"row": r, "label": ws.cell(row=r, column=c).value, "values": {}}
                for cc in range(c+1, min((ws.max_column or 100)+1, 200)):
                    vv = ws.cell(row=r, column=cc).value
                    if isinstance(vv, (int, float)) and vv > 0:
                        row_data["values"][cc] = vv
                if row_data["values"]:
                    results.append(row_data)
    return results

# ── B&M / SC IF sheets (store-level data) ────────────────────────────────────
def extract_bm_sheet(ws):
    """B&M Anita sheets may have store-by-store or brand-level inflows."""
    results = {}
    for r in range(1, 30):
        row_data = {}
        for c in range(1, 10):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row_data[c] = str(v)
        if row_data:
            vstr = " ".join(row_data.values()).lower()
            if any(k in vstr for k in ["sears", "kmart", "store count", "# stores", "b&m", "brick"]):
                results[r] = row_data
    return results

# ── Main extraction ───────────────────────────────────────────────────────────
def analyze_file(label, partial_name):
    filepath = find_file(partial_name)
    if not filepath:
        return {"label": label, "error": f"File not found: {partial_name}"}

    filename = os.path.basename(filepath)
    print(f"\n{'─'*60}\nProcessing: {label}\n  {filename}")

    result = {
        "label": label,
        "filename": filename,
        "monthly_actuals": None,      # from Inflows Actuals
        "weekly_forecast": None,      # from Inflows Forecasting
        "historical_monthly": None,   # from Historicals view / BU beta
        "store_counts": [],
        "bm_sheet_data": {},
        "territory_notes": [],
        "column_map": {},
        "errors": [],
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        sheets = wb.sheetnames
        print(f"  Sheets: {sheets[:10]}{'...' if len(sheets)>10 else ''}")

        for sn in sheets:
            try:
                ws = wb[sn]
                sn_lower = sn.lower()

                # Inflows Actuals
                if "inflows actuals" in sn_lower or "inflow actuals" in sn_lower:
                    monthly, sc, kc, hr = extract_inflows_actuals(ws, label)
                    if monthly:
                        result["monthly_actuals"] = monthly
                        result["column_map"]["actuals_sears_col"] = sc
                        result["column_map"]["actuals_kmart_col"] = kc
                        print(f"  ✓ Inflows Actuals: {len(monthly)} months, Sears col={sc}, Kmart col={kc}")

                # Inflows Forecasting
                if "inflows forecasting" in sn_lower or "inflow forecast" in sn_lower:
                    weekly = extract_inflows_forecasting(ws)
                    if weekly:
                        result["weekly_forecast"] = weekly
                        print(f"  ✓ Inflows Forecasting: {len(weekly)} weeks")

                # Historicals / BU beta
                if any(k in sn_lower for k in ["historicals", "historica", "bu beta"]):
                    hist = extract_historicals(ws, label)
                    if hist:
                        result["historical_monthly"] = hist
                        print(f"  ✓ Historical/BU beta: {len(hist)} months")

                # Store count scan on all sheets
                store_counts = extract_store_counts(ws)
                if store_counts:
                    result["store_counts"].extend([{"sheet": sn, **sc} for sc in store_counts])
                    print(f"  ✓ Store counts found in '{sn}': {[sc['label'] for sc in store_counts]}")

                # B&M sheets
                if "b&m" in sn_lower or "anita" in sn_lower or "sc if" in sn_lower:
                    bm = extract_bm_sheet(ws)
                    if bm:
                        result["bm_sheet_data"][sn] = bm
                        print(f"  ✓ B&M/SC IF sheet '{sn}': {len(bm)} relevant rows")

                # Territory check in all sheets
                for r in range(1, min((ws.max_row or 10)+1, 30)):
                    for c in range(1, min((ws.max_column or 10)+1, 100)):
                        v = normalize(ws.cell(row=r, column=c).value)
                        if any(k in v for k in ["guam", "usvi", "u.s.v.i", "virgin island"]):
                            result["territory_notes"].append(f"{sn} R{r}C{c}: {ws.cell(row=r,column=c).value}")

            except Exception as e:
                result["errors"].append(f"Sheet '{sn}': {e}")

        wb.close()

    except Exception as e:
        result["errors"].append(f"File open error: {e}")

    return result


# ── Summary builder ───────────────────────────────────────────────────────────
def build_longitudinal_table(all_results):
    """Build a clean longitudinal table from all results."""
    table = []
    for r in all_results:
        label = r["label"]

        # Preferred source: historical_monthly (already monthly totals)
        # Fallback: monthly_actuals (sum from daily data)

        sears_by_month = {}
        kmart_by_month = {}
        guam_by_month = {}
        usvi_by_month = {}

        if r.get("historical_monthly"):
            for entry in r["historical_monthly"]:
                m = entry["month"]
                if entry["sears"] is not None:
                    sears_by_month[m] = entry["sears"]
                if entry["kmart"] is not None:
                    kmart_by_month[m] = entry["kmart"]
                if entry.get("guam") is not None:
                    guam_by_month[m] = entry["guam"]
                if entry.get("usvi") is not None:
                    usvi_by_month[m] = entry["usvi"]

        if r.get("monthly_actuals"):
            for ym, d in r["monthly_actuals"].items():
                if ym not in sears_by_month and d["sears"]:
                    sears_by_month[ym] = round(d["sears"], 4)
                if ym not in kmart_by_month and d["kmart"]:
                    kmart_by_month[ym] = round(d["kmart"], 4)
                if ym not in guam_by_month and d.get("guam"):
                    guam_by_month[ym] = round(d["guam"], 4)
                if ym not in usvi_by_month and d.get("usvi"):
                    usvi_by_month[ym] = round(d["usvi"], 4)

        row = {
            "bod_label": label,
            "filename": r.get("filename", ""),
            "data_source": ("historical" if r.get("historical_monthly") else
                           "actuals" if r.get("monthly_actuals") else "none"),
            "sears_monthly_inflows_M": sears_by_month,
            "kmart_monthly_inflows_M": kmart_by_month,
            "guam_monthly_M": guam_by_month,
            "usvi_monthly_M": usvi_by_month,
            "store_count_data": r.get("store_counts", []),
            "territory_notes": r.get("territory_notes", [])[:20],
            "weekly_forecast_sample": r.get("weekly_forecast", [])[:10] if r.get("weekly_forecast") else None,
            "errors": r.get("errors", [])[:5],
        }
        table.append(row)
    return table


def main():
    all_results = []
    for label, partial in BOD_FILES:
        r = analyze_file(label, partial)
        all_results.append(r)

    longitudinal = build_longitudinal_table(all_results)

    output = {
        "generated": "2026-05-11",
        "description": "Sears and Kmart retail inflow analysis across 16 BOD files",
        "units": "Millions USD",
        "longitudinal_table": longitudinal,
        "raw_results": all_results,
    }

    out_path = os.path.join(BASE_DIR, "analysis_retail_stores.json")
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\n\n{'='*70}")
    print(f"SAVED: {out_path}")
    print(f"\nLONGITUDINAL SUMMARY ($ in millions):")
    print(f"{'BOD Label':30s} | {'Data Source':11s} | {'Sears Months':12s} | {'Kmart Months':12s} | {'Territory':10s}")
    print("-"*90)

    for row in longitudinal:
        sears_months = len(row["sears_monthly_inflows_M"])
        kmart_months = len(row["kmart_monthly_inflows_M"])
        has_terr = bool(row["guam_monthly_M"] or row["usvi_monthly_M"])
        print(f"{row['bod_label']:30s} | {row['data_source']:11s} | {sears_months:12d} | {kmart_months:12d} | {'YES' if has_terr else 'no':10s}")

    return all_results, longitudinal


if __name__ == "__main__":
    main()
