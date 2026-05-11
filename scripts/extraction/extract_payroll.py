"""
Payroll trajectory extraction from SP_Analysis daily cash forecast files.
Extracts monthly payroll actuals and most-recent-week data from each target file.
"""

import openpyxl
import datetime
import json
import os

BASE_DIR = "/Users/josh/Downloads/SP_Analysis"

TARGET_FILES = [
    ("9.19.22_Sep BOD",        "9/19/2022",  "Daily Cash Fcst - 9.19.22_BU view_Sep BOD.xlsx"),
    ("12.7.22_DecBOD",         "12/7/2022",  "Daily Cash Fcst - 12.7.22_BU view_DecBOD Treasury_liquidity update final.xlsx"),
    ("4.25.23_AprBOD",         "4/25/2023",  "Daily Cash Fcst - 4.25.23_BU view_April 27 BOD.xlsx"),
    ("8.22.23_AugBOD",         "8/22/2023",  "Daily Cash Fcst - 8.22.23_BU view_AugBOD.xlsx"),
    ("1.23.24_BOD",            "1/23/2024",  "Daily Cash Fcst - 1.23.24_BU view_BOD.xlsx"),
    ("7.23.24_JulBOD",         "7/23/2024",  "Daily Cash Fcst - 7.23.24_deck_JulBOD_FY25.xlsx"),
    ("12.31.24",               "12/31/2024", "Daily Cash Fcst - 12.31.24_BU view.xlsx"),
    ("1.21.25_JanBOD",         "1/21/2025",  "Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx"),
    ("5.20.25_MayBOD",         "5/20/2025",  "Daily Cash Fcst - 5.20.25_BU view_deck version_MayBOD.xlsx"),
    ("10.21.25_OctBOD",        "10/21/2025", "Daily Cash Fcst - 10.21.25_BU view_deck version_OctBOD.xlsx"),
    ("12.30.25_year_end",      "12/30/2025", "Daily Cash Fcst - 12.30.25_BU view_deck version final.xlsx"),
    ("02.17.26_deck",          "2/17/2026",  "Daily Cash Fcst - 02.17.26_BU view_deck version.xlsx"),
    ("03.03.26_post-reductions","3/3/2026",  "Daily Cash Fcst - 03.03.26_post-reductions_deck version_MarBOD.xlsx"),
    ("04.21.26_AprBOD",        "4/21/2026",  "Daily Cash Fcst - 04.21.26_ BU version_deck version_AprBOD.xlsx"),
    ("05.08.26",               "5/8/2026",   "Daily Cash Fcst - 05.08.26_BU view.xlsx"),
]


def safe_float(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def extract_payroll_data(label, file_date_str, fname):
    """
    Extract monthly payroll totals from the Payroll sheet.
    Returns a dict with actuals by month and BU breakdown.
    """
    filepath = os.path.join(BASE_DIR, fname)
    result = {
        "label": label,
        "file_date": file_date_str,
        "filename": fname,
        "payroll_sheet": {},
        "errors": [],
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        result["errors"].append(f"Cannot open file: {e}")
        return result

    # Find payroll sheet - prefer 'Payroll' main tab
    payroll_sheet_name = None
    for sname in wb.sheetnames:
        if sname.lower() == "payroll":
            payroll_sheet_name = sname
            break
    if payroll_sheet_name is None:
        for sname in wb.sheetnames:
            if "payroll" in sname.lower():
                payroll_sheet_name = sname
                break

    if payroll_sheet_name is None:
        result["errors"].append("No Payroll sheet found")
        wb.close()
        return result

    result["payroll_sheet"]["sheet_name"] = payroll_sheet_name

    try:
        ws = wb[payroll_sheet_name]

        # Load all rows up to 600 (covers ~4 years of daily data)
        all_rows = []
        for row in ws.iter_rows(max_row=600, max_col=40, values_only=True):
            all_rows.append(row)

        # Row 2 (index 1) = headers
        header_row = all_rows[1] if len(all_rows) > 1 else []

        # Find the forecast Total column (last non-None header in first 20 cols)
        # and actuals Total column
        forecast_total_col = None
        actuals_total_col = None
        forecast_headers = {}
        actuals_headers = {}

        # Header row: col A=banking days, then BU columns, then blank, then actuals BU columns
        # Find 'Total' in first 20 cols
        first_total_idx = None
        second_total_idx = None
        for j, v in enumerate(header_row[:40]):
            if v == 'Total':
                if first_total_idx is None:
                    first_total_idx = j
                elif second_total_idx is None:
                    second_total_idx = j

        if first_total_idx is not None:
            forecast_total_col = first_total_idx  # 0-indexed
        if second_total_idx is not None:
            actuals_total_col = second_total_idx  # 0-indexed

        # Build BU name -> column index maps for forecast and actuals sections
        if forecast_total_col is not None:
            for j in range(1, forecast_total_col + 1):
                v = header_row[j] if j < len(header_row) else None
                if v and v != 'Total':
                    forecast_headers[v] = j

        if actuals_total_col is not None:
            # Actuals section starts 2 cols after forecast Total
            act_start = forecast_total_col + 2 if forecast_total_col is not None else 18
            for j in range(act_start, actuals_total_col + 1):
                v = header_row[j] if j < len(header_row) else None
                if v and v != 'Total':
                    actuals_headers[v] = j

        result["payroll_sheet"]["forecast_total_col"] = forecast_total_col + 1 if forecast_total_col is not None else None  # 1-indexed for reporting
        result["payroll_sheet"]["actuals_total_col"] = actuals_total_col + 1 if actuals_total_col is not None else None
        result["payroll_sheet"]["forecast_bu_headers"] = list(forecast_headers.keys())
        result["payroll_sheet"]["actuals_bu_headers"] = list(actuals_headers.keys())

        # Find month label rows and monthly total rows
        # Month label: col A is a string like 'Aug 22', 'Sept 22', 'Jan 23', etc.
        # Monthly total: col A is None, but forecast_total_col has a value
        month_labels = {}   # row_idx -> label string
        monthly_totals = {}  # row_idx -> (prev_label, forecast_total, actuals_total)

        last_month_label = None
        last_month_label_idx = None

        for i, row in enumerate(all_rows):
            col_a = row[0] if len(row) > 0 else None

            # Check if this is a month label row
            if isinstance(col_a, str) and col_a not in ("Banking days",) and "Treasury" not in col_a:
                last_month_label = col_a
                last_month_label_idx = i
                month_labels[i] = col_a
                continue

            # Check if this is a monthly total row (no date in col A, but has numeric total)
            if col_a is None and forecast_total_col is not None:
                fcst_total = row[forecast_total_col] if forecast_total_col < len(row) else None
                act_total = row[actuals_total_col] if actuals_total_col is not None and actuals_total_col < len(row) else None

                if isinstance(fcst_total, (int, float)) and fcst_total != 0:
                    # Get BU breakdown for actuals
                    act_bu = {}
                    if actuals_headers and act_total is not None:
                        for bu, col_idx in actuals_headers.items():
                            v = row[col_idx] if col_idx < len(row) else None
                            if isinstance(v, (int, float)):
                                act_bu[bu] = round(v, 2)

                    # Get BU breakdown for forecast
                    fcst_bu = {}
                    for bu, col_idx in forecast_headers.items():
                        v = row[col_idx] if col_idx < len(row) else None
                        if isinstance(v, (int, float)):
                            fcst_bu[bu] = round(v, 2)

                    monthly_totals[i] = {
                        "month_label": last_month_label,
                        "forecast_total": round(fcst_total, 2),
                        "actuals_total": round(act_total, 2) if isinstance(act_total, (int, float)) else None,
                        "forecast_bu": fcst_bu,
                        "actuals_bu": act_bu,
                    }

        result["payroll_sheet"]["monthly_totals"] = monthly_totals

        # Find the most recent week with non-zero daily payroll data
        # Daily rows have a datetime in col A and numeric total in forecast_total_col
        last_date_with_data = None
        last_week_rows = []  # rows from last week (M-F or adjacent days)
        all_daily_rows = []

        for i, row in enumerate(all_rows):
            col_a = row[0] if len(row) > 0 else None
            if isinstance(col_a, datetime.datetime) and forecast_total_col is not None:
                fcst_total = row[forecast_total_col] if forecast_total_col < len(row) else None
                if isinstance(fcst_total, (int, float)) and fcst_total != 0:
                    all_daily_rows.append((col_a, i, row))

        if all_daily_rows:
            # Most recent date with data
            most_recent_date, most_recent_idx, most_recent_row = all_daily_rows[-1]
            last_date_with_data = most_recent_date.strftime("%Y-%m-%d")

            # Collect all rows within 7 days of most recent date
            week_start = most_recent_date - datetime.timedelta(days=6)
            recent_week = [(d, idx, r) for d, idx, r in all_daily_rows if d >= week_start]

            # Sum the week
            week_total_fcst = sum(r[forecast_total_col] for d, idx, r in recent_week
                                  if isinstance(r[forecast_total_col], (int, float)))
            week_total_act = None
            if actuals_total_col is not None:
                act_vals = [r[actuals_total_col] for d, idx, r in recent_week
                            if actuals_total_col < len(r) and isinstance(r[actuals_total_col], (int, float))]
                if act_vals:
                    week_total_act = sum(act_vals)

            # BU breakdown for the week (actuals if available, else forecast)
            week_bu_fcst = {}
            week_bu_act = {}
            for bu, col_idx in forecast_headers.items():
                total = sum(r[col_idx] for d, idx, r in recent_week
                            if col_idx < len(r) and isinstance(r[col_idx], (int, float)))
                if total != 0:
                    week_bu_fcst[bu] = round(total, 2)
            if actuals_headers and week_total_act is not None:
                for bu, col_idx in actuals_headers.items():
                    total = sum(r[col_idx] for d, idx, r in recent_week
                                if col_idx < len(r) and isinstance(r[col_idx], (int, float)))
                    if total != 0:
                        week_bu_act[bu] = round(total, 2)

            result["payroll_sheet"]["most_recent_week"] = {
                "week_end_date": last_date_with_data,
                "week_start_date": week_start.strftime("%Y-%m-%d"),
                "days_in_week": len(recent_week),
                "dates": [d.strftime("%Y-%m-%d") for d, idx, r in recent_week],
                "weekly_total_forecast": round(week_total_fcst, 2),
                "weekly_total_actuals": round(week_total_act, 2) if week_total_act is not None else None,
                "bu_breakdown_forecast": week_bu_fcst,
                "bu_breakdown_actuals": week_bu_act,
            }
            result["payroll_sheet"]["last_date_with_data"] = last_date_with_data

        # Also extract: monthly totals organized by month label with actuals preferred
        monthly_summary = []
        for row_idx in sorted(monthly_totals.keys()):
            entry = monthly_totals[row_idx]
            monthly_summary.append({
                "month": entry["month_label"],
                "forecast_monthly": entry["forecast_total"],
                "actuals_monthly": entry["actuals_total"],
                "preferred_monthly": entry["actuals_total"] if entry["actuals_total"] is not None else entry["forecast_total"],
                "actuals_bu": entry["actuals_bu"],
                "forecast_bu": entry["forecast_bu"],
            })

        result["payroll_sheet"]["monthly_summary"] = monthly_summary

        # Key: find the most recent month that has ACTUALS (not None)
        months_with_actuals = [m for m in monthly_summary if m["actuals_monthly"] is not None]
        if months_with_actuals:
            most_recent_actual_month = months_with_actuals[-1]
            result["payroll_sheet"]["most_recent_actual_month"] = most_recent_actual_month

        # And the as-of-date month (could be forecast if actuals not yet available)
        if monthly_summary:
            result["payroll_sheet"]["latest_month_available"] = monthly_summary[-1]

    except Exception as e:
        import traceback
        result["errors"].append(f"Extraction error: {e}\n{traceback.format_exc()}")

    wb.close()
    return result


def main():
    all_results = []

    for label, file_date, fname in TARGET_FILES:
        print(f"Processing: {label}...")
        data = extract_payroll_data(label, file_date, fname)
        all_results.append(data)
        if data["errors"]:
            print(f"  ERRORS: {data['errors']}")

    # Build the payroll timeline
    print("\n" + "="*80)
    print("PAYROLL TIMELINE SUMMARY")
    print("="*80)

    timeline = []
    for r in all_results:
        label = r["label"]
        file_date = r["file_date"]
        ps = r.get("payroll_sheet", {})

        # Get the most recent actual month data
        most_recent_actual = ps.get("most_recent_actual_month", {})
        latest_available = ps.get("latest_month_available", {})
        most_recent_week = ps.get("most_recent_week", {})

        # Monthly payroll (actuals preferred)
        monthly_payroll = None
        monthly_payroll_source = None
        monthly_bu_breakdown = None
        monthly_month_label = None

        if most_recent_actual:
            monthly_payroll = most_recent_actual.get("actuals_monthly")
            monthly_month_label = most_recent_actual.get("month")
            monthly_payroll_source = "actuals"
            monthly_bu_breakdown = most_recent_actual.get("actuals_bu") or most_recent_actual.get("forecast_bu")
        elif latest_available:
            monthly_payroll = latest_available.get("forecast_monthly")
            monthly_month_label = latest_available.get("month")
            monthly_payroll_source = "forecast"
            monthly_bu_breakdown = latest_available.get("forecast_bu")

        # Weekly payroll
        weekly_total = None
        weekly_date_range = None
        weekly_bu = None
        if most_recent_week:
            weekly_total = most_recent_week.get("weekly_total_actuals") or most_recent_week.get("weekly_total_forecast")
            weekly_date_range = f"{most_recent_week.get('week_start_date')} to {most_recent_week.get('week_end_date')}"
            weekly_bu = most_recent_week.get("bu_breakdown_actuals") or most_recent_week.get("bu_breakdown_forecast")

        entry = {
            "label": label,
            "file_date": file_date,
            "monthly_payroll_month": monthly_month_label,
            "monthly_payroll_amount": monthly_payroll,
            "monthly_payroll_source": monthly_payroll_source,
            "monthly_bu_breakdown": monthly_bu_breakdown,
            "weekly_payroll_total": weekly_total,
            "weekly_date_range": weekly_date_range,
            "weekly_bu_breakdown": weekly_bu,
            "last_date_with_data": ps.get("last_date_with_data"),
            "errors": r.get("errors", []),
        }
        timeline.append(entry)

        # Print summary line
        mp_str = f"${abs(monthly_payroll)/1e6:.2f}M/mo" if monthly_payroll else "N/A"
        wp_str = f"${abs(weekly_total)/1e6:.2f}M/wk" if weekly_total else "N/A"
        print(f"\n{label} ({file_date})")
        print(f"  Most recent actual month: {monthly_month_label} -> {mp_str} [{monthly_payroll_source}]")
        print(f"  Most recent week ({weekly_date_range}): {wp_str}")
        if monthly_bu_breakdown:
            # Print top BUs
            sorted_bu = sorted(monthly_bu_breakdown.items(), key=lambda x: abs(x[1] or 0), reverse=True)
            print(f"  Top BU breakdown (monthly):")
            for bu, amt in sorted_bu[:8]:
                if amt:
                    print(f"    {bu}: ${abs(amt)/1e6:.2f}M")

    # Key questions analysis
    print("\n" + "="*80)
    print("KEY ANALYSIS: Sep 2022 vs May 2026")
    print("="*80)

    sep22 = next((t for t in timeline if t["label"] == "9.19.22_Sep BOD"), None)
    may26 = next((t for t in timeline if t["label"] == "05.08.26"), None)

    if sep22 and may26:
        sep22_monthly = sep22.get("monthly_payroll_amount")
        may26_monthly = may26.get("monthly_payroll_amount")

        print(f"\nSep 2022 monthly payroll ({sep22['monthly_payroll_month']}): ${abs(sep22_monthly)/1e6:.2f}M" if sep22_monthly else "Sep 2022: N/A")
        print(f"May 2026 monthly payroll ({may26['monthly_payroll_month']}): ${abs(may26_monthly)/1e6:.2f}M" if may26_monthly else "May 2026: N/A")

        if sep22_monthly and may26_monthly:
            reduction = abs(sep22_monthly) - abs(may26_monthly)
            pct = reduction / abs(sep22_monthly) * 100
            print(f"Reduction: ${reduction/1e6:.2f}M/mo ({pct:.1f}%)")

            # Implied headcount (assuming $70K avg annual salary + ~30% benefits = ~$91K total comp)
            # Monthly cost per employee = $91K / 12 = $7,583
            avg_annual_total_comp = 85000  # industry avg for retail/tech mix
            monthly_cost_per_head = avg_annual_total_comp / 12

            sep22_headcount = abs(sep22_monthly) / monthly_cost_per_head
            may26_headcount = abs(may26_monthly) / monthly_cost_per_head
            print(f"\nImplied headcount (@${avg_annual_total_comp:,} avg annual total comp):")
            print(f"  Sep 2022: ~{sep22_headcount:,.0f} employees")
            print(f"  May 2026: ~{may26_headcount:,.0f} employees")
            print(f"  Reduction: ~{sep22_headcount - may26_headcount:,.0f} employees")

    # Apr 2026 vs May 2026 comparison (pre vs post payroll change)
    print("\n" + "="*80)
    print("PRE vs POST PAYROLL CHANGE: Apr 21, 2026 vs May 8, 2026")
    print("="*80)
    apr26 = next((t for t in timeline if t["label"] == "04.21.26_AprBOD"), None)

    if apr26 and may26:
        apr26_weekly = apr26.get("weekly_payroll_total")
        may26_weekly = may26.get("weekly_payroll_total")
        print(f"\nApr 21, 2026 (PRE) weekly payroll ({apr26['weekly_date_range']}): ${abs(apr26_weekly)/1e6:.2f}M" if apr26_weekly else "Apr 26: N/A")
        print(f"May 8, 2026 (POST) weekly payroll ({may26['weekly_date_range']}): ${abs(may26_weekly)/1e6:.2f}M" if may26_weekly else "May 26: N/A")
        if apr26_weekly and may26_weekly:
            weekly_change = abs(apr26_weekly) - abs(may26_weekly)
            monthly_change = weekly_change * 52 / 12
            print(f"Weekly change: ${weekly_change/1e6:.2f}M reduction")
            print(f"Annualized: ${weekly_change*52/1e6:.2f}M | Monthly equivalent: ${monthly_change/1e6:.2f}M")

    # Build output JSON
    output = {
        "analysis_date": "2026-05-11",
        "description": "Payroll cost trajectory 2022-2026 extracted from daily cash forecast files",
        "payroll_timeline": timeline,
        "raw_file_data": all_results,
        "key_findings": {
            "sep_2022_monthly_payroll": sep22.get("monthly_payroll_amount") if sep22 else None,
            "sep_2022_month_label": sep22.get("monthly_payroll_month") if sep22 else None,
            "may_2026_monthly_payroll": may26.get("monthly_payroll_amount") if may26 else None,
            "may_2026_month_label": may26.get("monthly_payroll_month") if may26 else None,
            "total_reduction_dollars": (abs(sep22.get("monthly_payroll_amount") or 0) - abs(may26.get("monthly_payroll_amount") or 0)) if sep22 and may26 else None,
            "reduction_pct": None,
            "implied_headcount_note": "Based on $85K avg annual total compensation (salary + benefits)",
        }
    }

    if output["key_findings"]["sep_2022_monthly_payroll"] and output["key_findings"]["may_2026_monthly_payroll"]:
        sep_abs = abs(output["key_findings"]["sep_2022_monthly_payroll"])
        may_abs = abs(output["key_findings"]["may_2026_monthly_payroll"])
        output["key_findings"]["reduction_pct"] = round((sep_abs - may_abs) / sep_abs * 100, 1)

    out_path = os.path.join(BASE_DIR, "analysis_payroll_headcount.json")
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\nOutput saved to: {out_path}")
    return output


if __name__ == "__main__":
    main()
