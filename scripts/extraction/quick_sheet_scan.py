"""Quick scan to understand sheet structure and find Sears/Kmart columns in first 2 files."""
import openpyxl
import os

BASE_DIR = "/Users/josh/Downloads/SP_Analysis"

def scan_first_rows(filepath, label, max_rows=10, max_cols=80):
    print(f"\n{'='*70}")
    print(f"FILE: {label}")
    print(f"  {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    print(f"  Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  --- Sheet: {sn} ---")
        for row in range(1, min(max_rows+1, (ws.max_row or 10)+1)):
            row_vals = []
            for col in range(1, min(max_cols+1, (ws.max_column or 10)+1)):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    row_vals.append(f"[{col}]{v}")
            if row_vals:
                line = " | ".join(str(x) for x in row_vals[:20])
                print(f"    R{row}: {line[:200]}")
    wb.close()

# Scan a few files
files = [
    ("9.19.22_Sep BOD", "9.19.22_BU view_Sep BOD"),
    ("05.08.26_BU view", "05.08.26_BU view"),
    ("7.22.25_JulBOD", "7.22.25_BU view_deck version_condensed_JulBOD"),
]

for label, partial in files:
    for fn in os.listdir(BASE_DIR):
        if partial.lower() in fn.lower() and fn.endswith(".xlsx"):
            scan_first_rows(os.path.join(BASE_DIR, fn), label)
            break
