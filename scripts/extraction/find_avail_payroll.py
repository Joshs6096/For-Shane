"""
Find Available Cash and payroll data in the Cash Flow sheet for all 9 files.
The Cash Flow sheet has daily rows with Available Cash column and weekly/monthly subtotals.
"""
import openpyxl
from pathlib import Path

BASE = Path("/Users/josh/Downloads/SP_Analysis")

FILES = [
    ("2024-09-24", "Daily Cash Fcst - 9.24.24_BU view_deck version_BOD.xlsx"),
    ("2024-10-15", "Daily Cash Fcst - 10.15.24_BU view_deck version_OctBOD.xlsx"),
    ("2024-11-12", "Daily Cash Fcst - 11.12.24_BU view_deck version_NovBOD.xlsx"),
    ("2024-12-10", "Daily Cash Fcst - 12.10.24_BU view_deck version_BOD_SC wo 3 RE Interco IF_$9M.xlsx"),
    ("2024-12-31", "Daily Cash Fcst - 12.31.24_BU view.xlsx"),
    ("2025-01-21", "Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx"),
    ("2025-03-18", "Daily Cash Fcst - 3.18.25_BU view_deck version_MarBOD.xlsx"),
    ("2025-04-18", "Daily Cash Fcst - 4.18.25_BU view_prelim AprBOD.xlsx"),
    ("2026-05-08", "Daily Cash Fcst - 05.08.26_BU view.xlsx"),
]

print("="*70)
print("CASH FLOW SHEET: Headers and first 3 data rows (to understand structure)")

for date_str, filename in FILES:
    filepath = BASE / filename
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    print(f"\n{date_str}: {filename[:45]}")

    # Look at Cash Flow sheet structure
    cf_name = None
    for sname in wb.sheetnames:
        if sname.lower() == "cash flow":
            cf_name = sname
            break

    if not cf_name:
        print("  No 'Cash Flow' sheet found")
        wb.close()
        continue

    ws = wb[cf_name]

    # Print first 5 rows (headers)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        non_none = [str(v)[:20] for v in row[:20] if v is not None]
        if non_none:
            print(f"  R{i}: {non_none[:12]}")

    # Look for "Available Cash", "Total Cash" in any row - scan more broadly
    # These files have the Cash Flow in COLUMNS not rows for each BU
    # Let's look at what row labels exist in col A
    print(f"  Col A labels (rows 1-15):")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=5, values_only=True), 1):
        if any(v is not None for v in row):
            print(f"    R{i}: {[str(v)[:25] for v in row if v is not None]}")

    # Look for payroll rows — scan rows with monthly summary-like data
    # In the Cash Flow sheet, there are monthly summary rows
    print(f"  Searching for payroll/available rows:")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=300, values_only=True), 1):
        if row is None: continue
        text = " ".join(str(v).lower() for v in row[:15] if v is not None and isinstance(v, str))
        if any(kw in text for kw in ["available", "total cash", "unavailable", "payroll/bens", "pyrl/bens", "payroll & bens"]):
            nums = [round(float(v), 3) for v in row if isinstance(v, (int, float)) and abs(float(v)) > 0.001]
            print(f"    R{i}: [{text[:60]}] -> {nums[:10]}")

    wb.close()

# Also check 2026 file for payroll
print("\n\n" + "="*70)
print("MAY 2026: All sheets with 'payroll' in name")
wb = openpyxl.load_workbook(BASE / "Daily Cash Fcst - 05.08.26_BU view.xlsx", data_only=True, read_only=True)
payroll_sheets = [s for s in wb.sheetnames if "payroll" in s.lower() or "pay" in s.lower() or "bens" in s.lower()]
print(f"Payroll-related sheets: {payroll_sheets}")

# Also check SHS Cash Changes sheet
shs_sheets = [s for s in wb.sheetnames if "shs" in s.lower() or "cash changes" in s.lower()]
print(f"SHS Cash Changes sheets: {shs_sheets}")

for sname in shs_sheets[:2]:
    ws = wb[sname]
    print(f"\n  Sheet: {sname}")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True), 1):
        if row is None: continue
        text = " ".join(str(v).lower() for v in row[:10] if v is not None and isinstance(v, str))
        nums = [round(float(v), 3) for v in row if isinstance(v, (int, float)) and abs(float(v)) > 0.1]
        if "payroll" in text or "bens" in text:
            print(f"    R{i}: [{text[:80]}] -> {nums[:10]}")

wb.close()
