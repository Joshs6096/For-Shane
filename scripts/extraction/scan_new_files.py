"""Scan structure of newer BOD files (2024-2026) to understand format changes."""
import openpyxl
import os

BASE_DIR = "/Users/josh/Downloads/SP_Analysis"

def scan_first_rows(filepath, label, max_rows=15, max_cols=80):
    print(f"\n{'='*70}")
    print(f"FILE: {label}")
    print(f"  {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    print(f"  Sheets: {wb.sheetnames}")
    # Focus on first 3 sheets and any with "inflow", "retail", "kmart", "sears" in name
    interesting = []
    for sn in wb.sheetnames:
        norm = sn.lower()
        if any(k in norm for k in ["inflow", "retail", "kmart", "sears", "store", "historica", "actuals", "detail", "bu beta", "summary"]):
            interesting.append(sn)
    # Always show first sheet
    first_sheets = wb.sheetnames[:2] + [s for s in interesting if s not in wb.sheetnames[:2]]
    for sn in first_sheets[:6]:
        ws = wb[sn]
        print(f"\n  --- Sheet: {sn} ---")
        for row in range(1, min(max_rows+1, (ws.max_row or 10)+1)):
            row_vals = []
            for col in range(1, min(max_cols+1, (ws.max_column or 10)+1)):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    row_vals.append(f"[{col}]{v}")
            if row_vals:
                line = " | ".join(str(x) for x in row_vals[:20])
                print(f"    R{row}: {line[:250]}")
    wb.close()

# Scan newer files
files = [
    ("05.08.26_BU view", "05.08.26_BU view"),
    ("04.21.26_AprBOD", "04.21.26_ BU version_deck version_AprBOD"),
    ("01.16.26_JanBOD", "01.16.26_BU view new format_asset sales_JanBOD"),
    ("12.15.25_DecBOD", "12.15.25_BU view_DecBOD"),
    ("7.22.25_JulBOD", "7.22.25_BU view_deck version_condensed_JulBOD"),
    ("3.18.25_MarBOD", "3.18.25_BU view_deck version_MarBOD"),
    ("12.31.24_BU view", "12.31.24_BU view"),
    ("10.15.24_OctBOD", "10.15.24_BU view_deck version_OctBOD"),
    ("6.18.24_JunBOD", "6.18.24_BU view_JunBOD"),
    ("2.20.24_BOD", "2.20.24_BU view_BOD"),
    ("11.14.23_NovBOD", "11.14.23_BU view_adUPDATE"),
    ("7.25.23_JulBOD", "7.25.23_BU view_JulBOD"),
    ("4.25.23_AprBOD", "4.25.23_BU view_April 27 BOD"),
    ("1.17.23_JanBOD", "1.17.23_BU view_JanBOD"),
    ("12.7.22_DecBOD", "12.7.22_BU view_DecBOD"),
]

for label, partial in files:
    for fn in os.listdir(BASE_DIR):
        if partial.lower() in fn.lower() and fn.endswith(".xlsx"):
            scan_first_rows(os.path.join(BASE_DIR, fn), label)
            break
    else:
        print(f"\n{'='*70}")
        print(f"FILE NOT FOUND: {label} (pattern: {partial})")
