import openpyxl
import json
import os
from openpyxl import load_workbook

BASE = "/Users/josh/Downloads/SP_Analysis/"

FILES = {
    "1.17.23_JanBOD": BASE + "Daily Cash Fcst - 1.17.23_BU view_JanBOD.xlsx",
    "1.31.23_BohemiaAdj": BASE + "Daily Cash Fcst - 1.31.23_BU view_Bohemia Adj New Methodology_Use for FY23a.xlsx",
    "6.18.24_JunBOD": BASE + "Daily Cash Fcst - 6.18.24_BU view_JunBOD.xlsx",
    "7.23.24_JulBOD_FY25": BASE + "Daily Cash Fcst - 7.23.24_deck_JulBOD_FY25.xlsx",
    "12.30.25_year_end": BASE + "Daily Cash Fcst - 12.30.25_BU view_deck version final.xlsx",
    "01.16.26_JanBOD_new_format": BASE + "Daily Cash Fcst - 01.16.26_BU view new format_asset sales_JanBOD.xlsx",
    "05.08.26_BU_view": BASE + "Daily Cash Fcst - 05.08.26_BU view.xlsx",
}

def safe_val(cell):
    v = cell.value
    if v is None:
        return None
    return str(v).strip() if isinstance(v, str) else v

def get_sheet_info(wb, ws):
    """Extract row1, row2 headers and scan for date range and row count."""
    row1 = []
    row2 = []
    max_col = ws.max_column or 0
    max_row = ws.max_row or 0
    
    # Cap at 300 columns
    max_col = min(max_col, 300)
    
    for col in range(1, max_col + 1):
        r1 = safe_val(ws.cell(row=1, column=col))
        r2 = safe_val(ws.cell(row=2, column=col))
        row1.append(r1)
        row2.append(r2)
    
    # Find non-None entries
    row1_clean = [v for v in row1 if v is not None]
    row2_clean = [v for v in row2 if v is not None]
    
    return {
        "row1": row1,
        "row2": row2,
        "row1_clean": row1_clean,
        "row2_clean": row2_clean,
        "max_col": max_col,
        "max_row": max_row,
    }

def find_fy_file_sheet(wb):
    """Find a sheet that looks like the FY File."""
    candidates = []
    for name in wb.sheetnames:
        ws = wb[name]
        state = ws.sheet_state if hasattr(ws, 'sheet_state') else 'visible'
        nl = name.lower()
        if 'fy' in nl or 'file' in nl or 'master' in nl or 'spine' in nl:
            candidates.append((name, state))
    return candidates

def search_term_all_sheets(wb, terms):
    """Search all sheets for specific terms. Returns list of hits."""
    hits = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        mr = min(ws.max_row or 0, 500)
        mc = min(ws.max_column or 0, 200)
        for row in ws.iter_rows(min_row=1, max_row=mr, min_col=1, max_col=mc):
            for cell in row:
                v = cell.value
                if v is not None:
                    vs = str(v).lower()
                    for term in terms:
                        if term.lower() in vs:
                            hits.append({
                                "sheet": sheet_name,
                                "row": cell.row,
                                "col": cell.column,
                                "value": str(v)[:200],
                                "term_matched": term
                            })
    return hits

def analyze_file(key, path):
    print(f"\n=== Analyzing: {key} ===")
    result = {
        "path": path,
        "all_sheets": [],
        "hidden_sheets": [],
        "visible_sheets": [],
        "fy_file_candidates": [],
        "fy_file_data": {},
    }
    
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        result["error"] = str(e)
        return result
    
    for name in wb.sheetnames:
        ws = wb[name]
        state = getattr(ws, 'sheet_state', 'visible')
        result["all_sheets"].append({"name": name, "state": state})
        if state == 'hidden':
            result["hidden_sheets"].append(name)
        else:
            result["visible_sheets"].append(name)
    
    print(f"  Sheets: {[s['name'] for s in result['all_sheets']]}")
    
    # Find FY File sheet candidates
    fy_candidates = find_fy_file_sheet(wb)
    result["fy_file_candidates"] = [{"name": n, "state": s} for n, s in fy_candidates]
    
    # Also check for sheet literally named "FY File" or similar
    for sheet_info in result["all_sheets"]:
        name = sheet_info["name"]
        nl = name.lower()
        if 'fy' in nl or 'file' in nl:
            if name not in [c["name"] for c in result["fy_file_candidates"]]:
                result["fy_file_candidates"].append({"name": name, "state": sheet_info["state"]})
    
    # Analyze each FY File candidate
    for cand in result["fy_file_candidates"]:
        sname = cand["name"]
        try:
            ws = wb[sname]
            info = get_sheet_info(wb, ws)
            result["fy_file_data"][sname] = info
            print(f"  FY File candidate '{sname}': {len(info['row1_clean'])} col headers, {info['max_row']} rows")
        except Exception as e:
            result["fy_file_data"][sname] = {"error": str(e)}
    
    wb.close()
    return result

def analyze_bohemia(path_old, path_new):
    """Deep dive on Bohemia references and methodology change."""
    result = {}
    
    for label, path in [("old_1.17.23", path_old), ("new_1.31.23", path_new)]:
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            hits = search_term_all_sheets(wb, ["bohemia", "methodology", "new method", "adj"])
            result[label] = {
                "sheet_names": wb.sheetnames,
                "bohemia_hits": [h for h in hits if "bohemia" in h["term_matched"].lower()],
                "methodology_hits": [h for h in hits if "methodology" in h["term_matched"].lower()],
            }
            wb.close()
        except Exception as e:
            result[label] = {"error": str(e)}
    
    return result

def analyze_fy25_intro(path_jun, path_jul):
    """Find when FY25 appears."""
    result = {}
    
    for label, path in [("jun_2024", path_jun), ("jul_2024", path_jul)]:
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            hits = search_term_all_sheets(wb, ["FY25", "FY 25", "2025", "fy2025"])
            result[label] = {
                "sheet_names": wb.sheetnames,
                "fy25_hits": hits[:50],  # cap
            }
            # Also get first 3 rows of each sheet to see structure
            sheet_structures = {}
            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = []
                for r in range(1, min(4, (ws.max_row or 0) + 1)):
                    row_vals = []
                    for c in range(1, min(20, (ws.max_column or 0) + 1)):
                        row_vals.append(safe_val(ws.cell(row=r, column=c)))
                    rows.append(row_vals)
                sheet_structures[sname] = rows
            result[label]["sheet_structures_top3rows"] = sheet_structures
            wb.close()
        except Exception as e:
            result[label] = {"error": str(e)}
    
    return result

def analyze_asset_sales(path_dec, path_jan):
    """Find asset sale references and cash flow structural changes."""
    result = {}
    
    for label, path in [("dec_2025", path_dec), ("jan_2026", path_jan)]:
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            hits = search_term_all_sheets(wb, ["asset sale", "asset sales", "proceeds", "disposition", "real estate", "RE sale"])
            
            # Get cash flow sheet structure
            cf_data = {}
            for sname in wb.sheetnames:
                nl = sname.lower()
                if 'cash' in nl or 'cf' in nl or 'flow' in nl:
                    ws = wb[sname]
                    rows = []
                    for r in range(1, min(60, (ws.max_row or 0) + 1)):
                        row_vals = []
                        for c in range(1, min(10, (ws.max_column or 0) + 1)):
                            row_vals.append(safe_val(ws.cell(row=r, column=c)))
                        rows.append(row_vals)
                    cf_data[sname] = rows
            
            result[label] = {
                "sheet_names": wb.sheetnames,
                "asset_sale_hits": hits[:60],
                "cash_flow_sheet_data": cf_data,
            }
            wb.close()
        except Exception as e:
            result[label] = {"error": str(e)}
    
    return result

def get_full_fy_file_columns(path, label):
    """Get complete column headers from FY File sheet."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        result = {"label": label, "sheets_checked": []}
        
        for sname in wb.sheetnames:
            nl = sname.lower()
            ws = wb[sname]
            state = getattr(ws, 'sheet_state', 'visible')
            
            if 'fy' in nl or 'file' in nl or 'master' in nl:
                mc = min(ws.max_column or 0, 300)
                mr = ws.max_row or 0
                
                r1 = [safe_val(ws.cell(row=1, column=c)) for c in range(1, mc+1)]
                r2 = [safe_val(ws.cell(row=2, column=c)) for c in range(1, mc+1)]
                
                # Scan column A for row labels
                col_a = [safe_val(ws.cell(row=r, column=1)) for r in range(1, min(mr+1, 200))]
                
                result["sheets_checked"].append({
                    "sheet_name": sname,
                    "state": state,
                    "max_col": mc,
                    "max_row": mr,
                    "row1_headers": r1,
                    "row2_headers": r2,
                    "col_a_labels": col_a,
                })
        
        wb.close()
        return result
    except Exception as e:
        return {"error": str(e)}

# Main analysis
print("Starting analysis...")

output = {
    "analysis_date": "2026-05-11",
    "files": {},
    "pair1_bohemia_analysis": {},
    "pair2_fy25_analysis": {},
    "pair3_asset_sales_analysis": {},
    "fy_file_evolution": {},
}

# Analyze each file
for key, path in FILES.items():
    output["files"][key] = analyze_file(key, path)

# Pair 1: Bohemia deep dive
print("\n=== Pair 1: Bohemia Analysis ===")
output["pair1_bohemia_analysis"] = analyze_bohemia(
    FILES["1.17.23_JanBOD"],
    FILES["1.31.23_BohemiaAdj"]
)

# Pair 2: FY25 introduction
print("\n=== Pair 2: FY25 Introduction ===")
output["pair2_fy25_analysis"] = analyze_fy25_intro(
    FILES["6.18.24_JunBOD"],
    FILES["7.23.24_JulBOD_FY25"]
)

# Pair 3: Asset sales / new format
print("\n=== Pair 3: Asset Sales / New Format ===")
output["pair3_asset_sales_analysis"] = analyze_asset_sales(
    FILES["12.30.25_year_end"],
    FILES["01.16.26_JanBOD_new_format"]
)

# FY File column evolution
print("\n=== FY File Column Evolution ===")
output["fy_file_evolution"] = {
    "earliest_1.17.23": get_full_fy_file_columns(FILES["1.17.23_JanBOD"], "1.17.23_JanBOD"),
    "latest_05.08.26": get_full_fy_file_columns(FILES["05.08.26_BU_view"], "05.08.26_BU_view"),
}

# Save
out_path = BASE + "analysis_methodology_evolution.json"
with open(out_path, "w") as f:
    json.dump(output, f, indent=2, default=str)

print(f"\n=== SAVED TO: {out_path} ===")
print(f"Total size: {os.path.getsize(out_path):,} bytes")
