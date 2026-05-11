"""
Inspect the 'Net Cash Flow' sheet row-by-row to find Available Cash labels.
"""
import openpyxl
from pathlib import Path

BASE = Path("/Users/josh/Downloads/SP_Analysis")

FILES = [
    ("2024-09-24", "Daily Cash Fcst - 9.24.24_BU view_deck version_BOD.xlsx"),
    ("2024-12-10", "Daily Cash Fcst - 12.10.24_BU view_deck version_BOD_SC wo 3 RE Interco IF_$9M.xlsx"),
    ("2026-05-08", "Daily Cash Fcst - 05.08.26_BU view.xlsx"),
]

for date_str, filename in FILES:
    filepath = BASE / filename
    print(f"\n{'='*60}")
    print(f"{date_str}: {filename[:50]}")
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    print(f"Sheets: {wb.sheetnames[:5]}")

    # Focus on Net Cash Flow sheet
    target_sheets = [s for s in wb.sheetnames if "net cash" in s.lower() or "liquidity" in s.lower() or "treasury summ" in s.lower()]
    if not target_sheets:
        target_sheets = [wb.sheetnames[0]]
    print(f"Target sheets: {target_sheets}")

    for sheet_name in target_sheets[:2]:
        ws = wb[sheet_name]
        print(f"\n  Sheet: {sheet_name}")
        # print(f"  Dims: {ws.dimensions}")

        # Read first 80 rows, first 100 cols
        rows_list = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=80, values_only=True), 1):
            rows_list.append((i, list(row)))

        for i, row in rows_list:
            # Check if row has any interesting text
            text_vals = [v for v in row[:10] if v is not None and isinstance(v, str)]
            num_vals = [v for v in row[5:] if v is not None and isinstance(v, (int, float))]
            if text_vals or (num_vals and any(abs(v) > 0.001 for v in num_vals)):
                # Print label cells and first 15 numeric values
                label = " | ".join(str(v)[:30] for v in row[:5] if v is not None)
                nums = [round(float(v), 3) for v in row[5:] if v is not None and isinstance(v, (int, float))][:15]
                print(f"    R{i}: [{label[:60]}] nums={nums}")

    wb.close()
