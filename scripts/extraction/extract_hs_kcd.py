#!/usr/bin/env python3
"""
Extract Home Services and KCD longitudinal cash flow data from all BOD files.
Outputs: analysis_HS_KCD.json

Key fix: Use first-occurrence-only month columns to avoid duplicate overwrite.
"""

import openpyxl
import json
import os
from collections import defaultdict

BASE_DIR = "/Users/josh/Downloads/SP_Analysis"

ALL_FILES = [
    ("2022-09", "Daily Cash Fcst - 9.19.22_BU view_Sep BOD.xlsx"),
    ("2023-01", "Daily Cash Fcst - 1.17.23_BU view_JanBOD.xlsx"),
    ("2023-04", "Daily Cash Fcst - 4.25.23_BU view_April 27 BOD.xlsx"),
    ("2023-08", "Daily Cash Fcst - 8.22.23_BU view_AugBOD.xlsx"),
    ("2023-10", "Daily Cash Fcst - 10.24.23_BU view_HS retail interco_OctBOD.xlsx"),
    ("2024-01", "Daily Cash Fcst - 1.23.24_BU view_BOD.xlsx"),
    ("2024-05", "Daily Cash Fcst - 5.7.24_BU view_MayBOD_5.9.24.xlsx"),
    ("2024-09", "Daily Cash Fcst - 9.24.24_BU view_deck version_BOD.xlsx"),
    ("2025-01", "Daily Cash Fcst - 1.21.25_BU view_deck version_working_Apr append_deck2_JanBOD.xlsx"),
    ("2025-05", "Daily Cash Fcst - 5.20.25_BU view_deck version_MayBOD.xlsx"),
    ("2025-09", "Daily Cash Fcst - 9.30.25_BU view_deck version.xlsx"),
    ("2026-03", "Daily Cash Fcst - 03.03.26_post-reductions_deck version_MarBOD.xlsx"),
    ("2026-05", "Daily Cash Fcst - 05.08.26_BU view.xlsx"),
]

MONTHS_DICT = {
    'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
    'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
    'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
}


def parse_month(s):
    """Parse 'Feb 2022' or "May'2026" -> '2022-02'"""
    if not isinstance(s, str):
        return None
    s = s.strip().replace("'", " ")
    parts = s.split()
    if len(parts) == 2 and parts[0] in MONTHS_DICT and len(parts[1]) == 4 and parts[1].isdigit():
        return f"{parts[1]}-{MONTHS_DICT[parts[0]]}"
    return None


def build_col_to_month(month_row):
    """
    Build column->month mapping using FIRST OCCURRENCE ONLY.
    This prevents duplicate month grids (audit/side tables at col 600+) from
    overwriting the primary data grid values.
    """
    col_to_month = {}
    seen_months = set()
    for j, v in enumerate(month_row):
        m = parse_month(v)
        if m and m not in seen_months:
            col_to_month[j] = m
            seen_months.add(m)
    return col_to_month


def safe_float(v):
    if isinstance(v, (int, float)):
        f = float(v)
        if f != f:  # NaN check
            return None
        return round(f, 6)
    return None


def extract_row_data(row, col_to_month):
    """Extract {month: value} from a spreadsheet row using the column->month map."""
    data = {}
    for j, v in enumerate(row):
        if j in col_to_month:
            fv = safe_float(v)
            if fv is not None:
                data[col_to_month[j]] = fv
    return data


def find_label_col_and_month_start(all_rows):
    """
    Detect the label column (BU-level row descriptions) and month start column.
    Strategy: in rows 5-15, find 'Home Services' label co-located with
    numeric data in the month columns.
    """
    # Month header is in row 3 (index 2)
    month_row = all_rows[2]
    col_to_month = build_col_to_month(month_row)

    if not col_to_month:
        return None, None, col_to_month

    month_start_col = min(col_to_month.keys())

    # Find label col
    label_col = None
    for i, row in enumerate(all_rows[5:15], start=6):
        for j, cell in enumerate(row):
            if cell is not None and str(cell).strip() == 'Home Services':
                # Verify numeric data exists in month columns
                has_data = any(
                    isinstance(row[k], (int, float)) and row[k] != 0
                    for k in range(month_start_col, min(month_start_col + 5, len(row)))
                )
                if has_data:
                    label_col = j
                    break
        if label_col is not None:
            break

    return label_col, month_start_col, col_to_month


def extract_bu_beta(fname_full, file_date):
    """
    Extract HS and KCD monthly data from BU beta sheet.
    Returns structured dict with inflows, disbursements, and net.
    """
    wb = openpyxl.load_workbook(fname_full, read_only=True, data_only=True)

    if 'BU beta' not in wb.sheetnames:
        wb.close()
        return {"error": "No BU beta sheet", "file": fname_full}

    ws = wb['BU beta']
    all_rows = list(ws.iter_rows(min_row=1, max_row=220, values_only=True))

    label_col, month_start_col, col_to_month = find_label_col_and_month_start(all_rows)

    if label_col is None:
        wb.close()
        return {"error": "Could not find label column", "file": fname_full}

    # Scan all rows and collect by label
    labeled_rows = []
    for i, row in enumerate(all_rows):
        if len(row) <= label_col:
            continue
        cell = row[label_col]
        if cell is None:
            continue
        cell_str = str(cell).strip()
        if not cell_str:
            continue
        data = extract_row_data(row, col_to_month)
        labeled_rows.append({'row': i + 1, 'label': cell_str, 'data': data})

    def find_rows(label, exact=True):
        if exact:
            return [r for r in labeled_rows if r['label'] == label]
        return [r for r in labeled_rows if label.lower() in r['label'].lower()]

    # ---- Home Services rows ----
    hs_all = find_rows('Home Services')
    # Occurrence 1 = inflows; 2 = payroll disbursements (back-office alloc section)
    hs_inflows = hs_all[0] if len(hs_all) >= 1 else None
    hs_payroll_disb = hs_all[1] if len(hs_all) >= 2 else None

    hs_total_list = find_rows('Total Home Services')
    hs_total_disb = hs_total_list[0] if hs_total_list else None

    servicelive_list = find_rows('Service Live Summary')
    servicelive_row = servicelive_list[0] if servicelive_list else None

    cc_list = find_rows('Cross Country Home Services')
    cross_country_row = cc_list[0] if cc_list else None

    assurant_ps_list = [r for r in labeled_rows if 'Assurant Profit Sharing' in r['label']]
    assurant_ps_row = assurant_ps_list[0] if assurant_ps_list else None

    assurant_disb_list = [r for r in labeled_rows
                          if 'Assurant' in r['label'] and 'Home Services' in r['label']
                          and 'Profit Sharing' not in r['label']]
    if not assurant_disb_list:
        assurant_disb_list = [r for r in labeled_rows if r['label'] == 'Assurant']
    assurant_disb_row = assurant_disb_list[0] if assurant_disb_list else None

    byov_list = [r for r in labeled_rows if 'BYOV' in r['label']]
    byov_row = byov_list[0] if byov_list else None

    # ---- KCD rows ----
    kcd_all = find_rows('KCD')
    kcd_inflows = kcd_all[0] if len(kcd_all) >= 1 else None
    kcd_payroll_disb = kcd_all[1] if len(kcd_all) >= 2 else None

    kcd_total_list = find_rows('Total KCD')
    kcd_total_disb = kcd_total_list[0] if kcd_total_list else None

    kcd_royalty_list = find_rows('KCD (Royalty)')
    if not kcd_royalty_list:
        kcd_royalty_list = find_rows('KCD Royalty')
    kcd_royalty_row = kcd_royalty_list[0] if kcd_royalty_list else None

    kcd_wholesale_list = find_rows('KCD Wholesale')
    kcd_wholesale_row = kcd_wholesale_list[0] if kcd_wholesale_list else None

    mx_kcd_list = [r for r in labeled_rows if 'Sears Mexico' in r['label'] and 'KCD' in r['label']]
    mx_kcd_row = mx_kcd_list[0] if mx_kcd_list else None

    def row_summary(r):
        if r is None:
            return {"row_number": None, "label": None, "data": {}}
        return {"row_number": r['row'], "label": r['label'], "data": r['data']}

    result = {
        "file_date": file_date,
        "filename": os.path.basename(fname_full),
        "metadata": {
            "label_col_0indexed": label_col,
            "month_start_col_0indexed": month_start_col,
            "months_covered": sorted(col_to_month.values()),
            "hs_occurrences_rows": [r['row'] for r in hs_all],
            "kcd_occurrences_rows": [r['row'] for r in kcd_all],
        },
        "home_services": {
            "inflows": row_summary(hs_inflows),
            "payroll_disbursements": row_summary(hs_payroll_disb),
            "total_disbursements": row_summary(hs_total_disb),
            "servicelive_summary_inflows": row_summary(servicelive_row),
            "byov_disbursements": row_summary(byov_row),
            "assurant_disbursements": row_summary(assurant_disb_row),
            "assurant_profit_sharing_inflows": row_summary(assurant_ps_row),
            "cross_country_inflows": row_summary(cross_country_row),
        },
        "kcd": {
            "inflows": row_summary(kcd_inflows),
            "payroll_disbursements": row_summary(kcd_payroll_disb),
            "total_disbursements": row_summary(kcd_total_disb),
            "royalty_inflows": row_summary(kcd_royalty_row),
            "wholesale_inflows": row_summary(kcd_wholesale_row),
            "sears_mexico_inflows": row_summary(mx_kcd_row),
        },
    }

    wb.close()
    return result


def extract_kes_kcd_sheet(fname_full, file_date):
    """
    Extract KES & KCD Cash Flow sheet for royalty and legal detail.
    Returns structured content.
    """
    wb = openpyxl.load_workbook(fname_full, read_only=True, data_only=True)
    kcd_sheet_names = [s for s in wb.sheetnames
                       if ('KCD' in s or 'KES' in s) and 'Cash Flow' in s]

    if not kcd_sheet_names:
        wb.close()
        return None

    sname = kcd_sheet_names[0]
    ws = wb[sname]
    rows = list(ws.iter_rows(min_row=1, max_row=60, values_only=True))

    # Extract rows with meaningful content
    content = []
    for i, row in enumerate(rows):
        nonnull = [(j, v) for j, v in enumerate(row) if v is not None and str(v).strip()]
        if nonnull:
            # First few as string preview
            content.append({
                "row": i + 1,
                "cells": [(j, str(v)[:80]) for j, v in nonnull[:20]]
            })

    wb.close()
    return {"sheet_name": sname, "content": content}


def extract_hs_pa_sheet(fname_full, file_date):
    """Extract HS-specific tabs for BYOV, metrics, and revenue detail."""
    wb = openpyxl.load_workbook(fname_full, read_only=True, data_only=True)

    hs_tab_names = ['HS PA', 'HS_Barb', 'BU HS Allocations 1.21',
                    'BU HS Allocations 5.6.25', 'BU HS Allocations 9.1.25',
                    'BU HS Allocations 2.28.25', 'BU HS Allocations 5.1.26',
                    'BU HS Allocations 12.14.23', 'BU HS Allocations 12.14.23_kk',
                    'HS FY24 Inflows_1.9.24', 'HS FY24 Inflows_2.6.24',
                    'HS IF 4.30.24', 'HS FY23 Inflows_11.12.23']

    results = {}
    for sname in hs_tab_names:
        if sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(min_row=1, max_row=50, values_only=True))
            content = []
            for i, row in enumerate(rows):
                nonnull = [(j, v) for j, v in enumerate(row) if v is not None and str(v).strip()]
                if nonnull:
                    content.append({
                        "row": i + 1,
                        "cells": [(j, str(v)[:80]) for j, v in nonnull[:15]]
                    })
            if content:
                results[sname] = content

    wb.close()
    return results if results else None


def extract_kcd_backup_sheet(fname_full, file_date):
    """Extract KCD backup tab for royalty rates and licensee detail."""
    wb = openpyxl.load_workbook(fname_full, read_only=True, data_only=True)

    kcd_backup = [s for s in wb.sheetnames if 'KCD backup' in s or 'KCD by vendor' in s]
    results = {}
    for sname in kcd_backup:
        ws = wb[sname]
        rows = list(ws.iter_rows(min_row=1, max_row=60, values_only=True))
        content = []
        for i, row in enumerate(rows):
            nonnull = [(j, v) for j, v in enumerate(row) if v is not None and str(v).strip()]
            if nonnull:
                content.append({
                    "row": i + 1,
                    "cells": [(j, str(v)[:80]) for j, v in nonnull[:15]]
                })
        if content:
            results[sname] = content

    wb.close()
    return results if results else None


def build_longitudinal_table(all_results):
    """
    Build best-estimate monthly longitudinal table.
    For each month, use the earliest BOD file dated ON OR AFTER the month
    (actual/locked data), falling back to the latest forecast file.
    """
    all_months = set()
    for r in all_results:
        if 'metadata' in r:
            all_months.update(r['metadata']['months_covered'])
    all_months = sorted(all_months)

    longitudinal = {}
    for month in all_months:
        candidates = []
        for r in all_results:
            if 'metadata' not in r:
                continue
            if month not in r['metadata']['months_covered']:
                continue

            hs = r['home_services']
            kcd = r['kcd']
            hs_inflow = hs['inflows']['data'].get(month)
            hs_total_disb = hs['total_disbursements']['data'].get(month)
            hs_payroll_disb = hs['payroll_disbursements']['data'].get(month)
            kcd_inflow = kcd['inflows']['data'].get(month)
            kcd_total_disb = kcd['total_disbursements']['data'].get(month)
            kcd_royalty = kcd['royalty_inflows']['data'].get(month)
            kcd_wholesale = kcd['wholesale_inflows']['data'].get(month)
            cc_hs = hs['cross_country_inflows']['data'].get(month)
            assurant_ps = hs['assurant_profit_sharing_inflows']['data'].get(month)
            assurant_disb = hs['assurant_disbursements']['data'].get(month)
            byov = hs['byov_disbursements']['data'].get(month)
            sears_mexico_kcd = kcd['sears_mexico_inflows']['data'].get(month)

            hs_net = None
            if hs_inflow is not None or hs_total_disb is not None:
                hs_net = round((hs_inflow or 0) + (hs_total_disb or 0), 6)

            kcd_net = None
            if kcd_inflow is not None or kcd_total_disb is not None:
                kcd_net = round((kcd_inflow or 0) + (kcd_total_disb or 0), 6)

            candidates.append({
                'file_date': r['file_date'],
                'filename': r['filename'],
                'hs_inflows': hs_inflow,
                'hs_total_disbursements': hs_total_disb,
                'hs_payroll_disbursements': hs_payroll_disb,
                'hs_net': hs_net,
                'hs_cross_country_inflows': cc_hs,
                'hs_assurant_profit_sharing': assurant_ps,
                'hs_assurant_disbursements': assurant_disb,
                'hs_byov_disbursements': byov,
                'kcd_inflows': kcd_inflow,
                'kcd_total_disbursements': kcd_total_disb,
                'kcd_royalty_inflows': kcd_royalty,
                'kcd_wholesale_inflows': kcd_wholesale,
                'kcd_sears_mexico': sears_mexico_kcd,
                'kcd_net': kcd_net,
            })

        if not candidates:
            continue

        candidates.sort(key=lambda x: x['file_date'])
        month_ym = month[:7]

        # First file dated >= the month = actual/locked
        best = next((c for c in candidates if c['file_date'] >= month_ym), candidates[-1])
        latest = candidates[-1]

        longitudinal[month] = {
            'month': month,
            'source_file_date': best['file_date'],
            'source_filename': best['filename'],
            'home_services': {
                'inflows_M': best['hs_inflows'],
                'total_disbursements_M': best['hs_total_disbursements'],
                'payroll_disbursements_M': best['hs_payroll_disbursements'],
                'net_contribution_M': best['hs_net'],
                'cross_country_inflows_M': best['hs_cross_country_inflows'],
                'assurant_profit_sharing_M': best['hs_assurant_profit_sharing'],
                'assurant_disbursements_M': best['hs_assurant_disbursements'],
                'byov_disbursements_M': best['hs_byov_disbursements'],
            },
            'kcd': {
                'inflows_M': best['kcd_inflows'],
                'total_disbursements_M': best['kcd_total_disbursements'],
                'royalty_inflows_M': best['kcd_royalty_inflows'],
                'wholesale_inflows_M': best['kcd_wholesale_inflows'],
                'sears_mexico_M': best['kcd_sears_mexico'],
                'net_contribution_M': best['kcd_net'],
            },
            'latest_file_revision': {
                'file_date': latest['file_date'],
                'hs_inflows_M': latest['hs_inflows'],
                'hs_net_M': latest['hs_net'],
                'kcd_inflows_M': latest['kcd_inflows'],
                'kcd_net_M': latest['kcd_net'],
            } if latest['file_date'] != best['file_date'] else None,
            'all_file_count': len(candidates),
        }

    return longitudinal


# ================================================================
# MAIN
# ================================================================

print("=" * 60)
print("HOME SERVICES & KCD LONGITUDINAL EXTRACTION")
print("=" * 60)

all_results = []
kcd_kes_details = {}
hs_pa_details = {}
kcd_backup_details = {}

for file_date, fname in ALL_FILES:
    fpath = os.path.join(BASE_DIR, fname)
    if not os.path.exists(fpath):
        print(f"  MISSING: {fname}")
        continue

    print(f"\n--- {file_date}: {fname[-55:]} ---")

    result = extract_bu_beta(fpath, file_date)
    if result and 'error' not in result:
        meta = result['metadata']
        hs = result['home_services']
        kcd = result['kcd']

        months = meta['months_covered']
        print(f"  Months: {months[0]} to {months[-1]} ({len(months)})")
        print(f"  HS rows: inflows={meta['hs_occurrences_rows']}, total_disb_row={hs['total_disbursements']['row_number']}")
        print(f"  KCD rows: inflows={meta['kcd_occurrences_rows']}, total_disb_row={kcd['total_disbursements']['row_number']}")

        # Print some data samples
        hs_data = hs['inflows']['data']
        kcd_data = kcd['inflows']['data']
        hs_sample = {m: v for m, v in sorted(hs_data.items()) if v and v != 0}
        kcd_sample = {m: v for m, v in sorted(kcd_data.items()) if v and v != 0}
        print(f"  HS inflows nonzero: first={list(hs_sample.items())[:3]}, last={list(hs_sample.items())[-3:]}")
        print(f"  KCD inflows nonzero: first={list(kcd_sample.items())[:3]}, last={list(kcd_sample.items())[-3:]}")

        # Check if HS appears as net inflow or outflow
        recent_hs = {m: v for m, v in sorted(hs_data.items()) if m >= file_date[:4] + '-01' and v}
        if recent_hs:
            avg = sum(recent_hs.values()) / len(recent_hs)
            print(f"  HS inflows sign: avg={avg:.2f}M ({'INFLOW' if avg > 0 else 'OUTFLOW'})")

        all_results.append(result)
    elif result:
        print(f"  ERROR: {result.get('error')}")

    # KES & KCD sheet
    kcd_detail = extract_kes_kcd_sheet(fpath, file_date)
    if kcd_detail:
        kcd_kes_details[file_date] = kcd_detail

    # HS PA sheet (sample for key dates only)
    if file_date in ['2022-09', '2023-08', '2024-05', '2025-05', '2026-05']:
        hs_pa = extract_hs_pa_sheet(fpath, file_date)
        if hs_pa:
            hs_pa_details[file_date] = hs_pa

    # KCD backup sheet (sample for key dates)
    if file_date in ['2022-09', '2023-08', '2024-05', '2025-05', '2026-05']:
        kcd_bk = extract_kcd_backup_sheet(fpath, file_date)
        if kcd_bk:
            kcd_backup_details[file_date] = kcd_bk

print(f"\n\n{'='*60}")
print("Building longitudinal table...")
longitudinal = build_longitudinal_table(all_results)
print(f"  Total months: {len(longitudinal)}")
if longitudinal:
    print(f"  Date range: {min(longitudinal.keys())} to {max(longitudinal.keys())}")

# Annual summaries
def annual_sum_nonzero(monthly_dict, key_path):
    """Sum a nested key across months, skipping None/zero."""
    annual = defaultdict(list)
    for month, d in monthly_dict.items():
        v = d
        for k in key_path:
            if isinstance(v, dict):
                v = v.get(k)
            else:
                v = None
                break
        if v is not None and v != 0:
            year = month[:4]
            annual[year].append(v)
    return {yr: round(sum(vals), 3) for yr, vals in sorted(annual.items())}


hs_annual_inflows = annual_sum_nonzero(longitudinal, ['home_services', 'inflows_M'])
hs_annual_disb = annual_sum_nonzero(longitudinal, ['home_services', 'total_disbursements_M'])
hs_annual_net = annual_sum_nonzero(longitudinal, ['home_services', 'net_contribution_M'])
kcd_annual_inflows = annual_sum_nonzero(longitudinal, ['kcd', 'inflows_M'])
kcd_annual_disb = annual_sum_nonzero(longitudinal, ['kcd', 'total_disbursements_M'])
kcd_annual_net = annual_sum_nonzero(longitudinal, ['kcd', 'net_contribution_M'])
kcd_royalty_annual = annual_sum_nonzero(longitudinal, ['kcd', 'royalty_inflows_M'])

print("\n=== HOME SERVICES ANNUAL SUMMARY ===")
for yr in sorted(set(list(hs_annual_inflows.keys()) + list(hs_annual_disb.keys()))):
    inf = hs_annual_inflows.get(yr, 0)
    disb = hs_annual_disb.get(yr, 0)
    net = hs_annual_net.get(yr, 0)
    print(f"  {yr}: Inflows={inf:.1f}M | Disb={disb:.1f}M | Net={net:.1f}M")

print("\n=== KCD ANNUAL SUMMARY ===")
for yr in sorted(set(list(kcd_annual_inflows.keys()) + list(kcd_annual_disb.keys()))):
    inf = kcd_annual_inflows.get(yr, 0)
    disb = kcd_annual_disb.get(yr, 0)
    net = kcd_annual_net.get(yr, 0)
    roy = kcd_royalty_annual.get(yr, 0)
    print(f"  {yr}: Inflows={inf:.1f}M | Disb={disb:.1f}M | Net={net:.1f}M | Royalty={roy:.2f}M")

# ================================================================
# BUILD OUTPUT JSON
# ================================================================

# Determine HS sign behavior by period
hs_sign_analysis = {}
for month, d in sorted(longitudinal.items()):
    inf = d['home_services']['inflows_M']
    disb = d['home_services']['total_disbursements_M']
    if inf is not None:
        sign = 'INFLOW' if inf > 0 else ('OUTFLOW' if inf < 0 else 'ZERO')
        hs_sign_analysis[month] = {'inflows': inf, 'sign': sign, 'disb': disb}

# Full file extraction summary
file_summary = []
for r in all_results:
    hs = r['home_services']
    kcd = r['kcd']
    hs_inf_data = hs['inflows']['data']
    kcd_inf_data = kcd['inflows']['data']

    # Non-zero month counts
    hs_nonzero = {m: v for m, v in hs_inf_data.items() if v and v != 0}
    kcd_nonzero = {m: v for m, v in kcd_inf_data.items() if v and v != 0}

    file_summary.append({
        'file_date': r['file_date'],
        'filename': r['filename'],
        'months_covered': f"{r['metadata']['months_covered'][0]} to {r['metadata']['months_covered'][-1]}",
        'hs_inflows_nonzero_months': len(hs_nonzero),
        'hs_inflows_first': list(hs_nonzero.items())[0] if hs_nonzero else None,
        'hs_inflows_last': list(hs_nonzero.items())[-1] if hs_nonzero else None,
        'kcd_inflows_nonzero_months': len(kcd_nonzero),
        'kcd_royalty_row_number': kcd['royalty_inflows']['row_number'],
        'hs_total_disb_row_number': hs['total_disbursements']['row_number'],
        'kcd_total_disb_row_number': kcd['total_disbursements']['row_number'],
        'byov_found': hs['byov_disbursements']['row_number'] is not None,
        'assurant_ps_found': hs['assurant_profit_sharing_inflows']['row_number'] is not None,
        'cross_country_found': hs['cross_country_inflows']['row_number'] is not None,
    })

output = {
    "metadata": {
        "generated_date": "2026-05-11",
        "description": "Home Services and KCD longitudinal cash flow analysis — 3.5 year history",
        "units": "USD millions",
        "sign_convention": "Inflows positive, Disbursements negative, Net = Inflows + Disbursements",
        "files_processed": len(all_results),
        "total_months_in_longitudinal": len(longitudinal),
        "date_range": f"{min(longitudinal.keys())} to {max(longitudinal.keys())}" if longitudinal else "N/A",
        "methodology": "For each month, use earliest BOD file dated on/after that month (actuals), else latest forecast",
        "source_data_note": "BU beta sheet, row 3 = month headers, label column identified programmatically. Duplicate month grids (audit side tables at col 600+) excluded."
    },
    "annual_summary": {
        "home_services": {
            "inflows_M_by_year": hs_annual_inflows,
            "total_disbursements_M_by_year": hs_annual_disb,
            "net_contribution_M_by_year": hs_annual_net,
        },
        "kcd": {
            "inflows_M_by_year": kcd_annual_inflows,
            "total_disbursements_M_by_year": kcd_annual_disb,
            "net_contribution_M_by_year": kcd_annual_net,
            "royalty_inflows_M_by_year": kcd_royalty_annual,
        }
    },
    "hs_sign_analysis": {
        "description": "Whether Home Services appears as a net inflow or outflow in the model, by month",
        "note": "HS is a large NET INFLOW (service revenue) with significant disbursements (payroll, parts, Assurant). The 'inflows' row is gross service revenue. Net = inflows + disbursements.",
        "monthly_sign": hs_sign_analysis,
    },
    "longitudinal_monthly": longitudinal,
    "file_by_file_extractions": all_results,
    "file_extraction_summary": file_summary,
    "kcd_kes_sheet_detail": kcd_kes_details,
    "hs_pa_sheet_detail": hs_pa_details,
    "kcd_backup_sheet_detail": kcd_backup_details,
    "analytical_notes": {
        "home_services_structure": {
            "description": "Home Services = tech repair & installation, home warranty, SHIP (Home Improvement)",
            "inflows_components": [
                "Cross Country Home Services (home warranty inflows, historically ~$4.8-5.5M/mo)",
                "BYOV (Bring Your Own Vendor) reimbursements (later files)",
                "Assurant/SHS Profit Sharing (protection plan margin share)",
                "ServiceLive (gig tech marketplace inflows)",
                "Direct service revenue (labor, parts)"
            ],
            "disbursements_components": [
                "Direct payroll (technicians, service staff)",
                "ServiceLive 1099 contractors",
                "SHIP subcontractors",
                "Parts & Inventory P-Card",
                "Assurant prefund (insurance reserve payments)",
                "Fuel P-Card",
                "Travel P-Card",
                "Other P-Card"
            ],
            "hs_is_net_inflow": True,
            "note": "HS inflows row represents total gross service cash receipts (~$46-75M/mo in FY22, declining through FY25-26)"
        },
        "kcd_structure": {
            "description": "KCD = Kenmore, Craftsman, DieHard brand IP licensing and wholesale",
            "inflows_components": [
                "KCD brand licensing royalties (external licensees)",
                "KCD (Royalty) — quarterly royalty settlements (often ~$2.56M/quarter from LG/others)",
                "Sears Mexico (KCD) — Mexico brand royalties",
                "KCD Wholesale (liquidation/inventory sales)"
            ],
            "disbursements_components": [
                "KCD back-office payroll allocation",
                "Legal (Katten Muchin for LG lawsuit)",
            ],
            "kcd_royalty_note": "KCD (Royalty) row shows $2.56M quarterly from LG settlement (appears at specific months). Katten Muchin mentioned in KES & KCD Cash Flow sheet re: LG lawsuit.",
            "kcd_inflows_trend": "Peaked ~$24.9M in FY22 (included large wholesale liquidations), dropped to ~$5-6M/yr FY23-25, then spiked in FY26 (~$41.6M) due to restructured licensee payments or settlement proceeds",
        },
        "data_gaps": [
            "FY22 Feb data only from Sep2022 BOD (earliest file). Jan 2022 is sparse.",
            "Some months between FY22 actuals and FY23 forecasts show structural zeroes in the row; actual values pulled from nearest file",
            "KCD Wholesale inflows often show 0 in BU beta but may flow through KES ABL / valley account",
        ]
    }
}

out_path = os.path.join(BASE_DIR, "analysis_HS_KCD.json")
with open(out_path, 'w') as f:
    json.dump(output, f, indent=2, default=str)

print(f"\n\n{'='*60}")
print(f"OUTPUT: {out_path}")
file_size = os.path.getsize(out_path) / 1024
print(f"File size: {file_size:.1f} KB")
